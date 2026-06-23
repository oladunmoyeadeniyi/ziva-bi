"""
ZivaBI — M8 master data configuration router.

Registered at prefix /api/config.

Endpoints:
  Dimensions:
    GET    /api/config/dimensions                              List tenant dimensions
    GET    /api/config/dimensions/org-structure-preview        Cost center nodes for add form
    POST   /api/config/dimensions                              Create dimension
    GET    /api/config/dimensions/{id}/inline-values           Merged values from all sources
    PATCH  /api/config/dimensions/{id}                         Update dimension
    DELETE /api/config/dimensions/{id}                         Soft delete (is_active=false)
    PATCH  /api/config/dimensions/{id}/reactivate              Reactivate a soft-deleted dimension
    DELETE /api/config/dimensions/{id}/permanent               Hard delete a custom dimension + its values
    POST   /api/config/dimensions/{id}/reorder                 Update sort_order

  Dimension Values:
    GET    /api/config/dimensions/{id}/values          List values for a dimension
    POST   /api/config/dimensions/{id}/values          Add single value
    PATCH  /api/config/dimensions/{id}/values/{vid}    Update value
    DELETE /api/config/dimensions/{id}/values/{vid}    Soft delete
    POST   /api/config/dimensions/{id}/values/upload   Bulk upload (xlsx/csv)

  Chart of Accounts:
    GET    /api/config/coa                             List GL accounts (paginated, searchable)
    POST   /api/config/coa                             Create single GL account
    PATCH  /api/config/coa/{id}                        Update GL account
    DELETE /api/config/coa/{id}                        Soft delete
    GET    /api/config/coa/template                    Download dynamically generated xlsx template
    POST   /api/config/coa/upload                      Bulk upload (xlsx/csv)
    PATCH  /api/config/coa/{id}/dimensions             Set dimension requirements for a GL account

  Expense Categories:
    GET    /api/config/categories                      List full category tree with GL mappings
    POST   /api/config/categories                      Create category or subcategory
    PATCH  /api/config/categories/{id}                 Update
    DELETE /api/config/categories/{id}                 Soft delete (cascades to subcategories)
    POST   /api/config/categories/{id}/gl-mappings     Add GL account to subcategory
    DELETE /api/config/categories/{id}/gl-mappings/{gl_id}  Remove GL mapping
    PATCH  /api/config/categories/{id}/gl-mappings/{gl_id}  Set/unset as default

  GL Search (M9):
    GET    /api/config/gl/search?q=...&limit=20        Search GL accounts with dimension requirements

All endpoints are tenant-scoped and require authentication.
Admin-only operations require is_tenant_admin or is_super_admin.
"""

import csv
import io
import logging
import re
import traceback
import uuid
from typing import Optional

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.middleware.auth import CurrentUser, require_auth, block_if_readonly_impersonation
from app.models.expenses import ExpenseCategory
from app.models.master_data import (
    CategoryGLMapping,
    ChartOfAccount,
    DimensionValue,
    Employee,
    GlCodeRemap,
    GLDimensionRequirement,
    TenantDimension,
)
from app.models.setup import OrgStructureNode, TenantOrgConfig
from app.schemas.config import (
    BulkActionRequest,
    BulkActionResult,
    CategoryCreate,
    CategoryGLMappingCreate,
    CategoryGLMappingResponse,
    CategoryUpdate,
    CategoryWithMappingsResponse,
    CoACreate,
    CoADimensionsUpdate,
    CoAListItem,
    CoAResponse,
    CoAUpdate,
    DimensionCreate,
    DimensionReorder,
    DimensionResponse,
    DimensionUpdate,
    DimensionValueCreate,
    DimensionValueResponse,
    DimensionValueUpdate,
    BulkRemapResult,
    BulkRemapRow,
    GlRemapHistoryEntry,
    GLSearchResult,
    InlineNewAccountFields,
    RemapRequest,
    RemapResult,
    RemapResultEntry,
    UploadResult,
    _generate_code,
)

router = APIRouter(prefix="/api/config", tags=["config"])


# ── Shared helpers ────────────────────────────────────────────────────────────

def _require_tenant(current_user: CurrentUser) -> uuid.UUID:
    if current_user.tenant_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="This feature is only available on business accounts.",
        )
    return current_user.tenant_id


def _require_admin(current_user: CurrentUser) -> None:
    if not current_user.is_tenant_admin and not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Tenant Admins can modify configuration.",
        )
    block_if_readonly_impersonation(current_user)


async def _parse_upload(file: UploadFile) -> tuple[list[str], list[list[str]]]:
    """
    Parse an uploaded CSV or XLSX file into (headers, data_rows).

    Returns headers as a list of strings and rows as lists of strings.
    Raises HTTPException for unsupported formats.
    """
    content = await file.read()
    fname = (file.filename or "").lower()

    if fname.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            return [], []
        return [h.strip() for h in all_rows[0]], [
            [cell.strip() for cell in row] for row in all_rows[1:]
        ]

    elif fname.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="openpyxl is not installed. Run: pip install openpyxl",
            ) from exc
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        if not rows:
            return [], []
        return rows[0], rows[1:]

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .csv and .xlsx files are supported.",
        )


# ═══════════════════════════════════════════════════════════════════════════════
# DIMENSIONS
# ═══════════════════════════════════════════════════════════════════════════════

# Standard dimensions seeded for every new tenant on first Dimensions page visit
STANDARD_DIMENSIONS = [
    {
        "name": "Cost center",
        "code": "cost_center",
        "is_required": True,
        "value_source": "org_structure",
        "dimension_sources": [{"source_type": "org_structure", "filter": None}],
        "description": "Tracks costs by organisational cost center.",
        "icon": "building-community",
        "sort_order": 1,
    },
    {
        "name": "Material / Product (SKU)",
        "code": "material",
        "is_required": False,
        "value_source": "product_master",
        "dimension_sources": [{"source_type": "product_master", "filter": None}],
        "description": "Tags transactions with product or SKU codes. Auto-syncs from product master when Inventory active.",
        "icon": "barcode",
        "sort_order": 2,
    },
    {
        "name": "Statistical internal order",
        "code": "statistical_internal_order",
        "is_required": False,
        "value_source": "hybrid",
        "dimension_sources": [{"source_type": "employee_master", "filter": None}],
        "description": "Employee codes auto-synced + manual codes for campaigns, vehicles, assets.",
        "icon": "git-branch",
        "sort_order": 3,
    },
    {
        "name": "Real internal order",
        "code": "real_internal_order",
        "is_required": False,
        "value_source": "manual",
        "dimension_sources": [],
        "description": "Tracks actual costs by project or initiative.",
        "icon": "git-commit",
        "sort_order": 4,
    },
    {
        "name": "Customer order",
        "code": "customer_order",
        "is_required": False,
        "value_source": "customer_order",
        "dimension_sources": [{"source_type": "customer_master", "filter": None}],
        "description": "Customer categories/segments. Manual now, auto-syncs from customer master when AR active.",
        "icon": "users-group",
        "sort_order": 5,
    },
    {
        "name": "Trading partner",
        "code": "trading_partner",
        "is_required": False,
        "value_source": "manual",
        "dimension_sources": [{"source_type": "group_structure", "filter": None}],
        "description": "Intercompany entities — subsidiaries, branches, related parties. Manual now, auto-syncs from group structure when Intercompany module is active.",
        "icon": "building-bridge",
        "sort_order": 6,
    },
]

# Correct value_source + dimension_sources for dimensions created before these columns existed
_SOURCE_FIXES = {
    "cost_center": ("org_structure", [{"source_type": "org_structure", "filter": None}]),
    "material": ("product_master", [{"source_type": "product_master", "filter": None}]),
    "statistical_internal_order": ("hybrid", [{"source_type": "employee_master", "filter": None}]),
    "real_internal_order": ("manual", []),
    "customer_order": ("customer_order", [{"source_type": "customer_master", "filter": None}]),
    "trading_partner": ("manual", [{"source_type": "group_structure", "filter": None}]),
}

# Codes that identify standard (non-custom) dimensions — cannot be permanently deleted
STANDARD_DIMENSION_CODES = {
    "cost_center", "material", "statistical_order", "statistical_internal_order",
    "real_order", "real_internal_order", "customer_order", "employee",
    "brand", "region", "channel", "project", "trading_partner",
}


@router.post("/dimensions/seed-standard", status_code=200)
async def seed_standard_dimensions(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Seed the 5 standard dimensions for a tenant if they don't exist yet.

    Also fixes value_source for dimensions created before the value_source column
    was added (which defaulted to 'manual'). Idempotent — safe to call on every
    page load.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    seeded = 0
    fixed = 0

    existing_result = await db.execute(
        select(TenantDimension).where(TenantDimension.tenant_id == tenant_id)
    )
    existing_dims = existing_result.scalars().all()
    existing_codes = {d.code for d in existing_dims}

    # Fix existing dimensions with wrong value_source or missing dimension_sources
    for dim in existing_dims:
        fix = _SOURCE_FIXES.get(dim.code)
        if fix:
            correct_source, correct_sources = fix
            changed = False
            if dim.value_source != correct_source:
                dim.value_source = correct_source
                changed = True
            if not dim.dimension_sources:
                dim.dimension_sources = correct_sources
                changed = True
            if changed:
                fixed += 1

    # Seed missing standard dimensions
    for std in STANDARD_DIMENSIONS:
        if std["code"] not in existing_codes:
            db.add(TenantDimension(
                tenant_id=tenant_id,
                name=std["name"],
                code=std["code"],
                is_required=std["is_required"],
                value_source=std["value_source"],
                dimension_sources=std["dimension_sources"],
                description=std["description"],
                icon=std["icon"],
                sort_order=std["sort_order"],
                is_active=True,
            ))
            seeded += 1

    await db.commit()
    return {"seeded": seeded, "fixed": fixed}


@router.get("/dimensions/org-structure-preview")
async def get_org_structure_preview(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """
    Return cost center nodes from the org structure for the dimension add form checklist.

    Used by the frontend when the user picks 'Org structure' as the value source for
    a new dimension, so they can see and optionally exclude specific cost centers.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(OrgStructureNode)
        .where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Cost center",
            OrgStructureNode.is_active.is_(True),
        )
        .order_by(OrgStructureNode.sort_order)
    )
    nodes = result.scalars().all()
    return [
        {
            "id": str(n.id),
            "name": n.name,
            "code": n.code,
            "cost_center_code": n.cost_center_code,
            "parent_id": str(n.parent_id) if n.parent_id else None,
        }
        for n in nodes
    ]


@router.get("/dimensions", response_model=list[DimensionResponse])
async def list_dimensions(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[DimensionResponse]:
    """List all dimensions for the tenant, ordered by sort_order then name."""
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(TenantDimension)
        .where(TenantDimension.tenant_id == tenant_id)
        .order_by(TenantDimension.sort_order, TenantDimension.name)
    )
    return [DimensionResponse.from_orm(d) for d in result.scalars().all()]


@router.post("/dimensions", response_model=DimensionResponse, status_code=status.HTTP_201_CREATED)
async def create_dimension(
    data: DimensionCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionResponse:
    """Create a new dimension for the tenant. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    code = data.code or _generate_code(data.name)

    # Check uniqueness
    existing = await db.execute(
        select(TenantDimension).where(
            TenantDimension.tenant_id == tenant_id,
            TenantDimension.code == code,
            TenantDimension.is_active == True,  # noqa: E712
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"A dimension with code '{code}' already exists.",
        )

    dim = TenantDimension(
        tenant_id=tenant_id,
        name=data.name,
        code=code,
        is_required=data.is_required,
        value_source=data.value_source,
        dimension_sources=data.dimension_sources,
        display_name=data.display_name,
        description=data.description,
        icon=data.icon,
    )
    db.add(dim)
    await db.flush()
    await db.refresh(dim)
    return DimensionResponse.from_orm(dim)


@router.patch("/dimensions/{dimension_id}", response_model=DimensionResponse)
async def update_dimension(
    dimension_id: uuid.UUID,
    data: DimensionUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionResponse:
    """Update a dimension (PATCH semantics). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dimension not found.")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(dim, field, value)

    if data.dimension_sources is not None:
        dim.dimension_sources = data.dimension_sources

    await db.flush()
    await db.refresh(dim)
    return DimensionResponse.from_orm(dim)


@router.delete("/dimensions/{dimension_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_dimension(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete a dimension (set is_active=false). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dimension not found.")

    dim.is_active = False
    await db.flush()


@router.patch("/dimensions/{dimension_id}/reactivate", response_model=DimensionResponse)
async def reactivate_dimension(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionResponse:
    """Reactivate a soft-deleted dimension. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dimension not found."
        )
    dim.is_active = True
    await db.flush()
    await db.refresh(dim)
    return DimensionResponse.from_orm(dim)


@router.delete("/dimensions/{dimension_id}/permanent", status_code=status.HTTP_204_NO_CONTENT)
async def hard_delete_dimension(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """
    Permanently delete a custom dimension and all its values.
    Standard dimensions cannot be permanently deleted — use soft delete instead.
    Admin only.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dimension not found."
        )


    # Delete all values first (cascade would handle this but being explicit)
    await db.execute(
        delete(DimensionValue).where(
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    await db.delete(dim)
    await db.flush()


@router.post("/dimensions/{dimension_id}/reorder", response_model=DimensionResponse)
async def reorder_dimension(
    dimension_id: uuid.UUID,
    data: DimensionReorder,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionResponse:
    """Update the sort_order of a dimension. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dimension not found.")

    dim.sort_order = data.sort_order
    await db.flush()
    await db.refresh(dim)
    return DimensionResponse.from_orm(dim)


# ═══════════════════════════════════════════════════════════════════════════════
# DIMENSION VALUES
# ═══════════════════════════════════════════════════════════════════════════════

async def _get_dimension_or_404(
    dimension_id: uuid.UUID, tenant_id: uuid.UUID, db: AsyncSession
) -> TenantDimension:
    result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.id == dimension_id,
            TenantDimension.tenant_id == tenant_id,
        )
    )
    dim = result.scalar_one_or_none()
    if not dim:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Dimension not found.")
    return dim


@router.get("/dimensions/{dimension_id}/inline-values")
async def get_dimension_inline_values(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """
    Return all values for a dimension in one call, merging auto-synced and manual sources.

    Used by the Master data / values tab to show inline values without jumping to source.
    Returns a flat list with source labels for each entry.
    """
    tenant_id = _require_tenant(current_user)
    dim = await _get_dimension_or_404(dimension_id, tenant_id, db)
    results: list[dict] = []

    sources = dim.dimension_sources or []
    source_types = [s.get("source_type") for s in sources]

    if "org_structure" in source_types or dim.value_source == "org_structure":
        nodes_result = await db.execute(
            select(OrgStructureNode).where(
                OrgStructureNode.tenant_id == tenant_id,
                OrgStructureNode.node_type == "Cost center",
                OrgStructureNode.is_active.is_(True),
            ).order_by(OrgStructureNode.sort_order)
        )
        for node in nodes_result.scalars().all():
            results.append({
                "id": str(node.id),
                "code": getattr(node, "cost_center_code", None) or getattr(node, "code", "") or "",
                "name": node.name,
                "source": "org_structure",
                "editable": False,
                "parent_id": str(node.parent_id) if node.parent_id else None,
                "node_id": str(node.id),
            })

    if "employee_master" in source_types or dim.value_source in ("employee_master", "hybrid"):
        emp_result = await db.execute(
            select(Employee).where(
                Employee.tenant_id == tenant_id,
                Employee.is_active.is_(True),
            ).order_by(Employee.last_name)
        )
        for emp in emp_result.scalars().all():
            results.append({
                "id": str(emp.id),
                "code": getattr(emp, "employee_code", None) or str(emp.id)[:8],
                "name": f"{emp.first_name or ''} {emp.last_name or ''}".strip(),
                "source": "employee_master",
                "editable": False,
            })

    vals_result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
            DimensionValue.is_active.is_(True),
        ).order_by(DimensionValue.code)
    )
    for val in vals_result.scalars().all():
        results.append({
            "id": str(val.id),
            "code": val.code,
            "name": val.name,
            "source": "manual",
            "editable": True,
        })

    return results


@router.get("/dimensions/{dimension_id}/values", response_model=list[DimensionValueResponse])
async def list_dimension_values(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[DimensionValueResponse]:
    """List all values for a dimension, ordered by sort_order then code."""
    tenant_id = _require_tenant(current_user)
    await _get_dimension_or_404(dimension_id, tenant_id, db)

    result = await db.execute(
        select(DimensionValue)
        .where(
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
        .order_by(DimensionValue.sort_order, DimensionValue.code)
    )
    return [DimensionValueResponse.from_orm(v) for v in result.scalars().all()]


@router.post(
    "/dimensions/{dimension_id}/values",
    response_model=DimensionValueResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_dimension_value(
    dimension_id: uuid.UUID,
    data: DimensionValueCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionValueResponse:
    """Add a single value to a dimension. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    dim = await _get_dimension_or_404(dimension_id, tenant_id, db)

    existing = await db.execute(
        select(DimensionValue).where(
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
            DimensionValue.code == data.code,
            DimensionValue.is_active == True,  # noqa: E712
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"A value with code '{data.code}' already exists for this dimension.",
        )

    val = DimensionValue(
        tenant_id=tenant_id,
        dimension_id=dim.id,
        code=data.code,
        name=data.name,
        sort_order=data.sort_order,
        value_type=data.value_type,
        cascade_dimension_id=data.cascade_dimension_id,
        cascade_value_id=data.cascade_value_id,
        valid_from=data.valid_from,
        valid_to=data.valid_to,
    )
    db.add(val)
    await db.flush()
    await db.refresh(val)
    return DimensionValueResponse.from_orm(val)


@router.delete(
    "/dimensions/{dimension_id}/values/{value_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_dimension_value(
    dimension_id: uuid.UUID,
    value_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete a dimension value. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.id == value_id,
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    val = result.scalar_one_or_none()
    if not val:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Value not found.")

    await db.delete(val)
    await db.flush()


@router.patch(
    "/dimensions/{dimension_id}/values/{value_id}/toggle",
    response_model=DimensionValueResponse,
)
async def toggle_dimension_value(
    dimension_id: uuid.UUID,
    value_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DimensionValueResponse:
    """Toggle is_active on a single dimension value. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.id == value_id,
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    val = result.scalar_one_or_none()
    if not val:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Value not found.")
    val.is_active = not val.is_active
    await db.flush()
    return DimensionValueResponse.from_orm(val)


@router.patch("/dimensions/{dimension_id}/values/{value_id}")
async def update_dimension_value(
    dimension_id: uuid.UUID,
    value_id: uuid.UUID,
    payload: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Update name, description, valid_from, valid_to, or is_active on a single dimension value. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.id == value_id,
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    val = result.scalar_one_or_none()
    if not val:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Value not found.")

    if "name" in payload and payload["name"]:
        val.name = payload["name"].strip()
    if "description" in payload:
        val.description = payload["description"]
    if "is_active" in payload:
        val.is_active = bool(payload["is_active"])

    from datetime import datetime as _dt

    def parse_date(raw: str):
        if not raw or not raw.strip():
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return _dt.strptime(raw.strip(), fmt).date()
            except ValueError:
                continue
        raise HTTPException(
            status_code=400,
            detail=f"Invalid date format: '{raw}'. Use DD/MM/YYYY."
        )

    if "valid_from" in payload:
        val.valid_from = parse_date(payload["valid_from"]) if payload["valid_from"] else None
    if "valid_to" in payload:
        val.valid_to = parse_date(payload["valid_to"]) if payload["valid_to"] else None

    await db.flush()
    return {
        "id": str(val.id),
        "code": val.code,
        "name": val.name,
        "description": val.description,
        "is_active": val.is_active,
        "valid_from": val.valid_from.strftime("%d/%m/%Y") if val.valid_from else None,
        "valid_to": val.valid_to.strftime("%d/%m/%Y") if val.valid_to else None,
    }


@router.get("/dimensions/{dimension_id}/values/template")
async def download_dimension_values_template(
    dimension_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Download a blank XLSX template for bulk uploading dimension values.

    Columns: Code (required), Name (required), Description (optional).
    Includes a sample row in italic grey so users can see the expected format.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    await _get_dimension_or_404(dimension_id, tenant_id, db)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    try:
        # Fetch all active manual dimensions for dropdown
        dim_result = await db.execute(
            select(TenantDimension).where(
                TenantDimension.tenant_id == tenant_id,
                TenantDimension.is_active == True,
            )
        )
        all_dims = dim_result.scalars().all()
        manual_dims = [d for d in all_dims if (d.value_source or "manual") in ("manual", "hybrid")]
        dim_names = [d.display_name or d.name for d in manual_dims]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Values"

        hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_align = Alignment(horizontal="center")

        headers = ["Dimension", "Code *", "Name *", "Description", "Valid From (dd/mm/yyyy)", "Valid To (dd/mm/yyyy)", "Is Active"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 12

        instr_font = Font(name="Arial", size=10, italic=True, color="555555")
        ws.cell(row=2, column=1, value="Select dimension from dropdown").font = instr_font
        ws.cell(row=2, column=2, value="Unique code e.g. NG_AEKHALAMA").font = instr_font
        ws.cell(row=2, column=3, value="Full name").font = instr_font
        ws.cell(row=2, column=4, value="Optional description").font = instr_font
        ws.cell(row=2, column=5, value="e.g. 01/01/2025 (optional)").font = instr_font
        ws.cell(row=2, column=6, value="e.g. 31/12/2025 (optional)").font = instr_font
        ws.cell(row=2, column=7, value="Yes or No (default: Yes)").font = instr_font

        example_font = Font(name="Arial", size=10, italic=True, color="888888")
        ws.cell(row=3, column=1, value=dim_names[0] if dim_names else "").font = example_font
        ws.cell(row=3, column=2, value="NG_AEKHALAMA").font = example_font
        ws.cell(row=3, column=3, value="Khalamanja").font = example_font
        ws.cell(row=3, column=4, value="Optional description").font = example_font
        ws.cell(row=3, column=5, value="01/01/2025").font = example_font
        ws.cell(row=3, column=6, value="").font = example_font
        ws.cell(row=3, column=7, value="Yes").font = example_font

        if dim_names:
            dim_formula = '"' + ",".join(dim_names) + '"'
            dv = DataValidation(
                type="list",
                formula1=dim_formula,
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid dimension",
                error="Please select a dimension from the dropdown list.",
            )
            dv.sqref = "A4:A1000"
            ws.add_data_validation(dv)

        dv_active = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
        dv_active.sqref = "G4:G1000"
        ws.add_data_validation(dv_active)

        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=dimension_values_template.xlsx"},
        )
    except Exception as exc:
        logger.error("download_dimension_values_template failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Template generation failed: {exc}") from exc


@router.post(
    "/dimensions/{dimension_id}/values/upload",
    response_model=UploadResult,
)
async def upload_dimension_values(
    dimension_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> UploadResult:
    """
    Bulk upload dimension values from .xlsx or .csv file.

    Expected columns: code (required), name (required), sort_order (optional).
    Skips duplicates by code (does not overwrite existing active values).
    Returns import summary: imported, skipped, errors.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    dim = await _get_dimension_or_404(dimension_id, tenant_id, db)

    headers, rows = await _parse_upload(file)
    headers_lower = [h.lower().replace("*", "").strip() for h in headers]

    def col(name: str) -> Optional[int]:
        try:
            return headers_lower.index(name)
        except ValueError:
            return None

    has_dimension_col = "dimension" in headers_lower
    dim_col_idx = col("dimension") if has_dimension_col else None
    _c = col("code"); code_col = _c if _c is not None else col("value code")
    _n = col("name"); name_col = _n if _n is not None else col("value name")
    order_col = col("sort_order")
    type_col = col("value type")
    from_col = col("valid from (dd/mm/yyyy)") or col("valid from")
    to_col = col("valid to (dd/mm/yyyy)") or col("valid to")
    active_col = col("is active")

    if code_col is None or name_col is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must have 'code' (or 'Value Code') and 'name' (or 'Value Name') columns.",
        )

    # Build dimension name lookup if Dimension column is present
    dim_by_name: dict[str, TenantDimension] = {}
    if has_dimension_col:
        all_dims_result = await db.execute(
            select(TenantDimension).where(
                TenantDimension.tenant_id == tenant_id,
                TenantDimension.is_active == True,
            )
        )
        for d in all_dims_result.scalars().all():
            key_display = (d.display_name or "").strip().lower()
            key_name = (d.name or "").strip().lower()
            if key_display:
                dim_by_name[key_display] = d
            if key_name:
                dim_by_name[key_name] = d

    from datetime import datetime as _dt2

    imported = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):
        if not any((cell or "").strip() for cell in row):
            continue

        def get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        # Determine target dimension for this row
        if has_dimension_col:
            dim_name_val = get(dim_col_idx)
            if not dim_name_val:
                skipped += 1
                continue
            target_dim = dim_by_name.get(dim_name_val.lower())
            if target_dim is None:
                errors.append({"row": i, "reason": f"Unknown dimension: {dim_name_val}"})
                continue
        else:
            target_dim = dim

        code = get(code_col)
        name = get(name_col)
        sort_str = get(order_col)
        value_type = get(type_col) or None
        valid_from_str = get(from_col)
        valid_to_str = get(to_col)
        is_active_str = get(active_col).lower()

        if not code:
            errors.append({"row": i, "reason": "Missing code."})
            continue
        if not name:
            errors.append({"row": i, "reason": "Missing name."})
            continue

        sort_order = 0
        if sort_str:
            try:
                sort_order = int(float(sort_str))
            except ValueError:
                errors.append({"row": i, "reason": f"Invalid sort_order: '{sort_str}'."})
                continue

        is_active = is_active_str not in ("no", "false", "0")

        valid_from = None
        valid_to = None
        for date_str, dest in [(valid_from_str, "Valid From"), (valid_to_str, "Valid To")]:
            if date_str:
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        parsed = _dt2.strptime(date_str, fmt).date()
                        if dest == "Valid From":
                            valid_from = parsed
                        else:
                            valid_to = parsed
                        break
                    except ValueError:
                        continue
                else:
                    errors.append({"row": i, "reason": f"Invalid date in {dest}: '{date_str}'."})

        existing = await db.execute(
            select(DimensionValue).where(
                DimensionValue.dimension_id == target_dim.id,
                DimensionValue.tenant_id == tenant_id,
                DimensionValue.code == code,
            )
        )
        dv_obj = existing.scalar_one_or_none()
        if dv_obj:
            dv_obj.name = name
            dv_obj.is_active = is_active
            dv_obj.sort_order = sort_order
            if value_type is not None:
                dv_obj.value_type = value_type
            if valid_from is not None:
                dv_obj.valid_from = valid_from
            if valid_to is not None:
                dv_obj.valid_to = valid_to
            updated += 1
        else:
            db.add(DimensionValue(
                tenant_id=tenant_id,
                dimension_id=target_dim.id,
                code=code,
                name=name,
                sort_order=sort_order,
                is_active=is_active,
                value_type=value_type,
                valid_from=valid_from,
                valid_to=valid_to,
            ))
            imported += 1

    await db.flush()
    return UploadResult(imported=imported, updated=updated, skipped=skipped, errors=errors)


@router.get("/dimensions/template/universal")
async def download_universal_dimension_values_template(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Download a universal XLSX template for bulk uploading values across ALL manual dimensions.

    Includes a Dimension column (first) with a dropdown of all active manual dimensions.
    Instruction row in row 2, example row in row 3. Data starts at row 4.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    try:
        dim_result = await db.execute(
            select(TenantDimension).where(
                TenantDimension.tenant_id == tenant_id,
                TenantDimension.is_active == True,
            )
        )
        all_dims = dim_result.scalars().all()
        manual_dims = [d for d in all_dims if (d.value_source or "manual") in ("manual", "hybrid")]
        dim_names = [d.display_name or d.name for d in manual_dims]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Values"

        hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_align = Alignment(horizontal="center")

        headers = ["Dimension", "Code *", "Name *", "Description", "Valid From (dd/mm/yyyy)", "Valid To (dd/mm/yyyy)", "Is Active"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 12

        instr_font = Font(name="Arial", size=10, italic=True, color="555555")
        ws.cell(row=2, column=1, value="Select dimension from dropdown").font = instr_font
        ws.cell(row=2, column=2, value="Unique code e.g. NG_AEKHALAMA").font = instr_font
        ws.cell(row=2, column=3, value="Full name").font = instr_font
        ws.cell(row=2, column=4, value="Optional description").font = instr_font
        ws.cell(row=2, column=5, value="e.g. 01/01/2025 (optional)").font = instr_font
        ws.cell(row=2, column=6, value="e.g. 31/12/2025 (optional)").font = instr_font
        ws.cell(row=2, column=7, value="Yes or No (default: Yes)").font = instr_font

        example_font = Font(name="Arial", size=10, italic=True, color="888888")
        ws.cell(row=3, column=1, value=dim_names[0] if dim_names else "").font = example_font
        ws.cell(row=3, column=2, value="NG_AEKHALAMA").font = example_font
        ws.cell(row=3, column=3, value="Khalamanja").font = example_font
        ws.cell(row=3, column=4, value="Optional description").font = example_font
        ws.cell(row=3, column=5, value="01/01/2025").font = example_font
        ws.cell(row=3, column=6, value="").font = example_font
        ws.cell(row=3, column=7, value="Yes").font = example_font

        if dim_names:
            dim_formula = '"' + ",".join(dim_names) + '"'
            dv = DataValidation(
                type="list",
                formula1=dim_formula,
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=True,
                errorTitle="Invalid dimension",
                error="Please select a dimension from the dropdown list.",
            )
            dv.sqref = "A4:A1000"
            ws.add_data_validation(dv)

        dv_active = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
        dv_active.sqref = "G4:G1000"
        ws.add_data_validation(dv_active)

        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=dimension_values_universal_template.xlsx"},
        )
    except Exception as exc:
        logger.error("download_universal_dimension_values_template failed: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Template generation failed: {exc}") from exc


@router.post("/dimensions/upload/universal", response_model=UploadResult)
async def upload_universal_dimension_values(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> UploadResult:
    """
    Bulk upload dimension values for multiple dimensions from a single .xlsx file.

    The file must have a Dimension column (first) whose values match active dimension
    names or display names for the tenant. Each row is routed to its named dimension.
    Returns: { imported, updated, skipped, errors }.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    headers, rows = await _parse_upload(file)
    headers_lower = [h.lower().replace("*", "").strip() for h in headers]

    def col(name: str) -> Optional[int]:
        try:
            return headers_lower.index(name)
        except ValueError:
            return None

    dim_col_idx = col("dimension")
    _c = col("code"); code_col = _c if _c is not None else col("value code")
    _n = col("name"); name_col = _n if _n is not None else col("value name")
    order_col = col("sort_order")
    type_col = col("value type")
    from_col = col("valid from (dd/mm/yyyy)") or col("valid from")
    to_col = col("valid to (dd/mm/yyyy)") or col("valid to")
    active_col = col("is active")

    if dim_col_idx is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Universal upload requires a 'Dimension' column. Use the universal template.",
        )
    if code_col is None or name_col is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must have 'code' and 'name' columns.",
        )

    # Build dimension name → TenantDimension lookup
    all_dims_result = await db.execute(
        select(TenantDimension).where(
            TenantDimension.tenant_id == tenant_id,
            TenantDimension.is_active == True,
        )
    )
    dim_by_name: dict[str, TenantDimension] = {}
    for d in all_dims_result.scalars().all():
        if d.display_name:
            dim_by_name[d.display_name.strip().lower()] = d
        dim_by_name[d.name.strip().lower()] = d

    from datetime import datetime as _dt2

    imported = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):
        if not any((cell or "").strip() for cell in row):
            continue

        def get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        dim_name_val = get(dim_col_idx)
        if not dim_name_val:
            skipped += 1
            continue
        target_dim = dim_by_name.get(dim_name_val.lower())
        if target_dim is None:
            errors.append({"row": i, "reason": f"Unknown dimension: {dim_name_val}"})
            continue

        code = get(code_col)
        name = get(name_col)
        sort_str = get(order_col)
        value_type = get(type_col) or None
        valid_from_str = get(from_col)
        valid_to_str = get(to_col)
        is_active_str = get(active_col).lower()

        if not code:
            errors.append({"row": i, "reason": "Missing code."})
            continue
        if not name:
            errors.append({"row": i, "reason": "Missing name."})
            continue

        sort_order = 0
        if sort_str:
            try:
                sort_order = int(float(sort_str))
            except ValueError:
                errors.append({"row": i, "reason": f"Invalid sort_order: '{sort_str}'."})
                continue

        is_active = is_active_str not in ("no", "false", "0")

        valid_from = None
        valid_to = None
        for date_str, dest in [(valid_from_str, "Valid From"), (valid_to_str, "Valid To")]:
            if date_str:
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        parsed = _dt2.strptime(date_str, fmt).date()
                        if dest == "Valid From":
                            valid_from = parsed
                        else:
                            valid_to = parsed
                        break
                    except ValueError:
                        continue
                else:
                    errors.append({"row": i, "reason": f"Invalid date in {dest}: '{date_str}'."})

        existing = await db.execute(
            select(DimensionValue).where(
                DimensionValue.dimension_id == target_dim.id,
                DimensionValue.tenant_id == tenant_id,
                DimensionValue.code == code,
            )
        )
        dv_obj = existing.scalar_one_or_none()
        if dv_obj:
            dv_obj.name = name
            dv_obj.is_active = is_active
            dv_obj.sort_order = sort_order
            if value_type is not None:
                dv_obj.value_type = value_type
            if valid_from is not None:
                dv_obj.valid_from = valid_from
            if valid_to is not None:
                dv_obj.valid_to = valid_to
            updated += 1
        else:
            db.add(DimensionValue(
                tenant_id=tenant_id,
                dimension_id=target_dim.id,
                code=code,
                name=name,
                sort_order=sort_order,
                is_active=is_active,
                value_type=value_type,
                valid_from=valid_from,
                valid_to=valid_to,
            ))
            imported += 1

    await db.flush()
    return UploadResult(imported=imported, updated=updated, skipped=skipped, errors=errors)


# ═══════════════════════════════════════════════════════════════════════════════
# CHART OF ACCOUNTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/coa/template")
async def download_coa_template(
    is_active: bool = Query(default=True),
    gl_group: bool = Query(default=True),
    gl_subgroup: bool = Query(default=True),
    gl_sub_subgroup: bool = Query(default=True),
    fs_head: bool = Query(default=True),
    fs_note: bool = Query(default=True),
    tb_mapping: bool = Query(default=True),
    group_account: bool = Query(default=True),
    account_classification: bool = Query(default=True),
    category: bool = Query(default=True),
    subcategory: bool = Query(default=True),
    is_default_gl: bool = Query(default=True),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Generate and download the CoA upload template (.xlsx).

    Column selection is controlled by query parameters. Mandatory columns
    (GL Number, GL Name, Account Type, Is Active) are always included.
    Group Account columns only appear for subsidiary/branch tenants.
    Sheet 1 rows 1-3 are protected with password "ziva".
    Sheet 2 dimension name column uses a dropdown data validation.
    """
    tenant_id = _require_tenant(current_user)

    # Fetch active dimensions and tenant info
    dim_result = await db.execute(
        select(TenantDimension)
        .where(TenantDimension.tenant_id == tenant_id, TenantDimension.is_active == True)  # noqa: E712
        .order_by(TenantDimension.sort_order, TenantDimension.name)
    )
    dimensions = list(dim_result.scalars().all())

    from app.models.auth import Tenant  # local import to avoid circulars
    tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant_obj = tenant_result.scalar_one_or_none()
    tenant_name = tenant_obj.name if tenant_obj else "Company"

    # Check if tenant is subsidiary — determines if group account columns show
    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()
    is_subsidiary = False
    if org and org.org_configuration:
        structure_type = org.org_configuration.get("structure_type", "")
        is_subsidiary = structure_type.lower() in ("subsidiary", "branch")

    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        ) from exc

    from datetime import date as _date
    today_str = _date.today().strftime("%Y%m%d")

    wb = openpyxl.Workbook()

    # ── Style helpers ─────────────────────────────────────────────────────────
    req_fill = PatternFill("solid", fgColor="DBEAFE")   # light blue — required
    opt_fill = PatternFill("solid", fgColor="F3F4F6")   # light grey — optional
    header_font = Font(bold=True, size=10)
    instr_font = Font(italic=True, color="6B7280", size=9)
    example_font = Font(color="374151", size=10)

    def _write_headers(ws_target, headers: list[tuple[str, bool]], start_col: int = 1) -> None:
        """Write headers with required (blue) / optional (grey) fill."""
        for i, (h, is_req) in enumerate(headers):
            col = start_col + i
            cell = ws_target.cell(row=1, column=col, value=h)
            cell.fill = req_fill if is_req else opt_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws_target.row_dimensions[1].height = 30
            ws_target.column_dimensions[get_column_letter(col)].width = max(16, len(h) + 2)

    def _write_row(ws_target, row_num: int, values: list, font=None) -> None:
        for i, v in enumerate(values, 1):
            cell = ws_target.cell(row=row_num, column=i, value=v)
            if font:
                cell.font = font

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — GL Accounts
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "GL Accounts"
    ws1.freeze_panes = "A4"

    # Build column list per-column from individual params
    # Each entry: (header_name, instruction_text)
    all_cols: list[tuple[str, str]] = [
        ("GL Number*", "Unique GL number e.g. 733060"),
        ("GL Name*", "Full GL description"),
        ("Account Type*", "PL = Profit & Loss (SOCI) | BS = Balance Sheet (SOFP)"),
    ]

    if is_active:         all_cols.append(("Is Active", "Yes or No"))
    if gl_group:          all_cols.append(("GL Group", "Top-level GL grouping"))
    if gl_subgroup:       all_cols.append(("GL Subgroup", "Second-level grouping"))
    if gl_sub_subgroup:   all_cols.append(("GL Sub-subgroup", "Third-level (optional)"))
    if fs_head:           all_cols.append(("FS Head", "Financial statement face line"))
    if fs_note:           all_cols.append(("FS Note", "FS note reference"))
    if tb_mapping:        all_cols.append(("TB Mapping", "Trial balance roll-up group"))
    if group_account:
        all_cols.append(("Group Account Number", "Optional — parent group GL number"))
        all_cols.append(("Group Account Name", "Optional — parent group GL name"))
    if account_classification:
        all_cols.append(("Account Classification", "Classification for module behaviour"))
    if category:          all_cols.append(("Category", "Expense category name"))
    if subcategory:       all_cols.append(("Subcategory", "Subcategory name"))
    if is_default_gl:
        all_cols.append(("Is Default GL for Subcategory", "Yes if default GL for subcategory"))

    # Dimension columns — always included, one per active dimension
    for dim in dimensions:
        display = dim.display_name or dim.name
        all_cols.append((display, "Required / Optional / N/A"))

    # Mandatory cols (first 3) get required fill; the rest get optional fill
    MANDATORY_COUNT = 3
    header_row = [c[0] for c in all_cols]
    instruction_row = [c[1] for c in all_cols]
    headers_for_write = [(c[0], i < MANDATORY_COUNT) for i, c in enumerate(all_cols)]

    _write_headers(ws1, headers_for_write)

    # Cell comments on header row — one example value per column
    examples: dict[str, str] = {
        "GL Number": "e.g. 400000",
        "GL Name": "e.g. Sales domestic",
        "Account Type": "e.g. PL or BS",
        "Is Active": "e.g. Yes",
        "GL Group": "e.g. PL2",
        "GL Subgroup": "e.g. N.S.V",
        "GL Sub-subgroup": "e.g. GSV",
        "FS Head": "e.g. Revenue",
        "FS Note": "e.g. Note 1 - Revenue",
        "TB Mapping": "e.g. Domestic sales",
        "Group Account Number": "e.g. 4000",
        "Group Account Name": "e.g. Net Revenue",
        "Account Classification": "e.g. Revenue",
    }
    for col_idx, (header, _instruction) in enumerate(all_cols, 1):
        cell = ws1.cell(row=1, column=col_idx)
        example = examples.get(header.replace("*", "").strip(), "")
        if example:
            comment = Comment(example, "Ziva BI")
            comment.width = 200
            comment.height = 60
            cell.comment = comment

    # Add comment and data validation to Account Type header cell
    at_col_idx = header_row.index("Account Type*") + 1  # 1-based
    at_cell = ws1.cell(row=1, column=at_col_idx)
    at_comment = Comment(
        text=(
            "Accepted values:\n"
            "  PL  =  Profit & Loss\n"
            "         (SOCI — Statement of Comprehensive Income)\n\n"
            "  BS  =  Balance Sheet\n"
            "         (SOFP — Statement of Financial Position)\n\n"
            "Both PL/BS and SOCI/SOFP are accepted on upload."
        ),
        author="Ziva BI"
    )
    at_comment.width = 280
    at_comment.height = 120
    at_cell.comment = at_comment

    # Instruction sub-row (row 2 — no example row)
    for ci, instr in enumerate(instruction_row, 1):
        cell = ws1.cell(row=2, column=ci, value=instr)
        cell.font = instr_font
        cell.fill = opt_fill

    # Data validations on Sheet 1
    # Account Type dropdown — dynamic column, PL/BS values
    at_col_letter = ws1.cell(row=1, column=at_col_idx).column_letter
    dv_type = DataValidation(
        type="list",
        formula1='"PL,BS"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid account type",
        error="Please enter PL (Profit & Loss) or BS (Balance Sheet).",
        showInputMessage=True,
        promptTitle="Account Type",
        prompt="PL = Profit & Loss (SOCI)\nBS = Balance Sheet (SOFP)",
    )
    dv_type.sqref = f"{at_col_letter}4:{at_col_letter}5000"
    ws1.add_data_validation(dv_type)

    # Is Active dropdown — dynamic column (may not be present if unchecked)
    if "Is Active" in header_row:
        active_col_letter = get_column_letter(header_row.index("Is Active") + 1)
        dv_active = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
        ws1.add_data_validation(dv_active)
        dv_active.sqref = f"{active_col_letter}4:{active_col_letter}10000"

    # Is Default GL for Subcategory dropdown — dynamic column
    try:
        default_gl_idx = header_row.index("Is Default GL for Subcategory") + 1
        dv_default_gl = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
        ws1.add_data_validation(dv_default_gl)
        col_letter = get_column_letter(default_gl_idx)
        dv_default_gl.sqref = f"{col_letter}4:{col_letter}10000"
    except ValueError:
        pass

    # Dimension requirement dropdowns — dimensions are the last N columns
    dim_start_idx = len(all_cols) - len(dimensions) + 1
    for i in range(len(dimensions)):
        col_letter = get_column_letter(dim_start_idx + i)
        dv_dim = DataValidation(
            type="list",
            formula1='"Required,Optional,N/A"',
            allow_blank=True,
            showDropDown=False,
        )
        ws1.add_data_validation(dv_dim)
        dv_dim.sqref = f"{col_letter}4:{col_letter}10000"

    # Marker row (row 3) — visual cue showing where data entry starts
    marker_font = Font(name="Arial", bold=True, size=10, color="1E3A5F", italic=True)
    marker_fill = PatternFill("solid", fgColor="E8F0FE")

    marker_cell = ws1.cell(row=3, column=1, value="→ Enter your GL accounts from row 4 onwards")
    marker_cell.font = marker_font
    marker_cell.fill = marker_fill
    marker_cell.alignment = Alignment(horizontal="left")

    for col_num in range(2, len(all_cols) + 1):
        ws1.cell(row=3, column=col_num).fill = marker_fill

    # Amber fill on rows 1-3 to visually indicate they are reference rows
    ref_fill = PatternFill("solid", fgColor="FFF3CD")

    for row_num in range(1, 4):
        for col_num in range(1, len(all_cols) + 2):
            cell = ws1.cell(row=row_num, column=col_num)
            if cell.fill.fgColor.rgb == "00000000" or not cell.fill.patternType:
                cell.fill = ref_fill  # only apply if no existing fill

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Instructions
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Instructions")
    instruction_rows = [
        ["COLUMN", "REQUIRED", "DESCRIPTION"],
        ["--- GL ACCOUNTS (Sheet 1) ---", "", ""],
        ["GL Number", "Yes", "Unique identifier for this GL account (e.g. 400000). Max 50 chars."],
        ["GL Name", "Yes", "Full descriptive name for the GL account."],
        ["Account Type", "Yes", "SOCI = Statement of Comprehensive Income (P&L items). SOFP = Statement of Financial Position (Balance Sheet)."],
        ["Is Active", "No", "Yes or No. Defaults to Yes if omitted."],
        ["GL Group", "No", "Top-level GL hierarchy grouping (e.g. PL3 - Marketing)."],
        ["GL Subgroup", "No", "Second level grouping (e.g. Sponsoring)."],
        ["GL Sub-subgroup", "No", "Third level grouping (e.g. Sport Events)."],
        ["FS Head", "No", "Line on the face of the financial statement (e.g. Revenue, Operating Expenses)."],
        ["FS Note", "No", "Note number/name in the financial statement notes (e.g. Note 5 - Staff Costs)."],
        ["TB Mapping", "No", "Trial balance grouping this GL rolls up to (e.g. OPEX, CAPEX, Revenue)."],
        ["Group Account Number", "No", "Parent group's equivalent GL number (for subsidiaries reporting to a group)."],
        ["Group Account Name", "No", "Parent group's GL name."],
        ["Account Classification", "No", "Account classification used by Tax Engine, AP, AR, Payroll, Reporting."],
        ["Category", "No", "Maps this GL to a top-level expense category (e.g. Travel Cost)."],
        ["Subcategory", "No", "Maps this GL to a subcategory (e.g. Hotel)."],
        ["Is Default GL for Subcategory", "No", "Yes if this GL is pre-selected when the subcategory is chosen by an employee."],
        *[(f"{dim.display_name or dim.name} (Dimension)", "No", f"Requirement for dimension '{dim.display_name or dim.name}'. Values: Required / Optional / N/A. Empty = Optional.") for dim in dimensions],
        ["", "", ""],
        ["--- NOTES ---", "", ""],
        ["Duplicate GL numbers", "", "If a GL number already exists, its record will be updated (not duplicated)."],
        ["Required columns marked *", "", "Required columns have a light blue header. Optional columns have a light grey header."],
        ["Locked rows", "", "Rows 1-3 (header, instructions, marker) are protected. Enter your data from row 4 onwards."],
    ]
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 80
    for ri, row_vals in enumerate(instruction_rows, 1):
        for ci, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True)
                cell.fill = req_fill
            elif val and val.startswith("---"):
                cell.font = Font(bold=True, color="1E40AF")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", tenant_name)
    filename = f"{safe_name}_CoA_Template_{today_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/coa", response_model=list[CoAListItem])
async def list_coa(
    search: str = Query(default="", description="Search GL number or name"),
    active_only: bool = Query(default=True),
    is_retired: Optional[bool] = Query(default=None, description="Filter by retired status. None = no filter, True = only retired, False = exclude retired"),
    account_type: str = Query(default="", description="Filter by SOCI or SOFP"),
    classification: str = Query(default="", description="Filter by account_classification"),
    limit: int = Query(default=200, description="Max results to return"),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[CoAListItem]:
    """List GL accounts for the tenant. Searchable by GL number or name.
    Optionally filtered by account_type (SOCI/SOFP), classification, and is_retired.
    active_only=true excludes both plain-inactive and retired accounts.
    is_retired=true returns only retired accounts regardless of active_only."""
    tenant_id = _require_tenant(current_user)

    q = select(ChartOfAccount).where(ChartOfAccount.tenant_id == tenant_id)
    if active_only:
        q = q.where(ChartOfAccount.is_active == True)  # noqa: E712
    if is_retired is not None:
        q = q.where(ChartOfAccount.is_retired == is_retired)  # noqa: E712
    if search.strip():
        term = f"%{search.strip()}%"
        q = q.where(
            ChartOfAccount.gl_number.ilike(term) | ChartOfAccount.gl_name.ilike(term)
        )
    if account_type.strip():
        at = account_type.strip().upper().replace("/", "").replace("&", "").replace(" ", "")
        # Normalise to PL variants
        if at in ("SOCI", "PL", "PANDL", "PROFITLOSS", "PROFITANDLOSS", "PLV"):
            q = q.where(
                ChartOfAccount.account_type.in_(
                    ["PL", "SOCI", "P/L", "P&L", "Profit & Loss", "Profit and Loss"]
                )
            )
        # Normalise to BS variants
        elif at in ("SOFP", "BS", "BALANCESHEET", "BALANCE"):
            q = q.where(
                ChartOfAccount.account_type.in_(
                    ["BS", "SOFP", "B/S", "Balance Sheet"]
                )
            )
        else:
            q = q.where(ChartOfAccount.account_type.ilike(f"%{account_type}%"))
    if classification.strip():
        # Support comma-separated list e.g. "Finance income,Finance cost"
        from sqlalchemy import or_
        classes = [c.strip() for c in classification.split(",") if c.strip()]
        if len(classes) == 1:
            q = q.where(ChartOfAccount.account_classification.ilike(f"%{classes[0]}%"))
        else:
            q = q.where(or_(*[
                ChartOfAccount.account_classification.ilike(f"%{c}%")
                for c in classes
            ]))
    q = q.order_by(ChartOfAccount.gl_number).limit(limit)

    result = await db.execute(q)
    return [CoAListItem.from_orm(g) for g in result.scalars().all()]


@router.get("/coa/fs-mappings")
async def get_fs_mappings(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
    account_type: Optional[str] = Query(None),
    fs_head: Optional[str] = Query(None),
) -> list[dict]:
    """Return all active GL accounts with FS mapping fields.
    Accounts with fs_head set come first; unmapped accounts last."""
    tenant_id = _require_tenant(current_user)
    q = select(ChartOfAccount).where(
        ChartOfAccount.tenant_id == tenant_id,
        ChartOfAccount.is_active == True,  # noqa: E712
    )
    if account_type:
        q = q.where(ChartOfAccount.account_type == account_type)
    if fs_head:
        q = q.where(ChartOfAccount.fs_head == fs_head)
    q = q.order_by(
        ChartOfAccount.fs_head.is_(None),
        ChartOfAccount.fs_head,
        ChartOfAccount.gl_number,
    )
    result = await db.execute(q)
    return [
        {
            "id": str(a.id),
            "gl_number": a.gl_number,
            "gl_name": a.gl_name,
            "account_type": a.account_type,
            "fs_head": a.fs_head,
            "fs_note": a.fs_note,
            "tb_mapping": a.tb_mapping,
        }
        for a in result.scalars().all()
    ]


@router.get("/coa/dimension-matrix")
async def get_coa_dimension_matrix(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Return all GL accounts with their dimension requirements as a matrix.

    Used by the CoA Dimensions tab to render a GL × dimension grid where
    each cell shows Required / Optional / N/A. Missing rows default to
    'optional'.
    """
    tenant_id = _require_tenant(current_user)

    dims_result = await db.execute(
        select(TenantDimension)
        .where(TenantDimension.tenant_id == tenant_id,
               TenantDimension.is_active == True)  # noqa: E712
        .order_by(TenantDimension.sort_order)
    )
    dimensions = dims_result.scalars().all()

    gl_result = await db.execute(
        select(ChartOfAccount)
        .where(ChartOfAccount.tenant_id == tenant_id)
        .order_by(ChartOfAccount.gl_number)
    )
    accounts = gl_result.scalars().all()

    req_result = await db.execute(
        select(GLDimensionRequirement)
        .where(GLDimensionRequirement.tenant_id == tenant_id)
    )
    requirements = req_result.scalars().all()

    req_map: dict[str, dict[str, str]] = {}
    for r in requirements:
        gl_key = str(r.gl_id)
        if gl_key not in req_map:
            req_map[gl_key] = {}
        req_map[gl_key][str(r.dimension_id)] = r.requirement

    return {
        "dimensions": [
            {"id": str(d.id), "name": d.display_name or d.name}
            for d in dimensions
        ],
        "accounts": [
            {
                "id": str(a.id),
                "gl_number": a.gl_number,
                "gl_name": a.gl_name,
                "account_type": a.account_type,
                "gl_group": a.gl_group,
                "is_active": a.is_active,
                "requirements": {
                    str(d.id): req_map.get(str(a.id), {}).get(str(d.id), "optional")
                    for d in dimensions
                },
            }
            for a in accounts
        ],
    }


@router.patch("/coa/dimension-requirements/bulk")
async def bulk_update_dimension_requirements(
    data: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Bulk-set a dimension requirement for multiple GL accounts at once.

    Payload: { gl_ids: [uuid, ...], dimension_id: uuid, requirement: required|optional|na }
    Upserts rows in gl_dimension_requirements. Admin only.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    gl_ids = [uuid.UUID(i) for i in data["gl_ids"]]
    dimension_id = uuid.UUID(data["dimension_id"])
    requirement = data["requirement"]

    if requirement not in ("required", "optional", "na"):
        raise HTTPException(status_code=400, detail="Invalid requirement value.")

    for gl_id in gl_ids:
        req_result = await db.execute(
            select(GLDimensionRequirement).where(
                GLDimensionRequirement.gl_id == gl_id,
                GLDimensionRequirement.dimension_id == dimension_id,
                GLDimensionRequirement.tenant_id == tenant_id,
            )
        )
        req_row = req_result.scalar_one_or_none()
        if req_row:
            req_row.requirement = requirement
        else:
            db.add(GLDimensionRequirement(
                tenant_id=tenant_id,
                gl_id=gl_id,
                dimension_id=dimension_id,
                requirement=requirement,
            ))

    await db.commit()
    return {"updated": len(gl_ids)}


@router.get("/coa/remap-template")
async def download_remap_template_early(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """See download_remap_template (defined below) — forward registration to avoid {gl_id} shadowing."""
    return await download_remap_template(current_user, db)


@router.get("/coa/remap-history", response_model=list[GlRemapHistoryEntry])
async def list_remap_history_early(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[GlRemapHistoryEntry]:
    """See list_remap_history (defined below) — forward registration to avoid {gl_id} shadowing."""
    return await list_remap_history(current_user, db)


@router.get("/coa/{gl_id}", response_model=CoAResponse)
async def get_coa(
    gl_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CoAResponse:
    """Return a single GL account by ID."""
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id == gl_id,
            ChartOfAccount.tenant_id == tenant_id,
        )
    )
    gl = result.scalar_one_or_none()
    if not gl:
        raise HTTPException(status_code=404, detail="GL account not found.")
    return CoAResponse.from_orm(gl)


@router.post("/coa", response_model=CoAListItem, status_code=status.HTTP_201_CREATED)
async def create_coa(
    data: CoACreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CoAListItem:
    """Create a single GL account. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    existing = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.tenant_id == tenant_id,
            ChartOfAccount.gl_number == data.gl_number,
            ChartOfAccount.is_active == True,  # noqa: E712
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"GL account '{data.gl_number}' already exists.",
        )

    gl = ChartOfAccount(
        tenant_id=tenant_id,
        gl_number=data.gl_number,
        gl_name=data.gl_name,
        account_type=data.account_type,
        account_classification=data.account_classification,
        is_foreign_currency=data.is_foreign_currency or False,
        foreign_currency_code=data.foreign_currency_code,
        revalue_at_period_end=data.revalue_at_period_end or False,
    )
    db.add(gl)
    await db.flush()
    await db.refresh(gl)
    return CoAListItem.from_orm(gl)


@router.patch("/coa/{gl_id}", response_model=CoAListItem)
async def update_coa(
    gl_id: uuid.UUID,
    data: CoAUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CoAListItem:
    """Update a GL account (PATCH semantics). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id == gl_id,
            ChartOfAccount.tenant_id == tenant_id,
        )
    )
    gl = result.scalar_one_or_none()
    if not gl:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="GL account not found.")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(gl, field, value)

    await db.flush()
    await db.refresh(gl)
    return CoAListItem.from_orm(gl)


@router.delete("/coa/{gl_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_coa(
    gl_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete a GL account. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id == gl_id,
            ChartOfAccount.tenant_id == tenant_id,
        )
    )
    gl = result.scalar_one_or_none()
    if not gl:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="GL account not found.")

    gl.is_active = False
    await db.flush()


@router.post("/coa/upload")
async def upload_coa(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Bulk upload Chart of Accounts from .xlsx (multi-sheet) or .csv (Sheet 1 only).

    For .xlsx:
      Sheet 2 (Dimensions Setup) is processed first — upsert dimension values.
      Sheet 1 (GL Accounts) is processed second — upsert GL accounts + hierarchy + requirements.

    Returns: {"sheet1": UploadResult, "sheet2": UploadResult}
    For .csv returns: {"sheet1": UploadResult, "sheet2": {"imported": 0, ...}}
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    content = await file.read()
    fname = (file.filename or "").lower()

    # Fetch active dimensions once
    dim_result = await db.execute(
        select(TenantDimension)
        .where(TenantDimension.tenant_id == tenant_id, TenantDimension.is_active == True)  # noqa: E712
    )
    dimensions = list(dim_result.scalars().all())
    dim_by_name = {d.name.lower(): d for d in dimensions}

    sheet1_rows: list[list[str]] = []
    sheet1_headers: list[str] = []
    sheet2_rows: list[list[str]] = []
    sheet2_headers: list[str] = []

    if fname.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="openpyxl is not installed. Run: pip install openpyxl",
            ) from exc
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheet_names = wb.sheetnames

        def _load_sheet(ws_name: str) -> tuple[list[str], list[list[str]]]:
            if ws_name not in sheet_names:
                return [], []
            ws_obj = wb[ws_name]
            all_rows: list[list[str]] = []
            for row in ws_obj.iter_rows(values_only=True):
                all_rows.append([str(c).strip() if c is not None else "" for c in row])
            if not all_rows:
                return [], []
            headers_row = all_rows[0]

            data_rows = all_rows[3:]  # skip header, instructions, marker (no example row)

            return [h.strip() for h in headers_row], data_rows

        sheet1_headers, sheet1_rows = _load_sheet("GL Accounts")
        if not sheet1_headers:
            sheet1_headers, sheet1_rows = _load_sheet(sheet_names[0])
        sheet2_headers, sheet2_rows = _load_sheet("Dimensions Setup")

    elif fname.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if all_rows:
            sheet1_headers = [h.strip() for h in all_rows[0]]
            sheet1_rows = [[c.strip() for c in r] for r in all_rows[1:]]
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .csv and .xlsx files are supported.",
        )

    # ── Process Sheet 2 — Dimension Values ───────────────────────────────────
    s2_result = UploadResult(imported=0, skipped=0, errors=[])

    if sheet2_headers:
        s2h = [h.lower().strip("*").strip() for h in sheet2_headers]

        def s2col(name: str) -> Optional[int]:
            try:
                return s2h.index(name)
            except ValueError:
                return None

        s2_dim_col = s2col("dimension name")
        s2_code_col = s2col("value code")
        s2_name_col = s2col("value name")
        s2_type_col = s2col("value type")
        s2_from_col = s2col("valid from (dd/mm/yyyy)") or s2col("valid from")
        s2_to_col = s2col("valid to (dd/mm/yyyy)") or s2col("valid to")
        s2_active_col = s2col("is active")

        if s2_dim_col is not None and s2_code_col is not None and s2_name_col is not None:
            from datetime import datetime as _dt
            for i, row in enumerate(sheet2_rows, start=4):
                if not any(c.strip() for c in row if c):
                    continue

                def s2get(idx: Optional[int]) -> str:
                    if idx is None or idx >= len(row):
                        return ""
                    return row[idx].strip() if row[idx] else ""

                dim_name = s2get(s2_dim_col)
                val_code = s2get(s2_code_col)
                val_name = s2get(s2_name_col)
                value_type = s2get(s2_type_col) or None
                valid_from_str = s2get(s2_from_col)
                valid_to_str = s2get(s2_to_col)
                is_active_str = s2get(s2_active_col).lower()

                if not dim_name:
                    continue
                if not val_code:
                    s2_result.errors.append({"row": i, "reason": "Missing Value Code."})
                    continue
                if not val_name:
                    s2_result.errors.append({"row": i, "reason": "Missing Value Name."})
                    continue

                dim_obj = dim_by_name.get(dim_name.lower())
                if not dim_obj:
                    s2_result.errors.append({"row": i, "reason": f"Dimension '{dim_name}' not found."})
                    s2_result.skipped += 1
                    continue

                is_active = is_active_str not in ("no", "false", "0")

                valid_from = None
                valid_to = None
                for date_str, dest_name in [(valid_from_str, "Valid From"), (valid_to_str, "Valid To")]:
                    if date_str:
                        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                            try:
                                parsed = _dt.strptime(date_str, fmt).date()
                                if dest_name == "Valid From":
                                    valid_from = parsed
                                else:
                                    valid_to = parsed
                                break
                            except ValueError:
                                continue
                        else:
                            s2_result.errors.append({"row": i, "reason": f"Invalid date in {dest_name}: '{date_str}'."})

                # Upsert dimension value
                existing_dv = await db.execute(
                    select(DimensionValue).where(
                        DimensionValue.dimension_id == dim_obj.id,
                        DimensionValue.tenant_id == tenant_id,
                        DimensionValue.code == val_code,
                    )
                )
                dv_obj = existing_dv.scalar_one_or_none()
                if dv_obj:
                    dv_obj.name = val_name
                    dv_obj.is_active = is_active
                    if value_type is not None:
                        dv_obj.value_type = value_type
                    if valid_from is not None:
                        dv_obj.valid_from = valid_from
                    if valid_to is not None:
                        dv_obj.valid_to = valid_to
                    s2_result.updated += 1
                else:
                    db.add(DimensionValue(
                        tenant_id=tenant_id,
                        dimension_id=dim_obj.id,
                        code=val_code,
                        name=val_name,
                        is_active=is_active,
                        value_type=value_type,
                        valid_from=valid_from,
                        valid_to=valid_to,
                    ))
                    s2_result.imported += 1

            await db.flush()

    # ── Process Sheet 1 — GL Accounts ────────────────────────────────────────
    s1_result = UploadResult(imported=0, skipped=0, errors=[])

    if not sheet1_headers:
        s1_result.errors.append({"row": 0, "reason": "Empty GL Accounts sheet."})
        return {"sheet1": s1_result.model_dump(), "sheet2": s2_result.model_dump()}

    s1h = [h.lower().strip("*").strip() for h in sheet1_headers]

    def s1col(name: str) -> Optional[int]:
        try:
            return s1h.index(name)
        except ValueError:
            return None

    gl_number_col = s1col("gl number")
    gl_name_col = s1col("gl name")
    account_type_col = s1col("account type")
    is_active_col = s1col("is active")
    gl_group_col = s1col("gl group")
    gl_subgroup_col = s1col("gl subgroup")
    gl_sub_subgroup_col = s1col("gl sub-subgroup") or s1col("gl subsubgroup")
    fs_head_col = s1col("fs head")
    fs_note_col = s1col("fs note")
    tb_mapping_col = s1col("tb mapping")
    group_acct_num_col = s1col("group account number")
    group_acct_name_col = s1col("group account name")
    category_col = s1col("category")
    subcategory_col = s1col("subcategory")

    if gl_number_col is None or gl_name_col is None or account_type_col is None:
        s1_result.errors.append({"row": 0, "reason": "File must have 'GL Number', 'GL Name', and 'Account Type' columns."})
        return {"sheet1": s1_result.model_dump(), "sheet2": s2_result.model_dump()}

    # Map dimension name → (dimension_id, column_index)
    dim_cols: list[tuple[uuid.UUID, int]] = []
    for dim in dimensions:
        idx = s1col(dim.name.lower())
        if idx is not None:
            dim_cols.append((dim.id, idx))

    VALID_REQUIREMENTS = {"required", "optional", "na", "n/a"}
    seen_gl_numbers: set[str] = set()

    for i, row in enumerate(sheet1_rows, start=4):
        if not any((c.strip() if c else "") for c in row):
            continue

        def s1get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        gl_number = s1get(gl_number_col)
        gl_name = s1get(gl_name_col)
        account_type_raw = s1get(account_type_col)
        is_active_raw = s1get(is_active_col).lower()

        if not gl_number:
            s1_result.errors.append({"row": i, "reason": "Missing GL Number."})
            continue
        if not gl_name:
            s1_result.errors.append({"row": i, "reason": "Missing GL Name."})
            continue

        # Normalise account type: accept SOCI/SOFP/PL/BS/P&L/B/S
        at = account_type_raw.strip().upper()
        at_map = {"PL": "SOCI", "P&L": "SOCI", "BS": "SOFP", "B/S": "SOFP"}
        account_type = at_map.get(at, at)
        if account_type not in ("SOCI", "SOFP"):
            s1_result.errors.append({"row": i, "reason": f"Invalid Account Type: '{account_type_raw}'. Use SOCI or SOFP."})
            continue

        is_active = is_active_raw not in ("no", "false", "0")

        if gl_number in seen_gl_numbers:
            s1_result.skipped += 1
            s1_result.skipped_rows.append({"row": i, "gl_number": gl_number, "reason": "Duplicate GL number in file"})
            continue
        seen_gl_numbers.add(gl_number)

        # Upsert GL record
        existing_result = await db.execute(
            select(ChartOfAccount).where(
                ChartOfAccount.tenant_id == tenant_id,
                ChartOfAccount.gl_number == gl_number,
            )
        )
        gl_obj = existing_result.scalar_one_or_none()
        if gl_obj:
            gl_obj.gl_name = gl_name
            gl_obj.account_type = account_type
            gl_obj.is_active = is_active
            gl_obj.gl_group = s1get(gl_group_col) or gl_obj.gl_group
            gl_obj.gl_subgroup = s1get(gl_subgroup_col) or gl_obj.gl_subgroup
            gl_obj.gl_sub_subgroup = s1get(gl_sub_subgroup_col) or gl_obj.gl_sub_subgroup
            gl_obj.fs_head = s1get(fs_head_col) or gl_obj.fs_head
            gl_obj.fs_note = s1get(fs_note_col) or gl_obj.fs_note
            gl_obj.tb_mapping = s1get(tb_mapping_col) or gl_obj.tb_mapping
            gl_obj.group_account_number = s1get(group_acct_num_col) or gl_obj.group_account_number
            gl_obj.group_account_name = s1get(group_acct_name_col) or gl_obj.group_account_name
            s1_result.updated += 1
        else:
            gl_obj = ChartOfAccount(
                tenant_id=tenant_id,
                gl_number=gl_number,
                gl_name=gl_name,
                account_type=account_type,
                is_active=is_active,
                gl_group=s1get(gl_group_col) or None,
                gl_subgroup=s1get(gl_subgroup_col) or None,
                gl_sub_subgroup=s1get(gl_sub_subgroup_col) or None,
                fs_head=s1get(fs_head_col) or None,
                fs_note=s1get(fs_note_col) or None,
                tb_mapping=s1get(tb_mapping_col) or None,
                group_account_number=s1get(group_acct_num_col) or None,
                group_account_name=s1get(group_acct_name_col) or None,
            )
            db.add(gl_obj)
            s1_result.imported += 1

        await db.flush()
        await db.refresh(gl_obj)

        # Handle dimension requirement columns
        row_errors: list[str] = []
        for dim_id, col_idx in dim_cols:
            req_raw = s1get(col_idx)
            req = req_raw.lower().replace("/", "") if req_raw else "optional"
            if req == "na":
                req = "na"
            elif not req:
                req = "optional"
            if req not in VALID_REQUIREMENTS:
                row_errors.append(f"Invalid requirement '{req_raw}' for dimension col {col_idx + 1}.")
                continue
            req_norm = "na" if req in ("na", "n/a") else req

            req_result = await db.execute(
                select(GLDimensionRequirement).where(
                    GLDimensionRequirement.gl_id == gl_obj.id,
                    GLDimensionRequirement.dimension_id == dim_id,
                )
            )
            req_row = req_result.scalar_one_or_none()
            if req_row:
                req_row.requirement = req_norm
            else:
                db.add(GLDimensionRequirement(
                    tenant_id=tenant_id,
                    gl_id=gl_obj.id,
                    dimension_id=dim_id,
                    requirement=req_norm,
                ))

        if row_errors:
            s1_result.errors.append({"row": i, "reason": "; ".join(row_errors)})

    await db.flush()
    return {"sheet1": s1_result.model_dump(), "sheet2": s2_result.model_dump()}


@router.post("/coa/bulk-action", response_model=BulkActionResult)
async def bulk_action_coa(
    data: BulkActionRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> BulkActionResult:
    """Bulk activate / deactivate / delete GL accounts by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id.in_(data.ids),
            ChartOfAccount.tenant_id == tenant_id,
        )
    )
    records = list(result.scalars().all())
    affected = 0
    errors: list[dict] = []

    for rec in records:
        if data.action == "activate":
            rec.is_active = True
            affected += 1
        elif data.action == "deactivate":
            rec.is_active = False
            affected += 1
        elif data.action == "delete":
            await db.delete(rec)
            affected += 1

    await db.flush()
    return BulkActionResult(action=data.action, affected=affected, skipped=len(data.ids) - affected, errors=errors)


@router.post("/coa/replace-all")
async def replace_all_coa(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Deactivate ALL existing GL accounts for the tenant, then import the uploaded file.

    Lifecycle gate: only allowed when tenant.lifecycle_status == "in_implementation".
    Once the tenant goes live, Replace All is permanently unavailable — use the
    Remap workflow instead to preserve historical journal integrity.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    # Lifecycle gate — in_implementation only
    from app.models.auth import Tenant
    t_res = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    t = t_res.scalar_one_or_none()
    if not t or t.lifecycle_status != "in_implementation":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Replace All is only available during implementation "
                f"(current status: {t.lifecycle_status if t else 'unknown'}). "
                "Use the Remap workflow to retire old codes once your tenant is live."
            ),
        )

    # Deactivate all existing GL accounts
    existing_result = await db.execute(
        select(ChartOfAccount).where(ChartOfAccount.tenant_id == tenant_id)
    )
    for gl in existing_result.scalars().all():
        gl.is_active = False
    await db.flush()

    # Delegate to upload handler in the same transaction
    file.file.seek(0)
    return await upload_coa(file, current_user, db)


@router.post("/coa/remap", response_model=RemapResult)
async def remap_coa(
    data: RemapRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RemapResult:
    """
    Retire one or more old GL accounts and remap them many:1 to a new or existing code.

    Lifecycle gate: only allowed when tenant.lifecycle_status == "live".
    During implementation use Replace All instead; Remap is for post-go-live corrections
    where journal history must be preserved.

    Rules enforced:
    - All old_account_ids must exist for this tenant, be active, and not already retired.
    - Exactly one of new_account_id or new_account must be provided.
    - All old accounts must share the same account_type (PL or BS).
    - The new account's account_type must match the old accounts.
    - On success: old accounts are marked is_active=False + is_retired=True;
      a gl_code_remaps row is created per old account; an audit log entry is written.
    - Historical journal lines continue pointing at old account IDs unchanged.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    # Lifecycle gate — live only
    from app.models.auth import Tenant
    t_res = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    t = t_res.scalar_one_or_none()
    if not t or t.lifecycle_status != "live":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Remap is only available once your tenant is live "
                f"(current status: {t.lifecycle_status if t else 'unknown'}). "
                "Use Replace All during implementation to restructure your Chart of Accounts."
            ),
        )

    if not data.old_account_ids:
        raise HTTPException(status_code=400, detail="At least one old_account_id is required.")
    if data.new_account_id is None and data.new_account is None:
        raise HTTPException(status_code=400, detail="Provide either new_account_id or new_account (inline creation).")
    if data.new_account_id is not None and data.new_account is not None:
        raise HTTPException(status_code=400, detail="Provide new_account_id OR new_account, not both.")

    # Load and validate old accounts
    old_result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id.in_(data.old_account_ids),
            ChartOfAccount.tenant_id == tenant_id,
        )
    )
    old_accounts = old_result.scalars().all()

    found_ids = {a.id for a in old_accounts}
    missing = set(data.old_account_ids) - found_ids
    if missing:
        raise HTTPException(status_code=404, detail=f"GL accounts not found: {[str(m) for m in missing]}")

    errors = []
    for a in old_accounts:
        if not a.is_active:
            errors.append(f"{a.gl_number} is already inactive.")
        if a.is_retired:
            errors.append(f"{a.gl_number} is already retired.")
    if errors:
        raise HTTPException(status_code=422, detail={"message": "Cannot retire these accounts.", "errors": errors})

    # Check account_type consistency — normalize PL→SOCI and BS→SOFP before comparing
    _normalize_type = {"PL": "SOCI", "BS": "SOFP", "P/L": "SOCI", "P&L": "SOCI",
                       "B/S": "SOFP", "Profit & Loss": "SOCI", "Balance Sheet": "SOFP"}
    def _norm(t: str) -> str:
        return _normalize_type.get(t, t)

    old_types_raw = {a.account_type for a in old_accounts}
    old_types_norm = {_norm(t) for t in old_types_raw}
    if len(old_types_norm) > 1:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "All accounts being retired must have the same account_type.",
                "conflicting_types": list(old_types_raw),
                "accounts": [f"{a.gl_number} ({a.account_type})" for a in old_accounts],
            },
        )
    required_type_norm = next(iter(old_types_norm))

    # Resolve or create the new account
    new_account_created = False
    if data.new_account_id is not None:
        new_res = await db.execute(
            select(ChartOfAccount).where(
                ChartOfAccount.id == data.new_account_id,
                ChartOfAccount.tenant_id == tenant_id,
            )
        )
        new_acct = new_res.scalar_one_or_none()
        if not new_acct:
            raise HTTPException(status_code=404, detail="New account not found.")
        if not new_acct.is_active or new_acct.is_retired:
            raise HTTPException(status_code=422, detail="New account must be active and not retired.")
        if new_acct.id in found_ids:
            raise HTTPException(status_code=422, detail="New account cannot be one of the accounts being retired.")
        if _norm(new_acct.account_type) != required_type_norm:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "New account type must match old accounts.",
                    "new_account_type": new_acct.account_type,
                    "old_account_type": next(iter(old_types_raw)),
                },
            )
    else:
        # Create new account inline
        naf: InlineNewAccountFields = data.new_account  # type: ignore[assignment]
        if _norm(naf.account_type) != required_type_norm:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "Inline new account type must match old accounts.",
                    "new_account_type": naf.account_type,
                    "old_account_type": next(iter(old_types_raw)),
                },
            )
        # Check gl_number uniqueness
        dup = await db.execute(
            select(ChartOfAccount).where(
                ChartOfAccount.tenant_id == tenant_id,
                ChartOfAccount.gl_number == naf.gl_number,
            )
        )
        if dup.scalar_one_or_none():
            raise HTTPException(status_code=422, detail=f"GL number '{naf.gl_number}' already exists for this tenant.")
        new_acct = ChartOfAccount(
            tenant_id=tenant_id,
            gl_number=naf.gl_number,
            gl_name=naf.gl_name,
            account_type=naf.account_type,
            is_active=True,
            is_retired=False,
            gl_group=naf.gl_group,
            gl_subgroup=naf.gl_subgroup,
            gl_sub_subgroup=naf.gl_sub_subgroup,
            fs_head=naf.fs_head,
            fs_note=naf.fs_note,
            tb_mapping=naf.tb_mapping,
            account_classification=naf.account_classification,
        )
        db.add(new_acct)
        await db.flush()  # get new_acct.id
        new_account_created = True

    # Retire old accounts + create remap rows
    remap_entries: list[RemapResultEntry] = []
    for old in old_accounts:
        old.is_active = False
        old.is_retired = True
        db.add(GlCodeRemap(
            tenant_id=tenant_id,
            old_account_id=old.id,
            new_account_id=new_acct.id,
            remapped_by=current_user.user_id,
            reason=data.reason,
        ))
        remap_entries.append(RemapResultEntry(
            old_gl_number=old.gl_number,
            old_gl_name=old.gl_name,
            new_gl_number=new_acct.gl_number,
            new_gl_name=new_acct.gl_name,
        ))

    # Audit log
    from app.models.auth import AuditLog
    db.add(AuditLog(
        event_type="coa.remap",
        user_id=current_user.user_id,
        tenant_id=tenant_id,
        log_metadata={
            "retired_accounts": [{"id": str(a.id), "gl_number": a.gl_number} for a in old_accounts],
            "new_account": {"id": str(new_acct.id), "gl_number": new_acct.gl_number},
            "new_account_created": new_account_created,
            "reason": data.reason,
        },
    ))

    await db.flush()
    return RemapResult(
        remapped=remap_entries,
        new_account_created=new_account_created,
        new_gl_number=new_acct.gl_number,
        new_gl_name=new_acct.gl_name,
    )


@router.get("/coa/remap-template")
async def download_remap_template(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Download the bulk remap upload template (.xlsx).

    Column structure: Old GL Number | New GL Number | Reason (optional)
    One row per old→new pair.  Multiple rows can share the same New GL Number
    for many-to-one remaps.  New GL Number must already exist as an active account.
    Retired-code creation inline is not supported in bulk — use the single-screen
    remap form for that.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl not installed.") from exc

    # Fetch active GL accounts for reference
    active_res = await db.execute(
        select(ChartOfAccount)
        .where(ChartOfAccount.tenant_id == tenant_id, ChartOfAccount.is_active == True)  # noqa: E712
        .order_by(ChartOfAccount.gl_number)
    )
    active_accounts = active_res.scalars().all()

    from openpyxl.comments import Comment as XLComment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remap"
    ws.freeze_panes = "A2"

    hdr_fill = PatternFill("solid", fgColor="FEF3C7")  # amber
    hdr_font = Font(bold=True, size=10)
    data_start_fill = PatternFill("solid", fgColor="FFFBEB")  # very light amber data-start hint

    headers = ["Old GL Number *", "New GL Number *", "Reason (optional)"]
    header_comments = [
        "Must exactly match an active, non-retired GL number in your CoA.\nExample: 5010",
        "Must exactly match an active, non-retired GL number to map old code under.\nExample: 5015",
        "Optional explanation for the audit trail.\nExample: Q3 consolidation",
    ]
    col_widths = [20, 20, 40]

    # Row 1: headers with cell comments (no inline instruction row per template format standard)
    for ci, (h, comment_text, w) in enumerate(zip(headers, header_comments, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
        comment = XLComment(comment_text, "Ziva BI")
        comment.width = 200; comment.height = 60
        cell.comment = comment
    ws.row_dimensions[1].height = 28

    # Row 2: data-start marker (visual hint per template format standard)
    ws.cell(row=2, column=1).fill = data_start_fill
    ws.cell(row=2, column=2).fill = data_start_fill
    ws.cell(row=2, column=3).fill = data_start_fill

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 70
    instr_rows = [
        ["COLUMN", "DESCRIPTION"],
        ["Old GL Number", "The GL number being retired. Must be active and not already retired."],
        ["New GL Number", "The GL number to remap under. Must exist and be active. Cannot be the same as the old code."],
        ["Reason", "Optional. Appears in the remap audit log."],
        ["", ""],
        ["DATA START ROW", "Enter your data starting at row 2. Row 1 is the header."],
        ["MANY-TO-ONE", "To remap multiple old codes to one new code, add one row per old code, all with the same New GL Number."],
        ["TYPE CHECK", "Old and new codes must all have the same account type (PL/SOCI or BS/SOFP). Rows with mismatched types are rejected."],
        ["EFFECT", "Retired codes are marked inactive. Historical journals still reference them. Reports can show data by old OR new code."],
    ]
    for ri, row_vals in enumerate(instr_rows, 1):
        for ci, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True)
                cell.fill = hdr_fill

    # Reference sheet with active GL accounts
    ws3 = wb.create_sheet("Active GL Accounts (Reference)")
    ws3.append(["GL Number", "GL Name", "Account Type"])
    ws3.cell(1, 1).font = Font(bold=True)
    ws3.cell(1, 2).font = Font(bold=True)
    ws3.cell(1, 3).font = Font(bold=True)
    for acct in active_accounts:
        ws3.append([acct.gl_number, acct.gl_name, acct.account_type])
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=coa_remap_template.xlsx"},
    )


@router.post("/coa/remap-bulk", response_model=BulkRemapResult)
async def bulk_remap_coa(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> BulkRemapResult:
    """
    Bulk remap GL accounts from an uploaded xlsx/csv file.

    Column structure: Old GL Number | New GL Number | Reason (optional)
    One row per old→new pair.  Groups rows by New GL Number; validates
    account_type consistency per group.  Applies valid groups, reports
    row-level errors for invalid ones.  Valid and invalid groups in the same
    file are handled independently — valid groups are applied even when other
    groups fail.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    # Lifecycle gate — live only
    from app.models.auth import Tenant
    t_res = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    t = t_res.scalar_one_or_none()
    if not t or t.lifecycle_status != "live":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=(
                "Remap is only available once your tenant is live "
                f"(current status: {t.lifecycle_status if t else 'unknown'}). "
                "Use Replace All during implementation."
            ),
        )

    content = await file.read()
    fname = (file.filename or "").lower()

    headers: list[str] = []
    rows: list[list[str]] = []

    if fname.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="openpyxl not installed.") from exc
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.worksheets[0]  # first sheet
        all_rows = [[str(c).strip() if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
        if all_rows:
            headers = all_rows[0]
            rows = all_rows[1:]  # row 1 = instructions, skip
    elif fname.endswith(".csv"):
        import csv
        all_csv = list(csv.reader(io.StringIO(content.decode("utf-8", errors="replace"))))
        if all_csv:
            headers = [h.strip() for h in all_csv[0]]
            rows = [[c.strip() for c in r] for r in all_csv[1:]]
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported.")

    h = [hdr.lower().strip("*").strip() for hdr in headers]

    def col(name: str) -> Optional[int]:
        try: return h.index(name)
        except ValueError: return None

    # Use 'is not None' pattern — col() returns 0 for the first column, which is falsy
    _oc = col("old gl number"); old_col    = _oc if _oc is not None else col("old gl number *")
    _nc = col("new gl number"); new_col    = _nc if _nc is not None else col("new gl number *")
    _rc = col("reason (optional)"); reason_col = _rc if _rc is not None else col("reason")

    if old_col is None or new_col is None:
        raise HTTPException(status_code=400, detail="File must have 'Old GL Number' and 'New GL Number' columns.")

    # Load all active CoA for this tenant
    all_res = await db.execute(
        select(ChartOfAccount).where(ChartOfAccount.tenant_id == tenant_id)
    )
    all_accounts = {a.gl_number.upper(): a for a in all_res.scalars().all()}

    result_rows: list[BulkRemapRow] = []
    remapped_count = 0
    error_count = 0

    # Parse all rows first, group by new_gl_number
    from collections import defaultdict
    groups: dict[str, list[tuple[int, str, str]]] = defaultdict(list)  # new_gl → [(row_no, old_gl, reason)]
    parse_errors: list[BulkRemapRow] = []

    for i, row in enumerate(rows, start=2):  # row 1 = header, row 2 = first instruction
        if not any((c or "").strip() for c in row):
            continue

        def get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row): return ""
            return (row[idx] or "").strip()

        old_gl    = get(old_col).upper()
        new_gl    = get(new_col).upper()
        reason    = get(reason_col) or None

        if not old_gl:
            parse_errors.append(BulkRemapRow(row=i, old_gl_number="", new_gl_number=new_gl, status="error", reason="Missing Old GL Number."))
            continue
        if not new_gl:
            parse_errors.append(BulkRemapRow(row=i, old_gl_number=old_gl, new_gl_number="", status="error", reason="Missing New GL Number."))
            continue
        groups[new_gl].append((i, old_gl, reason))

    error_count += len(parse_errors)
    result_rows.extend(parse_errors)

    # Validate and apply each group
    for new_gl, group_rows in groups.items():
        new_acct = all_accounts.get(new_gl)
        if not new_acct or not new_acct.is_active or new_acct.is_retired:
            for row_no, old_gl, _ in group_rows:
                result_rows.append(BulkRemapRow(row=row_no, old_gl_number=old_gl, new_gl_number=new_gl, status="error",
                                                reason=f"New GL '{new_gl}' does not exist or is not active."))
                error_count += 1
            continue

        # Validate old accounts in group
        group_errors: list[BulkRemapRow] = []
        valid_pairs: list[tuple[int, ChartOfAccount, str | None]] = []

        for row_no, old_gl, reason in group_rows:
            if old_gl == new_gl:
                group_errors.append(BulkRemapRow(row=row_no, old_gl_number=old_gl, new_gl_number=new_gl, status="error",
                                                  reason="Old and new GL number cannot be the same."))
                continue
            old_acct = all_accounts.get(old_gl)
            if not old_acct:
                group_errors.append(BulkRemapRow(row=row_no, old_gl_number=old_gl, new_gl_number=new_gl, status="error",
                                                  reason=f"GL '{old_gl}' not found."))
                continue
            if not old_acct.is_active:
                group_errors.append(BulkRemapRow(row=row_no, old_gl_number=old_gl, new_gl_number=new_gl, status="error",
                                                  reason=f"GL '{old_gl}' is already inactive."))
                continue
            if old_acct.is_retired:
                group_errors.append(BulkRemapRow(row=row_no, old_gl_number=old_gl, new_gl_number=new_gl, status="error",
                                                  reason=f"GL '{old_gl}' is already retired."))
                continue
            valid_pairs.append((row_no, old_acct, reason))

        if group_errors:
            error_count += len(group_errors)
            result_rows.extend(group_errors)
            # Still process any valid pairs from the same group? No — reject whole group on any error
            for row_no, old_acct, _ in valid_pairs:
                result_rows.append(BulkRemapRow(row=row_no, old_gl_number=old_acct.gl_number, new_gl_number=new_gl, status="error",
                                                reason="Group rejected due to other errors in the same new-GL group."))
                error_count += 1
            continue

        # Check account_type consistency — normalize PL/BS to SOCI/SOFP before comparing
        _nt = {"PL": "SOCI", "BS": "SOFP", "P/L": "SOCI", "P&L": "SOCI",
               "B/S": "SOFP", "Profit & Loss": "SOCI", "Balance Sheet": "SOFP"}
        def _n(t: str) -> str: return _nt.get(t, t)
        norm_types = {_n(old_acct.account_type) for _, old_acct, _ in valid_pairs}
        norm_types.add(_n(new_acct.account_type))
        if len(norm_types) > 1:
            raw_types = {old_acct.account_type for _, old_acct, _ in valid_pairs} | {new_acct.account_type}
            for row_no, old_acct, _ in valid_pairs:
                result_rows.append(BulkRemapRow(row=row_no, old_gl_number=old_acct.gl_number, new_gl_number=new_gl, status="error",
                                                reason=f"account_type mismatch within group: {list(raw_types)}"))
                error_count += 1
            continue

        # Apply the group
        from app.models.auth import AuditLog
        retired_meta = []
        for row_no, old_acct, reason in valid_pairs:
            old_acct.is_active = False
            old_acct.is_retired = True
            db.add(GlCodeRemap(
                tenant_id=tenant_id,
                old_account_id=old_acct.id,
                new_account_id=new_acct.id,
                remapped_by=current_user.user_id,
                reason=reason,
            ))
            result_rows.append(BulkRemapRow(row=row_no, old_gl_number=old_acct.gl_number, new_gl_number=new_gl, status="remapped"))
            remapped_count += 1
            retired_meta.append({"gl_number": old_acct.gl_number})

        db.add(AuditLog(
            event_type="coa.remap.bulk_group",
            user_id=current_user.user_id,
            tenant_id=tenant_id,
            log_metadata={"new_gl_number": new_gl, "retired": retired_meta},
        ))

    await db.flush()
    return BulkRemapResult(remapped=remapped_count, errors=error_count, rows=result_rows)


@router.get("/coa/remap-history", response_model=list[GlRemapHistoryEntry])
async def list_remap_history(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[GlRemapHistoryEntry]:
    """Return all GL code remap history for this tenant, newest first."""
    tenant_id = _require_tenant(current_user)
    from sqlalchemy.orm import selectinload
    res = await db.execute(
        select(GlCodeRemap)
        .where(GlCodeRemap.tenant_id == tenant_id)
        .options(
            selectinload(GlCodeRemap.old_account),
            selectinload(GlCodeRemap.new_account),
        )
        .order_by(GlCodeRemap.remapped_at.desc())
    )
    entries = []
    for r in res.scalars().all():
        entries.append(GlRemapHistoryEntry(
            id=str(r.id),
            old_gl_number=r.old_account.gl_number,
            old_gl_name=r.old_account.gl_name,
            new_gl_number=r.new_account.gl_number,
            new_gl_name=r.new_account.gl_name,
            remapped_at=r.remapped_at,
            reason=r.reason,
        ))
    return entries


@router.post("/dimensions/{dimension_id}/values/bulk-action", response_model=BulkActionResult)
async def bulk_action_dimension_values(
    dimension_id: uuid.UUID,
    data: BulkActionRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> BulkActionResult:
    """Bulk activate / deactivate / delete dimension values by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    await _get_dimension_or_404(dimension_id, tenant_id, db)

    result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.id.in_(data.ids),
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    records = list(result.scalars().all())
    affected = 0

    for rec in records:
        if data.action == "activate":
            rec.is_active = True
            affected += 1
        elif data.action == "deactivate":
            rec.is_active = False
            affected += 1
        elif data.action == "delete":
            await db.delete(rec)
            affected += 1

    await db.flush()
    return BulkActionResult(action=data.action, affected=affected, skipped=len(data.ids) - affected)


@router.post("/dimensions/{dimension_id}/values/bulk-deactivate")
async def bulk_deactivate_dimension_values(
    dimension_id: uuid.UUID,
    payload: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Deactivate dimension values by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    ids = [uuid.UUID(str(i)) for i in payload.get("ids", [])]
    await db.execute(
        update(DimensionValue)
        .where(
            DimensionValue.id.in_(ids),
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
        .values(is_active=False)
    )
    await db.flush()
    return {"updated": len(ids)}


@router.post("/dimensions/{dimension_id}/values/bulk-reactivate")
async def bulk_reactivate_dimension_values(
    dimension_id: uuid.UUID,
    payload: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Reactivate dimension values by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    ids = [uuid.UUID(str(i)) for i in payload.get("ids", [])]
    await db.execute(
        update(DimensionValue)
        .where(
            DimensionValue.id.in_(ids),
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
        .values(is_active=True)
    )
    await db.flush()
    return {"updated": len(ids)}


@router.post("/dimensions/{dimension_id}/values/bulk-delete")
async def bulk_delete_dimension_values(
    dimension_id: uuid.UUID,
    payload: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Hard-delete dimension values by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    ids = [uuid.UUID(str(i)) for i in payload.get("ids", [])]
    result = await db.execute(
        select(DimensionValue).where(
            DimensionValue.id.in_(ids),
            DimensionValue.dimension_id == dimension_id,
            DimensionValue.tenant_id == tenant_id,
        )
    )
    records = list(result.scalars().all())
    for rec in records:
        await db.delete(rec)
    await db.flush()
    return {"deleted": len(records)}


@router.post("/categories/bulk-action", response_model=BulkActionResult)
async def bulk_action_categories(
    data: BulkActionRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> BulkActionResult:
    """Bulk activate / deactivate / delete expense categories by ID list. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(ExpenseCategory).where(
            ExpenseCategory.id.in_(data.ids),
            ExpenseCategory.tenant_id == tenant_id,
        )
    )
    records = list(result.scalars().all())
    affected = 0

    for rec in records:
        if data.action == "activate":
            rec.is_active = True
            affected += 1
        elif data.action == "deactivate":
            rec.is_active = False
            affected += 1
        elif data.action == "delete":
            # Cascade to subcategories
            sub_result = await db.execute(
                select(ExpenseCategory).where(
                    ExpenseCategory.parent_id == rec.id,
                    ExpenseCategory.tenant_id == tenant_id,
                )
            )
            for sub in sub_result.scalars().all():
                await db.delete(sub)
            await db.delete(rec)
            affected += 1

    await db.flush()
    return BulkActionResult(action=data.action, affected=affected, skipped=len(data.ids) - affected)


@router.get("/coa/{gl_id}/dimensions")
async def get_coa_dimensions(
    gl_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """Return saved dimension requirements for a GL account."""
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(GLDimensionRequirement).where(
            GLDimensionRequirement.gl_id == gl_id,
            GLDimensionRequirement.tenant_id == tenant_id,
        )
    )
    reqs = result.scalars().all()
    return [
        {
            "dimension_id": str(r.dimension_id),
            "requirement": r.requirement,
        }
        for r in reqs
    ]


@router.patch("/coa/{gl_id}/dimensions", response_model=CoAResponse)
async def set_gl_dimensions(
    gl_id: uuid.UUID,
    data: CoADimensionsUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CoAResponse:
    """
    Set dimension requirements for a GL account.

    Replaces all existing requirements for this GL account with the provided list.
    Admin only.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(ChartOfAccount)
        .where(ChartOfAccount.id == gl_id, ChartOfAccount.tenant_id == tenant_id)
        .options(selectinload(ChartOfAccount.dimension_requirements).selectinload(GLDimensionRequirement.dimension))
    )
    gl = result.scalar_one_or_none()
    if not gl:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="GL account not found.")

    # Delete existing requirements
    await db.execute(
        delete(GLDimensionRequirement).where(GLDimensionRequirement.gl_id == gl_id)
    )
    await db.flush()

    for req_item in data.requirements:
        db.add(GLDimensionRequirement(
            tenant_id=tenant_id,
            gl_id=gl_id,
            dimension_id=req_item.dimension_id,
            requirement=req_item.requirement,
        ))

    await db.flush()

    # Reload with fresh requirements
    result2 = await db.execute(
        select(ChartOfAccount)
        .where(ChartOfAccount.id == gl_id)
        .options(selectinload(ChartOfAccount.dimension_requirements).selectinload(GLDimensionRequirement.dimension))
    )
    gl2 = result2.scalar_one()
    return CoAResponse.from_orm(gl2, gl2.dimension_requirements)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPENSE CATEGORIES (M8 config router — with GL mappings)
# ═══════════════════════════════════════════════════════════════════════════════

async def _get_category_or_404(
    category_id: uuid.UUID, tenant_id: uuid.UUID, db: AsyncSession
) -> ExpenseCategory:
    result = await db.execute(
        select(ExpenseCategory).where(
            ExpenseCategory.id == category_id,
            ExpenseCategory.tenant_id == tenant_id,
        )
    )
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Category not found.")
    return cat


async def _load_category_tree(tenant_id: uuid.UUID, db: AsyncSession) -> list[CategoryWithMappingsResponse]:
    """Load all active top-level categories with subcategories and GL mappings."""
    top_result = await db.execute(
        select(ExpenseCategory)
        .where(
            ExpenseCategory.tenant_id == tenant_id,
            ExpenseCategory.is_active == True,  # noqa: E712
            ExpenseCategory.parent_id.is_(None),
        )
        .order_by(ExpenseCategory.sort_order, ExpenseCategory.name)
    )
    top_cats = list(top_result.scalars().all())

    items = []
    for cat in top_cats:
        # Load subcategories
        sub_result = await db.execute(
            select(ExpenseCategory)
            .where(
                ExpenseCategory.tenant_id == tenant_id,
                ExpenseCategory.parent_id == cat.id,
                ExpenseCategory.is_active == True,  # noqa: E712
            )
            .order_by(ExpenseCategory.sort_order, ExpenseCategory.name)
        )
        subcats = list(sub_result.scalars().all())

        # Load GL mappings for subcategories
        enriched_subs = []
        for sub in subcats:
            mapping_result = await db.execute(
                select(CategoryGLMapping)
                .where(CategoryGLMapping.category_id == sub.id)
                .options(selectinload(CategoryGLMapping.gl_account))
            )
            mappings = list(mapping_result.scalars().all())
            enriched_subs.append(CategoryWithMappingsResponse.from_orm(sub, mappings=mappings))

        items.append(CategoryWithMappingsResponse(
            id=str(cat.id),
            tenant_id=str(cat.tenant_id),
            name=cat.name,
            code=cat.code,
            parent_id=None,
            is_active=cat.is_active,
            sort_order=cat.sort_order,
            created_at=cat.created_at,
            subcategories=enriched_subs,
        ))

    return items


@router.get("/categories", response_model=list[CategoryWithMappingsResponse])
async def list_categories(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[CategoryWithMappingsResponse]:
    """List full category tree with subcategories and GL mappings."""
    tenant_id = _require_tenant(current_user)
    return await _load_category_tree(tenant_id, db)


@router.post("/categories", response_model=CategoryWithMappingsResponse, status_code=status.HTTP_201_CREATED)
async def create_category(
    data: CategoryCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CategoryWithMappingsResponse:
    """Create a top-level category or subcategory. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    code = data.code or _generate_code(data.name)

    if data.parent_id is not None:
        parent = await _get_category_or_404(data.parent_id, tenant_id, db)
        if not parent.is_active:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Parent category is inactive.")
        if parent.parent_id is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only one level of subcategories is supported.",
            )

    cat = ExpenseCategory(
        tenant_id=tenant_id,
        name=data.name,
        code=code,
        parent_id=data.parent_id,
        sort_order=data.sort_order,
    )
    db.add(cat)
    await db.flush()
    await db.refresh(cat)
    return CategoryWithMappingsResponse.from_orm(cat)


@router.patch("/categories/{category_id}", response_model=CategoryWithMappingsResponse)
async def update_category(
    category_id: uuid.UUID,
    data: CategoryUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CategoryWithMappingsResponse:
    """Update a category (PATCH semantics). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    cat = await _get_category_or_404(category_id, tenant_id, db)

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(cat, field, value)

    await db.flush()
    await db.refresh(cat)
    return CategoryWithMappingsResponse.from_orm(cat)


@router.delete("/categories/{category_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_category(
    category_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete a category and all its subcategories. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    cat = await _get_category_or_404(category_id, tenant_id, db)

    cat.is_active = False

    # Cascade to subcategories
    sub_result = await db.execute(
        select(ExpenseCategory).where(
            ExpenseCategory.parent_id == category_id,
            ExpenseCategory.tenant_id == tenant_id,
        )
    )
    for sub in sub_result.scalars().all():
        sub.is_active = False

    await db.flush()


# ── Category GL Mappings ──────────────────────────────────────────────────────

@router.post(
    "/categories/{category_id}/gl-mappings",
    response_model=CategoryGLMappingResponse,
    status_code=status.HTTP_201_CREATED,
)
async def add_gl_mapping(
    category_id: uuid.UUID,
    data: CategoryGLMappingCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CategoryGLMappingResponse:
    """
    Add a GL account mapping to a subcategory. Admin only.

    Only active GL accounts from the tenant's CoA can be mapped.
    If is_default=True, clears any existing default mapping for this category first.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    cat = await _get_category_or_404(category_id, tenant_id, db)

    # Verify GL account exists and is active in this tenant's CoA
    gl_result = await db.execute(
        select(ChartOfAccount).where(
            ChartOfAccount.id == data.gl_id,
            ChartOfAccount.tenant_id == tenant_id,
            ChartOfAccount.is_active == True,  # noqa: E712
        )
    )
    gl = gl_result.scalar_one_or_none()
    if not gl:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="GL account not found or inactive. Upload your Chart of Accounts first.",
        )

    # Check not already mapped
    existing = await db.execute(
        select(CategoryGLMapping).where(
            CategoryGLMapping.category_id == category_id,
            CategoryGLMapping.gl_id == data.gl_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This GL account is already mapped to this category.",
        )

    # If setting as default, clear existing default
    if data.is_default:
        existing_defaults_result = await db.execute(
            select(CategoryGLMapping).where(
                CategoryGLMapping.category_id == category_id,
                CategoryGLMapping.is_default == True,  # noqa: E712
            )
        )
        for m in existing_defaults_result.scalars().all():
            m.is_default = False

    mapping = CategoryGLMapping(
        tenant_id=tenant_id,
        category_id=category_id,
        gl_id=data.gl_id,
        is_default=data.is_default,
    )
    db.add(mapping)
    await db.flush()
    await db.refresh(mapping)

    # Load gl_account for response
    mapping_result = await db.execute(
        select(CategoryGLMapping)
        .where(CategoryGLMapping.id == mapping.id)
        .options(selectinload(CategoryGLMapping.gl_account))
    )
    mapping_loaded = mapping_result.scalar_one()
    return CategoryGLMappingResponse.from_orm(mapping_loaded)


@router.delete(
    "/categories/{category_id}/gl-mappings/{gl_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def remove_gl_mapping(
    category_id: uuid.UUID,
    gl_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Remove a GL account mapping from a category. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(CategoryGLMapping).where(
            CategoryGLMapping.category_id == category_id,
            CategoryGLMapping.gl_id == gl_id,
            CategoryGLMapping.tenant_id == tenant_id,
        )
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="GL mapping not found.")

    await db.delete(mapping)
    await db.flush()


# ═══════════════════════════════════════════════════════════════════════════════
# GL SEARCH (M9 — expense form Level 3/4 GL picker)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/gl/search", response_model=list[GLSearchResult])
async def search_gl_accounts(
    q: str = Query(default="", description="Search term matched against GL number or name"),
    limit: int = Query(default=20, ge=1, le=100),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[GLSearchResult]:
    """
    Search active GL accounts by number or name for the expense form GL picker.

    Returns accounts with their per-dimension requirements so the form can render
    the correct dimension dropdowns immediately after GL selection.
    Tenant-scoped; available to all authenticated business users (not admin-only).
    """
    tenant_id = _require_tenant(current_user)

    query = (
        select(ChartOfAccount)
        .where(
            ChartOfAccount.tenant_id == tenant_id,
            ChartOfAccount.is_active == True,  # noqa: E712
        )
        .options(selectinload(ChartOfAccount.dimension_requirements))
        .order_by(ChartOfAccount.gl_number)
        .limit(limit)
    )

    if q.strip():
        term = f"%{q.strip()}%"
        query = query.where(
            ChartOfAccount.gl_number.ilike(term) | ChartOfAccount.gl_name.ilike(term)
        )

    result = await db.execute(query)
    return [GLSearchResult.from_orm(g) for g in result.scalars().all()]


@router.patch(
    "/categories/{category_id}/gl-mappings/{gl_id}",
    response_model=CategoryGLMappingResponse,
)
async def update_gl_mapping(
    category_id: uuid.UUID,
    gl_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CategoryGLMappingResponse:
    """
    Toggle is_default on a GL mapping.

    If setting this mapping as default, clears any existing default first.
    Admin only.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(CategoryGLMapping)
        .where(
            CategoryGLMapping.category_id == category_id,
            CategoryGLMapping.gl_id == gl_id,
            CategoryGLMapping.tenant_id == tenant_id,
        )
        .options(selectinload(CategoryGLMapping.gl_account))
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="GL mapping not found.")

    if not mapping.is_default:
        # Clear existing defaults
        defaults_result = await db.execute(
            select(CategoryGLMapping).where(
                CategoryGLMapping.category_id == category_id,
                CategoryGLMapping.is_default == True,  # noqa: E712
            )
        )
        for m in defaults_result.scalars().all():
            m.is_default = False
        mapping.is_default = True
    else:
        mapping.is_default = False

    await db.flush()
    return CategoryGLMappingResponse.from_orm(mapping)
