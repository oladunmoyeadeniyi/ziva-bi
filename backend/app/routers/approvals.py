"""
ZivaBI — approval workflow router (Milestones 4–5 + Approval Engine).

Implements the full expense approval chain with audit trail, snapshots,
refer-back enhancements, separation of duties, full email coverage, and the
new configurable routing engine (org_tree, requestor_selects, direct_to_hod).

Endpoints:
    POST   /api/approvals/matrix                                Create or update approval matrix (legacy)
    GET    /api/approvals/matrix                                Get current tenant's matrix (legacy)
    GET    /api/approvals/roles                                 List approver roles (auto-seeds defaults)
    POST   /api/approvals/roles                                 Create an approver role
    PATCH  /api/approvals/roles/{role_id}                       Update an approver role
    DELETE /api/approvals/roles                                 Delete ALL approver roles for tenant (single transaction)
    DELETE /api/approvals/roles/{role_id}                       Delete an approver role
    POST   /api/approvals/roles/bulk-upload                     Bulk-upload roles from Excel/CSV template
    GET    /api/approvals/policies                              List all policies for tenant
    POST   /api/approvals/policies                              Create or replace a module policy
    PATCH  /api/approvals/policies/{policy_id}                  Partial update a policy
    DELETE /api/approvals/policies/{policy_id}                  Delete a policy
    GET    /api/approvals/policies/{module}/chain-preview        Preview computed chain for a module
    GET    /api/approvals/delegations                           List my delegations
    POST   /api/approvals/delegations                           Create a delegation
    PATCH  /api/approvals/delegations/{delegation_id}           Update / revoke a delegation
    POST   /api/approvals/reports/{report_id}/submit            Submit report for approval
    GET    /api/approvals/queue                                  Reports pending current user's action
    GET    /api/approvals/rejected                               Reports rejected involving current user
    GET    /api/approvals/reports/{report_id}/audit-log          Chronological event trail
    GET    /api/approvals/reports/{report_id}/snapshot/{version} Snapshot of lines at submission
    GET    /api/approvals/reports/{report_id}                   All approval records for a report
    POST   /api/approvals/{approval_id}/approve                 Approve at current level
    POST   /api/approvals/{approval_id}/reject                  Reject with comment
    POST   /api/approvals/{approval_id}/refer-back              Refer back to lower approver or requestor
"""

import io
import logging
import smtplib
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from email.mime.text import MIMEText

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.database import get_db
from app.middleware.auth import CurrentUser, require_auth, block_if_readonly_impersonation
from app.models.approvals import (
    ApprovalDelegation,
    ApprovalMatrix,
    ApprovalPolicy,
    ApprovalRole,
    ApprovalRoleScope,
    ApprovalRoleThreshold,
    ExpenseApproval,
    FinanceReviewStep,
)
from app.models.setup import OrgStructureNode
from app.models.auth import AuditLog, User, UserTenant
from app.models.expenses import ExpenseReport, ExpenseReportSnapshot
from app.schemas.approvals import (
    EntityOption,
    ApprovalDelegationCreate,
    ApprovalDelegationResponse,
    ApprovalDelegationUpdate,
    ApprovalMatrixCreate,
    ApprovalMatrixResponse,
    ApprovalPolicyCreate,
    ApprovalPolicyResponse,
    ApprovalPolicyUpdate,
    ApprovalQueueItem,
    ApprovalRecordResponse,
    ApprovalRoleCreate,
    ApprovalRoleResponse,
    ApprovalRoleUpdate,
    ApproveRequest,
    AuditLogEntry,
    ChainPreviewStep,
    ReferBackRequest,
    RejectRequest,
    RoleBulkUploadResult,
    RoleScopeResponse,
    RoleScopeUpdate,
    SnapshotResponse,
    SubmitWithApproversRequest,
    FinanceReviewStepBulkSave,
    FinanceReviewStepResponse,
)
from app.schemas.expenses import ExpenseReportResponse
from app.services.account_determination import AccountMappingError
from app.services.approval_routing import (
    ApprovalChainHoldError,
    ApprovalRoutingError,
    compute_chain,
    get_policy,
    preview_chain,
)
from app.services.expense_posting import ExpensePostingError, PostingResult, post_expense_to_gl
from app.services.gl_posting import PostingError
from app.services.periods import is_date_postable

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/approvals", tags=["approvals"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _require_tenant(current_user: CurrentUser) -> uuid.UUID:
    """Raise 403 if the current user has no tenant (individual account)."""
    if current_user.tenant_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Approval workflow is a business-tier feature.",
        )
    return current_user.tenant_id


def _require_admin(current_user: CurrentUser) -> None:
    """Raise 403 if the current user is not a tenant admin or super admin."""
    if not current_user.is_tenant_admin and not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tenant admin access is required.",
        )
    block_if_readonly_impersonation(current_user)


async def _get_matrix(tenant_id: uuid.UUID, db: AsyncSession) -> ApprovalMatrix | None:
    """Fetch the approval matrix for a tenant, returning None if not configured."""
    result = await db.execute(
        select(ApprovalMatrix).where(ApprovalMatrix.tenant_id == tenant_id)
    )
    return result.scalar_one_or_none()


async def _get_report_or_404(
    report_id: uuid.UUID,
    tenant_id: uuid.UUID,
    db: AsyncSession,
) -> ExpenseReport:
    """Fetch an expense report by ID scoped to tenant, raising 404 if not found."""
    result = await db.execute(
        select(ExpenseReport)
        .where(ExpenseReport.id == report_id, ExpenseReport.tenant_id == tenant_id)
        .options(selectinload(ExpenseReport.lines))
    )
    report = result.scalar_one_or_none()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found.")
    return report


async def _reload_report(report_id: uuid.UUID, db: AsyncSession) -> ExpenseReport:
    """Re-fetch a report with lines after a mutation."""
    result = await db.execute(
        select(ExpenseReport)
        .where(ExpenseReport.id == report_id)
        .options(selectinload(ExpenseReport.lines))
        .execution_options(populate_existing=True)
    )
    return result.scalar_one()


async def _validate_approver(
    approver_id: uuid.UUID,
    tenant_id: uuid.UUID,
    db: AsyncSession,
) -> User:
    """Validate that the approver belongs to the same tenant."""
    result = await db.execute(
        select(User)
        .join(UserTenant, User.id == UserTenant.user_id)
        .where(
            User.id == approver_id,
            UserTenant.tenant_id == tenant_id,
            UserTenant.is_active.is_(True),
            User.is_active.is_(True),
        )
    )
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Approver {approver_id} not found in your company.",
        )
    return user


def _role_label_for_level(matrix: ApprovalMatrix, level: int) -> str:
    """Return the configured role label for a given approval level."""
    if level == 1:
        return matrix.level1_role
    if level == 2:
        return matrix.level2_role or f"Level {level}"
    return matrix.level3_role or f"Level {level}"


async def _write_audit_log(
    db: AsyncSession,
    event_type: str,
    user_id: uuid.UUID,
    tenant_id: uuid.UUID,
    metadata: dict,
) -> None:
    """Append an immutable audit log entry. Never raises — failures are logged only."""
    try:
        db.add(AuditLog(
            event_type=event_type,
            user_id=user_id,
            tenant_id=tenant_id,
            log_metadata=metadata,
        ))
    except Exception as exc:
        logger.error("Failed to write audit log %s: %s", event_type, exc)


async def _write_snapshot(
    report: ExpenseReport,
    tenant_id: uuid.UUID,
    db: AsyncSession,
) -> int:
    """
    Write an immutable snapshot of the report's current lines and header.

    Returns the snapshot version number (1-based, increments per resubmission).
    Must be called BEFORE the report status is changed so the lines are still
    in their pre-submission state.
    """
    version_result = await db.execute(
        select(func.count(ExpenseReportSnapshot.id)).where(
            ExpenseReportSnapshot.report_id == report.id
        )
    )
    version = (version_result.scalar_one() or 0) + 1

    lines_data = [
        {
            # Legacy text fields
            "line_number": ln.line_number,
            "gl_account": ln.gl_account,
            "pl_group": ln.pl_group,
            "io_dimension": ln.io_dimension,
            "cost_center": ln.cost_center,
            "location": ln.location,
            "invoice_date": str(ln.invoice_date) if ln.invoice_date else None,
            "invoice_number": ln.invoice_number,
            "description": ln.description,
            "amount": str(ln.amount),
            # M9 structured fields — must be captured for a legally complete audit record.
            # Capturing at submission time ensures the snapshot reflects the GL/dimension
            # coding as it existed when submitted, even if accounts are later remapped/retired.
            "gl_id": str(ln.gl_id) if ln.gl_id else None,
            "dimension_values": ln.dimension_values,
            "is_split_parent": ln.is_split_parent,
            "split_parent_id": str(ln.split_parent_id) if ln.split_parent_id else None,
            "flag_incorrect": ln.flag_incorrect,
            "flag_comment": ln.flag_comment,
        }
        for ln in (report.lines or [])
    ]

    db.add(ExpenseReportSnapshot(
        report_id=report.id,
        tenant_id=tenant_id,
        snapshot_data={
            "report_number": report.report_number,
            "employee_id": str(report.employee_id),
            "report_date": str(report.report_date),
            "currency": report.currency,
            "total_amount": str(report.total_amount),
            "lines": lines_data,
        },
        submitted_at=datetime.now(timezone.utc),
        version=version,
    ))
    return version


def _send_rejection_email(
    to_email: str,
    report_number: str,
    report_date: str,
    total_amount: Decimal,
    rejection_comment: str,
) -> None:
    """Send rejection notification; falls back to console log if SMTP not configured."""
    subject = f"Expense Report {report_number} Rejected"
    body = (
        f"Your expense report {report_number} dated {report_date} "
        f"for ₦{total_amount:,.2f} has been rejected.\n\n"
        f"Reason: {rejection_comment}\n\n"
        f"Please log in to Ziva BI to review and resubmit."
    )
    _smtp_send(to_email, subject, body)


def _send_approver_notification_email(
    to_email: str,
    report_number: str,
    report_date: str,
    total_amount: Decimal,
    employee_name: str,
    role_label: str,
) -> None:
    """Notify an approver that a report is awaiting their action."""
    subject = f"Action Required: Expense Report {report_number} awaiting your approval"
    body = (
        f"{employee_name} has submitted expense report {report_number} "
        f"dated {report_date} for ₦{total_amount:,.2f} requiring your "
        f"approval as {role_label}.\n\n"
        f"Please log in to Ziva BI to review and action."
    )
    _smtp_send(to_email, subject, body)


def _send_approval_complete_email(
    to_email: str,
    report_number: str,
    report_date: str,
    total_amount: Decimal,
) -> None:
    """Notify the requestor that their report has been fully approved."""
    subject = f"Approved: Expense Report {report_number}"
    body = (
        f"Your expense report {report_number} dated {report_date} "
        f"for ₦{total_amount:,.2f} has been fully approved.\n\n"
        f"Please log in to Ziva BI to view the approved report."
    )
    _smtp_send(to_email, subject, body)


def _send_refer_back_email(
    to_email: str,
    report_number: str,
    comment: str,
    referring_level: int,
) -> None:
    """Notify requestor that their report was referred back to them."""
    subject = f"Query on Expense Report {report_number}"
    body = (
        f"There is a query on your expense report {report_number}.\n\n"
        f"Query: {comment}\n\n"
        f"Please log in to Ziva BI to view the details."
    )
    _smtp_send(to_email, subject, body)


def _send_referred_approver_email(
    to_email: str,
    report_number: str,
    referring_approver_name: str,
    referring_level: int,
    comment: str,
) -> None:
    """Notify a lower approver that a report has been referred to them for consultation."""
    subject = f"Referred to you: Expense Report {report_number}"
    body = (
        f"Expense report {report_number} has been referred to you by "
        f"{referring_approver_name} (Level {referring_level}) for review.\n\n"
        f"Query: {comment}\n\n"
        f"Please log in to Ziva BI to respond."
    )
    _smtp_send(to_email, subject, body)


def _smtp_send(to_email: str, subject: str, body: str) -> None:
    """Shared SMTP send; logs to console when SMTP credentials are not configured."""
    if not all([settings.smtp_host, settings.smtp_user, settings.smtp_password]):
        logger.info(
            "[EMAIL SIMULATION]\nTo: %s\nSubject: %s\n\n%s",
            to_email, subject, body,
        )
        return
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = settings.smtp_from_email or settings.smtp_user
    msg["To"] = to_email
    try:
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as smtp:
            smtp.starttls()
            smtp.login(settings.smtp_user, settings.smtp_password)
            smtp.send_message(msg)
    except Exception as exc:
        logger.warning("Failed to send email to %s: %s", to_email, exc)


# ── Approval Roles ────────────────────────────────────────────────────────────

DEFAULT_ROLES = [
    ("Line Manager",      "Direct line manager of the employee submitting the transaction.", 0),
    ("Department Head",   "Head of the submitting employee's department or cost center.",   1),
    ("Finance Director",  "Approves from a finance control perspective.",                   2),
    ("CFO",               "Chief Financial Officer — final sign-off for high-value items.", 3),
]


@router.get("/roles", response_model=list[ApprovalRoleResponse])
async def list_approval_roles(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalRoleResponse]:
    """Return all approver roles for the current tenant with occupant employees."""
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    from app.models.master_data import Employee
    from app.schemas.approvals import RoleOccupant
    rows = (await db.execute(
        select(ApprovalRole)
        .options(
            selectinload(ApprovalRole.cost_center),
            selectinload(ApprovalRole.entity_node),
            selectinload(ApprovalRole.parent_role),
        )
        .where(ApprovalRole.tenant_id == current_user.tenant_id)
        .order_by(ApprovalRole.display_order, ApprovalRole.name)
    )).scalars().all()

    # Fetch all employees for this tenant that have an approval_role_id
    role_ids = [r.id for r in rows]
    occupant_rows = (await db.execute(
        select(Employee).where(
            Employee.tenant_id == current_user.tenant_id,
            Employee.approval_role_id.in_(role_ids),
            Employee.is_active.is_(True),
        )
    )).scalars().all()

    # Group occupants by role_id
    from collections import defaultdict
    occ_by_role: dict = defaultdict(list)
    for emp in occupant_rows:
        full_name = f"{emp.first_name} {emp.last_name}".strip()
        parts = full_name.split()
        initials = "".join(p[0].upper() for p in parts if p)[:2]
        occ_by_role[str(emp.approval_role_id)].append(
            RoleOccupant(
                id=str(emp.id),
                full_name=full_name,
                initials=initials,
                employee_code=emp.employee_code,
            )
        )

    return [ApprovalRoleResponse.from_orm(r, occ_by_role.get(str(r.id), [])) for r in rows]


@router.post("/roles", response_model=ApprovalRoleResponse, status_code=201)
async def create_approval_role(
    data: ApprovalRoleCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalRoleResponse:
    """Create a new approver role."""
    block_if_readonly_impersonation(current_user)
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    # Duplicate check: same role = same name + same cost centre + same entity +
    # same area + same sub_area + same employment_type.  NULL fields are treated
    # as equal to NULL (IS NOT DISTINCT FROM semantics via conditional filter).
    def _eq_or_null(col, val):  # type: ignore[no-untyped-def]
        return col.is_(None) if val is None else col == val

    emp_val = data.employment_type or "permanent"
    existing = (await db.execute(
        select(ApprovalRole).where(
            ApprovalRole.tenant_id == current_user.tenant_id,
            ApprovalRole.name == data.name.strip(),
            _eq_or_null(ApprovalRole.cost_center_id, data.cost_center_id),
            _eq_or_null(ApprovalRole.entity_node_id, data.entity_node_id),
            _eq_or_null(ApprovalRole.area, data.area),
            _eq_or_null(ApprovalRole.sub_area, data.sub_area),
            ApprovalRole.employment_type == emp_val,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=(
                "A role with this exact combination of name, cost centre, area, "
                "sub area, and employment type already exists."
            ),
        )
    role = ApprovalRole(
        tenant_id=current_user.tenant_id,
        name=data.name.strip(),
        description=data.description,
        display_order=data.display_order,
        parent_role_id=data.parent_role_id,
        cost_center_id=data.cost_center_id,
        entity_node_id=data.entity_node_id,
        max_occupants=data.max_occupants,
        designation=data.designation,
        area=data.area,
        sub_area=data.sub_area,
        employment_type=data.employment_type or "permanent",
        code=data.code,
        grade=data.grade,
    )
    db.add(role)
    await db.commit()
    await db.refresh(role)
    # reload with relationships eagerly so from_orm can read all names
    role = (await db.execute(
        select(ApprovalRole)
        .options(
            selectinload(ApprovalRole.cost_center),
            selectinload(ApprovalRole.entity_node),
            selectinload(ApprovalRole.parent_role),
        )
        .where(ApprovalRole.id == role.id)
    )).scalar_one()
    return ApprovalRoleResponse.from_orm(role)


@router.patch("/roles/{role_id}", response_model=ApprovalRoleResponse)
async def update_approval_role(
    role_id: uuid.UUID,
    data: ApprovalRoleUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalRoleResponse:
    """Update an existing approver role."""
    block_if_readonly_impersonation(current_user)
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    role = (await db.execute(
        select(ApprovalRole).where(
            ApprovalRole.id == role_id,
            ApprovalRole.tenant_id == current_user.tenant_id,
        )
    )).scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found.")
    if data.name is not None:
        role.name = data.name.strip()
    if data.description is not None:
        role.description = data.description
    if data.display_order is not None:
        role.display_order = data.display_order
    if data.is_active is not None:
        role.is_active = data.is_active
    if "parent_role_id" in data.model_fields_set:
        role.parent_role_id = data.parent_role_id  # allows clearing to None (drag-to-root)
    if "max_occupants" in data.model_fields_set:
        role.max_occupants = data.max_occupants  # allows setting to None (unlimited)
    if "cost_center_id" in data.model_fields_set:
        role.cost_center_id = data.cost_center_id  # allows clearing to None
    if "entity_node_id" in data.model_fields_set:
        role.entity_node_id = data.entity_node_id  # allows clearing to None
    if "designation" in data.model_fields_set:
        role.designation = data.designation
    if "area" in data.model_fields_set:
        role.area = data.area
    if "sub_area" in data.model_fields_set:
        role.sub_area = data.sub_area
    if "employment_type" in data.model_fields_set:
        role.employment_type = data.employment_type or "permanent"
    if data.code is not None:
        role.code = data.code
    if data.grade is not None:
        role.grade = data.grade
    await db.commit()
    # reload with relationships eagerly
    role = (await db.execute(
        select(ApprovalRole)
        .options(
            selectinload(ApprovalRole.cost_center),
            selectinload(ApprovalRole.entity_node),
            selectinload(ApprovalRole.parent_role),
        )
        .where(ApprovalRole.id == role.id)
    )).scalar_one()
    return ApprovalRoleResponse.from_orm(role)


@router.delete("/roles", status_code=204)
async def clear_all_approval_roles(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """
    Delete ALL approval roles for the current tenant in a single transaction.

    All FK references to approval_roles (parent_role_id, employees.approval_role_id,
    employee_position_assignments.approval_role_id) are SET NULL on delete, so no
    ordering constraint applies. One round-trip; no partial-failure risk.
    """
    _require_admin(current_user)
    block_if_readonly_impersonation(current_user)
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    await db.execute(
        delete(ApprovalRole).where(ApprovalRole.tenant_id == current_user.tenant_id)
    )
    await db.commit()


@router.delete("/roles/{role_id}", status_code=204)
async def delete_approval_role(
    role_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Delete an approver role."""
    block_if_readonly_impersonation(current_user)
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    role = (await db.execute(
        select(ApprovalRole).where(
            ApprovalRole.id == role_id,
            ApprovalRole.tenant_id == current_user.tenant_id,
        )
    )).scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found.")
    await db.delete(role)
    await db.commit()


@router.patch("/roles/{role_id}/permission-tier", response_model=ApprovalRoleResponse)
async def set_role_permission_tier(
    role_id: uuid.UUID,
    body: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalRoleResponse:
    """
    Set or clear the permission_tier on an org role.

    Body: { "permission_tier": "power_admin" | "functional_admin" | null }

    Every employee who holds this role will inherit the tier at their next login
    (union with any directly-assigned tier on their UserTenant record).
    """
    block_if_readonly_impersonation(current_user)
    if not current_user.tenant_id:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    role = (await db.execute(
        select(ApprovalRole)
        .options(selectinload(ApprovalRole.cost_center), selectinload(ApprovalRole.entity_node))
        .where(
            ApprovalRole.id == role_id,
            ApprovalRole.tenant_id == current_user.tenant_id,
        )
    )).scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found.")

    tier = body.get("permission_tier")
    if tier not in (None, "power_admin", "functional_admin"):
        raise HTTPException(status_code=422, detail="permission_tier must be 'power_admin', 'functional_admin', or null.")
    role.permission_tier = tier
    await db.commit()
    await db.refresh(role)
    return ApprovalRoleResponse.from_orm(role)


@router.get("/roles/{role_id}/scope", response_model=RoleScopeResponse)
async def get_role_scope(
    role_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RoleScopeResponse:
    """Return the section scope configuration for an org role."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    rows = (await db.execute(
        select(ApprovalRoleScope).where(
            ApprovalRoleScope.tenant_id == tenant_id,
            ApprovalRoleScope.role_id == role_id,
        )
    )).scalars().all()

    return RoleScopeResponse(
        role_id=str(role_id),
        sections=[{"section": r.section, "access_level": r.access_level} for r in rows],
    )


@router.patch("/roles/{role_id}/scope", response_model=RoleScopeResponse)
async def patch_role_scope(
    role_id: uuid.UUID,
    data: RoleScopeUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RoleScopeResponse:
    """Replace all scope sections for an org role (full replace, not merge)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    await db.execute(
        delete(ApprovalRoleScope).where(
            ApprovalRoleScope.tenant_id == tenant_id,
            ApprovalRoleScope.role_id == role_id,
        )
    )

    for item in data.sections:
        db.add(ApprovalRoleScope(
            tenant_id=tenant_id,
            role_id=role_id,
            section=item.section,
            access_level=item.access_level,
        ))

    await db.commit()

    return RoleScopeResponse(
        role_id=str(role_id),
        sections=[{"section": i.section, "access_level": i.access_level} for i in data.sections],
    )


@router.get("/roles/template")
async def download_roles_template(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Download an Excel (.xlsx) template for bulk role upload.

    Includes data-validation dropdown lists for:
    - Entity Code  — derived from org_structure Legal entity nodes.
    - Cost Center  — derived from org_structure Cost center nodes.
    - Capacity     — fixed list: single / unlimited / 2..10.
    """
    import io as _io
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")
    except Exception as _exc:
        raise HTTPException(status_code=500, detail=f"Template init error: {_exc}")

    tenant_id = _require_tenant(current_user)

    # Fetch dropdown data + existing roles
    cc_nodes = (await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Cost center",
            OrgStructureNode.is_active.is_(True),
        ).order_by(OrgStructureNode.name)
    )).scalars().all()
    entity_nodes = (await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type.in_(["Legal entity", "Parent company"]),
            OrgStructureNode.is_active.is_(True),
        ).order_by(OrgStructureNode.name)
    )).scalars().all()
    existing_roles = (await db.execute(
        select(ApprovalRole).where(
            ApprovalRole.tenant_id == tenant_id,
            ApprovalRole.is_active.is_(True),
        ).order_by(ApprovalRole.display_order, ApprovalRole.name)
    )).scalars().all()

    cc_codes = [n.cost_center_code or n.code for n in cc_nodes if (n.cost_center_code or n.code)]
    entity_codes = [n.entity_code or n.code for n in entity_nodes if (n.entity_code or n.code)]

    # Build reverse lookups for pre-populating existing roles
    cc_id_to_code: dict = {n.id: (n.cost_center_code or n.code) for n in cc_nodes}
    ent_id_to_code: dict = {n.id: (n.entity_code or n.code) for n in entity_nodes}
    role_id_to_name: dict = {r.id: r.name for r in existing_roles}

    def _cap_label(max_occ: int | None) -> str:
        if max_occ is None: return "unlimited"
        if max_occ == 1:    return "single"
        return str(max_occ)

    def _desig_label(d: str | None) -> str:
        if d == "head_of_entity":          return "Head of Entity"
        if d == "head_of_department":      return "Head of Department"
        if d in ("manager", "section_head"): return "Manager"      # section_head = legacy
        if d == "team_lead":               return "Team Lead"
        if d in ("individual_contributor", "regular"): return "Individual Contributor"
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roles"

    # ── Hidden ref sheets ─────────────────────────────────────────────────────
    # ── Instructions sheet ───────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Instructions", 0)  # insert as first sheet
    guide_title_fill = PatternFill("solid", fgColor="1D4ED8")
    guide_title_font = Font(bold=True, color="FFFFFF", size=13)
    guide_head_fill  = PatternFill("solid", fgColor="1E40AF")
    guide_head_font  = Font(bold=True, color="FFFFFF", size=11)
    guide_rows = [
        ("ROLE HIERARCHY UPLOAD GUIDE", "", ""),
        ("", "", ""),
        ("Column", "Required?", "Description & Accepted Values"),
        ("A  Role Name",        "REQUIRED",  "Display name on the org chart. Must be unique per Cost Center + Area combination. Same name is allowed across different Areas (e.g. multiple Regional Managers)."),
        ("B  Parent Role",      "optional",  "Name of the role this role reports to. Leave blank only for the top-level role. If multiple roles share a parent name, fill Area (col H) to disambiguate."),
        ("C  Entity Code",      "REQUIRED",  "Legal entity code from your Org Structure. Select from dropdown. Example: ENT01"),
        ("D  Cost Center",      "REQUIRED",  "Cost center code from your Org Structure. Select from dropdown. Example: CC001"),
        ("E  Capacity",         "REQUIRED",  "How many people can hold this role. Values: single | unlimited | 2 to 10"),
        ("F  Designation",      "REQUIRED",  "Authority level. Values: Head of Entity | Head of Department | Manager | Team Lead | Individual Contributor. Controls org chart appearance and approval routing."),
        ("G  Employment Type",  "REQUIRED",  "Engagement type. Values: Permanent | Contract | Outsourced"),
        ("H  Area / Location",  "optional",  "Primary scope (geography, channel, segment, etc.). Required when multiple roles share the same name. Example: North | South | Key Accounts"),
        ("I  Sub Area",         "optional",  "Narrower scope within the Area. Used so subordinate roles can inherit the parent's Sub Area as their own Area. Example: Abuja Central"),
        ("J  Description",      "optional",  "Plain-English summary of the role's purpose. Shown on the role detail view."),
        ("", "", ""),
        ("DESIGNATION GUIDE", "", ""),
        ("Head of Entity",        "", "Top-level leader of the entire entity (GM, MD, CEO, Country Manager). Only one per entity."),
        ("Head of Department",    "", "Leads a major functional department (Sales Director, Finance Director)."),
        ("Manager",               "", "Heads a unit or sub-department within a department (National On-Premise Manager, Brand Manager)."),
        ("Team Lead",             "", "Leads a small team with direct reports but is not a department/unit head (Senior Analyst, Team Coordinator)."),
        ("Individual Contributor","", "No direct reports; executes role independently (Analyst, Coordinator, Specialist, Officer). Default when left blank."),
        ("", "", ""),
        ("AREA & SUB AREA PATTERN", "", ""),
        ("Example:", "", "Regional Manager  -->  Area = North,     Sub Area = Abuja"),
        ("",         "", "Account Officer   -->  Area = Abuja     (inherited from Regional Manager Sub Area)"),
        ("", "", ""),
        ("TIP:", "", "Fill in the Roles sheet. Do not rename or delete columns. * = required field."),
    ]
    ws_guide.column_dimensions["A"].width = 26
    ws_guide.column_dimensions["B"].width = 12
    ws_guide.column_dimensions["C"].width = 90
    for r_idx, (col_a, col_b, col_c) in enumerate(guide_rows, start=1):
        ws_guide.cell(row=r_idx, column=1, value=col_a)
        ws_guide.cell(row=r_idx, column=2, value=col_b)
        ws_guide.cell(row=r_idx, column=3, value=col_c)
        if r_idx == 1:
            c = ws_guide.cell(row=r_idx, column=1)
            c.fill = guide_title_fill; c.font = guide_title_font
            ws_guide.merge_cells(f"A1:C1")
        elif col_a in ("Column", "DESIGNATION GUIDE", "AREA & SUB AREA PATTERN"):
            for ci in range(1, 4):
                cc = ws_guide.cell(row=r_idx, column=ci)
                cc.fill = guide_head_fill; cc.font = guide_head_font
        # wrap col C
        ws_guide.cell(row=r_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws_guide.row_dimensions[1].height = 22
    # make Roles the active sheet
    wb.active = wb["Roles"]

    ws_cc = wb.create_sheet("_cc")
    for i, code in enumerate(cc_codes, start=1):
        ws_cc.cell(row=i, column=1, value=code)
    ws_cc.sheet_state = "hidden"

    ws_ent = wb.create_sheet("_ent")
    for i, code in enumerate(entity_codes, start=1):
        ws_ent.cell(row=i, column=1, value=code)
    ws_ent.sheet_state = "hidden"

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        "Role Name *",
        "Parent Role",
        "Entity Code *",
        "Cost Center *",
        "Capacity *",
        "Designation *",
        "Employment Type *",
        "Area / Location",
        "Sub Area",
        "Description",
    ]
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 26
    ws.column_dimensions["I"].width = 26
    ws.column_dimensions["J"].width = 36
    ws.freeze_panes = "A2"

    def _emp_label(et: str | None) -> str:
        mapping = {"contract": "Contract", "outsourced": "Outsourced"}
        return mapping.get(et or "", "Permanent")

    # ── Pre-populate with existing tenant roles ───────────────────────────────
    for row_idx, role in enumerate(existing_roles, start=2):
        parent_name = role_id_to_name.get(role.parent_role_id, "") if role.parent_role_id else ""
        entity_code = ent_id_to_code.get(role.entity_node_id, "") if role.entity_node_id else ""
        cc_code     = cc_id_to_code.get(role.cost_center_id, "") if role.cost_center_id else ""
        row_data = [
            role.name,
            parent_name,
            entity_code,
            cc_code,
            _cap_label(role.max_occupants),
            _desig_label(role.designation if hasattr(role, "designation") else None),
            _emp_label(role.employment_type if hasattr(role, "employment_type") else None),
            role.area if hasattr(role, "area") and role.area else "",
            role.sub_area if hasattr(role, "sub_area") and role.sub_area else "",
            role.description or "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # ── Data validation ───────────────────────────────────────────────────────
    max_row = 1000

    if entity_codes:
        dv_ent = DataValidation(
            type="list",
            formula1=f"_ent!$A$1:$A${len(entity_codes)}",
            allow_blank=True,
            showErrorMessage=True,
            error="Select from the list of valid entity codes.",
            errorTitle="Invalid Entity Code",
        )
        ws.add_data_validation(dv_ent)
        dv_ent.sqref = f"C2:C{max_row}"

    if cc_codes:
        dv_cc = DataValidation(
            type="list",
            formula1=f"_cc!$A$1:$A${len(cc_codes)}",
            allow_blank=True,
            showErrorMessage=True,
            error="Select from the list of valid cost center codes.",
            errorTitle="Invalid Cost Center",
        )
        ws.add_data_validation(dv_cc)
        dv_cc.sqref = f"D2:D{max_row}"

    dv_desig = DataValidation(
        type="list",
        formula1='"Head of Entity,Head of Department,Manager,Team Lead,Individual Contributor"',
        allow_blank=True,
        showErrorMessage=True,
        error="Select: Head of Entity, Head of Department, Manager, Team Lead, Individual Contributor, or leave blank.",
        errorTitle="Invalid Designation",
    )
    ws.add_data_validation(dv_desig)
    dv_desig.sqref = f"F2:F{max_row}"

    cap_list = '"single,unlimited,2,3,4,5,6,7,8,9,10"'
    dv_cap = DataValidation(
        type="list",
        formula1=cap_list,
        allow_blank=True,
        showErrorMessage=True,
        error="Enter: single, unlimited, or a number 2-10.",
        errorTitle="Invalid Capacity",
    )
    ws.add_data_validation(dv_cap)
    dv_cap.sqref = f"E2:E{max_row}"

    dv_emp = DataValidation(
        type="list",
        formula1='"Permanent,Contract,Outsourced"',
        allow_blank=True,
        showErrorMessage=True,
        error="Select: Permanent, Contract, or Outsourced.",
        errorTitle="Invalid Employment Type",
    )
    ws.add_data_validation(dv_emp)
    dv_emp.sqref = f"G2:G{max_row}"

    # ── Stream ────────────────────────────────────────────────────────────────
    try:
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as _exc:
        raise HTTPException(status_code=500, detail=f"Template generation error: {_exc}")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=roles_template.xlsx", "Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@router.post("/roles/bulk-upload", response_model=RoleBulkUploadResult, status_code=200)
async def bulk_upload_roles(
    file: UploadFile = File(..., description="Excel (.xlsx) or CSV file from the roles template"),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RoleBulkUploadResult:
    """
    Bulk-create or update approval roles from an Excel/CSV template.

    Expected columns (case-insensitive):
        Role Name       | Required. Unique within the tenant.
        Parent Role     | Optional. Name of the parent role (must already exist or appear earlier in file).
        Cost Center     | Optional. Name of an OrgStructureNode with node_type='Cost center'.
        Capacity        | Optional. 'single', a number (e.g. '3'), or blank/unlimited.
        Description     | Optional. Free-text description.

    Two-pass logic:
        Pass 1: create / update all roles (without parent link).
        Pass 2: wire parent_role_id by name.

    Existing roles (matched by name) are updated; new names are created.
    """
    block_if_readonly_impersonation(current_user)
    tenant_id = _require_tenant(current_user)

    # ── Parse file ────────────────────────────────────────────────────────────
    content = await file.read()
    filename = (file.filename or "").lower()
    rows: list[dict] = []
    errors: list[dict] = []

    if filename.endswith(".csv"):
        import csv
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        rows = [dict(r) for r in reader]
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed — use CSV upload instead.")
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace("*", "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload an .xlsx or .csv file.")

    # ── Normalise column names ────────────────────────────────────────────────
    def _col(row: dict, *keys: str) -> str:
        for k in keys:
            for rk in row:
                if rk.lower().replace(" ", "") == k.lower().replace(" ", ""):
                    return (row[rk] or "").strip()
        return ""

    # ── Load existing roles, cost centers, entity nodes ─────────────────────
    existing_roles_q = (await db.execute(
        select(ApprovalRole).where(ApprovalRole.tenant_id == tenant_id)
    )).scalars().all()
    # Primary lookup: name-only (last-one-wins, used only as fallback)
    role_by_name: dict[str, ApprovalRole] = {r.name.lower(): r for r in existing_roles_q}
    # Composite lookup: (name, area) — correctly disambiguates same-named roles
    # (e.g. five DPMs each covering a different area). This is the preferred lookup.
    role_by_composite: dict[tuple[str, str, str], ApprovalRole] = {
        (r.name.lower(), (r.area or "").lower(), (r.sub_area or "").lower()): r for r in existing_roles_q
    }

    cost_centers_q = (await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Cost center",
        )
    )).scalars().all()
    cc_by_name: dict[str, uuid.UUID] = {c.name.lower(): c.id for c in cost_centers_q}
    cc_by_code: dict[str, uuid.UUID] = {
        (c.cost_center_code or c.code).lower(): c.id for c in cost_centers_q
    }

    entity_nodes_q = (await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Legal entity",
        )
    )).scalars().all()
    entity_by_code: dict[str, uuid.UUID] = {
        (n.entity_code or n.code).lower(): n.id for n in entity_nodes_q if (n.entity_code or n.code)
    }

    # ── Pass 1: upsert roles (no parent wiring yet) ──────────────────────────
    result = RoleBulkUploadResult()
    upserted: list[tuple[str, ApprovalRole]] = []  # (original_name, role)

    for i, row in enumerate(rows, start=2):
        role_name = _col(row, "RoleName", "Role Name", "name")
        if not role_name:
            continue  # skip blank rows

        cc_name = _col(row, "CostCenter", "Cost Center", "costcenter")
        entity_code_raw = _col(row, "EntityCode", "Entity Code", "entitycode")
        capacity_raw = _col(row, "Capacity", "MaxOccupants", "max_occupants")
        _desig = _col(row, "Designation", "designation").lower().replace(" ", "_")
        # Normalise legacy values + accept both old and new names
        _DESIG_ALIAS = {
            "section_head": "manager",          # renamed
            "regular": "individual_contributor", # renamed
        }
        _desig = _DESIG_ALIAS.get(_desig, _desig)
        _VALID_DESIG = ("head_of_entity", "head_of_department", "manager", "team_lead", "individual_contributor")
        designation_raw: str | None = _desig if _desig in _VALID_DESIG else None
        _emp = _col(row, "EmploymentType", "Employment Type", "employmenttype").lower().strip()
        employment_type_raw = _emp if _emp in ("contract", "outsourced") else "permanent"
        area_raw = _col(row, "Area", "Area / Location", "arealocation") or None
        sub_area_raw = _col(row, "SubArea", "Sub Area", "subarea") or None
        description = _col(row, "Description", "Desc")

        # Resolve cost center (by name or code)
        cc_id: uuid.UUID | None = None
        if cc_name:
            cc_id = cc_by_name.get(cc_name.lower()) or cc_by_code.get(cc_name.lower())
            if cc_id is None:
                errors.append({"row": i, "role": role_name, "error": f"Cost center '{cc_name}' not found."})
                result.skipped += 1
                continue

        # Resolve entity node (by entity_code or code)
        entity_node_id: uuid.UUID | None = None
        if entity_code_raw:
            entity_node_id = entity_by_code.get(entity_code_raw.lower())
            if entity_node_id is None:
                errors.append({"row": i, "role": role_name, "error": f"Entity code '{entity_code_raw}' not found."})
                result.skipped += 1
                continue

        # Parse capacity
        max_occupants: int | None = None
        if capacity_raw.lower() in ("", "unlimited", "none", "-"):
            max_occupants = None
        elif capacity_raw.lower() in ("single", "1", "solo"):
            max_occupants = 1
        else:
            try:
                max_occupants = int(capacity_raw)
            except ValueError:
                errors.append({"row": i, "role": role_name, "error": f"Invalid capacity '{capacity_raw}'."})
                result.skipped += 1
                continue

        # Look up by composite key first (handles multiple roles with same name)
        composite_key = (role_name.lower(), (area_raw or "").lower(), (sub_area_raw or "").lower())
        existing = role_by_composite.get(composite_key) or (
            role_by_name.get(role_name.lower()) if not area_raw else None
        )
        if existing:
            existing.description = description or existing.description
            existing.cost_center_id = cc_id if cc_id is not None else existing.cost_center_id
            existing.entity_node_id = entity_node_id if entity_node_id is not None else existing.entity_node_id
            existing.designation = designation_raw
            existing.max_occupants = max_occupants
            existing.employment_type = employment_type_raw
            existing.area = area_raw if area_raw is not None else existing.area
            existing.sub_area = sub_area_raw if sub_area_raw is not None else existing.sub_area
            upserted.append((role_name, existing))
            result.updated += 1
        else:
            new_role = ApprovalRole(
                tenant_id=tenant_id,
                name=role_name,
                description=description or None,
                cost_center_id=cc_id,
                entity_node_id=entity_node_id,
                max_occupants=max_occupants,
                designation=designation_raw,
                employment_type=employment_type_raw,
                area=area_raw,
                sub_area=sub_area_raw,
                display_order=0,
            )
            db.add(new_role)
            # Register in both dicts so later rows in this upload find it
            role_by_composite[composite_key] = new_role
            role_by_name[role_name.lower()] = new_role  # name-only fallback
            upserted.append((role_name, new_role))
            result.created += 1

    await db.flush()  # generate PKs so pass-2 can reference them

    # ── Pass 2: wire parent_role_id ───────────────────────────────────────────
    # role_by_composite keyed by (name, area, sub_area) is the primary lookup.
    # role_by_subarea keyed by (name, sub_area) handles the cascade pattern:
    #   DPM.area="Lagos", DPM.sub_area="Lagos Mainland"
    #   DPS.area="Lagos Mainland" → matches DPM via sub_area key.
    role_by_subarea: dict[tuple[str, str], ApprovalRole] = {
        (r.name.lower(), r.sub_area.lower()): r
        for r in role_by_composite.values() if r.sub_area
    }

    for i, row in enumerate(rows, start=2):
        role_name = _col(row, "RoleName", "Role Name", "name")
        parent_name = _col(row, "ParentRole", "Parent Role", "parentrole")
        if not role_name or not parent_name:
            continue
        row_area     = _col(row, "Area", "Area / Location", "arealocation").lower()
        row_sub_area = _col(row, "SubArea", "Sub Area", "subarea").lower()
        # Find THIS child by full (name, area, sub_area) key, then fall back
        child = (
            role_by_composite.get((role_name.lower(), row_area, row_sub_area))
            or role_by_composite.get((role_name.lower(), row_area, ""))
            or role_by_name.get(role_name.lower())
        )
        # Find parent: try area match (parent's area = child's area), then
        # sub_area cascade (child's area = parent's sub_area), then name-only
        parent = None
        if row_area:
            parent = (
                role_by_composite.get((parent_name.lower(), row_area, ""))
                or role_by_subarea.get((parent_name.lower(), row_area))
            )
        if not parent:
            parent = role_by_name.get(parent_name.lower())
        if child and parent:
            child.parent_role_id = parent.id
        elif child and not parent:
            errors.append({"row": i, "role": role_name, "error": f"Parent role '{parent_name}' not found."})

    await db.commit()
    return result


# ── Approval Policies ────────────────────────────────────────────────────────

@router.get("/policies", response_model=list[ApprovalPolicyResponse])
async def list_policies(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalPolicyResponse]:
    """List all approval policies for the current tenant."""
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(ApprovalPolicy)
        .options(
            selectinload(ApprovalPolicy.ceiling_role),
            selectinload(ApprovalPolicy.thresholds).selectinload(ApprovalRoleThreshold.role),
        )
        .where(ApprovalPolicy.tenant_id == tenant_id)
        .order_by(ApprovalPolicy.module)
    )
    policies = result.scalars().all()
    return [ApprovalPolicyResponse.from_orm(p) for p in policies]


@router.post("/policies", response_model=ApprovalPolicyResponse, status_code=201)
async def upsert_policy(
    data: ApprovalPolicyCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalPolicyResponse:
    """
    Create or replace the approval policy for a module.
    If a policy already exists for this module, it is replaced (upsert semantics).
    Thresholds are fully replaced when provided.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    # Check for existing policy for this module
    existing_result = await db.execute(
        select(ApprovalPolicy)
        .options(
            selectinload(ApprovalPolicy.ceiling_role),
            selectinload(ApprovalPolicy.thresholds).selectinload(ApprovalRoleThreshold.role),
        )
        .where(
            and_(ApprovalPolicy.tenant_id == tenant_id, ApprovalPolicy.module == data.module)
        )
    )
    policy = existing_result.scalar_one_or_none()

    def _to_uuid(v: str | None) -> uuid.UUID | None:
        return uuid.UUID(v) if v else None

    if policy:
        policy.routing_mode = data.routing_mode
        policy.selected_designations = data.selected_designations
        policy.ceiling_role_id = _to_uuid(data.ceiling_role_id)
        policy.vacant_seat_behavior = data.vacant_seat_behavior
        policy.fallback_approver_id = _to_uuid(data.fallback_approver_id)
        policy.requires_finance_review = data.requires_finance_review
        policy.finance_levels = data.finance_levels
        policy.finance_l1_role_id = _to_uuid(data.finance_l1_role_id)
        policy.finance_l2_role_id = _to_uuid(data.finance_l2_role_id)
        policy.finance_l3_role_id = _to_uuid(data.finance_l3_role_id)
        policy.finance_amount_threshold_l2 = data.finance_amount_threshold_l2
        policy.finance_amount_threshold_l3 = data.finance_amount_threshold_l3
        policy.is_active = True
        # Delete old thresholds and replace
        old_thresh = await db.execute(
            select(ApprovalRoleThreshold).where(ApprovalRoleThreshold.policy_id == policy.id)
        )
        for t in old_thresh.scalars().all():
            await db.delete(t)
        await db.flush()
    else:
        policy = ApprovalPolicy(
            tenant_id=tenant_id,
            module=data.module,
            routing_mode=data.routing_mode,
            selected_designations=data.selected_designations,
            ceiling_role_id=_to_uuid(data.ceiling_role_id),
            vacant_seat_behavior=data.vacant_seat_behavior,
            fallback_approver_id=_to_uuid(data.fallback_approver_id),
            requires_finance_review=data.requires_finance_review,
            finance_levels=data.finance_levels,
            finance_l1_role_id=_to_uuid(data.finance_l1_role_id),
            finance_l2_role_id=_to_uuid(data.finance_l2_role_id),
            finance_l3_role_id=_to_uuid(data.finance_l3_role_id),
            finance_amount_threshold_l2=data.finance_amount_threshold_l2,
            finance_amount_threshold_l3=data.finance_amount_threshold_l3,
        )
        db.add(policy)
        await db.flush()

    for t in data.thresholds:
        db.add(ApprovalRoleThreshold(
            policy_id=policy.id,
            approval_role_id=uuid.UUID(t.approval_role_id),
            max_amount=t.max_amount,
        ))

    await db.commit()

    # Reload with relationships
    reloaded = (await db.execute(
        select(ApprovalPolicy)
        .options(
            selectinload(ApprovalPolicy.ceiling_role),
            selectinload(ApprovalPolicy.thresholds).selectinload(ApprovalRoleThreshold.role),
        )
        .where(ApprovalPolicy.id == policy.id)
    )).scalar_one()
    return ApprovalPolicyResponse.from_orm(reloaded)


@router.patch("/policies/{policy_id}", response_model=ApprovalPolicyResponse)
async def update_policy(
    policy_id: uuid.UUID,
    data: ApprovalPolicyUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalPolicyResponse:
    """Partially update an approval policy. Pass thresholds to fully replace them."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    def _to_uuid(v: str | None) -> uuid.UUID | None:
        return uuid.UUID(v) if v else None

    policy = (await db.execute(
        select(ApprovalPolicy)
        .options(
            selectinload(ApprovalPolicy.ceiling_role),
            selectinload(ApprovalPolicy.thresholds).selectinload(ApprovalRoleThreshold.role),
        )
        .where(and_(ApprovalPolicy.id == policy_id, ApprovalPolicy.tenant_id == tenant_id))
    )).scalar_one_or_none()
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found.")

    if data.routing_mode is not None:
        policy.routing_mode = data.routing_mode
    if data.selected_designations is not None:
        policy.selected_designations = data.selected_designations
    if data.ceiling_role_id is not None:
        policy.ceiling_role_id = _to_uuid(data.ceiling_role_id)
    if data.vacant_seat_behavior is not None:
        policy.vacant_seat_behavior = data.vacant_seat_behavior
    if data.fallback_approver_id is not None:
        policy.fallback_approver_id = _to_uuid(data.fallback_approver_id)
    if data.requires_finance_review is not None:
        policy.requires_finance_review = data.requires_finance_review
    if data.finance_levels is not None:
        policy.finance_levels = data.finance_levels
    if data.finance_l1_role_id is not None:
        policy.finance_l1_role_id = _to_uuid(data.finance_l1_role_id)
    if data.finance_l2_role_id is not None:
        policy.finance_l2_role_id = _to_uuid(data.finance_l2_role_id)
    if data.finance_l3_role_id is not None:
        policy.finance_l3_role_id = _to_uuid(data.finance_l3_role_id)
    if data.finance_amount_threshold_l2 is not None:
        policy.finance_amount_threshold_l2 = data.finance_amount_threshold_l2
    if data.finance_amount_threshold_l3 is not None:
        policy.finance_amount_threshold_l3 = data.finance_amount_threshold_l3
    if data.is_active is not None:
        policy.is_active = data.is_active

    if data.thresholds is not None:
        old = await db.execute(
            select(ApprovalRoleThreshold).where(ApprovalRoleThreshold.policy_id == policy.id)
        )
        for t in old.scalars().all():
            await db.delete(t)
        await db.flush()
        for t in data.thresholds:
            db.add(ApprovalRoleThreshold(
                policy_id=policy.id,
                approval_role_id=uuid.UUID(t.approval_role_id),
                max_amount=t.max_amount,
            ))

    await db.commit()

    reloaded = (await db.execute(
        select(ApprovalPolicy)
        .options(
            selectinload(ApprovalPolicy.ceiling_role),
            selectinload(ApprovalPolicy.thresholds).selectinload(ApprovalRoleThreshold.role),
        )
        .where(ApprovalPolicy.id == policy.id)
    )).scalar_one()
    return ApprovalPolicyResponse.from_orm(reloaded)


@router.delete("/policies/{policy_id}", status_code=204)
async def delete_policy(
    policy_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Delete an approval policy."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    policy = (await db.execute(
        select(ApprovalPolicy).where(
            and_(ApprovalPolicy.id == policy_id, ApprovalPolicy.tenant_id == tenant_id)
        )
    )).scalar_one_or_none()
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found.")
    await db.delete(policy)
    await db.commit()


@router.get("/policies/{module}/chain-preview", response_model=list[ChainPreviewStep])
async def get_chain_preview(
    module: str,
    amount: Decimal = Query(default=Decimal("0")),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ChainPreviewStep]:
    """
    Preview the computed approval chain for the current user and a given amount.
    Used by the expense submission form to show who will be notified before submitting.
    """
    tenant_id = _require_tenant(current_user)
    steps = await preview_chain(
        submitter_user_id=current_user.user_id,
        tenant_id=tenant_id,
        module=module,
        total_amount=amount,
        db=db,
    )
    return [ChainPreviewStep(**s) for s in steps]


# ── Approval Delegations ──────────────────────────────────────────────────────

async def _load_delegation_response(d: ApprovalDelegation, db: AsyncSession) -> ApprovalDelegationResponse:
    """Load user names for delegation response."""
    delegator = (await db.execute(select(User).where(User.id == d.delegator_id))).scalar_one_or_none()
    delegate = (await db.execute(select(User).where(User.id == d.delegate_id))).scalar_one_or_none()
    return ApprovalDelegationResponse.from_orm(
        d,
        delegator_name=delegator.full_name if delegator else "Unknown",
        delegate_name=delegate.full_name if delegate else "Unknown",
    )


@router.get("/delegations", response_model=list[ApprovalDelegationResponse])
async def list_delegations(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalDelegationResponse]:
    """List delegations where the current user is the delegator."""
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(ApprovalDelegation)
        .where(
            and_(
                ApprovalDelegation.tenant_id == tenant_id,
                ApprovalDelegation.delegator_id == current_user.user_id,
            )
        )
        .order_by(ApprovalDelegation.created_at.desc())
    )
    delegations = result.scalars().all()
    return [await _load_delegation_response(d, db) for d in delegations]


@router.post("/delegations", response_model=ApprovalDelegationResponse, status_code=201)
async def create_delegation(
    data: ApprovalDelegationCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalDelegationResponse:
    """
    Create an approval delegation — delegate your approval authority to another user.
    Cannot delegate to yourself. end_date=null means open-ended until revoked.
    """
    block_if_readonly_impersonation(current_user)
    tenant_id = _require_tenant(current_user)

    delegate_id = uuid.UUID(data.delegate_id)
    if delegate_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="Cannot delegate approval authority to yourself.")

    # Validate delegate exists in this tenant
    delegate = (await db.execute(
        select(User)
        .join(UserTenant, UserTenant.user_id == User.id)
        .where(and_(User.id == delegate_id, UserTenant.tenant_id == tenant_id))
    )).scalar_one_or_none()
    if not delegate:
        raise HTTPException(status_code=404, detail="Delegate user not found in this tenant.")

    delegation = ApprovalDelegation(
        tenant_id=tenant_id,
        delegator_id=current_user.user_id,
        delegate_id=delegate_id,
        start_date=data.start_date,
        end_date=data.end_date,
        reason=data.reason,
        created_by_id=current_user.user_id,
    )
    db.add(delegation)
    await db.commit()
    await db.refresh(delegation)
    return await _load_delegation_response(delegation, db)


@router.patch("/delegations/{delegation_id}", response_model=ApprovalDelegationResponse)
async def update_delegation(
    delegation_id: uuid.UUID,
    data: ApprovalDelegationUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalDelegationResponse:
    """Update or revoke a delegation. Set is_active=false to revoke immediately."""
    block_if_readonly_impersonation(current_user)
    tenant_id = _require_tenant(current_user)

    delegation = (await db.execute(
        select(ApprovalDelegation).where(
            and_(
                ApprovalDelegation.id == delegation_id,
                ApprovalDelegation.tenant_id == tenant_id,
                ApprovalDelegation.delegator_id == current_user.user_id,
            )
        )
    )).scalar_one_or_none()
    if not delegation:
        raise HTTPException(status_code=404, detail="Delegation not found.")

    if data.end_date is not None:
        delegation.end_date = data.end_date
    if data.is_active is not None:
        delegation.is_active = data.is_active
    if data.reason is not None:
        delegation.reason = data.reason

    await db.commit()
    await db.refresh(delegation)
    return await _load_delegation_response(delegation, db)


# ── Approval Matrix (legacy) ──────────────────────────────────────────────────

@router.post("/matrix", response_model=ApprovalMatrixResponse)
async def upsert_approval_matrix(
    data: ApprovalMatrixCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalMatrixResponse:
    """Create or update the tenant's approval matrix. Tenant Admin only."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    matrix = await _get_matrix(tenant_id, db)
    if matrix:
        matrix.levels = data.levels
        matrix.level1_role = data.level1_role
        matrix.level2_role = data.level2_role if data.levels >= 2 else None
        matrix.level3_role = data.level3_role if data.levels >= 3 else None
        matrix.amount_threshold_l2 = data.amount_threshold_l2 if data.levels >= 2 else None
        matrix.amount_threshold_l3 = data.amount_threshold_l3 if data.levels >= 3 else None
    else:
        matrix = ApprovalMatrix(
            tenant_id=tenant_id,
            levels=data.levels,
            level1_role=data.level1_role,
            level2_role=data.level2_role if data.levels >= 2 else None,
            level3_role=data.level3_role if data.levels >= 3 else None,
            amount_threshold_l2=data.amount_threshold_l2 if data.levels >= 2 else None,
            amount_threshold_l3=data.amount_threshold_l3 if data.levels >= 3 else None,
        )
        db.add(matrix)

    await db.flush()
    return ApprovalMatrixResponse.from_orm(matrix)


@router.get("/matrix", response_model=ApprovalMatrixResponse | None)
async def get_approval_matrix(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ApprovalMatrixResponse | None:
    """Return the current tenant's approval matrix, or null if not configured."""
    tenant_id = _require_tenant(current_user)
    matrix = await _get_matrix(tenant_id, db)
    return ApprovalMatrixResponse.from_orm(matrix) if matrix else None


# ── Submit with Approvers ─────────────────────────────────────────────────────

@router.post("/reports/{report_id}/submit", response_model=ExpenseReportResponse)
async def submit_with_approvers(
    report_id: uuid.UUID,
    data: SubmitWithApproversRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ExpenseReportResponse:
    """
    Submit an expense report for approval.

    First-time submission: approver IDs must be provided via the request body.
    Resubmission (after rejection/referral): backend reuses original approver IDs
    and resumes the chain from rejected_at_level, skipping already-approved levels.

    Separation of duties: none of the selected approvers may be the same person
    as the report's employee (requestor).

    Writes an expense snapshot and audit log entry on every submission.
    Sends an email notification to the first active approver.
    """
    tenant_id = _require_tenant(current_user)
    block_if_readonly_impersonation(current_user)
    if current_user.is_tenant_admin and not current_user.has_non_admin_role:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Tenant administrators cannot submit expense reports.",
        )
    report = await _get_report_or_404(report_id, tenant_id, db)

    if report.status not in ("DRAFT", "REJECTED", "REFERRED_TO_REQUESTOR"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Only DRAFT, REJECTED, or REFERRED_TO_REQUESTOR reports can be submitted.",
        )
    if not report.lines:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="A report must have at least one expense line before submitting.",
        )

    # Period check at submission time (fix for audit finding #006).
    # Validates report_date is in an open period NOW, while the submitter can still
    # change the date. Without this, the report is locked at submission and a
    # DATE_NOT_POSTABLE error at final approval gives the approver no way to resolve it.
    postable, period_reason = await is_date_postable(
        tenant_id, report.report_date, db, module="expense"
    )
    if not postable:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Cannot submit: report date {report.report_date} is not in an open accounting period "
                f"({period_reason}). Please update the report date to a date in an open period before submitting."
            ),
        )

    # ── Determine routing mode: policy engine or legacy matrix ───────────────
    policy = await get_policy("expense", tenant_id, db)

    existing_result = await db.execute(
        select(ExpenseApproval)
        .where(ExpenseApproval.report_id == report.id)
        .order_by(ExpenseApproval.level.asc())
    )
    existing_approvals = existing_result.scalars().all()

    # Write snapshot before mutating report state
    snapshot_version = await _write_snapshot(report, tenant_id, db)

    if existing_approvals:
        # ── Resubmission — smart resume from rejected_at_level ────────────────
        # Resubmission reuses the original chain (same approvers, same levels).
        # The routing engine is NOT re-run here — the chain was already computed
        # on first submission and the approver ids are stored in expense_approvals.
        rejected_at = report.rejected_at_level or 1

        to_recreate = [a for a in existing_approvals if a.level >= rejected_at]
        level_to_record: dict[int, ExpenseApproval] = {a.level: a for a in to_recreate}

        for old in to_recreate:
            await db.delete(old)
        await db.flush()

        for level, old_rec in sorted(level_to_record.items()):
            db.add(ExpenseApproval(
                report_id=report.id,
                tenant_id=tenant_id,
                level=level,
                approver_id=old_rec.approver_id,
                delegated_from_id=old_rec.delegated_from_id,
                chain_type=old_rec.chain_type,
                role_label=old_rec.role_label,
                status="PENDING",
            ))

        start_level = rejected_at
        approver_ids_for_log = [str(r.approver_id) for r in level_to_record.values()]

        report.status = "PENDING_APPROVAL"
        report.current_approval_level = start_level
        report.submitted_at = datetime.now(timezone.utc)
        report.rejection_comment = None
        report.rejected_at_level = None
        report.referred_back_from_level = None
        report.referred_back_levels = None

        await db.flush()

        await _write_audit_log(db, "EXPENSE_RESUBMITTED", current_user.user_id, tenant_id, {
            "report_id": str(report.id),
            "report_number": report.report_number,
            "total_amount": str(report.total_amount),
            "resumed_from_level": start_level,
            "snapshot_version": snapshot_version,
        })

        # Notify the first active approver
        first_rec = level_to_record.get(start_level)
        if first_rec:
            approver_result = await db.execute(select(User).where(User.id == first_rec.approver_id))
            approver = approver_result.scalar_one_or_none()
            employee_result = await db.execute(select(User).where(User.id == report.employee_id))
            employee = employee_result.scalar_one_or_none()
            if approver and employee:
                role_label = first_rec.role_label or "Approver"
                _send_approver_notification_email(
                    to_email=approver.email,
                    report_number=report.report_number,
                    report_date=str(report.report_date),
                    total_amount=report.total_amount,
                    employee_name=employee.full_name,
                    role_label=role_label,
                )

        return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))

    # ── First-time submission ─────────────────────────────────────────────────

    if policy:
        # ── Routing engine path ───────────────────────────────────────────────
        try:
            chain_steps = await compute_chain(
                submitter_user_id=current_user.user_id,
                tenant_id=tenant_id,
                module="expense",
                total_amount=report.total_amount,
                db=db,
                requestor_selected_approver_id=data.selected_approver_id,
            )
        except ApprovalChainHoldError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e))
        except ApprovalRoutingError as e:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(e))

        # Separation of duties: no step may target the submitter
        for step in chain_steps:
            if step.approver_user_id == current_user.user_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="An approver in the computed chain is the same person as the submitter. "
                           "Check the reporting structure or delegation configuration.",
                )

        for step in chain_steps:
            db.add(ExpenseApproval(
                report_id=report.id,
                tenant_id=tenant_id,
                level=step.level,
                approver_id=step.approver_user_id,
                delegated_from_id=step.delegated_from_id,
                chain_type=step.chain_type,
                role_label=step.role_label,
                status="PENDING",
            ))

        report.status = "PENDING_APPROVAL"
        report.current_approval_level = 1
        report.submitted_at = datetime.now(timezone.utc)
        report.rejection_comment = None
        report.rejected_at_level = None
        report.referred_back_from_level = None
        report.referred_back_levels = None

        await db.flush()

        approver_ids_for_log = [str(s.approver_user_id) for s in chain_steps]
        await _write_audit_log(db, "EXPENSE_SUBMITTED", current_user.user_id, tenant_id, {
            "report_id": str(report.id),
            "report_number": report.report_number,
            "total_amount": str(report.total_amount),
            "employee_id": str(report.employee_id),
            "approver_ids": approver_ids_for_log,
            "routing_mode": policy.routing_mode,
            "snapshot_version": snapshot_version,
        })

        # Notify Level 1 approver
        l1 = chain_steps[0]
        approver_result = await db.execute(select(User).where(User.id == l1.approver_user_id))
        approver = approver_result.scalar_one_or_none()
        employee_result = await db.execute(select(User).where(User.id == report.employee_id))
        employee = employee_result.scalar_one_or_none()
        if approver and employee:
            _send_approver_notification_email(
                to_email=approver.email,
                report_number=report.report_number,
                report_date=str(report.report_date),
                total_amount=report.total_amount,
                employee_name=employee.full_name,
                role_label=l1.role_label,
            )

        return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))

    # ── Legacy matrix path (no policy configured) ─────────────────────────────
    matrix = await _get_matrix(tenant_id, db)
    if not matrix:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                "No approval policy or matrix is configured. "
                "Go to Setup → Approval Workflows to configure one."
            ),
        )

    if data.level1_approver_id is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Level 1 approver is required.",
        )

    applicable_levels: list[tuple[int, uuid.UUID]] = [(1, data.level1_approver_id)]

    if matrix.levels >= 2:
        threshold_l2 = matrix.amount_threshold_l2
        if threshold_l2 is None or report.total_amount > threshold_l2:
            if data.level2_approver_id is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Level 2 approver is required for this report.",
                )
            applicable_levels.append((2, data.level2_approver_id))

    if matrix.levels >= 3:
        threshold_l3 = matrix.amount_threshold_l3
        if threshold_l3 is None or report.total_amount > threshold_l3:
            if data.level3_approver_id is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="Level 3 approver is required for this report.",
                )
            applicable_levels.append((3, data.level3_approver_id))

    # Separation of duties
    for _, approver_id in applicable_levels:
        if approver_id == report.employee_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="An expense approver cannot be the same person as the requestor.",
            )

    for _, approver_id in applicable_levels:
        await _validate_approver(approver_id, tenant_id, db)

    for level, approver_id in applicable_levels:
        db.add(ExpenseApproval(
            report_id=report.id,
            tenant_id=tenant_id,
            level=level,
            approver_id=approver_id,
            role_label=_role_label_for_level(matrix, level),
            chain_type="management",
            status="PENDING",
        ))

    report.status = "PENDING_APPROVAL"
    report.current_approval_level = 1
    report.submitted_at = datetime.now(timezone.utc)
    report.rejection_comment = None
    report.rejected_at_level = None
    report.referred_back_from_level = None
    report.referred_back_levels = None

    await db.flush()

    approver_ids_for_log = [str(approver_id) for _, approver_id in applicable_levels]
    await _write_audit_log(db, "EXPENSE_SUBMITTED", current_user.user_id, tenant_id, {
        "report_id": str(report.id),
        "report_number": report.report_number,
        "total_amount": str(report.total_amount),
        "employee_id": str(report.employee_id),
        "approver_ids": approver_ids_for_log,
        "routing_mode": "legacy_matrix",
        "snapshot_version": snapshot_version,
    })

    l1_approver_id = applicable_levels[0][1]
    approver_result = await db.execute(select(User).where(User.id == l1_approver_id))
    approver = approver_result.scalar_one_or_none()
    employee_result = await db.execute(select(User).where(User.id == report.employee_id))
    employee = employee_result.scalar_one_or_none()
    if approver and employee:
        _send_approver_notification_email(
            to_email=approver.email,
            report_number=report.report_number,
            report_date=str(report.report_date),
            total_amount=report.total_amount,
            employee_name=employee.full_name,
            role_label=_role_label_for_level(matrix, 1),
        )

    return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))


# ── Approver Queue ────────────────────────────────────────────────────────────

@router.get("/queue", response_model=list[ApprovalQueueItem])
async def get_approval_queue(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalQueueItem]:
    """
    Return all expense reports currently awaiting the current user's approval.

    A report is in the queue when the approval record is PENDING and
    expense_reports.current_approval_level matches the record's level,
    enforcing sequential approval.
    """
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(ExpenseApproval, ExpenseReport, User)
        .join(ExpenseReport, ExpenseApproval.report_id == ExpenseReport.id)
        .join(User, ExpenseReport.employee_id == User.id)
        .where(
            ExpenseApproval.approver_id == current_user.user_id,
            ExpenseApproval.status == "PENDING",
            ExpenseReport.current_approval_level == ExpenseApproval.level,
            ExpenseApproval.tenant_id == tenant_id,
        )
        .order_by(ExpenseApproval.created_at.asc())
    )
    rows = result.all()
    matrix = await _get_matrix(tenant_id, db)

    return [
        ApprovalQueueItem(
            approval_id=str(approval.id),
            report_id=str(report.id),
            report_number=report.report_number,
            employee_name=employee.full_name,
            report_date=report.report_date,
            total_amount=report.total_amount,
            level=approval.level,
            level_label=_role_label_for_level(matrix, approval.level) if matrix else f"Level {approval.level}",
            created_at=approval.created_at,
        )
        for approval, report, employee in rows
    ]


# ── Rejected Reports (approver visibility) ───────────────────────────────────

@router.get("/rejected", response_model=list[ApprovalQueueItem])
async def get_rejected_reports(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalQueueItem]:
    """
    Return reports that were rejected and where the current user was an approver.

    Deduplicated — each report appears once even if assigned at multiple levels.
    """
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(ExpenseApproval, ExpenseReport, User)
        .join(ExpenseReport, ExpenseApproval.report_id == ExpenseReport.id)
        .join(User, ExpenseReport.employee_id == User.id)
        .where(
            ExpenseApproval.approver_id == current_user.user_id,
            ExpenseReport.status == "REJECTED",
            ExpenseApproval.tenant_id == tenant_id,
        )
        .order_by(ExpenseReport.created_at.desc(), ExpenseApproval.level.asc())
    )
    rows = result.all()
    matrix = await _get_matrix(tenant_id, db)

    seen: set[uuid.UUID] = set()
    items: list[ApprovalQueueItem] = []
    for approval, report, employee in rows:
        if report.id in seen:
            continue
        seen.add(report.id)
        items.append(ApprovalQueueItem(
            approval_id=str(approval.id),
            report_id=str(report.id),
            report_number=report.report_number,
            employee_name=employee.full_name,
            report_date=report.report_date,
            total_amount=report.total_amount,
            level=approval.level,
            level_label=_role_label_for_level(matrix, approval.level) if matrix else f"Level {approval.level}",
            created_at=approval.created_at,
            rejection_comment=report.rejection_comment,
        ))
    return items


# ── Audit Trail ───────────────────────────────────────────────────────────────

@router.get("/reports/{report_id}/audit-log", response_model=list[AuditLogEntry])
async def get_audit_log(
    report_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[AuditLogEntry]:
    """
    Return the full chronological audit trail for an expense report.

    Restricted to tenant admins and super admins.
    The report must belong to the current user's tenant.
    """
    tenant_id = _require_tenant(current_user)
    if not current_user.is_tenant_admin and not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Audit trail access requires Tenant Admin or Super Admin.",
        )

    # Verify report belongs to this tenant
    report_check = await db.execute(
        select(ExpenseReport.id).where(
            ExpenseReport.id == report_id,
            ExpenseReport.tenant_id == tenant_id,
        )
    )
    if not report_check.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found.")

    result = await db.execute(
        select(AuditLog, User)
        .join(User, AuditLog.user_id == User.id, isouter=True)
        .where(
            AuditLog.tenant_id == tenant_id,
            AuditLog.log_metadata["report_id"].astext == str(report_id),
            AuditLog.event_type.in_([
                "EXPENSE_SUBMITTED", "EXPENSE_APPROVED", "EXPENSE_REJECTED",
            ]),
        )
        .order_by(AuditLog.created_at.asc())
    )
    rows = result.all()

    return [
        AuditLogEntry(
            id=str(log.id),
            event_type=log.event_type,
            user_id=str(log.user_id) if log.user_id else None,
            actor_name=actor.full_name if actor else "Unknown",
            log_metadata=log.log_metadata or {},
            created_at=log.created_at,
        )
        for log, actor in rows
    ]


# ── Snapshot ──────────────────────────────────────────────────────────────────

@router.get("/reports/{report_id}/snapshot/{version}", response_model=SnapshotResponse)
async def get_snapshot(
    report_id: uuid.UUID,
    version: int,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> SnapshotResponse:
    """Return the expense report snapshot for a specific submission version."""
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(ExpenseReportSnapshot).where(
            ExpenseReportSnapshot.report_id == report_id,
            ExpenseReportSnapshot.tenant_id == tenant_id,
            ExpenseReportSnapshot.version == version,
        )
    )
    snapshot = result.scalar_one_or_none()
    if not snapshot:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Snapshot version {version} not found for this report.",
        )

    return SnapshotResponse(
        id=str(snapshot.id),
        report_id=str(snapshot.report_id),
        version=snapshot.version,
        submitted_at=snapshot.submitted_at,
        snapshot_data=snapshot.snapshot_data,
        created_at=snapshot.created_at,
    )


# ── Report Approval Chain ─────────────────────────────────────────────────────

@router.get("/reports/{report_id}", response_model=list[ApprovalRecordResponse])
async def get_report_approvals(
    report_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[ApprovalRecordResponse]:
    """Return all approval records for a given expense report."""
    tenant_id = _require_tenant(current_user)

    report_result = await db.execute(
        select(ExpenseReport).where(
            ExpenseReport.id == report_id,
            ExpenseReport.tenant_id == tenant_id,
        )
    )
    if not report_result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found.")

    result = await db.execute(
        select(ExpenseApproval, User)
        .join(User, ExpenseApproval.approver_id == User.id)
        .where(
            ExpenseApproval.report_id == report_id,
            ExpenseApproval.tenant_id == tenant_id,
        )
        .order_by(ExpenseApproval.level.asc(), ExpenseApproval.created_at.desc())
    )
    rows = result.all()
    matrix = await _get_matrix(tenant_id, db)

    seen_levels: set[int] = set()
    deduped: list[ApprovalRecordResponse] = []
    for approval, approver in rows:
        if approval.level in seen_levels:
            continue
        seen_levels.add(approval.level)
        level_label = (
            approval.role_label
            or (_role_label_for_level(matrix, approval.level) if matrix else f"Level {approval.level}")
        )
        deduped.append(ApprovalRecordResponse(
            id=str(approval.id),
            level=approval.level,
            level_label=level_label,
            approver_id=str(approval.approver_id),
            approver_name=approver.full_name,
            status=approval.status,
            comment=approval.comment,
            visible_to_requestor=approval.visible_to_requestor,
            response_comment=approval.response_comment,
            actioned_at=approval.actioned_at,
            created_at=approval.created_at,
        ))
    return deduped


# ── Approve ───────────────────────────────────────────────────────────────────

@router.post("/{approval_id}/approve", response_model=ExpenseReportResponse)
async def approve(
    approval_id: uuid.UUID,
    data: ApproveRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ExpenseReportResponse:
    """Approve an expense report at the current level."""
    tenant_id = _require_tenant(current_user)
    block_if_readonly_impersonation(current_user)

    result = await db.execute(
        select(ExpenseApproval).where(
            ExpenseApproval.id == approval_id,
            ExpenseApproval.tenant_id == tenant_id,
        )
    )
    approval = result.scalar_one_or_none()
    if not approval:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Approval record not found.")

    if approval.approver_id != current_user.user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN,
                            detail="You are not the designated approver for this record.")
    if approval.status != "PENDING":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval record has already been actioned.")

    report = await _get_report_or_404(approval.report_id, tenant_id, db)

    if report.current_approval_level != approval.level:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval level is not currently active for this report.")

    approval.status = "APPROVED"
    approval.comment = data.comment
    approval.response_comment = data.response_comment
    approval.actioned_at = datetime.now(timezone.utc)

    await _write_audit_log(db, "EXPENSE_APPROVED", current_user.user_id, tenant_id, {
        "report_id": str(report.id),
        "report_number": report.report_number,
        "level": approval.level,
        "approver_id": str(approval.approver_id),
        "comment": data.comment,
        "response_comment": data.response_comment,
        "total_amount": str(report.total_amount),
    })

    if report.referred_back_from_level is not None:
        referred_levels_queue = report.referred_back_levels or []

        if referred_levels_queue:
            next_level = referred_levels_queue[0]
            new_queue = referred_levels_queue[1:]

            next_result = await db.execute(
                select(ExpenseApproval).where(
                    ExpenseApproval.report_id == report.id,
                    ExpenseApproval.level == next_level,
                ).order_by(ExpenseApproval.created_at.desc())
            )
            next_target = next_result.scalars().first()
            if next_target:
                next_target.status = "PENDING"
                next_target.actioned_at = None

                next_approver_result = await db.execute(select(User).where(User.id == next_target.approver_id))
                next_approver = next_approver_result.scalar_one_or_none()
                referring_approver_result = await db.execute(select(User).where(User.id == approval.approver_id))
                referring_approver = referring_approver_result.scalar_one_or_none()
                if next_approver and referring_approver:
                    _send_referred_approver_email(
                        to_email=next_approver.email,
                        report_number=report.report_number,
                        referring_approver_name=referring_approver.full_name,
                        referring_level=report.referred_back_from_level,
                        comment=approval.comment or "",
                    )

            report.current_approval_level = next_level
            report.referred_back_levels = new_queue if new_queue else None

        else:
            referring_level = report.referred_back_from_level
            referring_result = await db.execute(
                select(ExpenseApproval).where(
                    ExpenseApproval.report_id == report.id,
                    ExpenseApproval.level == referring_level,
                    ExpenseApproval.status == "REFERRED_BACK",
                ).order_by(ExpenseApproval.created_at.desc())
            )
            referring_approval = referring_result.scalars().first()
            if referring_approval:
                referring_approval.status = "PENDING"
                referring_approval.actioned_at = None

            report.current_approval_level = referring_level
            report.referred_back_from_level = None
            report.referred_back_levels = None
    else:
        next_result = await db.execute(
            select(ExpenseApproval).where(
                ExpenseApproval.report_id == report.id,
                ExpenseApproval.level == approval.level + 1,
                ExpenseApproval.status == "PENDING",
            ).order_by(ExpenseApproval.created_at.desc())
        )
        next_approval = next_result.scalars().first()

        if next_approval:
            report.current_approval_level = next_approval.level

            next_approver_result = await db.execute(select(User).where(User.id == next_approval.approver_id))
            next_approver = next_approver_result.scalar_one_or_none()
            employee_result = await db.execute(select(User).where(User.id == report.employee_id))
            employee = employee_result.scalar_one_or_none()
            if next_approver and employee:
                role_label = next_approval.role_label or f"Level {next_approval.level}"
                _send_approver_notification_email(
                    to_email=next_approver.email,
                    report_number=report.report_number,
                    report_date=str(report.report_date),
                    total_amount=report.total_amount,
                    employee_name=employee.full_name,
                    role_label=role_label,
                )
        else:
            try:
                posting_result: PostingResult = await post_expense_to_gl(
                    db, tenant_id, report, current_user.user_id
                )
            except ExpensePostingError as exc:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)) from exc
            except AccountMappingError as exc:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"{exc} Configure the missing posting role(s) in Setup -> Account Mapping, then re-approve.",
                ) from exc
            except PostingError as exc:
                if exc.code == "DATE_NOT_POSTABLE":
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail=(
                            f"Cannot post: {exc.message} "
                            "To resolve: either reject the report so the submitter can correct the date, "
                            "or ask a Ziva consultant to reopen the period via Setup -> Periods."
                        ),
                    ) from exc
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=str(exc)) from exc

            report.status = "APPROVED"
            report.current_approval_level = None

            audit_event = (
                "EXPENSE_GL_POSTED"
                if posting_result.mode == "full_erp"
                else "EXPENSE_BATCH_QUEUED"
                if posting_result.mode == "connected"
                else "EXPENSE_APPROVED_LITE"
            )
            await _write_audit_log(db, audit_event, current_user.user_id, tenant_id, {
                "report_id": str(report.id),
                "report_number": report.report_number,
                "posting_mode": posting_result.mode,
                "posting_reference": posting_result.reference,
                "total_amount": str(report.total_amount),
            })

            employee_result = await db.execute(select(User).where(User.id == report.employee_id))
            employee = employee_result.scalar_one_or_none()
            if employee:
                _send_approval_complete_email(
                    to_email=employee.email,
                    report_number=report.report_number,
                    report_date=str(report.report_date),
                    total_amount=report.total_amount,
                )

    await db.flush()
    return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))


# ── Reject ────────────────────────────────────────────────────────────────────

@router.post("/{approval_id}/reject", response_model=ExpenseReportResponse)
async def reject(
    approval_id: uuid.UUID,
    data: RejectRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ExpenseReportResponse:
    """Reject an expense report with a mandatory comment."""
    tenant_id = _require_tenant(current_user)
    block_if_readonly_impersonation(current_user)

    result = await db.execute(
        select(ExpenseApproval).where(
            ExpenseApproval.id == approval_id,
            ExpenseApproval.tenant_id == tenant_id,
        )
    )
    approval = result.scalar_one_or_none()
    if not approval:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Approval record not found.")

    if approval.approver_id != current_user.user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN,
                            detail="You are not the designated approver for this record.")
    if approval.status != "PENDING":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval record has already been actioned.")

    report = await _get_report_or_404(approval.report_id, tenant_id, db)

    if report.current_approval_level != approval.level:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval level is not currently active for this report.")

    approval.status = "REJECTED"
    approval.comment = data.comment
    approval.actioned_at = datetime.now(timezone.utc)

    report.status = "REJECTED"
    report.rejection_comment = data.comment
    report.rejected_at_level = approval.level
    report.referred_back_from_level = None
    report.referred_back_levels = None
    report.current_approval_level = None

    await db.flush()

    await _write_audit_log(db, "EXPENSE_REJECTED", current_user.user_id, tenant_id, {
        "report_id": str(report.id),
        "report_number": report.report_number,
        "level": approval.level,
        "approver_id": str(approval.approver_id),
        "comment": data.comment,
        "total_amount": str(report.total_amount),
        "rejected_at_level": approval.level,
    })

    employee_result = await db.execute(select(User).where(User.id == report.employee_id))
    employee = employee_result.scalar_one_or_none()
    if employee:
        _send_rejection_email(
            to_email=employee.email,
            report_number=report.report_number,
            report_date=str(report.report_date),
            total_amount=report.total_amount,
            rejection_comment=data.comment,
        )

    return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))


# ── Refer Back ────────────────────────────────────────────────────────────────

@router.post("/{approval_id}/refer-back", response_model=ExpenseReportResponse)
async def refer_back(
    approval_id: uuid.UUID,
    data: ReferBackRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ExpenseReportResponse:
    """Refer back an expense report from the current approval level."""
    tenant_id = _require_tenant(current_user)
    block_if_readonly_impersonation(current_user)

    result = await db.execute(
        select(ExpenseApproval).where(
            ExpenseApproval.id == approval_id,
            ExpenseApproval.tenant_id == tenant_id,
        )
    )
    approval = result.scalar_one_or_none()
    if not approval:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Approval record not found.")

    if approval.approver_id != current_user.user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN,
                            detail="You are not the designated approver for this record.")
    if approval.status != "PENDING":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval record has already been actioned.")

    report = await _get_report_or_404(approval.report_id, tenant_id, db)

    if report.current_approval_level != approval.level:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail="This approval level is not currently active for this report.")

    approval.status = "REFERRED_BACK"
    approval.comment = data.comment
    approval.visible_to_requestor = data.visible_to_requestor
    approval.actioned_at = datetime.now(timezone.utc)

    if data.target_type == "requestor":
        report.status = "REFERRED_TO_REQUESTOR"
        report.rejection_comment = data.comment
        report.rejected_at_level = approval.level
        report.current_approval_level = None

        await _write_audit_log(db, "EXPENSE_REFERRED_BACK", current_user.user_id, tenant_id, {
            "report_id": str(report.id),
            "report_number": report.report_number,
            "level": approval.level,
            "referring_approver_id": str(approval.approver_id),
            "target_type": "requestor",
            "target_levels": [],
            "comment": data.comment,
            "visible_to_requestor": data.visible_to_requestor,
            "total_amount": str(report.total_amount),
        })

        if data.visible_to_requestor:
            employee_result = await db.execute(select(User).where(User.id == report.employee_id))
            employee = employee_result.scalar_one_or_none()
            if employee:
                _send_refer_back_email(
                    to_email=employee.email,
                    report_number=report.report_number,
                    comment=data.comment,
                    referring_level=approval.level,
                )

    else:
        if not data.target_levels:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="target_levels is required when target_type is \'approver\'.",
            )

        invalid = [lvl for lvl in data.target_levels if lvl >= approval.level]
        if invalid:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"All target levels must be lower than the current level ({approval.level}).",
            )

        sorted_levels = sorted(set(data.target_levels))

        target_approvals: dict[int, ExpenseApproval] = {}
        for lvl in sorted_levels:
            target_result = await db.execute(
                select(ExpenseApproval).where(
                    ExpenseApproval.report_id == report.id,
                    ExpenseApproval.level == lvl,
                ).order_by(ExpenseApproval.created_at.desc())
            )
            ta = target_result.scalars().first()
            if not ta:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"No approval record found for level {lvl}.",
                )
            target_approvals[lvl] = ta

        first_level = sorted_levels[0]
        remaining = sorted_levels[1:]

        first_target = target_approvals[first_level]
        first_target.status = "PENDING"
        first_target.actioned_at = None

        report.current_approval_level = first_level
        report.referred_back_from_level = approval.level
        report.referred_back_levels = remaining if remaining else None

        await _write_audit_log(db, "EXPENSE_REFERRED_BACK", current_user.user_id, tenant_id, {
            "report_id": str(report.id),
            "report_number": report.report_number,
            "level": approval.level,
            "referring_approver_id": str(approval.approver_id),
            "target_type": "approver",
            "target_levels": sorted_levels,
            "comment": data.comment,
            "visible_to_requestor": data.visible_to_requestor,
            "total_amount": str(report.total_amount),
        })

        referring_approver_result = await db.execute(select(User).where(User.id == approval.approver_id))
        referring_approver = referring_approver_result.scalar_one_or_none()
        first_approver_result = await db.execute(select(User).where(User.id == first_target.approver_id))
        first_approver = first_approver_result.scalar_one_or_none()
        if first_approver and referring_approver:
            _send_referred_approver_email(
                to_email=first_approver.email,
                report_number=report.report_number,
                referring_approver_name=referring_approver.full_name,
                referring_level=approval.level,
                comment=data.comment,
            )

    await db.flush()
    return ExpenseReportResponse.from_orm(await _reload_report(report.id, db))


# ── Finance Review Steps ──────────────────────────────────────────────────────

def _serialize_step(step, emp_name=None):
    return {
        "id": str(step.id),
        "policy_id": str(step.policy_id),
        "tenant_id": str(step.tenant_id),
        "level": step.level,
        "step_type": step.step_type,
        "label": step.label,
        "function_code": step.function_code,
        "assigned_employee_id": str(step.assigned_employee_id) if step.assigned_employee_id else None,
        "assigned_designation": step.assigned_designation,
        "min_amount": float(step.min_amount) if step.min_amount is not None else None,
        "can_send_back": step.can_send_back,
        "can_correct_gl": step.can_correct_gl,
        "is_required": step.is_required,
        "instructions": step.instructions,
        "created_at": step.created_at,
        "updated_at": step.updated_at,
        "assigned_employee_name": emp_name,
    }


@router.get("/policies/{policy_id}/finance-steps", response_model=list[FinanceReviewStepResponse])
async def list_finance_steps(
    policy_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Return all finance review steps for a policy, ordered by level."""
    tenant_id = current_user.tenant_id
    policy_result = await db.execute(
        select(ApprovalPolicy).where(ApprovalPolicy.id == policy_id, ApprovalPolicy.tenant_id == tenant_id)
    )
    if not policy_result.scalars().first():
        raise HTTPException(status_code=404, detail="Policy not found.")

    result = await db.execute(
        select(FinanceReviewStep)
        .where(FinanceReviewStep.policy_id == policy_id, FinanceReviewStep.tenant_id == tenant_id)
        .order_by(FinanceReviewStep.level)
    )
    steps = result.scalars().all()

    emp_ids = [s.assigned_employee_id for s in steps if s.assigned_employee_id]
    emp_name_map = {}
    if emp_ids:
        from app.models.master_data import Employee
        emp_result = await db.execute(select(Employee).where(Employee.id.in_(emp_ids)))
        for emp in emp_result.scalars().all():
            emp_name_map[emp.id] = f"{emp.first_name} {emp.last_name}"

    return [_serialize_step(s, emp_name_map.get(s.assigned_employee_id)) for s in steps]


@router.put("/policies/{policy_id}/finance-steps", response_model=list[FinanceReviewStepResponse])
async def bulk_save_finance_steps(
    policy_id: uuid.UUID,
    payload: FinanceReviewStepBulkSave,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
):
    """Replace all finance review steps for a policy (atomic bulk-save)."""
    tenant_id = current_user.tenant_id
    if current_user.role not in ("admin", "owner"):
        raise HTTPException(status_code=403, detail="Admin or Owner role required.")

    policy_result = await db.execute(
        select(ApprovalPolicy).where(ApprovalPolicy.id == policy_id, ApprovalPolicy.tenant_id == tenant_id)
    )
    if not policy_result.scalars().first():
        raise HTTPException(status_code=404, detail="Policy not found.")

    from app.models.master_data import Employee
    emp_ids_in = [uuid.UUID(s.assigned_employee_id) for s in payload.steps if s.assigned_employee_id]
    emp_name_map = {}
    if emp_ids_in:
        emp_result = await db.execute(
            select(Employee).where(Employee.id.in_(emp_ids_in), Employee.tenant_id == tenant_id)
        )
        found = emp_result.scalars().all()
        found_ids = {e.id for e in found}
        for e in found:
            emp_name_map[e.id] = f"{e.first_name} {e.last_name}"
        missing = [str(eid) for eid in emp_ids_in if eid not in found_ids]
        if missing:
            raise HTTPException(status_code=422, detail=f"Employee(s) not found: {', '.join(missing)}")

    await db.execute(
        delete(FinanceReviewStep).where(
            FinanceReviewStep.policy_id == policy_id, FinanceReviewStep.tenant_id == tenant_id
        )
    )

    sorted_steps = sorted(payload.steps, key=lambda s: s.level)
    new_rows = []
    for idx, step_in in enumerate(sorted_steps, start=1):
        emp_id = uuid.UUID(step_in.assigned_employee_id) if step_in.assigned_employee_id else None
        row = FinanceReviewStep(
            policy_id=policy_id,
            tenant_id=tenant_id,
            level=idx,
            step_type=step_in.step_type,
            label=step_in.label,
            function_code=step_in.function_code,
            assigned_employee_id=emp_id,
            assigned_designation=step_in.assigned_designation,
            min_amount=Decimal(str(step_in.min_amount)) if step_in.min_amount is not None else None,
            can_send_back=step_in.can_send_back,
            can_correct_gl=step_in.can_correct_gl,
            is_required=step_in.is_required,
            instructions=step_in.instructions,
        )
        db.add(row)
        new_rows.append(row)

    await db.commit()
    for row in new_rows:
        await db.refresh(row)

    return [
        _serialize_step(r, emp_name_map.get(r.assigned_employee_id) if r.assigned_employee_id else None)
        for r in new_rows
    ]
