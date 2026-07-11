"""
ZivaBI — M8.2/M8.3 Implementation Portal router.

Registered at prefix /api/setup.

Endpoints:
  GET   /api/setup/progress                   Setup dashboard completion state
  GET   /api/setup/org                        Organisation config
  PATCH /api/setup/org                        Update organisation config
  GET   /api/setup/org-structure              Org tree (recursive)
  POST  /api/setup/org-structure              Add a node
  PATCH /api/setup/org-structure/{id}         Update a node
  DELETE /api/setup/org-structure/{id}        Remove a node
  GET   /api/setup/org-structure/template     Download xlsx template
  POST  /api/setup/org-structure/upload       Upload structure from xlsx/csv
  POST   /api/setup/periods/generate                    Generate 12 monthly periods for a FY
  DELETE /api/setup/periods/fiscal-year/{fiscal_year}  Delete all periods for a FY (only if none are closed)
  GET   /api/setup/periods                    List accounting periods (filter by ?fiscal_year=)
  GET   /api/setup/periods/check              Postability check for a date (?date=YYYY-MM-DD)
  POST  /api/setup/periods/{id}/soft-close    Manually soft-close a period
  POST  /api/setup/periods/{id}/hard-close    Hard-close a period (sequential enforcement)
  POST  /api/setup/periods/{id}/reopen        Reopen a hard-closed period (consultant only; writes audit log)
  POST  /api/setup/periods/management-close   Stage 1 year-end close (Dec must be hard-closed)
  POST  /api/setup/periods/statutory-close    Stage 2 permanent year lock
  GET   /api/setup/periods/year-state         FiscalYearState (seeds OPEN if absent)
  PATCH /api/setup/periods/year-state/{fy}    Update audit grace months
  GET   /api/setup/periods/audit-log          Period audit log (filterable)
  GET   /api/setup/periods/checklist          List close checklist template items
  POST  /api/setup/periods/checklist          Add a checklist item
  PATCH /api/setup/periods/checklist/{id}     Update a checklist item
  DELETE /api/setup/periods/checklist/{id}    Soft-delete a checklist item
  GET   /api/setup/periods/{id}/checklist     Per-period checklist with completion state
  POST  /api/setup/periods/{id}/checklist/{item_id}/prepare  Mark item prepared
  POST  /api/setup/periods/{id}/checklist/{item_id}/approve  Mark item approved (preparer≠approver)
  GET   /api/setup/modules                    Activated modules list (with is_licensed)
  PATCH /api/setup/modules                    Update module activations (checks is_licensed)
  POST  /api/setup/dimensions/not-applicable   Mark tenant as not using dimensions
  GET   /api/setup/currencies                 FX config
  PATCH /api/setup/currencies                 Update FX config
  GET   /api/setup/tax                        Tax config
  PATCH /api/setup/tax                        Update tax config
  GET   /api/setup/roles/matrix               Permission matrix
  PATCH /api/setup/roles/matrix               Update permission matrix
  GET   /api/setup/roles/assignments          Role tier assignments
  POST  /api/setup/roles/assignments          Create role tier assignment
  PATCH /api/setup/roles/assignments/{id}     Update role tier assignment
  GET   /api/setup/documents                  Document rules (filter by ?module=)
  POST  /api/setup/documents                  Create document rule
  PATCH /api/setup/documents/{id}             Update document rule
  DELETE /api/setup/documents/{id}            Delete document rule
  POST  /api/setup/go-live                    Mark tenant as live (consultant only)

All endpoints are tenant-scoped and require authentication.
Admin-only: require is_tenant_admin or is_super_admin.
go-live / reopen: require is_super_admin (consultant role_tier stripped M9.3a).
"""

import calendar as _cal
import io
import uuid
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func as sqlfunc, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.middleware.auth import CurrentUser, require_auth, block_if_readonly_impersonation
from app.models.auth import AuditLog, UserTenant, User, Tenant
from app.models.account_mapping import TenantAccountMapping, PostingRole
from app.models.bank_account import BankAccount
from app.services.periods import (
    ALLOWED_MODULES,
    add_months,
    apply_auto_soft_close,
    check_audit_overdue,
    checklist_complete,
    compute_grace_expiry,
    generate_monthly_periods,
    get_matching_grace_row,
    initial_status_for,
    is_date_postable,
    parse_fy_start_year,
)
from app.models.expenses import TenantExpenseConfig
from app.models.master_data import ChartOfAccount, Employee, TenantDimension, TenantPermissionMatrix, UserFunctionalScope
from app.models.setup import (
    AccountingPeriod,
    CloseChecklistItem,
    DocumentRule,
    EmployeeOnboardingToken,
    FiscalYearState,
    FuturePostingException,
    OrgStructureNode,
    PeriodAuditLog,
    PeriodChecklistCompletion,
    PeriodGraceOverride,
    SystemFunctionMapping,
    TenantFxConfig,
    TenantModule,
    TenantOrgConfig,
    TenantTaxConfig,
)
from app.schemas.setup import (
    AccountingPeriodResponse,
    AuditGraceUpdate,
    BrandingUpdate,
    CloseChecklistItemCreate,
    CloseChecklistItemResponse,
    CloseChecklistItemUpdate,
    DocumentRuleCreate,
    DocumentRuleResponse,
    DocumentRuleUpdate,
    FiscalYearStateResponse,
    FiscalYearUpdate,
    FuturePostingExceptionCreate,
    FuturePostingExceptionResponse,
    FxConfigResponse,
    FxConfigUpdate,
    GeneratePeriodsRequest,
    GoLiveResponse,
    JournalBlockResponse,
    JournalBlockUpdate,
    ManagementCloseRequest,
    ModuleState,
    ModulesResponse,
    ModulesUpdate,
    OrgConfigResponse,
    OrgIdentityUpdate,
    OrgStructureNodeCreate,
    OrgStructureNodeResponse,
    OrgStructureNodeUpdate,
    OrgStructureTreeResponse,
    OrgStructureUpdate,
    OrgStructureUploadResult,
    PermissionMatrixResponse,
    PermissionMatrixUpdate,
    PeriodAuditLogResponse,
    PeriodChecklistCompletionResponse,
    PeriodChecklistEntryResponse,
    PeriodCheckResponse,
    PeriodGraceOverrideCreate,
    PeriodGraceOverrideResponse,
    PeriodGraceOverrideUpdate,
    ProgressResponse,
    ReopenRequest,
    FunctionalScopeItem,
    FunctionalScopeResponse,
    FunctionalScopeUpdate,
    RoleAssignmentCreate,
    RoleAssignmentResponse,
    RoleAssignmentUpdate,
    SectionStatus,
    StatutoryCloseRequest,
    TaxConfigResponse,
    TaxConfigUpdate,
    FunctionMappingItem,
    FunctionMappingUpsertItem,
    FunctionMappingsResponse,
    FunctionTeamMember,
    FunctionTeamResponse,
    FunctionUserItem,
    FunctionUsersResponse,
)

router = APIRouter(prefix="/api/setup", tags=["setup"])

# Country → functional currency map (ISO 4217)
COUNTRY_CURRENCY_MAP: dict[str, str] = {
    "NG": "NGN", "GH": "GHS", "KE": "KES", "ZA": "ZAR",
    "GB": "GBP", "US": "USD", "CA": "CAD", "AU": "AUD",
    "DE": "EUR", "FR": "EUR", "NL": "EUR", "AE": "AED",
    "SG": "SGD", "IN": "INR", "BR": "BRL", "JP": "JPY",
    "CN": "CNY", "EG": "EGP", "ET": "ETB", "RW": "RWF",
}

# ── Module catalogue (all 14 modules) ─────────────────────────────────────────

MODULE_CATALOGUE = [
    {"key": "expense",         "label": "Expense Management"},
    {"key": "ap",              "label": "Accounts Payable"},
    {"key": "ar",              "label": "Accounts Receivable"},
    {"key": "payroll",         "label": "Payroll & HR"},
    {"key": "inventory",       "label": "Inventory Management"},
    {"key": "fixed_assets",    "label": "Fixed Assets"},
    {"key": "posm",            "label": "POSM Management"},
    {"key": "vendor_portal",   "label": "Vendor Portal"},
    {"key": "customer_portal", "label": "Customer Portal"},
    {"key": "warehouse",       "label": "Warehouse / 3PL Portal"},
    {"key": "bank_recon",      "label": "Bank Reconciliation"},
    {"key": "budget",          "label": "Budget Engine"},
    {"key": "tax_engine",      "label": "Tax Engine"},
    {"key": "reporting",       "label": "Reporting & Analytics"},
]

MODULE_KEY_TO_LABEL = {m["key"]: m["label"] for m in MODULE_CATALOGUE}


# ── Helpers ────────────────────────────────────────────────────────────────────

# Tiers that pass the admin gate. consultant removed M9.3a (super admin impersonation
# replaces tenant consultant tier). functional_admin is intentionally excluded —
# it is function-scoped and must be checked at individual endpoints.
_ADMIN_TIERS: frozenset[str] = frozenset({"power_admin"})


def _require_admin(current_user: CurrentUser) -> None:
    """Raise 403 if the user lacks admin access.

    Passes for: is_super_admin OR role_tier == 'power_admin'.
    is_tenant_admin no longer grants config admin (M9.3c) — configuration is done
    by a super admin via impersonation or a power_admin tier user. is_tenant_admin
    is preserved in the JWT for potential future RBAC reintroduction.
    Also raises 403 for support-mode impersonation on live (read-only session).
    """
    if (
        not current_user.is_super_admin
        and current_user.role_tier not in _ADMIN_TIERS
    ):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required.",
        )
    block_if_readonly_impersonation(current_user)


def _require_consultant(current_user: CurrentUser) -> None:
    """Raise 403 if the user is not a super admin (Ziva consultant).

    M9.3a: consultant role_tier is stripped from tenant users. Only a super admin
    acting in implementation mode (or their native platform token) may call
    consultant-gated endpoints.
    """
    if not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Super admin (Ziva consultant) access required.",
        )


def _require_tenant(current_user: CurrentUser) -> uuid.UUID:
    """Return tenant_id or raise 400 if not in a tenant context."""
    if not current_user.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Tenant context required.",
        )
    return current_user.tenant_id


async def _get_or_create_org(tenant_id: uuid.UUID, db: AsyncSession) -> TenantOrgConfig:
    """Fetch the org config row, creating one seeded with functional currency if it does not exist."""
    result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = result.scalar_one_or_none()
    if org is None:
        # Derive functional currency from the tenant's country (fallback for legacy tenants)
        tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
        tenant = tenant_result.scalar_one_or_none()
        functional_currency = COUNTRY_CURRENCY_MAP.get(
            tenant.country if tenant else "", "USD"
        )
        org = TenantOrgConfig(
            tenant_id=tenant_id,
            functional_currency=functional_currency,
        )
        db.add(org)
        await db.flush()
    return org


async def _get_or_create_fx(tenant_id: uuid.UUID, db: AsyncSession) -> TenantFxConfig:
    """Fetch FX config, creating a blank row if needed."""
    result = await db.execute(
        select(TenantFxConfig).where(TenantFxConfig.tenant_id == tenant_id)
    )
    fx = result.scalar_one_or_none()
    if fx is None:
        fx = TenantFxConfig(tenant_id=tenant_id)
        db.add(fx)
        await db.flush()
    return fx


async def _get_or_create_tax(tenant_id: uuid.UUID, db: AsyncSession) -> TenantTaxConfig:
    """Fetch tax config, creating a blank row if needed."""
    result = await db.execute(
        select(TenantTaxConfig).where(TenantTaxConfig.tenant_id == tenant_id)
    )
    tax = result.scalar_one_or_none()
    if tax is None:
        tax = TenantTaxConfig(tenant_id=tenant_id)
        db.add(tax)
        await db.flush()
    return tax


def _org_to_response(org: Optional[TenantOrgConfig], tenant_id: uuid.UUID) -> OrgConfigResponse:
    """Convert ORM org config to response schema."""
    if org is None:
        return OrgConfigResponse(tenant_id=str(tenant_id))
    return OrgConfigResponse(
        tenant_id=str(org.tenant_id),
        legal_name=org.legal_name,
        rc_number=org.rc_number,
        date_of_registration=org.date_of_registration,
        commencement_date=org.commencement_date,
        company_type=org.company_type,
        industry=org.industry,
        tin=org.tin,
        vat_reg_number=org.vat_reg_number,
        country=org.country,
        registered_address=org.registered_address,
        operating_address=org.operating_address,
        company_phone=org.company_phone,
        company_email=org.company_email,
        website=org.website,
        external_auditor=org.external_auditor,
        group_structure=org.group_structure,
        parent_company_name=org.parent_company_name,
        functional_currency=org.functional_currency,
        reporting_currency=org.reporting_currency,
        authorised_share_capital=float(org.authorised_share_capital) if org.authorised_share_capital else None,
        first_fiscal_year_end=org.first_fiscal_year_end,
        fiscal_year_start_month=org.fiscal_year_start_month,
        fiscal_year_start_day=org.fiscal_year_start_day,
        fiscal_year_name_format=org.fiscal_year_name_format,
        period_closing_frequency=org.period_closing_frequency,
        branding=org.branding,
        org_configuration=org.org_configuration,
    )


def _build_tree(nodes: list[OrgStructureNode]) -> list[OrgStructureNodeResponse]:
    """Build a nested tree from a flat list of nodes (single pass)."""
    node_map: dict[str, OrgStructureNodeResponse] = {}
    for n in nodes:
        node_map[str(n.id)] = OrgStructureNodeResponse(
            id=str(n.id),
            parent_id=str(n.parent_id) if n.parent_id else None,
            node_type=n.node_type,
            name=n.name,
            code=n.code,
            cost_center_code=n.cost_center_code,
            entity_code=n.entity_code,
            is_active=n.is_active,
            sort_order=n.sort_order,
            children=[],
        )

    roots: list[OrgStructureNodeResponse] = []
    for n in nodes:
        resp = node_map[str(n.id)]
        if n.parent_id and str(n.parent_id) in node_map:
            node_map[str(n.parent_id)].children.append(resp)
        else:
            roots.append(resp)
    return roots


# ── Progress ───────────────────────────────────────────────────────────────────

@router.get("/progress", response_model=ProgressResponse)
async def get_progress(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ProgressResponse:
    """
    Return setup dashboard completion state for the current tenant.

    Fixed logic per M8.2 fixes brief:
    - Dimensions: complete if not_applicable flag set OR all dims have >= 1 value
    - Locked/unlocked sequence enforced
    - is_licensed enforced on module activation
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    # Fetch tenant flags
    tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = tenant_result.scalar_one_or_none()
    dims_not_applicable = bool(tenant and tenant.dimensions_not_applicable)
    docs_setup_complete = bool(tenant and tenant.documents_setup_complete)

    # Check org config
    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()
    org_config = org.org_configuration if org and org.org_configuration else {}
    use_multi_currency = org_config.get("use_multi_currency", True)
    posting_mode = org.posting_mode if org else "lite"
    # Sync with org_configuration — org_configuration is the source of truth
    if org_config:
        use_dimensions = org_config.get("use_dimensions")
        if use_dimensions is False:
            # Explicitly turned off — hide Dimensions
            dims_not_applicable = True
        elif use_dimensions is True:
            # Explicitly turned on — show Dimensions regardless of tenant flag
            dims_not_applicable = False
        # If use_dimensions is None (not set), fall back to tenant flag as-is
    org_complete = bool(org and org.legal_name and org.functional_currency)

    # Check modules (at least 1 active)
    mods_result = await db.execute(
        select(TenantModule).where(
            TenantModule.tenant_id == tenant_id, TenantModule.is_active.is_(True)
        )
    )
    active_modules = mods_result.scalars().all()
    modules_complete = len(active_modules) > 0

    # Check CoA
    coa_count_result = await db.execute(
        select(sqlfunc.count(ChartOfAccount.id)).where(
            ChartOfAccount.tenant_id == tenant_id,
            ChartOfAccount.is_active.is_(True),
        )
    )
    coa_count = coa_count_result.scalar_one() or 0
    coa_complete = coa_count > 0

    # Check dimensions
    dim_count_result = await db.execute(
        select(sqlfunc.count(TenantDimension.id)).where(
            TenantDimension.tenant_id == tenant_id,
            TenantDimension.is_active.is_(True),
        )
    )
    dim_count = dim_count_result.scalar_one() or 0
    dims_complete = dims_not_applicable or dim_count > 0
    dims_in_progress = not dims_not_applicable and dim_count > 0

    # Check employees
    emp_count_result = await db.execute(
        select(sqlfunc.count(Employee.id)).where(
            Employee.tenant_id == tenant_id,
            Employee.is_active.is_(True),
        )
    )
    emp_count = emp_count_result.scalar_one() or 0
    employees_complete = emp_count > 0

    # Check currencies (auto-complete if org functional_currency set)
    currencies_complete = bool(org and org.functional_currency)

    # Check tax (at least one rule configured)
    tax_result = await db.execute(
        select(TenantTaxConfig).where(TenantTaxConfig.tenant_id == tenant_id)
    )
    tax = tax_result.scalar_one_or_none()
    tax_complete = bool(tax and (tax.vat_config or tax.wht_config or tax.paye_config))

    # Check roles — at least 1 Power Admin assigned
    pa_result = await db.execute(
        select(sqlfunc.count(UserTenant.id)).where(
            UserTenant.tenant_id == tenant_id,
            UserTenant.role_tier == "power_admin",
            UserTenant.is_active.is_(True),
        )
    )
    pa_count = pa_result.scalar_one() or 0
    roles_complete = pa_count > 0

    # Check workflows — at least 1 expense approval matrix entry
    from app.models.approvals import ApprovalMatrix
    wf_result = await db.execute(
        select(sqlfunc.count(ApprovalMatrix.id)).where(
            ApprovalMatrix.tenant_id == tenant_id,
        )
    )
    wf_count = wf_result.scalar_one() or 0
    workflows_complete = wf_count > 0

    # Document rules (manually marked complete by consultant)
    docs_complete = docs_setup_complete

    # Module setup — expense config exists (proxy for at least 1 module configured)
    ec_result = await db.execute(
        select(TenantExpenseConfig).where(TenantExpenseConfig.tenant_id == tenant_id)
    )
    ec = ec_result.scalar_one_or_none()
    module_setup_complete = ec is not None

    # Account mapping: ALL system PostingRoles must be mapped.
    # am_count vs total_roles distinguishes not_started / in_progress / complete.
    # This prevents a tenant going live with gaps in the posting rule matrix.
    total_roles_result = await db.execute(select(sqlfunc.count(PostingRole.role_key)))
    total_roles = total_roles_result.scalar_one() or 0
    am_count_result = await db.execute(
        select(sqlfunc.count(TenantAccountMapping.id)).where(
            TenantAccountMapping.tenant_id == tenant_id
        )
    )
    am_count = am_count_result.scalar_one() or 0
    account_mapping_complete = total_roles > 0 and am_count >= total_roles
    account_mapping_in_progress = 0 < am_count < total_roles

    # Bank accounts (at least 1 active)
    ba_count_result = await db.execute(
        select(sqlfunc.count(BankAccount.id)).where(
            BankAccount.tenant_id == tenant_id,
            BankAccount.is_active.is_(True),
        )
    )
    ba_count = ba_count_result.scalar_one() or 0
    bank_accounts_complete = ba_count > 0

    # Accounting periods (at least 1 period generated)
    per_count_result = await db.execute(
        select(sqlfunc.count(AccountingPeriod.id)).where(
            AccountingPeriod.tenant_id == tenant_id
        )
    )
    per_count = per_count_result.scalar_one() or 0
    periods_complete = per_count > 0

    # Go-live blocking items — scope depends on posting_mode.
    # lite: no GL work needed; connected: CoA + account_mapping blocking;
    # full_erp: CoA + dims + account_mapping blocking.
    _base_blocking = [
        org_complete, modules_complete, employees_complete,
        tax_complete, roles_complete, workflows_complete,
    ]
    if posting_mode == "lite":
        blocking_complete = all(_base_blocking)
    elif posting_mode == "connected":
        blocking_complete = all(_base_blocking + [coa_complete, account_mapping_complete])
    else:  # full_erp
        blocking_complete = all(
            _base_blocking + [coa_complete, dims_complete, account_mapping_complete]
        )

    # Locked/unlocked sequence per brief
    def _s(
        key: str,
        label: str,
        complete: bool,
        subtitle: str,
        route: str,
        blocking: bool = True,
        locked: bool = False,
        in_progress: bool = False,
    ) -> SectionStatus:
        if locked:
            st = "locked"
        elif complete:
            st = "complete"
        elif in_progress:
            st = "in_progress"
        else:
            st = "not_started"
        return SectionStatus(
            key=key, label=label, status=st,
            subtitle=subtitle, route=route, blocking=blocking,
        )

    # Unlock sequence:
    # - Organisation: always
    # - Module activation: always
    # - Dimensions: unlocked after Organisation
    # - CoA: unlocked after Dimensions (or not_applicable)
    # - Employees: unlocked after CoA
    # - Currencies: unlocked after Organisation
    # - Tax: unlocked after Organisation
    # - Roles: unlocked after Employees
    # - Workflows: unlocked after Roles
    # - Documents: unlocked after Module activation
    # - Module setup: unlocked after CoA + Dimensions
    # - Go-live: unlocked when all blocking complete

    # GL-tier unlock logic (only relevant for connected / full_erp)
    dims_locked = not org_complete
    coa_locked = (
        (not (dims_complete or dims_not_applicable) or dims_locked)
        if posting_mode == "full_erp"
        else not org_complete
    )
    account_mapping_locked = not coa_complete or coa_locked
    bank_accounts_locked = not org_complete
    periods_locked = not org_complete
    # In lite mode employees unlock after org; in GL modes they unlock after CoA
    employees_locked = (
        (not coa_complete or coa_locked)
        if posting_mode != "lite"
        else not org_complete
    )
    currencies_locked = not org_complete
    tax_locked = not org_complete
    roles_locked = not employees_complete or employees_locked
    workflows_locked = not roles_complete or roles_locked
    docs_locked = not modules_complete
    module_setup_locked = (
        not (coa_complete and (dims_complete or dims_not_applicable))
        if posting_mode != "lite"
        else not modules_complete
    )
    golive_locked = not blocking_complete

    # ── Core sections (always shown) ─────────────────────────────────────────
    sections: list = [
        _s("organisation", "Organisation", org_complete,
           f"Legal name: {org.legal_name}" if org_complete else "Not configured",
           "/dashboard/business/setup/organisation"),
        _s("modules", "Module activation", modules_complete,
           f"{len(active_modules)} module(s) active" if modules_complete else "No modules activated",
           "/dashboard/business/setup/modules"),
    ]

    # ── GL sections (connected + full_erp only) ───────────────────────────────
    if posting_mode in ("connected", "full_erp"):
        if not dims_not_applicable and posting_mode == "full_erp":
            sections.append(
                _s("dimensions", "Dimensions", dims_complete,
                   f"{dim_count} dimension(s) configured" if dims_complete else "Not configured",
                   "/dashboard/business/settings/dimensions",
                   locked=dims_locked, in_progress=dims_in_progress)
            )
        sections.append(
            _s("coa", "Chart of accounts", coa_complete,
               f"{coa_count:,} GL accounts loaded" if coa_complete else (
                   "Requires Dimensions first" if coa_locked else "No accounts loaded"),
               "/dashboard/business/settings/chart-of-accounts",
               locked=coa_locked)
        )
        sections.append(
            _s("account_mapping", "Account mapping", account_mapping_complete,
               f"{am_count} of {total_roles} role(s) mapped" if am_count > 0 else (
                   "Requires CoA first" if account_mapping_locked else "No mappings configured"),
               "/dashboard/business/setup/account-mapping",
               locked=account_mapping_locked,
               in_progress=account_mapping_in_progress)
        )

    # ── full_erp-only GL sections ──────────────────────────────────────────────
    if posting_mode == "full_erp":
        sections.extend([
            _s("periods", "Accounting periods", periods_complete,
               f"{per_count} period(s) generated" if periods_complete else (
                   "Requires Organisation first" if periods_locked else "No periods generated"),
               "/dashboard/business/setup/periods", blocking=False,
               locked=periods_locked),
            _s("bank_accounts", "Bank accounts", bank_accounts_complete,
               f"{ba_count} account(s) configured" if bank_accounts_complete else (
                   "Requires Organisation first" if bank_accounts_locked else "Not configured"),
               "/dashboard/business/setup/bank-accounts", blocking=False,
               locked=bank_accounts_locked),
        ])

    # ── Remaining core sections ────────────────────────────────────────────────
    sections.extend([
        _s("employees", "Employees", employees_complete,
           f"{emp_count:,} employee(s) loaded" if employees_complete else (
               "Requires CoA first" if employees_locked and posting_mode != "lite"
               else ("Requires Organisation first" if employees_locked else "No employees loaded")),
           "/dashboard/business/settings/employees",
           locked=employees_locked),
        *([_s("currencies", "Currencies & FX", currencies_complete,
               f"Functional: {org.functional_currency}" if currencies_complete else (
                   "Requires Organisation first" if currencies_locked else "Not configured"),
               "/dashboard/business/setup/currencies", blocking=False,
               locked=currencies_locked),
        ] if use_multi_currency else []),
        _s("tax", "Tax & statutory", tax_complete,
           "Tax rules configured" if tax_complete else (
               "Requires Organisation first" if tax_locked else "Not configured"),
           "/dashboard/business/setup/tax",
           locked=tax_locked),
        _s("roles", "Roles & permissions", roles_complete,
           f"{pa_count} Power Admin(s) assigned" if roles_complete else (
               "Requires Employees first" if roles_locked else "No Power Admin assigned"),
           "/dashboard/business/setup/roles",
           locked=roles_locked),
        _s("workflows", "Approval workflows", workflows_complete,
           "Workflows configured" if workflows_complete else (
               "Requires Roles first" if workflows_locked else "Not configured"),
           "/dashboard/business/settings/approval-matrix",
           locked=workflows_locked),
        _s("documents", "Document rules", docs_complete,
           "Marked complete" if docs_complete else (
               "Requires Module activation first" if docs_locked else "Not marked complete"),
           "/dashboard/business/setup/documents", blocking=False,
           locked=docs_locked),
        _s("module_setup", "Module setup", module_setup_complete,
           "Modules configured" if module_setup_complete else (
               "Requires CoA & Dimensions first" if module_setup_locked and posting_mode != "lite"
               else ("Requires Module activation first" if module_setup_locked else "Not configured")),
           "/dashboard/business/settings/expense-config", blocking=False,
           locked=module_setup_locked),
        _s("golive", "Go-live", blocking_complete,
           "All blocking items complete" if blocking_complete else "Blocking items incomplete",
           "/dashboard/business/setup/go-live",
           locked=golive_locked),
    ])

    completed = sum(1 for s in sections if s.status == "complete")
    total = len(sections)
    pct = round(completed / total * 100) if total else 0

    return ProgressResponse(
        sections=sections, total=total, completed=completed, percentage=pct,
        lifecycle_status=tenant.lifecycle_status if tenant else "in_implementation",
        posting_mode=posting_mode,
    )


# ── Organisation ───────────────────────────────────────────────────────────────

@router.get("/org", response_model=OrgConfigResponse)
async def get_org(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgConfigResponse:
    """Return the tenant org configuration (all four tabs)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = result.scalar_one_or_none()
    return _org_to_response(org, tenant_id)


@router.patch("/org", response_model=OrgConfigResponse)
async def patch_org(
    data: OrgIdentityUpdate | OrgStructureUpdate | BrandingUpdate | FiscalYearUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgConfigResponse:
    """Update org config (any subset of fields)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    org = await _get_or_create_org(tenant_id, db)

    PROTECTED_ORG_FIELDS = {"functional_currency"}

    update_data = data.model_dump(exclude_unset=True)

    # ── first_fiscal_year_end: validate before applying ────────────────────────
    # When this field is provided, we validate it falls within one year of the
    # earlier of date_of_registration / commencement_date, then derive the
    # recurring fiscal_year_start_month and fiscal_year_start_day automatically.
    raw_fye = update_data.get("first_fiscal_year_end")
    if raw_fye is not None:
        fye: date = raw_fye
        # Prefer already-updated anchor values from the same payload
        reg_date = update_data.get("date_of_registration", org.date_of_registration)
        comm_date = update_data.get("commencement_date", org.commencement_date)
        anchor: Optional[date] = None
        if reg_date and comm_date:
            anchor = min(reg_date, comm_date)
        elif reg_date:
            anchor = reg_date
        elif comm_date:
            anchor = comm_date
        if anchor is None:
            raise HTTPException(
                status_code=400,
                detail="Set registration or commencement date before fiscal year end.",
            )
        max_fy_end = date(anchor.year + 2, anchor.month, anchor.day) - timedelta(days=1)
        if not (anchor <= fye <= max_fy_end):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"First fiscal year end must be between "
                    f"{anchor.strftime('%d/%m/%Y')} and "
                    f"{max_fy_end.strftime('%d/%m/%Y')}."
                ),
            )

    for field, value in update_data.items():
        if field in PROTECTED_ORG_FIELDS:
            continue
        if hasattr(org, field):
            setattr(org, field, value)

    # Derive fiscal_year_start_month/day from first_fiscal_year_end after applying update
    if raw_fye is not None:
        org.fiscal_year_start_month = (raw_fye.month % 12) + 1
        org.fiscal_year_start_day = 1

    # Sync dimensions toggle with tenant.dimensions_not_applicable
    org_config_update = update_data.get("org_configuration")
    if org_config_update is not None:
        use_dims = org_config_update.get("use_dimensions")
        if use_dims is not None:
            tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
            tenant = tenant_result.scalar_one_or_none()
            if tenant:
                tenant.dimensions_not_applicable = not use_dims

    await db.commit()
    await db.refresh(org)

    # ── Auto-generate current fiscal year on settings save ───────────────────
    # Trigger 1: when the tenant saves fiscal year settings, silently generate
    # periods for the current FY if they don't exist yet. This is the primary
    # discovery moment — the admin just configured the FY, so generate immediately.
    # All validations (registration floor, stub year, etc.) are inside the helper.
    current_year = date.today().year
    # Idempotency check is by date-range overlap, not by formatted label string —
    # the label (fiscal_year) embeds start_month/format and can change between
    # saves (e.g. start_month re-derived from a new first_fiscal_year_end), which
    # would otherwise make this check miss the existing year and double-generate.
    existing_current = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.start_date <= date(current_year, 12, 31),
            AccountingPeriod.end_date >= date(current_year, 1, 1),
        ).limit(1)
    )
    if not existing_current.scalar_one_or_none():
        created = await _generate_periods_for_year(db, tenant_id, current_year, org)
        if created:
            await db.commit()

    return _org_to_response(org, tenant_id)


# ── Org Structure ─────────────────────────────────────────────────────────────



@router.get("/entity-options", response_model=list[dict])
async def list_entity_options(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[dict]:
    """
    Return all Legal entity nodes for the tenant as lightweight dropdown options.
    Used by the approval-role form to associate roles with a specific entity.
    Response: [{id, name, code, entity_code}]
    """
    tenant_id = _require_tenant(current_user)
    result = await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Legal entity",
            OrgStructureNode.is_active.is_(True),
        ).order_by(OrgStructureNode.name)
    )
    nodes = result.scalars().all()
    return [
        {"id": str(n.id), "name": n.name, "code": n.code, "entity_code": n.entity_code}
        for n in nodes
    ]

@router.get("/org-structure", response_model=OrgStructureTreeResponse)
async def get_org_structure(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgStructureTreeResponse:
    """Return the full org hierarchy as a nested tree (single recursive query)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(OrgStructureNode)
        .where(OrgStructureNode.tenant_id == tenant_id, OrgStructureNode.is_active.is_(True))
        .order_by(OrgStructureNode.sort_order, OrgStructureNode.name)
    )
    flat = result.scalars().all()
    return OrgStructureTreeResponse(nodes=_build_tree(flat))


@router.post("/org-structure", response_model=OrgStructureNodeResponse, status_code=201)
async def create_org_node(
    data: OrgStructureNodeCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgStructureNodeResponse:
    """Add a new node to the org hierarchy."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    node = OrgStructureNode(
        tenant_id=tenant_id,
        parent_id=data.parent_id,
        node_type=data.node_type,
        name=data.name,
        code=data.code,
        cost_center_code=data.cost_center_code,
        entity_code=data.entity_code,
    )
    db.add(node)
    try:
        await db.commit()
    except Exception:
        await db.rollback()
        raise HTTPException(status_code=409, detail="A node with this code already exists.")

    await db.refresh(node)
    return OrgStructureNodeResponse(
        id=str(node.id),
        parent_id=str(node.parent_id) if node.parent_id else None,
        node_type=node.node_type,
        name=node.name,
        code=node.code,
        cost_center_code=node.cost_center_code,
        entity_code=node.entity_code,
        is_active=node.is_active,
        sort_order=node.sort_order,
    )


@router.patch("/org-structure/{node_id}", response_model=OrgStructureNodeResponse)
async def update_org_node(
    node_id: uuid.UUID,
    data: OrgStructureNodeUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgStructureNodeResponse:
    """Update an existing org node."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.id == node_id,
            OrgStructureNode.tenant_id == tenant_id,
        )
    )
    node = result.scalar_one_or_none()
    if not node:
        raise HTTPException(status_code=404, detail="Org node not found.")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(node, field, value)

    await db.commit()
    await db.refresh(node)
    return OrgStructureNodeResponse(
        id=str(node.id),
        parent_id=str(node.parent_id) if node.parent_id else None,
        node_type=node.node_type,
        name=node.name,
        code=node.code,
        cost_center_code=node.cost_center_code,
        entity_code=node.entity_code,
        is_active=node.is_active,
        sort_order=node.sort_order,
    )


@router.delete("/org-structure/{node_id}", status_code=204)
async def delete_org_node(
    node_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete an org node."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(OrgStructureNode).where(
            OrgStructureNode.id == node_id,
            OrgStructureNode.tenant_id == tenant_id,
        )
    )
    node = result.scalar_one_or_none()
    if not node:
        raise HTTPException(status_code=404, detail="Org node not found.")
    node.is_active = False
    await db.commit()


@router.get("/org-structure/template")
async def download_org_structure_template(
    current_user: CurrentUser = Depends(require_auth),
) -> StreamingResponse:
    """Generate and stream a .xlsx template for org structure upload."""
    _require_admin(current_user)
    _require_tenant(current_user)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    wb = openpyxl.Workbook()

    # ── Instructions sheet ────────────────────────────────────────────────────
    ins = wb.active
    ins.title = "Instructions"
    ins.column_dimensions["A"].width = 26
    ins.column_dimensions["B"].width = 62

    ins["A1"] = "Org Structure Upload — Instructions"
    ins["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ins["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
    ins["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ins.merge_cells("A1:B1")
    ins.row_dimensions[1].height = 36

    ins["A2"] = "Use the 'Org Structure' sheet to build your organisation hierarchy. Delete sample rows before uploading."
    ins["A2"].font = Font(name="Arial", size=10, color="555555")
    ins["A2"].fill = PatternFill("solid", fgColor="F7F9FC")
    ins.merge_cells("A2:B2")
    ins.row_dimensions[2].height = 22

    ins.append(["", ""])

    for col, val in [("A4", "Column"), ("B4", "Description")]:
        ins[col].value = val
        ins[col].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ins[col].fill = PatternFill("solid", fgColor="2D6A9F")
        ins[col].alignment = Alignment(vertical="center", indent=1)
    ins.row_dimensions[4].height = 24

    col_rows = [
        ("Node Type *",      "Required. Select from dropdown:\n• Parent company — top-level holding / group entity (e.g. Red Bull GMBH)\n• Legal entity — subsidiary registered in a specific country\n• Division / Business unit — major grouping\n• Department — functional unit\n• Cost center — lowest level where costs are posted"),
        ("Name *",           "Required. Full display name. Examples: Sales, Finance, Off Premise"),
        ("Code *",           "Required. Unique short code, no spaces. Examples: N22341SA, NG_FIN"),
        ("Parent Code",      "Code of this node's parent. Leave blank only for the top-level Legal entity.\nMust exactly match a Code in this file or already in the system."),
        ("Cost Center Code", "Required when Node Type = Cost center. Must match the dimension value code in Ziva BI.\nLeave blank for Legal entity, Division, and Department nodes.\nConditional formatting flags missing values in red."),
        ("Entity Code",      "Optional. Legal entity nodes only. Stores the ERP profit centre / entity code.\nExample: N22341 (Sage X3 profit centre for Red Bull Nigeria)."),
        ("Description",      "Optional. Any notes about this node."),
    ]

    section_fill = PatternFill("solid", fgColor="EBF2FF")
    alt_fill = PatternFill("solid", fgColor="F7F9FC")

    for i, (col, desc) in enumerate(col_rows):
        r = 5 + i
        c1 = ins.cell(row=r, column=1, value=col)
        c1.font = Font(name="Arial", bold=True, size=11)
        c1.fill = section_fill if i % 2 == 0 else alt_fill
        c1.alignment = Alignment(vertical="top", indent=1)
        c2 = ins.cell(row=r, column=2, value=desc)
        c2.font = Font(name="Arial", size=11)
        c2.fill = section_fill if i % 2 == 0 else alt_fill
        c2.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ins.row_dimensions[r].height = 56

    warn_row = 5 + len(col_rows) + 1
    ins.cell(row=warn_row, column=1, value="⚠  Important").font = Font(name="Arial", bold=True, size=10, color="7B4F00")
    ins.cell(row=warn_row, column=1).fill = PatternFill("solid", fgColor="FFF8E1")
    ins.cell(row=warn_row, column=1).alignment = Alignment(vertical="top", indent=1)
    ins.cell(row=warn_row, column=2, value=(
        "• Do not delete or rename column headers.\n"
        "• Do not add extra columns.\n"
        "• Codes must be unique — duplicates will be rejected.\n"
        "• Parent Code must reference a Code that exists in this file or already in the system.\n"
        "• Uploading replaces the existing structure. Back up before uploading.\n"
        "• Red highlighted cells = Cost center nodes missing a Cost Center Code. Fix before uploading."
    )).font = Font(name="Arial", size=10, color="7B4F00")
    ins.cell(row=warn_row, column=2).fill = PatternFill("solid", fgColor="FFF8E1")
    ins.cell(row=warn_row, column=2).alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    ins.row_dimensions[warn_row].height = 88

    # ── Org Structure sheet ───────────────────────────────────────────────────
    ws = wb.create_sheet("Org Structure")

    col_widths = [26, 28, 18, 18, 22, 18, 30]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Node Type *", "Name *", "Code *", "Parent Code",
               "Cost Center Code", "Entity Code", "Description"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Node Type dropdown
    dv = DataValidation(
        type="list",
        formula1='"Parent company,Legal entity,Division / Business unit,Department,Cost center"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid node type",
        error="Select from: Parent company, Legal entity, Division / Business unit, Department, Cost center",
    )
    ws.add_data_validation(dv)
    dv.sqref = "A2:A500"

    # Conditional formatting: amber row + red cell when Cost center missing code
    ws.conditional_formatting.add(
        "A2:G500",
        FormulaRule(
            formula=['AND($A2="Cost center",$E2="")'],
            fill=PatternFill("solid", fgColor="FFF3CD"),
        )
    )
    ws.conditional_formatting.add(
        "E2:E500",
        FormulaRule(
            formula=['AND(A2="Cost center",E2="")'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
            font=Font(name="Arial", size=10, color="CC0000"),
        )
    )
    ws.conditional_formatting.add(
        "E2:E500",
        FormulaRule(
            formula=['AND(A2="Cost center",E2<>"")'],
            fill=PatternFill("solid", fgColor="D4EDDA"),
        )
    )

    # Sample rows
    samples = [
        ("Legal entity",            "Acme Corporation", "ACME",   "",     "",         "ACM001", "Top-level legal entity"),
        ("Cost center",             "Finance",          "FIN",    "ACME", "FIN001",   "",       "Finance department"),
        ("Cost center",             "Sales",            "SAL",    "ACME", "SAL001",   "",       "Sales department"),
        ("Cost center",             "Off Premise",      "SAL_OP", "SAL",  "SAL_OP01", "",       "Off premise sales"),
        ("Cost center",             "Marketing",        "MKT",    "ACME", "",         "",       "Missing cost center code — intentional demo of red flag"),
    ]
    s_font = Font(name="Arial", size=10, color="444444", italic=True)
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = s_font
            cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 20

    note_row = len(samples) + 3
    ws.cell(row=note_row, column=1,
        value="↑ Delete all sample rows above before uploading. Marketing row intentionally has no Cost Center Code to demonstrate the red flag."
    ).font = Font(name="Arial", size=10, color="AA0000", italic=True)
    ws.merge_cells(f"A{note_row}:G{note_row}")

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=org_structure_template.xlsx"},
    )


@router.post("/org-structure/upload", response_model=OrgStructureUploadResult)
async def upload_org_structure(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> OrgStructureUploadResult:
    """Upload org structure from .xlsx or .csv. Upserts by code."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    content = await file.read()
    rows: list[dict] = []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        # Try "Org Structure" sheet first, fall back to active
        ws = wb["Org Structure"] if "Org Structure" in wb.sheetnames else wb.active

        VALID_NODE_TYPES = {"Parent company", "Legal entity", "Division / Business unit", "Department", "Cost center"}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            # Skip header rows and note rows
            first_cell = str(row[0]).strip() if row[0] else ""
            if first_cell.startswith("↑"):
                continue
            if first_cell.lower() in ("node type*", "node type *", "node type"):
                continue
            if first_cell not in VALID_NODE_TYPES:
                # Could be an instruction row — skip silently if name/code also empty
                if not row[1] and not row[2]:
                    continue
            rows.append({
                "node_type":        first_cell,
                "name":             str(row[1]).strip() if row[1] else "",
                "code":             str(row[2]).strip() if row[2] else "",
                "parent_code":      str(row[3]).strip() if row[3] else None,
                "cost_center_code": str(row[4]).strip() if row[4] else None,
                "entity_code":      str(row[5]).strip() if row[5] else None,
            })
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    VALID_TYPES = {"Parent company", "Legal entity", "Division / Business unit", "Department", "Cost center"}
    imported = updated = 0
    errors: list[dict] = []

    # Fetch existing nodes for this tenant
    existing_result = await db.execute(
        select(OrgStructureNode).where(OrgStructureNode.tenant_id == tenant_id)
    )
    existing_map = {n.code: n for n in existing_result.scalars().all()}

    # Pass 1: upsert all nodes without parent assignment
    code_to_row = {}
    for row_num, row in enumerate(rows, start=2):
        if not row["node_type"] or row["node_type"] not in VALID_TYPES:
            errors.append({"row": row_num, "reason": f"Invalid Node Type: '{row['node_type']}'"})
            continue
        if not row["name"] or not row["code"]:
            errors.append({"row": row_num, "reason": "Name and Code are required."})
            continue
        code_to_row[row["code"]] = (row_num, row)

        if row["code"] in existing_map:
            node = existing_map[row["code"]]
            node.name = row["name"]
            node.node_type = row["node_type"]
            node.cost_center_code = row["cost_center_code"]
            node.entity_code = row["entity_code"]
            node.is_active = True
            updated += 1
        else:
            node = OrgStructureNode(
                tenant_id=tenant_id,
                node_type=row["node_type"],
                name=row["name"],
                code=row["code"],
                parent_id=None,  # set in pass 2
                cost_center_code=row["cost_center_code"],
                entity_code=row["entity_code"],
            )
            db.add(node)
            existing_map[row["code"]] = node
            imported += 1

    # Flush so all nodes have IDs
    await db.flush()

    # Pass 2: assign parent_id now that all nodes exist
    for row_num, row in code_to_row.values():
        if not row["parent_code"]:
            continue
        node = existing_map.get(row["code"])
        parent = existing_map.get(row["parent_code"])
        if not parent:
            errors.append({"row": row_num, "reason": f"Parent code '{row['parent_code']}' not found."})
            continue
        if node:
            node.parent_id = parent.id

    await db.commit()
    return OrgStructureUploadResult(imported=imported, updated=updated, errors=errors)


# ── Accounting Periods (M8.3 Brief 1) ────────────────────────────────────────

def _build_fy_label(fmt: Optional[str], year: int, start_month: int = 1) -> str:
    """Build a fiscal year label from the tenant's name format template and a start year.

    Handles both the new format codes (YYYY, FYYYYY, YYYY/YYYY, YYYY-YYYY,
    MMM YYYY - MMM YYYY) and the legacy template codes ({year}, {nextyear}, MMM)
    for backward compatibility.

    Args:
        fmt:          The fiscal_year_name_format value from TenantOrgConfig.
                      Defaults to "FYYYYY" (e.g. "FY2026") when None.
        year:         The calendar year the fiscal year starts in.
        start_month:  The FY start month (1=Jan, …, 12=Dec). Used for the
                      "MMM YYYY - MMM YYYY" format to compute the end month.
    """
    template = fmt or "FYYYYY"
    next_year = year + 1

    # Compute month abbreviations for "MMM YYYY - MMM YYYY" format
    start_mon_abbr = _cal.month_abbr[start_month]
    end_mon_num = ((start_month - 2) % 12) + 1  # month immediately before start
    end_mon_abbr = _cal.month_abbr[end_mon_num]
    # The end year equals the start year when the FY end month falls later in
    # the calendar year than the start month (e.g. Jan→Dec within same year).
    # Otherwise the end falls in the following calendar year (e.g. Apr→Mar).
    end_year = year if end_mon_num >= start_month else next_year

    return (
        template
        # New format codes — order matters: FYYYYY before plain YYYY
        .replace("FYYYYY", f"FY{year}")
        .replace("YYYY/YYYY", f"{year}/{next_year}")
        .replace("YYYY-YYYY", f"{year}-{next_year}")
        .replace("MMM YYYY - MMM YYYY", f"{start_mon_abbr} {year} - {end_mon_abbr} {end_year}")
        .replace("YYYY", str(year))
        # Legacy template codes (backward compat for existing tenants)
        .replace("{year}", str(year))
        .replace("{nextyear}", str(next_year))
        .replace("MMM", _cal.month_abbr[1])
    )


async def _generate_periods_for_year(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    year: int,
    org: TenantOrgConfig,
    *,
    label: Optional[str] = None,
) -> list[AccountingPeriod]:
    """Insert monthly AccountingPeriod rows for the given fiscal year start year.

    Shared by the deprecated POST /periods/generate endpoint and both auto-generation
    triggers (on fiscal settings save and on last-period hard-close).

    Parameters
    ----------
    db        Async session. Caller is responsible for committing.
    tenant_id Tenant UUID.
    year      Calendar year of the FY start (e.g. 2026 for FY2026 / 2026-27).
    org       TenantOrgConfig — must be loaded and up to date.
    label     Fiscal year label stored on AccountingPeriod.fiscal_year. Defaults to
              _build_fy_label(org.fiscal_year_name_format, year).

    Returns the list of created (flushed) AccountingPeriod rows, or an empty list
    when org config is not ready (missing start month, non-monthly frequency, or the
    year is before the company's registration date).

    Validations applied
    -------------------
    - fiscal_year_start_month must be set.
    - period_closing_frequency must be monthly (or unset, which defaults to monthly).
    - year must be >= year(date_of_registration) when registration date is set.
    - Stub first year: if fy_start falls before date_of_registration in the same
      calendar year, fy_start is clamped to the registration date.
    - fy_end is always derived from tenant config (day before next FY starts),
      never from fy_start + 12 months.

    Does NOT check whether periods already exist — the caller decides.
    """
    if not org.fiscal_year_start_month:
        return []

    frequency = (org.period_closing_frequency or "monthly").lower()
    if frequency != "monthly":
        return []

    start_month: int = org.fiscal_year_start_month
    start_day: int = org.fiscal_year_start_day or 1
    fy_label = label or _build_fy_label(org.fiscal_year_name_format, year, start_month)

    # Anchor date: the earlier of date_of_registration and commencement_date.
    # Used as the fy_start for the first fiscal year and as the registration floor.
    anchor: Optional[date] = None
    if org.date_of_registration and org.commencement_date:
        anchor = min(org.date_of_registration, org.commencement_date)
    elif org.date_of_registration:
        anchor = org.date_of_registration
    elif org.commencement_date:
        anchor = org.commencement_date

    # Year-level floor: skip years before the anchor
    if anchor and year < anchor.year:
        return []

    # Determine fy_start
    if org.first_fiscal_year_end and anchor and year == anchor.year:
        # First fiscal year: start exactly from the anchor date.
        # first_fiscal_year_end confirms the tenant has configured this year as the
        # first FY, so we don't apply the normal start_month/day.
        fy_start = anchor
    else:
        # Subsequent fiscal years: start from the configured start_month/start_day.
        fy_start = date(year, start_month, start_day)
        # Legacy stub-year clamp when first_fiscal_year_end is not configured:
        # if the computed start is before the anchor (same year), clamp to anchor.
        if not org.first_fiscal_year_end and anchor and fy_start < anchor:
            if year == anchor.year:
                fy_start = anchor
            else:
                return []

    # Fiscal year end = day before next FY starts (derived from config, not +12 months)
    next_fy_start = date(year + 1, start_month, start_day)
    fy_end = next_fy_start - timedelta(days=1)

    period_specs = generate_monthly_periods(fy_start, num_periods=12, start_day=start_day)
    period_specs = [spec for spec in period_specs if spec[2] <= fy_end]

    created: list[AccountingPeriod] = []
    for period_no, period_name, start, end in period_specs:
        p = AccountingPeriod(
            tenant_id=tenant_id,
            fiscal_year=fy_label,
            period_no=period_no,
            period_name=period_name,
            start_date=start,
            end_date=end,
            status=initial_status_for(start, end),
        )
        db.add(p)
        created.append(p)

    if created:
        await db.flush()

    return created


def _period_to_response(p: AccountingPeriod) -> AccountingPeriodResponse:
    """Map an AccountingPeriod ORM row to the API response schema."""
    return AccountingPeriodResponse(
        id=str(p.id),
        tenant_id=str(p.tenant_id),
        fiscal_year=p.fiscal_year,
        period_no=p.period_no,
        period_name=p.period_name,
        start_date=p.start_date,
        end_date=p.end_date,
        status=p.status,
        hard_closed_at=p.hard_closed_at,
        hard_closed_by=str(p.hard_closed_by) if p.hard_closed_by else None,
        soft_closed_at=p.soft_closed_at,
        grace_expires_at=p.grace_expires_at,
        reopened_count=p.reopened_count,
    )


@router.post("/periods/generate", response_model=list[AccountingPeriodResponse], status_code=201)
async def generate_periods(
    data: GeneratePeriodsRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[AccountingPeriodResponse]:
    """Generate monthly accounting periods for the given fiscal year label.

    Deprecated: auto-generation now handles this — periods are created automatically
    on fiscal year settings save and when the last period of a fiscal year is
    hard-closed. This endpoint remains for backward compatibility and manual override.

    User-facing validations (all raise HTTPException):
    - org.fiscal_year_start_month must be configured.
    - Only monthly frequency is supported.
    - FY start year must be >= year(date_of_registration).
    - FY start year must be <= current calendar year.
    - No periods may already exist for the requested label (409).

    The actual period creation is delegated to _generate_periods_for_year, which
    is also used by the auto-generation triggers.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()
    if not org or not org.fiscal_year_start_month:
        raise HTTPException(
            status_code=422,
            detail="Fiscal year start month not configured. Set it on the Organisation page first.",
        )

    frequency = (org.period_closing_frequency or "monthly").lower()
    if frequency != "monthly":
        raise HTTPException(
            status_code=422,
            detail="Only monthly periods are supported in M8.3. Quarterly/annual support will be added later.",
        )

    label = data.fiscal_year_label
    start_year = parse_fy_start_year(label)
    if start_year is None:
        raise HTTPException(status_code=422, detail="Could not parse year from fiscal_year_label.")

    # Year must be >= year of date_of_registration
    if org.date_of_registration and start_year < org.date_of_registration.year:
        raise HTTPException(
            status_code=400,
            detail="Cannot create periods before the company's registration date.",
        )

    # Year must be <= current calendar year
    if start_year > date.today().year:
        raise HTTPException(
            status_code=400,
            detail="Cannot create periods for a future fiscal year.",
        )

    # Idempotent guard: raise 409 if any periods already exist for this label
    existing_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == label,
        )
    )
    if existing_result.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Fiscal year {label} has already been generated. "
                "Delete all periods for this year before regenerating."
            ),
        )

    # Delegate creation to shared helper (pass user-provided label explicitly so
    # periods carry exactly the label the user specified, not the format-derived one).
    created = await _generate_periods_for_year(db, tenant_id, start_year, org, label=label)

    await db.commit()
    for p in created:
        await db.refresh(p)

    return [_period_to_response(p) for p in created]


@router.delete("/periods/fiscal-year/{fiscal_year}", status_code=200)
async def delete_fiscal_year(
    fiscal_year: str,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Delete all periods for a fiscal year label, plus its FiscalYearState if present.

    Blocked (409) if any period is in a closed state (SOFT_CLOSED, OVERDUE, or
    HARD_CLOSED). Only FUTURE and OPEN periods may be deleted — this prevents
    accidental destruction of close history.

    Intended use: correct a misconfigured generation before any period is closed.
    After deletion the fiscal year label may be regenerated via POST /periods/generate.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    # Block if any period is in a closed state
    closed_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == fiscal_year,
            AccountingPeriod.status.in_(["SOFT_CLOSED", "OVERDUE", "HARD_CLOSED"]),
        )
    )
    if closed_result.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=409,
            detail="Cannot delete a fiscal year with closed periods.",
        )

    # Delete all periods for this fiscal year
    await db.execute(
        delete(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == fiscal_year,
        )
    )

    # Delete FiscalYearState if present
    await db.execute(
        delete(FiscalYearState).where(
            FiscalYearState.tenant_id == tenant_id,
            FiscalYearState.fiscal_year == fiscal_year,
        )
    )

    await db.commit()
    return {"deleted": fiscal_year}


@router.get("/periods", response_model=list[AccountingPeriodResponse])
async def list_periods(
    fiscal_year: Optional[str] = Query(default=None),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[AccountingPeriodResponse]:
    """
    List accounting periods for this tenant, optionally filtered by fiscal_year.

    Auto-soft-closes any OPEN periods whose end_date has passed, persisting the
    transition before returning results. Future: replace with a scheduled job.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    stmt = select(AccountingPeriod).where(AccountingPeriod.tenant_id == tenant_id)
    if fiscal_year:
        stmt = stmt.where(AccountingPeriod.fiscal_year == fiscal_year)
    stmt = stmt.order_by(AccountingPeriod.start_date)

    result = await db.execute(stmt)
    periods = list(result.scalars().all())

    # Auto-soft-close on read (no scheduler yet — future: move to cron/celery)
    changed = False
    for p in periods:
        if await apply_auto_soft_close(p, db):
            changed = True
    if changed:
        await db.commit()
        for p in periods:
            await db.refresh(p)

    return [_period_to_response(p) for p in periods]


@router.get("/periods/check", response_model=PeriodCheckResponse)
async def check_period(
    date_str: str = Query(alias="date", description="ISO date YYYY-MM-DD"),
    module: Optional[str] = Query(default=None, description="Posting module, e.g. expense | manual_journal"),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PeriodCheckResponse:
    """
    Check whether a specific date is open for posting by this tenant.

    Optional query params (Brief 2):
        module — enables module-specific grace rows and the manual-journal block check.

    Returns { postable: bool, reason: str }. The reason is empty when postable.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    try:
        target = date.fromisoformat(date_str)
    except ValueError:
        raise HTTPException(status_code=422, detail="date must be in YYYY-MM-DD format.")

    postable, reason = await is_date_postable(
        tenant_id, target, db,
        user_id=current_user.user_id,
        module=module,
        role_tier=current_user.role_tier,
    )
    return PeriodCheckResponse(postable=postable, reason=reason)


@router.post("/periods/{period_id}/soft-close", response_model=AccountingPeriodResponse)
async def soft_close_period(
    period_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> AccountingPeriodResponse:
    """
    Manually soft-close an OPEN period.

    Normally automatic (triggered on read when today > end_date), but exposed
    here for use cases where an admin wants to close ahead of time.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    if period.status == "HARD_CLOSED":
        raise HTTPException(status_code=409, detail="Period is already hard-closed.")
    if period.status == "SOFT_CLOSED":
        raise HTTPException(status_code=409, detail="Period is already soft-closed.")

    period.status = "SOFT_CLOSED"
    period.soft_closed_at = datetime.now(timezone.utc)
    db.add(period)
    await db.commit()
    await db.refresh(period)
    return _period_to_response(period)


@router.post("/periods/{period_id}/hard-close", response_model=AccountingPeriodResponse)
async def hard_close_period(
    period_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> AccountingPeriodResponse:
    """
    Hard-close a period. Sequential-close enforcement: any period with an earlier
    start_date (across all FYs for this tenant) must already be HARD_CLOSED.

    BRIEF-3: checklist gate — before allowing hard-close, call checklist_complete()
    here to verify all close checklist items are ticked. No logic yet; stub below.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    if period.status == "HARD_CLOSED":
        raise HTTPException(status_code=409, detail="Period is already hard-closed.")

    # ── Sequential-close enforcement ──────────────────────────────────────────
    earlier_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.start_date < period.start_date,
            AccountingPeriod.status != "HARD_CLOSED",
        )
    )
    if earlier_result.scalars().first() is not None:
        raise HTTPException(
            status_code=409,
            detail="Earlier periods must be hard-closed first.",
        )

    # ── Checklist gate (Brief 3) ──────────────────────────────────────────────
    ok, missing = await checklist_complete(period, db)
    if not ok:
        raise HTTPException(
            status_code=409,
            detail=f"Close checklist incomplete: {', '.join(missing)}",
        )

    period.status = "HARD_CLOSED"
    period.hard_closed_at = datetime.now(timezone.utc)
    period.hard_closed_by = current_user.user_id
    db.add(period)
    await db.commit()
    await db.refresh(period)

    # ── Auto-generate next fiscal year when the last period is hard-closed ───
    # Trigger 2: if every period in this FY is now HARD_CLOSED, the year is
    # fully closed — automatically generate the next FY so the team can
    # immediately start submitting for it.
    all_fy_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == period.fiscal_year,
        )
    )
    all_fy_periods = all_fy_result.scalars().all()
    if all_fy_periods and all(p.status == "HARD_CLOSED" for p in all_fy_periods):
        fy_start_year = parse_fy_start_year(period.fiscal_year)
        if fy_start_year is not None:
            next_year = fy_start_year + 1
            # Guard: only generate up to current year + 1 to avoid creating
            # far-future FUTURE-status periods with no operational value.
            if next_year <= date.today().year + 1:
                org_result = await db.execute(
                    select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
                )
                org = org_result.scalar_one_or_none()
                if org:
                    # Same date-range overlap check as Trigger 1 — avoids
                    # double-generation if the label format/start_month changed.
                    next_existing = await db.execute(
                        select(AccountingPeriod).where(
                            AccountingPeriod.tenant_id == tenant_id,
                            AccountingPeriod.start_date <= date(next_year, 12, 31),
                            AccountingPeriod.end_date >= date(next_year, 1, 1),
                        ).limit(1)
                    )
                    if not next_existing.scalar_one_or_none():
                        created = await _generate_periods_for_year(db, tenant_id, next_year, org)
                        if created:
                            await db.commit()

    return _period_to_response(period)


@router.post("/periods/{period_id}/reopen", response_model=AccountingPeriodResponse)
async def reopen_period(
    period_id: uuid.UUID,
    data: Optional[ReopenRequest] = Body(default=None),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> AccountingPeriodResponse:
    """
    Reopen a HARD_CLOSED period — restricted to consultant role.

    Sets status back to SOFT_CLOSED and increments reopened_count.
    Refuses (409) if the period's fiscal year is STATUTORY_CLOSED.
    Accepts an optional body: { reason: string } for the audit trail.
    Writes a PeriodAuditLog REOPEN entry (Brief 4 — fills the hook).
    """
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="Only a super admin (Ziva consultant) may reopen hard-closed periods.")
    block_if_readonly_impersonation(current_user)

    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    if period.status != "HARD_CLOSED":
        raise HTTPException(status_code=409, detail="Only HARD_CLOSED periods can be reopened.")

    # ── Statutory-closed guard (Brief 4) ─────────────────────────────────────
    fy_result = await db.execute(
        select(FiscalYearState).where(
            FiscalYearState.tenant_id == tenant_id,
            FiscalYearState.fiscal_year == period.fiscal_year,
        )
    )
    fy_state = fy_result.scalar_one_or_none()
    if fy_state and fy_state.status == "STATUTORY_CLOSED":
        raise HTTPException(
            status_code=409,
            detail="This period's fiscal year is permanently locked (statutory closed) and cannot be reopened.",
        )

    period.status = "SOFT_CLOSED"
    period.hard_closed_at = None
    period.hard_closed_by = None
    period.reopened_count += 1
    db.add(period)

    # ── Audit log (Brief 4 — fills # BRIEF-4: audit log on reopen hook) ──────
    audit = PeriodAuditLog(
        tenant_id=tenant_id,
        period_id=period.id,
        fiscal_year=period.fiscal_year,
        action="REOPEN",
        actor_id=current_user.user_id,
        detail=data.reason if data and data.reason else None,
    )
    db.add(audit)
    await db.commit()
    await db.refresh(period)
    return _period_to_response(period)


# ── Year-end helpers (M8.3 Brief 4) ─────────────────────────────────────────

async def _get_or_seed_fy_state(
    tenant_id: uuid.UUID,
    fiscal_year: str,
    db: AsyncSession,
    org: Optional[TenantOrgConfig] = None,
) -> FiscalYearState:
    """Return the FiscalYearState for this tenant+FY, seeding an OPEN row if absent."""
    result = await db.execute(
        select(FiscalYearState).where(
            FiscalYearState.tenant_id == tenant_id,
            FiscalYearState.fiscal_year == fiscal_year,
        )
    )
    fy = result.scalar_one_or_none()
    if fy is None:
        grace = org.default_audit_grace_months if org else 3
        fy = FiscalYearState(
            tenant_id=tenant_id,
            fiscal_year=fiscal_year,
            status="OPEN",
            audit_grace_months=grace,
        )
        db.add(fy)
        await db.commit()
        await db.refresh(fy)
    return fy


def _fy_state_to_response(fy: FiscalYearState) -> FiscalYearStateResponse:
    """Convert a FiscalYearState ORM row to its response schema."""
    return FiscalYearStateResponse(
        id=str(fy.id),
        tenant_id=str(fy.tenant_id),
        fiscal_year=fy.fiscal_year,
        status=fy.status,
        management_closed_at=fy.management_closed_at,
        management_closed_by=str(fy.management_closed_by) if fy.management_closed_by else None,
        audit_grace_months=fy.audit_grace_months,
        audit_grace_expires_at=fy.audit_grace_expires_at,
        statutory_closed_at=fy.statutory_closed_at,
        statutory_closed_by=str(fy.statutory_closed_by) if fy.statutory_closed_by else None,
        retained_earnings_rolled=fy.retained_earnings_rolled,
        created_at=fy.created_at,
    )


# ── Year-end endpoints (M8.3 Brief 4) ────────────────────────────────────────

@router.post("/periods/management-close", response_model=FiscalYearStateResponse)
async def management_close(
    data: ManagementCloseRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FiscalYearStateResponse:
    """
    Stage 1 of year-end close: management close of a fiscal year.

    Precondition: the FINAL period of the FY (highest period_no — December in a
    normal 12-month year, but an earlier month for a registration-truncated stub
    first year) must be HARD_CLOSED. Never hardcode period_no == 12 here: a stub
    year can have fewer than 12 periods and would never satisfy that check,
    permanently blocking management close for that year.
    Transitions FiscalYearState OPEN → AUDIT_PENDING. Sets audit_grace_expires_at.
    Sets retained_earnings_rolled=True; actual GL roll-forward journal is stubbed below.
    Writes a MANAGEMENT_CLOSE PeriodAuditLog entry.
    Idempotent: returns 409 if already AUDIT_PENDING or STATUTORY_CLOSED.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    fiscal_year = data.fiscal_year_label.strip()

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()

    fy = await _get_or_seed_fy_state(tenant_id, fiscal_year, db, org)

    if fy.status in ("AUDIT_PENDING", "AUDIT_OVERDUE", "STATUTORY_CLOSED"):
        raise HTTPException(
            status_code=409,
            detail=f"Fiscal year {fiscal_year} is already in status {fy.status}.",
        )

    # The FINAL period of the FY (highest period_no, NOT a hardcoded 12 — a
    # stub first year clamped to the registration date can have fewer than 12
    # periods) must be HARD_CLOSED before management close.
    last_period_result = await db.execute(
        select(AccountingPeriod)
        .where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == fiscal_year,
        )
        .order_by(AccountingPeriod.period_no.desc())
        .limit(1)
    )
    last_period = last_period_result.scalar_one_or_none()
    if not last_period or last_period.status != "HARD_CLOSED":
        period_label = last_period.period_name if last_period else "the final period"
        raise HTTPException(
            status_code=409,
            detail=f"{period_label} must be hard-closed before management close.",
        )

    now = datetime.now(timezone.utc)
    fy.status = "AUDIT_PENDING"
    fy.management_closed_at = now
    fy.management_closed_by = current_user.user_id
    fy.audit_grace_expires_at = add_months(now, fy.audit_grace_months)
    fy.retained_earnings_rolled = True
    db.add(fy)

    audit = PeriodAuditLog(
        tenant_id=tenant_id,
        fiscal_year=fiscal_year,
        action="MANAGEMENT_CLOSE",
        actor_id=current_user.user_id,
        detail=f"Management close of {fiscal_year}; audit grace {fy.audit_grace_months} months.",
    )
    db.add(audit)
    await db.commit()
    await db.refresh(fy)

    # M8.x: post retained-earnings roll-forward journal here (posting engine required)

    return _fy_state_to_response(fy)


@router.post("/periods/statutory-close", response_model=FiscalYearStateResponse)
async def statutory_close(
    data: StatutoryCloseRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FiscalYearStateResponse:
    """
    Stage 2 of year-end close: statutory (permanent) close of a fiscal year.

    Precondition: FiscalYearState must be AUDIT_PENDING or AUDIT_OVERDUE.
    Once STATUTORY_CLOSED: is_date_postable refuses any date in this FY; periods cannot be reopened.

    Audit-artifact prerequisites (audited TB, signed AFS, CFO sign-off) are gated in M8.4.
    This brief leaves the artifact gate open with a clearly-marked stub.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    fiscal_year = data.fiscal_year_label.strip()

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()

    fy = await _get_or_seed_fy_state(tenant_id, fiscal_year, db, org)

    if fy.status not in ("AUDIT_PENDING", "AUDIT_OVERDUE"):
        raise HTTPException(
            status_code=409,
            detail=(
                f"Statutory close requires AUDIT_PENDING or AUDIT_OVERDUE status; "
                f"current: {fy.status}."
            ),
        )

    # M8.4: require audited TB + signed AFS + CFO sign-off before allowing statutory close
    # (artifact checks will be inserted here in M8.4 — gate is open in this brief)

    now = datetime.now(timezone.utc)
    fy.status = "STATUTORY_CLOSED"
    fy.statutory_closed_at = now
    fy.statutory_closed_by = current_user.user_id
    db.add(fy)

    audit = PeriodAuditLog(
        tenant_id=tenant_id,
        fiscal_year=fiscal_year,
        action="STATUTORY_CLOSE",
        actor_id=current_user.user_id,
        detail=f"Statutory close of {fiscal_year}. Year permanently locked.",
    )
    db.add(audit)
    await db.commit()
    await db.refresh(fy)
    return _fy_state_to_response(fy)


@router.get("/periods/year-state", response_model=FiscalYearStateResponse)
async def get_year_state(
    fiscal_year: str = Query(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FiscalYearStateResponse:
    """
    Return the FiscalYearState for a given fiscal_year label.
    Seeds an OPEN row if none exists (so the UI is never gated on prior management-close).
    Auto-transitions AUDIT_PENDING → AUDIT_OVERDUE on read if grace has expired.
    # FUTURE: move overdue check to a scheduled job (cron/celery).
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()

    fy = await _get_or_seed_fy_state(tenant_id, fiscal_year.strip(), db, org)
    changed = await check_audit_overdue(fy, db)
    if changed:
        await db.commit()
    return _fy_state_to_response(fy)


@router.patch("/periods/year-state/{fiscal_year}", response_model=FiscalYearStateResponse)
async def update_year_state(
    fiscal_year: str,
    data: AuditGraceUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FiscalYearStateResponse:
    """
    Update audit_grace_months for a fiscal year (consultant or admin).
    Recomputes audit_grace_expires_at if the year is already management-closed.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()

    fy = await _get_or_seed_fy_state(tenant_id, fiscal_year.strip(), db, org)
    fy.audit_grace_months = data.audit_grace_months

    if fy.management_closed_at:
        fy.audit_grace_expires_at = add_months(fy.management_closed_at, data.audit_grace_months)

    db.add(fy)
    await db.commit()
    await db.refresh(fy)
    return _fy_state_to_response(fy)


@router.get("/periods/audit-log", response_model=list[PeriodAuditLogResponse])
async def get_period_audit_log(
    fiscal_year: Optional[str] = Query(default=None),
    period_id: Optional[uuid.UUID] = Query(default=None),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[PeriodAuditLogResponse]:
    """List period audit log entries; filterable by fiscal_year and/or period_id."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    stmt = (
        select(PeriodAuditLog)
        .where(PeriodAuditLog.tenant_id == tenant_id)
        .order_by(PeriodAuditLog.created_at.desc())
    )
    if fiscal_year:
        stmt = stmt.where(PeriodAuditLog.fiscal_year == fiscal_year)
    if period_id:
        stmt = stmt.where(PeriodAuditLog.period_id == period_id)

    result = await db.execute(stmt)
    logs = result.scalars().all()

    return [
        PeriodAuditLogResponse(
            id=str(lg.id),
            tenant_id=str(lg.tenant_id),
            fiscal_year=lg.fiscal_year,
            period_id=str(lg.period_id) if lg.period_id else None,
            action=lg.action,
            actor_id=str(lg.actor_id),
            detail=lg.detail,
            created_at=lg.created_at,
        )
        for lg in logs
    ]


# ── Grace overrides (M8.3 Brief 2) ───────────────────────────────────────────

_VALID_APPLIES_TO = {"all", "role", "user"}
_VALID_PERIOD_TYPES = {"regular", "year_end"}
_VALID_GRACE_UNITS = {"workdays", "calendar_days"}
_VALID_ROLE_TIERS = {"consultant", "power_admin", "functional_admin"}


async def _seed_default_grace_row(tenant_id: uuid.UUID, db: AsyncSession) -> PeriodGraceOverride:
    """Create and persist the default grace row (3 workdays, regular, all) if absent."""
    row = PeriodGraceOverride(
        tenant_id=tenant_id,
        module="default",
        applies_to_type="all",
        applies_to_role=None,
        applies_to_user_id=None,
        period_type="regular",
        grace_value=3,
        grace_unit="workdays",
        is_default=True,
    )
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


async def _resolve_grace_response(
    row: PeriodGraceOverride, db: AsyncSession
) -> PeriodGraceOverrideResponse:
    """Build the response schema for a grace row, resolving user_id → name."""
    user_name: Optional[str] = None
    if row.applies_to_user_id:
        u_result = await db.execute(
            select(User).where(User.id == row.applies_to_user_id)
        )
        user = u_result.scalar_one_or_none()
        if user:
            user_name = getattr(user, "full_name", None) or getattr(user, "email", None)

    return PeriodGraceOverrideResponse(
        id=str(row.id),
        tenant_id=str(row.tenant_id),
        module=row.module,
        applies_to_type=row.applies_to_type,
        applies_to_role=row.applies_to_role,
        applies_to_user_id=str(row.applies_to_user_id) if row.applies_to_user_id else None,
        applies_to_user_name=user_name,
        period_type=row.period_type,
        grace_value=row.grace_value,
        grace_unit=row.grace_unit,
        is_default=row.is_default,
        created_at=row.created_at,
    )


@router.get("/periods/grace", response_model=list[PeriodGraceOverrideResponse])
async def list_grace_overrides(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[PeriodGraceOverrideResponse]:
    """
    List all grace override rows for this tenant. Seeds the default row if none exist.

    The default row (is_default=True) is always present after first access:
        module="default", applies_to_type="all", period_type="regular",
        grace_value=3, grace_unit="workdays".
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(PeriodGraceOverride)
        .where(PeriodGraceOverride.tenant_id == tenant_id)
        .order_by(PeriodGraceOverride.created_at)
    )
    rows = list(result.scalars().all())

    if not rows:
        default_row = await _seed_default_grace_row(tenant_id, db)
        await db.commit()
        await db.refresh(default_row)
        rows = [default_row]

    return [await _resolve_grace_response(r, db) for r in rows]


@router.post("/periods/grace", response_model=PeriodGraceOverrideResponse, status_code=201)
async def create_grace_override(
    data: PeriodGraceOverrideCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PeriodGraceOverrideResponse:
    """
    Add a grace override row. Validates module + enums + conditional required fields.
    Refuses exact-duplicate rows (same module + applies_to + period_type).
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    if data.module not in ALLOWED_MODULES:
        raise HTTPException(status_code=422, detail=f"Invalid module '{data.module}'. Allowed: {sorted(ALLOWED_MODULES)}")
    if data.applies_to_type not in _VALID_APPLIES_TO:
        raise HTTPException(status_code=422, detail=f"Invalid applies_to_type '{data.applies_to_type}'.")
    if data.period_type not in _VALID_PERIOD_TYPES:
        raise HTTPException(status_code=422, detail=f"Invalid period_type '{data.period_type}'.")
    if data.grace_unit not in _VALID_GRACE_UNITS:
        raise HTTPException(status_code=422, detail=f"Invalid grace_unit '{data.grace_unit}'.")

    if data.applies_to_type == "role":
        if not data.applies_to_role:
            raise HTTPException(status_code=422, detail="applies_to_role is required when applies_to_type='role'.")
        if data.applies_to_role not in _VALID_ROLE_TIERS:
            raise HTTPException(status_code=422, detail=f"Invalid applies_to_role '{data.applies_to_role}'.")
    if data.applies_to_type == "user" and not data.applies_to_user_id:
        raise HTTPException(status_code=422, detail="applies_to_user_id is required when applies_to_type='user'.")

    # Duplicate check (same module + applies_to_type + applies_to_role/user + period_type)
    dup_stmt = select(PeriodGraceOverride).where(
        PeriodGraceOverride.tenant_id == tenant_id,
        PeriodGraceOverride.module == data.module,
        PeriodGraceOverride.applies_to_type == data.applies_to_type,
        PeriodGraceOverride.period_type == data.period_type,
    )
    if data.applies_to_role:
        dup_stmt = dup_stmt.where(PeriodGraceOverride.applies_to_role == data.applies_to_role)
    if data.applies_to_user_id:
        try:
            uid = uuid.UUID(data.applies_to_user_id)
        except ValueError:
            raise HTTPException(status_code=422, detail="applies_to_user_id must be a valid UUID.")
        dup_stmt = dup_stmt.where(PeriodGraceOverride.applies_to_user_id == uid)

    dup_result = await db.execute(dup_stmt)
    if dup_result.scalar_one_or_none() is not None:
        raise HTTPException(status_code=409, detail="An identical grace override row already exists.")

    user_uuid: Optional[uuid.UUID] = None
    if data.applies_to_user_id:
        user_uuid = uuid.UUID(data.applies_to_user_id)

    row = PeriodGraceOverride(
        tenant_id=tenant_id,
        module=data.module,
        applies_to_type=data.applies_to_type,
        applies_to_role=data.applies_to_role if data.applies_to_type == "role" else None,
        applies_to_user_id=user_uuid if data.applies_to_type == "user" else None,
        period_type=data.period_type,
        grace_value=data.grace_value,
        grace_unit=data.grace_unit,
        is_default=False,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return await _resolve_grace_response(row, db)


@router.patch("/periods/grace/{grace_id}", response_model=PeriodGraceOverrideResponse)
async def update_grace_override(
    grace_id: uuid.UUID,
    data: PeriodGraceOverrideUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PeriodGraceOverrideResponse:
    """
    Edit a grace override row.

    The default row's module, applies_to, and period_type are locked — only
    grace_value and grace_unit may be changed. All fields editable on non-default rows.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(PeriodGraceOverride).where(
            PeriodGraceOverride.id == grace_id,
            PeriodGraceOverride.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Grace override not found.")

    if row.is_default:
        # Locked fields on the default row.
        for locked in ("module", "applies_to_type", "applies_to_role", "applies_to_user_id", "period_type"):
            if getattr(data, locked) is not None:
                raise HTTPException(
                    status_code=409,
                    detail=f"Field '{locked}' cannot be changed on the default grace row.",
                )

    if data.grace_value is not None:
        if data.grace_value < 0:
            raise HTTPException(status_code=422, detail="grace_value must be non-negative.")
        row.grace_value = data.grace_value
    if data.grace_unit is not None:
        if data.grace_unit not in _VALID_GRACE_UNITS:
            raise HTTPException(status_code=422, detail=f"Invalid grace_unit '{data.grace_unit}'.")
        row.grace_unit = data.grace_unit

    if not row.is_default:
        if data.module is not None:
            if data.module not in ALLOWED_MODULES:
                raise HTTPException(status_code=422, detail=f"Invalid module '{data.module}'.")
            row.module = data.module
        if data.applies_to_type is not None:
            if data.applies_to_type not in _VALID_APPLIES_TO:
                raise HTTPException(status_code=422, detail=f"Invalid applies_to_type '{data.applies_to_type}'.")
            row.applies_to_type = data.applies_to_type
        if data.applies_to_role is not None:
            row.applies_to_role = data.applies_to_role
        if data.applies_to_user_id is not None:
            try:
                row.applies_to_user_id = uuid.UUID(data.applies_to_user_id)
            except ValueError:
                raise HTTPException(status_code=422, detail="applies_to_user_id must be a valid UUID.")
        if data.period_type is not None:
            if data.period_type not in _VALID_PERIOD_TYPES:
                raise HTTPException(status_code=422, detail=f"Invalid period_type '{data.period_type}'.")
            row.period_type = data.period_type

    db.add(row)
    await db.commit()
    await db.refresh(row)
    return await _resolve_grace_response(row, db)


@router.delete("/periods/grace/{grace_id}", status_code=204)
async def delete_grace_override(
    grace_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Remove a grace override row. Refuses deletion of the default row (409)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(PeriodGraceOverride).where(
            PeriodGraceOverride.id == grace_id,
            PeriodGraceOverride.tenant_id == tenant_id,
        )
    )
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Grace override not found.")
    if row.is_default:
        raise HTTPException(status_code=409, detail="The default grace row cannot be deleted.")

    await db.delete(row)
    await db.commit()


# ── Manual-journal block toggle (M8.3 Brief 2) ───────────────────────────────

@router.get("/periods/journal-block", response_model=JournalBlockResponse)
async def get_journal_block(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> JournalBlockResponse:
    """Return the current manual-journal block setting for this tenant."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = result.scalar_one_or_none()
    enabled = org.block_journal_into_open_prior if org else True
    return JournalBlockResponse(enabled=enabled)


@router.patch("/periods/journal-block", response_model=JournalBlockResponse)
async def set_journal_block(
    data: JournalBlockUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> JournalBlockResponse:
    """
    Enable or disable the manual-journal block for this tenant.

    When enabled (default): blocks manual journal entries into a period while
    any earlier period remains open. When disabled: allows free-form journal dating.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = result.scalar_one_or_none()
    if not org:
        raise HTTPException(status_code=404, detail="Organisation not configured.")

    org.block_journal_into_open_prior = data.enabled
    db.add(org)
    await db.commit()
    return JournalBlockResponse(enabled=data.enabled)


# ── Future-dated posting exception (M8.3 Brief 2) ────────────────────────────

def _can_post_future(current_user: CurrentUser, grace_rows: list[PeriodGraceOverride]) -> bool:
    """
    Return True if the current user is permitted to grant a future-dated exception.

    Default: consultant only. Override: any grace row with module="future_exception"
    and applies_to matching this user/role/all.
    """
    if current_user.is_super_admin:
        return True
    for row in grace_rows:
        if row.module != "future_exception":
            continue
        if row.applies_to_type == "all":
            return True
        if row.applies_to_type == "role" and row.applies_to_role == current_user.role_tier:
            return True
        if row.applies_to_type == "user" and row.applies_to_user_id == current_user.user_id:
            return True
    return False


@router.post("/periods/future-exception", response_model=FuturePostingExceptionResponse, status_code=201)
async def create_future_exception(
    data: FuturePostingExceptionCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FuturePostingExceptionResponse:
    """
    Record a permission grant for a future-dated posting.

    Access: consultant by default, or any user/role covered by a
    future_exception grace row. Returns 403 otherwise.

    This endpoint records the *intent* only — it does NOT itself create a journal entry.
    The posting engine (a later brief) will verify a valid exception exists before
    allowing a journal entry dated in a FUTURE period.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    if data.module not in ALLOWED_MODULES:
        raise HTTPException(status_code=422, detail=f"Invalid module '{data.module}'.")

    # Load grace rows for permission check.
    grace_result = await db.execute(
        select(PeriodGraceOverride).where(PeriodGraceOverride.tenant_id == tenant_id)
    )
    grace_rows = list(grace_result.scalars().all())

    if not _can_post_future(current_user, grace_rows):
        raise HTTPException(
            status_code=403,
            detail="You are not permitted to create future-dated posting exceptions.",
        )

    exc = FuturePostingException(
        tenant_id=tenant_id,
        created_by=current_user.user_id,
        target_date=data.target_date,
        module=data.module,
        reason=data.reason,
    )
    db.add(exc)
    await db.commit()
    await db.refresh(exc)

    return FuturePostingExceptionResponse(
        id=str(exc.id),
        tenant_id=str(exc.tenant_id),
        created_by=str(exc.created_by),
        target_date=exc.target_date,
        module=exc.module,
        reason=exc.reason,
        created_at=exc.created_at,
    )


# ── Close checklist template CRUD (M8.3 Brief 3) ─────────────────────────────

@router.get("/periods/checklist", response_model=list[CloseChecklistItemResponse])
async def list_checklist_items(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[CloseChecklistItemResponse]:
    """Return all close checklist template items for this tenant, ordered by sort_order."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(CloseChecklistItem)
        .where(CloseChecklistItem.tenant_id == tenant_id)
        .order_by(CloseChecklistItem.sort_order, CloseChecklistItem.created_at)
    )
    items = result.scalars().all()

    return [
        CloseChecklistItemResponse(
            id=str(i.id),
            tenant_id=str(i.tenant_id),
            label=i.label,
            description=i.description,
            applies_to=i.applies_to,
            sort_order=i.sort_order,
            is_active=i.is_active,
            created_at=i.created_at,
        )
        for i in items
    ]


@router.post("/periods/checklist", response_model=CloseChecklistItemResponse, status_code=201)
async def create_checklist_item(
    data: CloseChecklistItemCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CloseChecklistItemResponse:
    """Add a close checklist template item. applies_to must be 'every_close' or 'year_end_only'."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    if data.applies_to not in ("every_close", "year_end_only"):
        raise HTTPException(
            status_code=422,
            detail="applies_to must be 'every_close' or 'year_end_only'.",
        )

    item = CloseChecklistItem(
        tenant_id=tenant_id,
        label=data.label,
        description=data.description,
        applies_to=data.applies_to,
        sort_order=data.sort_order,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)

    return CloseChecklistItemResponse(
        id=str(item.id),
        tenant_id=str(item.tenant_id),
        label=item.label,
        description=item.description,
        applies_to=item.applies_to,
        sort_order=item.sort_order,
        is_active=item.is_active,
        created_at=item.created_at,
    )


@router.patch("/periods/checklist/{item_id}", response_model=CloseChecklistItemResponse)
async def update_checklist_item(
    item_id: uuid.UUID,
    data: CloseChecklistItemUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CloseChecklistItemResponse:
    """Edit a close checklist template item. Setting is_active=False is the soft-delete path."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(CloseChecklistItem).where(
            CloseChecklistItem.id == item_id,
            CloseChecklistItem.tenant_id == tenant_id,
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item not found.")

    if data.applies_to is not None and data.applies_to not in ("every_close", "year_end_only"):
        raise HTTPException(
            status_code=422,
            detail="applies_to must be 'every_close' or 'year_end_only'.",
        )

    if data.label is not None:
        item.label = data.label
    if data.description is not None:
        item.description = data.description
    if data.applies_to is not None:
        item.applies_to = data.applies_to
    if data.sort_order is not None:
        item.sort_order = data.sort_order
    if data.is_active is not None:
        item.is_active = data.is_active

    db.add(item)
    await db.commit()
    await db.refresh(item)

    return CloseChecklistItemResponse(
        id=str(item.id),
        tenant_id=str(item.tenant_id),
        label=item.label,
        description=item.description,
        applies_to=item.applies_to,
        sort_order=item.sort_order,
        is_active=item.is_active,
        created_at=item.created_at,
    )


@router.delete("/periods/checklist/{item_id}", status_code=204)
async def delete_checklist_item(
    item_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """
    Soft-delete a close checklist item (sets is_active=False).

    Completion history in period_checklist_completions is preserved intact because the
    FK has no CASCADE and item_label_snapshot holds the label at time of sign-off.
    Hard-delete is intentionally NOT used.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(CloseChecklistItem).where(
            CloseChecklistItem.id == item_id,
            CloseChecklistItem.tenant_id == tenant_id,
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item not found.")

    item.is_active = False
    db.add(item)
    await db.commit()


# ── Per-period checklist completion (M8.3 Brief 3) ───────────────────────────

@router.get(
    "/periods/{period_id}/checklist",
    response_model=list[PeriodChecklistEntryResponse],
)
async def get_period_checklist(
    period_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[PeriodChecklistEntryResponse]:
    """
    Return applicable checklist items for this period, with their current completion state.

    "Applicable" = active items where applies_to=='every_close' OR
    (applies_to=='year_end_only' AND this is the FY's final period).

    "Final period" is the period with the highest period_no for that fiscal_year,
    NOT a hardcoded period_no==12 — a registration-truncated stub first year can
    have fewer than 12 periods, and its real year-end period must still pick up
    year_end_only checklist items.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    period_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    max_period_no_result = await db.execute(
        select(sqlfunc.max(AccountingPeriod.period_no)).where(
            AccountingPeriod.tenant_id == tenant_id,
            AccountingPeriod.fiscal_year == period.fiscal_year,
        )
    )
    max_period_no = max_period_no_result.scalar_one_or_none()
    is_year_end = max_period_no is not None and period.period_no == max_period_no
    items_stmt = (
        select(CloseChecklistItem)
        .where(
            CloseChecklistItem.tenant_id == tenant_id,
            CloseChecklistItem.is_active == True,  # noqa: E712
        )
        .order_by(CloseChecklistItem.sort_order, CloseChecklistItem.created_at)
    )
    if not is_year_end:
        items_stmt = items_stmt.where(CloseChecklistItem.applies_to == "every_close")

    items_result = await db.execute(items_stmt)
    items = items_result.scalars().all()

    entries: list[PeriodChecklistEntryResponse] = []
    for item in items:
        comp_result = await db.execute(
            select(PeriodChecklistCompletion).where(
                PeriodChecklistCompletion.period_id == period_id,
                PeriodChecklistCompletion.checklist_item_id == item.id,
            )
        )
        comp = comp_result.scalar_one_or_none()

        entries.append(
            PeriodChecklistEntryResponse(
                checklist_item_id=str(item.id),
                label=item.label,
                description=item.description,
                applies_to=item.applies_to,
                sort_order=item.sort_order,
                completion_id=str(comp.id) if comp else None,
                status=comp.status if comp else "pending",
                prepared_by=str(comp.prepared_by) if comp and comp.prepared_by else None,
                prepared_at=comp.prepared_at if comp else None,
                approved_by=str(comp.approved_by) if comp and comp.approved_by else None,
                approved_at=comp.approved_at if comp else None,
            )
        )

    return entries


@router.post(
    "/periods/{period_id}/checklist/{item_id}/prepare",
    response_model=PeriodChecklistCompletionResponse,
)
async def prepare_checklist_item(
    period_id: uuid.UUID,
    item_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PeriodChecklistCompletionResponse:
    """
    Mark a checklist item as prepared for this period.

    Creates the completion row if it doesn't exist (snapshotting the item label).
    If the item is already approved, returns 409.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    period_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    item_result = await db.execute(
        select(CloseChecklistItem).where(
            CloseChecklistItem.id == item_id,
            CloseChecklistItem.tenant_id == tenant_id,
        )
    )
    item = item_result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item not found.")

    comp_result = await db.execute(
        select(PeriodChecklistCompletion).where(
            PeriodChecklistCompletion.period_id == period_id,
            PeriodChecklistCompletion.checklist_item_id == item_id,
        )
    )
    comp = comp_result.scalar_one_or_none()

    if comp and comp.status == "approved":
        raise HTTPException(status_code=409, detail="Item is already approved and cannot be re-prepared.")

    now = datetime.now(timezone.utc)
    if comp is None:
        comp = PeriodChecklistCompletion(
            tenant_id=tenant_id,
            period_id=period_id,
            checklist_item_id=item_id,
            item_label_snapshot=item.label,
            prepared_by=current_user.user_id,
            prepared_at=now,
            status="prepared",
        )
        db.add(comp)
    else:
        comp.prepared_by = current_user.user_id
        comp.prepared_at = now
        comp.status = "prepared"
        db.add(comp)

    await db.commit()
    await db.refresh(comp)

    return PeriodChecklistCompletionResponse(
        id=str(comp.id),
        period_id=str(comp.period_id),
        checklist_item_id=str(comp.checklist_item_id),
        item_label_snapshot=comp.item_label_snapshot,
        status=comp.status,
        prepared_by=str(comp.prepared_by) if comp.prepared_by else None,
        prepared_at=comp.prepared_at,
        approved_by=str(comp.approved_by) if comp.approved_by else None,
        approved_at=comp.approved_at,
        created_at=comp.created_at,
    )


@router.post(
    "/periods/{period_id}/checklist/{item_id}/approve",
    response_model=PeriodChecklistCompletionResponse,
)
async def approve_checklist_item(
    period_id: uuid.UUID,
    item_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PeriodChecklistCompletionResponse:
    """
    Mark a checklist item as approved for this period.

    Segregation of duties: the approver must be a different user than the preparer.
    Returns 409 if the item is not yet prepared, or if the approver == preparer.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    period_result = await db.execute(
        select(AccountingPeriod).where(
            AccountingPeriod.id == period_id,
            AccountingPeriod.tenant_id == tenant_id,
        )
    )
    period = period_result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Period not found.")

    item_result = await db.execute(
        select(CloseChecklistItem).where(
            CloseChecklistItem.id == item_id,
            CloseChecklistItem.tenant_id == tenant_id,
        )
    )
    if not item_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Checklist item not found.")

    comp_result = await db.execute(
        select(PeriodChecklistCompletion).where(
            PeriodChecklistCompletion.period_id == period_id,
            PeriodChecklistCompletion.checklist_item_id == item_id,
        )
    )
    comp = comp_result.scalar_one_or_none()

    if comp is None or comp.status == "pending":
        raise HTTPException(status_code=409, detail="Item must be prepared before approval.")

    if comp.status == "approved":
        raise HTTPException(status_code=409, detail="Item is already approved.")

    if comp.prepared_by == current_user.user_id:
        raise HTTPException(
            status_code=409,
            detail="Segregation of duties: the approver must be a different user than the preparer.",
        )

    now = datetime.now(timezone.utc)
    comp.approved_by = current_user.user_id
    comp.approved_at = now
    comp.status = "approved"
    db.add(comp)
    await db.commit()
    await db.refresh(comp)

    return PeriodChecklistCompletionResponse(
        id=str(comp.id),
        period_id=str(comp.period_id),
        checklist_item_id=str(comp.checklist_item_id),
        item_label_snapshot=comp.item_label_snapshot,
        status=comp.status,
        prepared_by=str(comp.prepared_by) if comp.prepared_by else None,
        prepared_at=comp.prepared_at,
        approved_by=str(comp.approved_by) if comp.approved_by else None,
        approved_at=comp.approved_at,
        created_at=comp.created_at,
    )


# ── Modules ────────────────────────────────────────────────────────────────────

@router.get("/modules", response_model=ModulesResponse)
async def get_modules(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ModulesResponse:
    """Return all 14 modules with their activation and licensing state."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(TenantModule).where(TenantModule.tenant_id == tenant_id)
    )
    existing = {m.module_key: m for m in result.scalars().all()}

    modules = [
        ModuleState(
            module_key=m["key"],
            label=m["label"],
            is_active=existing[m["key"]].is_active if m["key"] in existing else False,
            is_licensed=existing[m["key"]].is_licensed if m["key"] in existing else False,
        )
        for m in MODULE_CATALOGUE
    ]
    return ModulesResponse(modules=modules)


@router.patch("/modules", response_model=ModulesResponse)
async def patch_modules(
    data: ModulesUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ModulesResponse:
    """
    Activate or deactivate modules for the tenant.

    Enforces: a module can only be activated if is_licensed = true.
    Returns 403 if attempting to activate an unlicensed module.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(TenantModule).where(TenantModule.tenant_id == tenant_id)
    )
    existing_map = {m.module_key: m for m in result.scalars().all()}

    now = datetime.now(timezone.utc)
    for key, active in data.modules.items():
        if key not in MODULE_KEY_TO_LABEL:
            continue

        existing = existing_map.get(key)
        if existing is None:
            # Create with is_licensed=False by default (consultant must license it)
            existing = TenantModule(
                tenant_id=tenant_id,
                module_key=key,
                is_active=False,
                is_licensed=False,
            )
            db.add(existing)
            existing_map[key] = existing

        if active and not existing.is_licensed:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Module '{MODULE_KEY_TO_LABEL[key]}' is not licensed for this tenant.",
            )

        if active and not existing.is_active:
            existing.activated_at = now
            existing.activated_by = current_user.user_id
        existing.is_active = active

    await db.commit()
    return await get_modules(current_user=current_user, db=db)


@router.patch("/modules/{module_key}/license", response_model=ModuleState)
async def set_module_license(
    module_key: str,
    is_licensed: bool,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> ModuleState:
    """Set is_licensed for a module (super admin only)."""
    if not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only consultants can set module licensing.",
        )
    tenant_id = _require_tenant(current_user)

    if module_key not in MODULE_KEY_TO_LABEL:
        raise HTTPException(status_code=404, detail="Unknown module key.")

    result = await db.execute(
        select(TenantModule).where(
            TenantModule.tenant_id == tenant_id,
            TenantModule.module_key == module_key,
        )
    )
    mod = result.scalar_one_or_none()
    if mod is None:
        mod = TenantModule(tenant_id=tenant_id, module_key=module_key, is_active=False, is_licensed=is_licensed)
        db.add(mod)
    else:
        mod.is_licensed = is_licensed

    await db.commit()
    await db.refresh(mod)
    return ModuleState(
        module_key=mod.module_key,
        label=MODULE_KEY_TO_LABEL[mod.module_key],
        is_active=mod.is_active,
        is_licensed=mod.is_licensed,
    )


# ── Dimensions not-applicable flag ────────────────────────────────────────────

@router.post("/dimensions/not-applicable", status_code=204)
async def set_dimensions_not_applicable(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Mark this tenant as not using analytical dimensions.

    Sets dimensions_not_applicable=True on the tenant record, which causes
    the setup progress endpoint to count dimensions as complete and unlocks
    the Chart of Accounts section.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    tenant_result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = tenant_result.scalar_one_or_none()
    if tenant is None:
        raise HTTPException(status_code=404, detail="Tenant not found.")
    tenant.dimensions_not_applicable = True
    await db.commit()


# ── Currencies & FX ───────────────────────────────────────────────────────────

@router.get("/currencies", response_model=FxConfigResponse)
async def get_currencies(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FxConfigResponse:
    """Return the tenant's currency config and FX mechanics.

    Single source of truth for the currency list:
      functional_currency  — from tenant_org_config (protected; set at signup)
      enabled_currencies   — from tenant_org_config (sorted ISO-code list)
      reporting_currency   — from tenant_org_config

    FX mechanics (rates, revaluation rules) come from tenant_fx_config.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    org_result = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org = org_result.scalar_one_or_none()

    fx_result = await db.execute(
        select(TenantFxConfig).where(TenantFxConfig.tenant_id == tenant_id)
    )
    fx = fx_result.scalar_one_or_none()

    return FxConfigResponse(
        functional_currency=org.functional_currency if org else None,
        enabled_currencies=org.enabled_currencies if org else None,
        reporting_currency=org.reporting_currency if org else None,
        fx_rates=fx.fx_rates if fx else None,
        revaluation_rules=fx.revaluation_rules if fx else None,
    )


@router.patch("/currencies", response_model=FxConfigResponse)
async def patch_currencies(
    data: FxConfigUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FxConfigResponse:
    """Update FX config for the tenant.

    enabled_currencies and reporting_currency are written to tenant_org_config
    (single source of truth for which currencies exist).
    fx_rates and revaluation_rules are written to tenant_fx_config.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    payload = data.model_dump(exclude_unset=True)

    # enabled_currencies and reporting_currency belong on org_config
    org_fields = {}
    if "enabled_currencies" in payload:
        org_fields["enabled_currencies"] = payload.pop("enabled_currencies")
    if "reporting_currency" in payload:
        org_fields["reporting_currency"] = payload.pop("reporting_currency")

    if org_fields:
        org = await _get_or_create_org(tenant_id, db)
        for field, value in org_fields.items():
            setattr(org, field, value)

    # fx_rates and revaluation_rules belong on fx_config
    if payload:
        fx = await _get_or_create_fx(tenant_id, db)
        for field, value in payload.items():
            setattr(fx, field, value)

    await db.commit()
    return await get_currencies(current_user=current_user, db=db)


@router.post("/currencies/fx-rates")
async def add_fx_rate(
    data: dict,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Add a single FX rate entry to the rate history."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    fx = await _get_or_create_fx(tenant_id, db)

    existing_rates = fx.fx_rates or []
    new_rate = {
        "id": str(uuid.uuid4()),
        "from_currency": data.get("from_currency"),
        "to_currency": data.get("to_currency"),
        "rate_type": data.get("rate_type", "mid"),
        "rate": data.get("rate"),
        "source": data.get("source", "manual"),
        "effective_date": data.get("effective_date"),
        "period": data.get("period"),
        "entered_by": current_user.email,
        "entered_at": datetime.utcnow().isoformat(),
        "proof_required": data.get("proof_required", False),
        "proof_reference": data.get("proof_reference"),
    }
    fx.fx_rates = existing_rates + [new_rate]
    await db.commit()
    return new_rate


# ── Tax & Statutory ───────────────────────────────────────────────────────────

@router.get("/tax", response_model=TaxConfigResponse)
async def get_tax(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> TaxConfigResponse:
    """Return tax config for the tenant (all four tabs)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    tax = await _get_or_create_tax(tenant_id, db)
    await db.commit()
    return TaxConfigResponse(
        vat_config=tax.vat_config,
        wht_config=tax.wht_config,
        paye_config=tax.paye_config,
        other_statutory=tax.other_statutory,
    )


@router.patch("/tax", response_model=TaxConfigResponse)
async def patch_tax(
    data: TaxConfigUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> TaxConfigResponse:
    """Update tax config for the tenant."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)
    tax = await _get_or_create_tax(tenant_id, db)

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(tax, field, value)

    await db.commit()
    await db.refresh(tax)
    return TaxConfigResponse(
        vat_config=tax.vat_config,
        wht_config=tax.wht_config,
        paye_config=tax.paye_config,
        other_statutory=tax.other_statutory,
    )


# ── Roles & Permissions ───────────────────────────────────────────────────────

_MATRIX_SECTIONS = [
    "Organisation", "Module activation", "Chart of accounts", "Dimensions",
    "Employees", "Currencies & FX", "Tax & statutory", "Roles & permissions",
    "Approval workflows", "Document rules", "Module setup",
]

_MATRIX_DEFAULTS: dict[tuple[str, str], str] = {
    (sec, "consultant"):       "full"
    for sec in _MATRIX_SECTIONS
} | {
    (sec, "power_admin"):      "read_only"
    for sec in _MATRIX_SECTIONS
} | {
    (sec, "functional_admin"): "read_only"
    for sec in _MATRIX_SECTIONS
}


@router.get("/roles/matrix", response_model=PermissionMatrixResponse)
async def get_roles_matrix(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PermissionMatrixResponse:
    """Return the permission matrix (DB overrides + hardcoded defaults)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    rows = (await db.execute(
        select(TenantPermissionMatrix).where(TenantPermissionMatrix.tenant_id == tenant_id)
    )).scalars().all()

    overrides: dict[tuple[str, str], str] = {
        (r.section, r.role_tier): r.access_level for r in rows
    }

    cells = []
    for sec in _MATRIX_SECTIONS:
        for tier in ("consultant", "power_admin", "functional_admin"):
            cells.append({
                "section": sec,
                "role_tier": tier,
                "access_level": overrides.get((sec, tier), _MATRIX_DEFAULTS[(sec, tier)]),
            })

    return PermissionMatrixResponse(cells=cells)


@router.patch("/roles/matrix", response_model=PermissionMatrixResponse)
async def patch_roles_matrix(
    data: PermissionMatrixUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PermissionMatrixResponse:
    """Upsert permission matrix cells. Super admin (consultant) only."""
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="Only ZivaBI Consultants can modify the permission matrix.")
    tenant_id = _require_tenant(current_user)

    for cell in data.cells:
        existing = (await db.execute(
            select(TenantPermissionMatrix).where(
                TenantPermissionMatrix.tenant_id == tenant_id,
                TenantPermissionMatrix.section == cell.section,
                TenantPermissionMatrix.role_tier == cell.role_tier,
            )
        )).scalar_one_or_none()

        if existing:
            existing.access_level = cell.access_level
        else:
            db.add(TenantPermissionMatrix(
                tenant_id=tenant_id,
                section=cell.section,
                role_tier=cell.role_tier,
                access_level=cell.access_level,
            ))

    await db.commit()

    # Return the full matrix after save
    rows = (await db.execute(
        select(TenantPermissionMatrix).where(TenantPermissionMatrix.tenant_id == tenant_id)
    )).scalars().all()
    overrides = {(r.section, r.role_tier): r.access_level for r in rows}
    cells_out = []
    for sec in _MATRIX_SECTIONS:
        for tier in ("consultant", "power_admin", "functional_admin"):
            cells_out.append({
                "section": sec,
                "role_tier": tier,
                "access_level": overrides.get((sec, tier), _MATRIX_DEFAULTS[(sec, tier)]),
            })
    return PermissionMatrixResponse(cells=cells_out)


@router.get("/roles/assignments", response_model=list[RoleAssignmentResponse])
async def get_role_assignments(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[RoleAssignmentResponse]:
    """List all users in this tenant with their role tier assignments."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(UserTenant, User)
        .join(User, UserTenant.user_id == User.id)
        .where(UserTenant.tenant_id == tenant_id, UserTenant.is_active.is_(True))
    )
    return [
        RoleAssignmentResponse(
            id=str(ut.id),
            user_id=str(u.id),
            user_tenant_id=str(ut.id),
            full_name=u.full_name,
            email=u.email,
            role_tier=ut.role_tier,
            is_active=ut.is_active,
        )
        for ut, u in result.all()
    ]


@router.post("/roles/assignments", response_model=RoleAssignmentResponse)
async def create_role_assignment(
    data: RoleAssignmentCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RoleAssignmentResponse:
    """Assign a role tier to a user within this tenant."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(UserTenant, User)
        .join(User, UserTenant.user_id == User.id)
        .where(UserTenant.id == data.user_tenant_id, UserTenant.tenant_id == tenant_id)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="User not found in this tenant.")

    ut, u = row
    ut.role_tier = data.role_tier
    await db.commit()
    await db.refresh(ut)
    return RoleAssignmentResponse(
        id=str(ut.id), user_id=str(u.id), user_tenant_id=str(ut.id),
        full_name=u.full_name, email=u.email,
        role_tier=ut.role_tier, is_active=ut.is_active,
    )


@router.patch("/roles/assignments/{assignment_id}", response_model=RoleAssignmentResponse)
async def update_role_assignment(
    assignment_id: uuid.UUID,
    data: RoleAssignmentUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> RoleAssignmentResponse:
    """Update the role tier for an existing assignment."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(UserTenant, User)
        .join(User, UserTenant.user_id == User.id)
        .where(UserTenant.id == assignment_id, UserTenant.tenant_id == tenant_id)
    )
    row = result.first()
    if not row:
        raise HTTPException(status_code=404, detail="Assignment not found.")

    ut, u = row
    ut.role_tier = data.role_tier
    await db.commit()
    await db.refresh(ut)
    return RoleAssignmentResponse(
        id=str(ut.id), user_id=str(u.id), user_tenant_id=str(ut.id),
        full_name=u.full_name, email=u.email,
        role_tier=ut.role_tier, is_active=ut.is_active,
    )


@router.get("/roles/assignments/{assignment_id}/scope", response_model=FunctionalScopeResponse)
async def get_functional_scope(
    assignment_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionalScopeResponse:
    """Return the list of sections granted to a functional admin."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    rows = (await db.execute(
        select(UserFunctionalScope).where(
            UserFunctionalScope.tenant_id == tenant_id,
            UserFunctionalScope.user_tenant_id == assignment_id,
        )
    )).scalars().all()

    return FunctionalScopeResponse(
        user_tenant_id=str(assignment_id),
        sections=[{"section": r.section, "access_level": r.access_level} for r in rows],
    )


@router.patch("/roles/assignments/{assignment_id}/scope", response_model=FunctionalScopeResponse)
async def patch_functional_scope(
    assignment_id: uuid.UUID,
    data: FunctionalScopeUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionalScopeResponse:
    """Replace all scope sections for a functional admin (full replace, not merge)."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    # Delete existing scope for this user
    await db.execute(
        delete(UserFunctionalScope).where(
            UserFunctionalScope.tenant_id == tenant_id,
            UserFunctionalScope.user_tenant_id == assignment_id,
        )
    )

    # Insert new scope rows with individual access levels
    for item in data.sections:
        db.add(UserFunctionalScope(
            tenant_id=tenant_id,
            user_tenant_id=assignment_id,
            section=item.section,
            access_level=item.access_level,
        ))

    await db.commit()

    return FunctionalScopeResponse(
        user_tenant_id=str(assignment_id),
        sections=[{"section": i.section, "access_level": i.access_level} for i in data.sections],
    )


# ── Document Rules ─────────────────────────────────────────────────────────────

@router.get("/documents", response_model=list[DocumentRuleResponse])
async def get_documents(
    module: Optional[str] = Query(None),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[DocumentRuleResponse]:
    """List document rules, optionally filtered by module."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    q = select(DocumentRule).where(
        DocumentRule.tenant_id == tenant_id, DocumentRule.is_active.is_(True)
    )
    if module:
        q = q.where(DocumentRule.module == module)
    q = q.order_by(DocumentRule.module, DocumentRule.transaction_type)

    result = await db.execute(q)
    return [
        DocumentRuleResponse(
            id=str(r.id), module=r.module, transaction_type=r.transaction_type,
            document_name=r.document_name, is_required=r.is_required,
            track_expiry=r.track_expiry, ocr_template=r.ocr_template,
            max_size_mb=r.max_size_mb, allowed_formats=r.allowed_formats,
            max_files=r.max_files, is_active=r.is_active,
        )
        for r in result.scalars().all()
    ]


@router.post("/documents", response_model=DocumentRuleResponse, status_code=201)
async def create_document_rule(
    data: DocumentRuleCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DocumentRuleResponse:
    """Create a new document rule."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    rule = DocumentRule(
        tenant_id=tenant_id, module=data.module, transaction_type=data.transaction_type,
        document_name=data.document_name, is_required=data.is_required,
        track_expiry=data.track_expiry, ocr_template=data.ocr_template,
        max_size_mb=data.max_size_mb, allowed_formats=data.allowed_formats,
        max_files=data.max_files,
    )
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return DocumentRuleResponse(
        id=str(rule.id), module=rule.module, transaction_type=rule.transaction_type,
        document_name=rule.document_name, is_required=rule.is_required,
        track_expiry=rule.track_expiry, ocr_template=rule.ocr_template,
        max_size_mb=rule.max_size_mb, allowed_formats=rule.allowed_formats,
        max_files=rule.max_files, is_active=rule.is_active,
    )


@router.patch("/documents/{rule_id}", response_model=DocumentRuleResponse)
async def update_document_rule(
    rule_id: uuid.UUID,
    data: DocumentRuleUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> DocumentRuleResponse:
    """Update a document rule."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(DocumentRule).where(DocumentRule.id == rule_id, DocumentRule.tenant_id == tenant_id)
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="Document rule not found.")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(rule, field, value)

    await db.commit()
    await db.refresh(rule)
    return DocumentRuleResponse(
        id=str(rule.id), module=rule.module, transaction_type=rule.transaction_type,
        document_name=rule.document_name, is_required=rule.is_required,
        track_expiry=rule.track_expiry, ocr_template=rule.ocr_template,
        max_size_mb=rule.max_size_mb, allowed_formats=rule.allowed_formats,
        max_files=rule.max_files, is_active=rule.is_active,
    )


@router.delete("/documents/{rule_id}", status_code=204)
async def delete_document_rule(
    rule_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Soft-delete a document rule."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(DocumentRule).where(DocumentRule.id == rule_id, DocumentRule.tenant_id == tenant_id)
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="Document rule not found.")
    rule.is_active = False
    await db.commit()


# ── System Function Mappings ───────────────────────────────────────────────────
# Maps business functions (Finance, HR, Procurement, …) to OrgStructureNodes so
# that downstream features (Finance Review Chain, payroll scoping, etc.) know
# which cost centres belong to which department type.

# Catalogue of supported functions and which module activates them (None = always shown).
_FUNCTION_CATALOGUE: dict[str, dict] = {
    "finance":     {"label": "Finance & Accounting", "module": None},
    "hr":          {"label": "Human Resources",      "module": "payroll"},
    "procurement": {"label": "Procurement",          "module": "accounts_payable"},
    "sales":       {"label": "Sales & Revenue",      "module": "accounts_receivable"},
    "operations":  {"label": "Operations",           "module": "inventory"},
    "audit":       {"label": "Internal Audit",       "module": None},
}


@router.get("/function-mappings", response_model=FunctionMappingsResponse)
async def get_function_mappings(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionMappingsResponse:
    """
    Return all system function → cost-centre mappings for this tenant.

    Each row includes the resolved cost-centre name and code from org_structure.
    """
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(SystemFunctionMapping, OrgStructureNode)
        .join(OrgStructureNode, SystemFunctionMapping.cost_center_id == OrgStructureNode.id)
        .where(SystemFunctionMapping.tenant_id == tenant_id)
        .order_by(SystemFunctionMapping.function_code)
    )
    rows = result.all()

    return FunctionMappingsResponse(
        mappings=[
            FunctionMappingItem(
                id=str(m.id),
                function_code=m.function_code,
                cost_center_id=str(m.cost_center_id),
                cost_center_name=node.name,
                cost_center_code=node.code,
                is_primary=m.is_primary,
            )
            for m, node in rows
        ]
    )


@router.put("/function-mappings", response_model=FunctionMappingsResponse)
async def upsert_function_mappings(
    body: list[FunctionMappingUpsertItem],
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionMappingsResponse:
    """
    Replace all function mappings for this tenant with the supplied list.

    Strategy: delete all existing rows for the tenant then insert the new set.
    This is a full replacement (not partial), so send the complete desired state.

    Validates:
    - function_code must be one of the known catalogue keys.
    - cost_center_id must belong to this tenant's org_structure.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    valid_codes = set(_FUNCTION_CATALOGUE.keys())
    for item in body:
        if item.function_code not in valid_codes:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown function_code '{item.function_code}'. Valid: {sorted(valid_codes)}",
            )

    # Verify all cost_center_ids belong to this tenant
    cc_ids = [uuid.UUID(item.cost_center_id) for item in body]
    if cc_ids:
        cc_result = await db.execute(
            select(OrgStructureNode.id).where(
                OrgStructureNode.id.in_(cc_ids),
                OrgStructureNode.tenant_id == tenant_id,
            )
        )
        found_ids = {row[0] for row in cc_result.all()}
        missing = [str(i) for i in cc_ids if i not in found_ids]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"cost_center_id(s) not found in this tenant's org: {missing}",
            )

    # Full replace
    await db.execute(
        delete(SystemFunctionMapping).where(SystemFunctionMapping.tenant_id == tenant_id)
    )
    new_rows = [
        SystemFunctionMapping(
            tenant_id=tenant_id,
            function_code=item.function_code,
            cost_center_id=uuid.UUID(item.cost_center_id),
            is_primary=item.is_primary,
        )
        for item in body
    ]
    db.add_all(new_rows)
    await db.commit()

    # Re-fetch with joined node names
    result = await db.execute(
        select(SystemFunctionMapping, OrgStructureNode)
        .join(OrgStructureNode, SystemFunctionMapping.cost_center_id == OrgStructureNode.id)
        .where(SystemFunctionMapping.tenant_id == tenant_id)
        .order_by(SystemFunctionMapping.function_code)
    )
    rows = result.all()

    return FunctionMappingsResponse(
        mappings=[
            FunctionMappingItem(
                id=str(m.id),
                function_code=m.function_code,
                cost_center_id=str(m.cost_center_id),
                cost_center_name=node.name,
                cost_center_code=node.code,
                is_primary=m.is_primary,
            )
            for m, node in rows
        ]
    )


@router.get("/functions/{function_code}/team", response_model=FunctionTeamResponse)
async def get_function_team(
    function_code: str,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionTeamResponse:
    """
    Return all OrgRoles whose cost centre is mapped to the given function code.

    Used by the Finance Review Chain step builder (and similar features) to
    populate the Named Assignee dropdown with only Finance staff.
    """
    from app.models.approvals import OrgRole

    tenant_id = _require_tenant(current_user)

    if function_code not in _FUNCTION_CATALOGUE:
        raise HTTPException(status_code=404, detail=f"Unknown function code: {function_code}")

    mapping_result = await db.execute(
        select(SystemFunctionMapping, OrgStructureNode)
        .join(OrgStructureNode, SystemFunctionMapping.cost_center_id == OrgStructureNode.id)
        .where(
            SystemFunctionMapping.tenant_id == tenant_id,
            SystemFunctionMapping.function_code == function_code,
        )
    )
    mapping_rows = mapping_result.all()

    if not mapping_rows:
        return FunctionTeamResponse(
            function_code=function_code,
            function_label=_FUNCTION_CATALOGUE[function_code]["label"],
            cost_centers=[],
            team=[],
        )

    cc_ids = [m.cost_center_id for m, _ in mapping_rows]
    cc_names = [node.name for _, node in mapping_rows]

    roles_result = await db.execute(
        select(OrgRole).where(
            OrgRole.tenant_id == tenant_id,
            OrgRole.cost_center_id.in_(cc_ids),
        ).order_by(OrgRole.name)
    )
    roles = roles_result.scalars().all()

    team: list[FunctionTeamMember] = []
    for role in roles:
        node_result = await db.execute(
            select(OrgStructureNode).where(OrgStructureNode.id == role.cost_center_id)
        )
        node = node_result.scalar_one_or_none()
        team.append(
            FunctionTeamMember(
                role_id=str(role.id),
                role_name=role.name,
                designation=role.designation,
                cost_center_id=str(role.cost_center_id) if role.cost_center_id else "",
                cost_center_name=node.name if node else "",
                occupants=role.occupants or [],
            )
        )

    return FunctionTeamResponse(
        function_code=function_code,
        function_label=_FUNCTION_CATALOGUE[function_code]["label"],
        cost_centers=cc_names,
        team=team,
    )


@router.get("/functions/{function_code}/users", response_model=FunctionUsersResponse)
async def get_function_users(
    function_code: str,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FunctionUsersResponse:
    """
    Return all active Ziva users whose Employee record sits in a cost centre mapped to function_code.

    Bridge: Employee.email == User.email (Employee has no direct user_id FK).
    Returns mapped=False if no cost-centre mapping has been configured for this function.
    Used by Finance Review Chain and similar features to pre-filter assignee dropdowns.
    """
    tenant_id = _require_tenant(current_user)

    if function_code not in _FUNCTION_CATALOGUE:
        raise HTTPException(status_code=404, detail=f"Unknown function code: {function_code}")

    # Step 1: cost centres mapped to this function
    mapping_result = await db.execute(
        select(SystemFunctionMapping, OrgStructureNode)
        .join(OrgStructureNode, SystemFunctionMapping.cost_center_id == OrgStructureNode.id)
        .where(
            SystemFunctionMapping.tenant_id == tenant_id,
            SystemFunctionMapping.function_code == function_code,
        )
    )
    mapping_rows = mapping_result.all()

    if not mapping_rows:
        return FunctionUsersResponse(
            function_code=function_code,
            function_label=_FUNCTION_CATALOGUE[function_code]["label"],
            mapped=False,
            users=[],
        )

    cc_ids = [m.cost_center_id for m, _ in mapping_rows]

    # Step 2: active employees in those cost centres who have active Ziva user accounts
    emp_result = await db.execute(
        select(Employee, User, OrgStructureNode)
        .join(User, User.email == Employee.email)
        .outerjoin(OrgStructureNode, OrgStructureNode.id == Employee.cost_center_id)
        .where(
            Employee.tenant_id == tenant_id,
            Employee.cost_center_id.in_(cc_ids),
            Employee.is_active.is_(True),
            User.is_active.is_(True),
        )
        .order_by(User.full_name)
    )
    emp_rows = emp_result.all()

    users = [
        FunctionUserItem(
            user_id=str(user.id),
            full_name=user.full_name,
            email=user.email,
            employee_code=emp.employee_code,
            cost_center_name=node.name if node else "",
            cost_center_code=node.code if node else "",

        )
        for emp, user, node in emp_rows
    ]

    return FunctionUsersResponse(
        function_code=function_code,
        function_label=_FUNCTION_CATALOGUE[function_code]["label"],
        mapped=True,
        users=users,
    )


# ── Go-live ──────────────────────────────────────────────────────────────────────────────

@router.post("/go-live", response_model=GoLiveResponse)
async def mark_go_live(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> GoLiveResponse:
    """
    Mark this tenant as live.

    Requires is_super_admin (consultant role_tier stripped M9.3a).
    All blocking sections must be complete.
    """
    if not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only consultants and super admins can mark a tenant as live.",
        )

    tenant_id = _require_tenant(current_user)

    progress = await get_progress(current_user=current_user, db=db)
    blocking_incomplete = [s.label for s in progress.sections if s.blocking and s.status != "complete"]
    if blocking_incomplete:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Blocking items incomplete: {', '.join(blocking_incomplete)}",
        )

    result = await db.execute(select(Tenant).where(Tenant.id == tenant_id))
    tenant = result.scalar_one_or_none()
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant not found.")

    if tenant.environment != "live":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "This tenant has no live environment yet. Live is created by "
                "promoting validated test configuration: POST "
                "/api/platform/tenants/{tenant_id}/promotion/diff then "
                ".../promotion/apply (super-admin only), not by go-live directly."
            ),
        )

    old_lifecycle = tenant.lifecycle_status

    tenant.is_active = True
    tenant.lifecycle_status = "live"

    db.add(AuditLog(
        event_type="platform.lifecycle.updated",
        user_id=current_user.user_id,
        tenant_id=tenant_id,
        log_metadata={
            "from": old_lifecycle,
            "to": "live",
            "via": "go_live",
        },
    ))

    await db.commit()

    return GoLiveResponse(
        message="Tenant is now live. Welcome emails will be sent to all Power Admins.",
        tenant_id=str(tenant_id),
    )
