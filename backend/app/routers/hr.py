"""
ZivaBI — M8.1 / M8.2 HR module router.

Registered at prefix /api/hr.

Endpoints:
  Employees:
    GET    /api/hr/employees                         List employees (paginated, searchable)
    POST   /api/hr/employees                         Create single employee
    PATCH  /api/hr/employees/{id}                    Update employee
    DELETE /api/hr/employees/{id}                    Soft delete (deactivate)
    POST   /api/hr/employees/upload                  Bulk upload via xlsx/csv
    GET    /api/hr/employees/template                Download employee upload template
    POST   /api/hr/employees/{id}/transfer           Transfer to new cost center
    POST   /api/hr/employees/{id}/update-code        Update employee code (retro/progressive)
    GET    /api/hr/employees/{id}/history            View code change + transfer history
    POST   /api/hr/employees/invite                  Send self-onboarding invite (M8.2)
    POST   /api/hr/employees/{id}/approve-onboarding Approve HR self-onboarding (M8.2)
    POST   /api/hr/employees/{id}/reject-onboarding  Reject HR self-onboarding with comment (M8.2)

  Cost Center Config:
    GET    /api/hr/cost-centers                      List cost centers with head assignments
    PUT    /api/hr/cost-centers/{id}/head            Set or update cost center head

  Finance Review Config:
    GET    /api/hr/finance-review                    List finance reviewers by module
    POST   /api/hr/finance-review                    Add a reviewer
    PATCH  /api/hr/finance-review/{id}               Update reviewer level/scope
    DELETE /api/hr/finance-review/{id}               Remove reviewer

Public (no auth required):
  GET    /onboard/{token}                            Validate token and return employee stub
  POST   /onboard/{token}                            New hire submits self-onboarding form

All authenticated endpoints are tenant-scoped and require authentication.
Admin-only operations require is_tenant_admin or is_super_admin.
"""

import csv
import io
import re
import secrets
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.database import get_db
from app.middleware.auth import CurrentUser, require_auth, block_if_readonly_impersonation
from app.models.master_data import (
    CostCenterConfig,
    Employee,
    EmployeeCodeHistory,
    EmployeePositionAssignment,
    EmployeeTransfer,
    FinanceReviewConfig,
)
from app.models.approvals import ApprovalRole as OrgRole
from app.models.setup import OrgStructureNode
from app.models.setup import EmployeeOnboardingToken
from app.models.auth import Tenant
from app.schemas.setup import EmployeeInviteCreate, SelfOnboardingSubmit, SelfOnboardingTokenResponse
from app.schemas.hr import (
    CodeUpdateRequest,
    CodeHistoryResponse,
    CostCenterConfigResponse,
    CostCenterHeadUpdate,
    CostCenterOption,
    EmployeeAssignmentResponse,
    EmployeeAssignRequest,
    EmployeeCreate,
    EmployeeHistoryResponse,
    EmployeeListItem,
    EmployeeResponse,
    EmployeeUpdate,
    EmployeeUploadResult,
    FinanceReviewConfigResponse,
    FinanceReviewerCreate,
    FinanceReviewerUpdate,
    PositionCreate,
    PositionMoveRequest,
    PositionOccupant,
    PositionResponse,
    PositionUpdate,
    TransferCreate,
    TransferResponse,
)

router = APIRouter(prefix="/api/hr", tags=["hr"])


# ── Shared helpers ────────────────────────────────────────────────────────────

def _require_tenant(current_user: CurrentUser) -> uuid.UUID:
    if current_user.tenant_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="This feature is only available on business accounts.",
        )
    return current_user.tenant_id


def _require_admin(current_user: CurrentUser) -> None:
    if not current_user.is_tenant_admin and not current_user.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only Tenant Admins can perform this action.",
        )
    block_if_readonly_impersonation(current_user)


async def _ensure_portal_account(
    email: str,
    full_name: str,
    first_name: str | None,
    tenant_id: uuid.UUID,
    db: AsyncSession,
    employee: "Employee | None" = None,
) -> None:
    """
    Idempotently ensure that a Ziva portal account (User + UserTenant) exists
    for the given employee email. Supports rehire: if the UserTenant was previously
    deactivated (e.g. employee resigned), it is reactivated.

    Also sets employee.user_id (direct FK) when an Employee object is supplied,
    and marks the UserTenant as user_type='employee'.

    Called every time an employee is created or imported.
    """
    from app.models.auth import User as UserModel, UserTenant as UserTenantModel, AccountType
    from app.core.security import hash_password as _hash

    result = await db.execute(select(UserModel).where(UserModel.email == email))
    user = result.scalar_one_or_none()

    if not user:
        user = UserModel(
            email=email,
            full_name=full_name,
            first_name=first_name,
            account_type=AccountType.business,
        )
        db.add(user)
        await db.flush()

    # Link employee → user (direct FK, replaces email-based joins)
    if employee is not None and employee.user_id != user.id:
        employee.user_id = user.id

    ut_result = await db.execute(
        select(UserTenantModel).where(
            UserTenantModel.user_id == user.id,
            UserTenantModel.tenant_id == tenant_id,
        )
    )
    existing_ut = ut_result.scalar_one_or_none()
    if existing_ut is None:
        user_tenant = UserTenantModel(
            user_id=user.id,
            tenant_id=tenant_id,
            user_type="employee",
            # Random unusable password — employee activates account via forgot-password.
            password_hash=_hash(secrets.token_hex(32)),
        )
        db.add(user_tenant)
    else:
        # Rehire: reactivate a previously deactivated UserTenant
        if not existing_ut.is_active:
            existing_ut.is_active = True
        existing_ut.user_type = "employee"
    await db.flush()


async def _get_employee_or_404(
    employee_id: uuid.UUID, tenant_id: uuid.UUID, db: AsyncSession
) -> Employee:
    result = await db.execute(
        select(Employee)
        .where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
        .options(
            selectinload(Employee.cost_center),
            selectinload(Employee.line_manager),
            selectinload(Employee.approval_role),
        )
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found.")
    return emp


async def _next_employee_code(tenant_id: uuid.UUID, db: AsyncSession) -> str:
    """
    Generate the next auto employee code for the tenant.

    Uses prefix EMP- with 5-digit zero-padded sequence based on existing count.
    Prefix/format customisation is out of scope for M8.1 — placeholder pattern.
    """
    result = await db.execute(
        select(Employee).where(
            Employee.tenant_id == tenant_id,
            Employee.employee_code.isnot(None),
            Employee.employee_code_auto_generated == True,  # noqa: E712
        )
    )
    count = len(result.scalars().all())
    return f"EMP-{count + 1:05d}"


# ═══════════════════════════════════════════════════════════════════════════════
# EMPLOYEES
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/employees/template")
async def download_employee_template(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """Download the employee bulk upload template (.xlsx)."""
    tenant_id = _require_tenant(current_user)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl is not installed.",
        ) from exc

    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment as XLComment

    # Fetch cost center nodes from org_structure (single source of truth)
    cc_nodes_res = await db.execute(_cc_nodes_query(current_user.tenant_id))
    cc_nodes_all = cc_nodes_res.scalars().all()
    cc_codes_for_template = [_cc_node_code(n) for n in cc_nodes_all if _cc_node_code(n)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.freeze_panes = "A2"

    req_fill = PatternFill("solid", fgColor="DBEAFE")   # blue — required
    opt_fill = PatternFill("solid", fgColor="F3F4F6")   # grey — optional
    data_start_fill = PatternFill("solid", fgColor="EFF6FF")  # very light blue data-start hint
    hdr_font = Font(bold=True, size=10)

    # Load org roles for template dropdown
    from app.models.approvals import ApprovalRole as AR
    role_rows_q = await db.execute(
        select(AR.id, AR.name, AR.cost_center_id, AR.area, AR.sub_area).where(AR.tenant_id == tenant_id, AR.is_active.is_(True)).order_by(AR.display_order, AR.name)
    )
    role_rows = role_rows_q.all()
    # Fetch CC codes for disambiguation (name collisions across departments)
    cc_code_map: dict[uuid.UUID, str] = {}
    if cc_nodes_all:
        cc_code_map = {n.id: (_cc_node_code(n) or "") for n in cc_nodes_all if _cc_node_code(n)}
    # Build display names: append CC code when name is not unique
    from collections import Counter as _Counter
    cc_name_map: dict[uuid.UUID, str] = {}
    if cc_nodes_all:
        cc_name_map = {n.id: n.name for n in cc_nodes_all}
    from collections import Counter as _Counter
    role_display_names: dict[uuid.UUID, str] = {}
    for r in role_rows:
        label = r.name
        if r.cost_center_id:
            cc_n = cc_name_map.get(r.cost_center_id, "")
            if cc_n:
                label = f"{r.name} — {cc_n}"
        parts: list[str] = []
        if r.area:
            parts.append(r.area)
        if r.sub_area:
            parts.append(r.sub_area)
        if parts:
            label = f"{label} [{' > '.join(parts)}]"
        role_display_names[r.id] = label
    role_names_for_template = list(role_display_names.values())

    headers = [
        ("Org Role*", True),
        ("First Name*", True), ("Last Name*", True), ("Email*", True),
        ("Other Name", False), ("Preferred Name", False),
        ("Employee Code", False), ("Phone", False),
        ("Cost Center Code", False),
        ("Resumption Date (dd/mm/yyyy)*", True),
    ]
    # Examples go in cell comments on header cells
    header_examples = [
        f"Select the employee's org chart role (mandatory).\ne.g. {role_names_for_template[0] if role_names_for_template else 'Finance Manager'}\nControls approval authority and org-chart placement.",
        "e.g. Amara",
        "e.g. Okafor",
        "e.g. amara.okafor@company.com",
        "e.g. Grace (middle or other name)",
        "e.g. Amy (display name on expense forms)",
        "e.g. EMP-00001 (leave blank to auto-generate)",
        "e.g. +234-801-234-5678",
        f"Optional if the role has a fixed cost centre (auto-assigns).\ne.g. {cc_codes_for_template[0] if cc_codes_for_template else 'CC001'}",
        "Date the employee started (mandatory). Format: dd/mm/yyyy.\ne.g. 01/01/2024",
    ]

    # Row 1: headers with cell comments (no inline instruction/example rows per format standard)
    for ci, ((h, is_req), example_text) in enumerate(zip(headers, header_examples), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = req_fill if is_req else opt_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(18, len(h) + 2)
        comment = XLComment(example_text, "Ziva BI")
        comment.width = 220; comment.height = 70
        cell.comment = comment
    ws.row_dimensions[1].height = 30

    # Row 2: data-start marker (visual cue per template format standard; data entry starts here)
    for ci in range(1, len(headers) + 1):
        ws.cell(row=2, column=ci).fill = data_start_fill

    # ── Data validations — ranges start at row 2 (data row 1) ────────────────

    # Cost Center Code column (H = col 8): always add the DV, use a hidden sheet when
    # many codes exist (formula1 has a 255-char limit when embedded as a string).
    # When cc_codes_for_template is empty the DV is still created but with an empty
    # list — Excel ignores empty-list DVs gracefully (no dropdown appears, no block).
    if cc_codes_for_template:
        cc_formula = '"' + ",".join(cc_codes_for_template) + '"'
        if len(cc_formula) <= 255:
            dv_cc = DataValidation(
                type="list",
                formula1=cc_formula,
                showDropDown=False,
                showErrorMessage=True,
                error="Please select a valid Cost Center Code from the dropdown.",
                errorTitle="Invalid Cost Center",
                prompt="Select from the list of valid cost center codes.",
                promptTitle="Cost Center Code",
            )
            ws.add_data_validation(dv_cc)
            dv_cc.sqref = "I2:I10002"
        else:
            # Too many CC codes for an inline formula — use a hidden helper sheet
            ws_cc = wb.create_sheet("_CC_Codes")
            for ri, code in enumerate(cc_codes_for_template, 1):
                ws_cc.cell(row=ri, column=1, value=code)
            ws_cc.sheet_state = "hidden"
            dv_cc = DataValidation(
                type="list",
                formula1=f"_CC_Codes!$A$1:$A${len(cc_codes_for_template)}",
                showDropDown=False,
                showErrorMessage=True,
                error="Please select a valid Cost Center Code.",
                errorTitle="Invalid Cost Center",
            )
            ws.add_data_validation(dv_cc)
            dv_cc.sqref = "I2:I10002"

    # Org Role column (A = col 1; first column): dropdown from role names.
    # Same hidden-sheet guard as CC codes — inline DV formula1 is capped at 255 chars.
    if role_names_for_template:
        role_inline = '"' + ",".join(role_names_for_template) + '"'
        if len(role_inline) <= 255:
            role_dv = DataValidation(
                type="list", formula1=role_inline,
                allow_blank=True, showDropDown=False,
                showErrorMessage=True,
                error="Please select a valid Org Role from the dropdown.",
                errorTitle="Invalid Org Role",
            )
            ws.add_data_validation(role_dv)
            role_dv.sqref = "A2:A10002"
        else:
            ws_roles = wb.create_sheet("_Role_Names")
            for ri, rname in enumerate(role_names_for_template, 1):
                ws_roles.cell(row=ri, column=1, value=rname)
            ws_roles.sheet_state = "hidden"
            role_dv = DataValidation(
                type="list",
                formula1=f"_Role_Names!$A$1:$A${len(role_names_for_template)}",
                allow_blank=True, showDropDown=False,
                showErrorMessage=True,
                error="Please select a valid Org Role from the dropdown.",
                errorTitle="Invalid Org Role",
            )
            ws.add_data_validation(role_dv)
            role_dv.sqref = "A2:A10002"

    # Sheet 2 — Instructions
    ws2 = wb.create_sheet("Instructions")
    rows = [
        ["COLUMN", "REQUIRED", "DESCRIPTION"],
        ["Org Role", "Yes", "Position in the role hierarchy this employee occupies. Must match an existing role exactly. Controls approval authority and org-chart placement."],
        ["First Name", "Yes", "Employee's first name."],
        ["Last Name", "Yes", "Employee's last name."],
        ["Email", "Yes", "Employee's work email. Must be unique per company."],
        ["Other Name", "No", "Middle name or other name."],
        ["Preferred Name", "No", "Name to display on expense forms (defaults to First Name)."],
        ["Employee Code", "No", "Required if auto-generate is turned off in HR settings."],
        ["Phone", "No", "Phone number including country code."],
        ["Cost Center Code", "No", "Leave blank to auto-assign from the role's cost centre. If provided, must match the cost centre assigned to the selected role."],
        ["Resumption Date", "Yes", "Date the employee started (mandatory). Format: dd/mm/yyyy. Determines when the account becomes active."],
        ["", "", ""],
        ["DATA START ROW", "", "Enter your data starting at row 2. Row 1 is the header row."],
        ["CELL COMMENTS", "", "Hover over each header cell (row 1) for example values and guidance."],
        ["CAPACITY LIMIT", "", "Upload is blocked when a role's headcount limit is reached. Check Role Hierarchy for capacity settings."],
    ]
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 70
    for ri, row_vals in enumerate(rows, 1):
        for ci, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True)
                cell.fill = req_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"},
    )


@router.get("/employees", response_model=list[EmployeeListItem])
async def list_employees(
    search: str = Query(default="", description="Search name, email, or employee code"),
    cost_center_id: Optional[uuid.UUID] = Query(default=None),
    active_only: bool = Query(default=True),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[EmployeeListItem]:
    """List employees for the tenant (paginated, searchable)."""
    tenant_id = _require_tenant(current_user)

    q = (
        select(Employee)
        .where(Employee.tenant_id == tenant_id)
        .options(
            selectinload(Employee.cost_center),
            selectinload(Employee.line_manager),
            selectinload(Employee.approval_role),
        )
        .order_by(Employee.last_name, Employee.first_name)
        .limit(limit)
        .offset(offset)
    )
    if active_only:
        q = q.where(Employee.is_active == True)  # noqa: E712
    if cost_center_id:
        q = q.where(Employee.cost_center_id == cost_center_id)
    if search.strip():
        term = f"%{search.strip()}%"
        q = q.where(
            Employee.first_name.ilike(term)
            | Employee.last_name.ilike(term)
            | Employee.email.ilike(term)
            | Employee.employee_code.ilike(term)
        )

    result = await db.execute(q)
    rows = result.scalars().all()
    items = [EmployeeListItem.from_orm(e) for e in rows]

    # M9.3b: batch-resolve email → user_id so impersonation entry point works.
    if items:
        from app.models.auth import User as UserModel
        emails = [it.email for it in items]
        uid_rows = await db.execute(
            select(UserModel.id, UserModel.email).where(UserModel.email.in_(emails))
        )
        email_to_uid = {row.email: str(row.id) for row in uid_rows}
        for item in items:
            item.user_id = email_to_uid.get(item.email)

    return items


@router.post("/employees", response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
async def create_employee(
    data: EmployeeCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeResponse:
    """Create a single employee. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    # Check email uniqueness
    existing_email = await db.execute(
        select(Employee).where(
            Employee.tenant_id == tenant_id,
            Employee.email == data.email,
        )
    )
    if existing_email.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"An employee with email '{data.email}' already exists.",
        )

    employee_code = data.employee_code
    auto_generated = False
    if not employee_code:
        employee_code = await _next_employee_code(tenant_id, db)
        auto_generated = True

    emp = Employee(
        tenant_id=tenant_id,
        first_name=data.first_name,
        last_name=data.last_name,
        other_name=data.other_name,
        preferred_name=data.preferred_name,
        email=data.email,
        phone=data.phone,
        employee_code=employee_code,
        employee_code_auto_generated=auto_generated,
        cost_center_id=data.cost_center_id,
        line_manager_id=data.line_manager_id,
        resumption_date=data.resumption_date,
        approval_role_id=data.approval_role_id,
    )
    db.add(emp)
    await db.flush()
    await db.refresh(emp)

    # Record initial code in history
    db.add(EmployeeCodeHistory(
        tenant_id=tenant_id,
        employee_id=emp.id,
        old_code=None,
        new_code=employee_code,
        change_type="progressive",
        effective_date=data.resumption_date or __import__("datetime").date.today(),
        changed_by=current_user.user_id,
    ))
    await db.flush()

    # Auto-provision portal account so the employee can log in (and SA can impersonate them)
    # without a separate manual invite step.
    full_name = f"{data.first_name} {data.last_name}".strip()
    await _ensure_portal_account(data.email, full_name, data.first_name, tenant_id, db, employee=emp)

    orm_obj = await _get_employee_or_404(emp.id, tenant_id, db)
    return EmployeeResponse.from_orm(orm_obj)


@router.patch("/employees/{employee_id}", response_model=EmployeeResponse)
async def update_employee(
    employee_id: uuid.UUID,
    data: EmployeeUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeResponse:
    """Update an employee (PATCH semantics). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    emp = await _get_employee_or_404(employee_id, tenant_id, db)

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(emp, field, value)

    await db.flush()
    orm_obj = await _get_employee_or_404(employee_id, tenant_id, db)
    return EmployeeResponse.from_orm(orm_obj)


async def _cascade_employee_deactivate(
    emp: Employee, tenant_id: uuid.UUID, db: AsyncSession
) -> None:
    """
    Deactivate the portal user account linked to an employee.
    Also revokes all active sessions for that user in this tenant.
    Safe to call even if emp.user_id is None.

    Email fallback: if emp.user_id is not set (employee created before the M9.3b
    user_id backfill ran, or never linked explicitly), we resolve the user by
    matching emp.email against users in this tenant.  This prevents deleted employees
    whose user_id column was NULL from silently keeping their portal account active.
    """
    from app.models.auth import UserTenant as UserTenantModel
    from app.models.auth import Session as SessionModel, User as UserModel

    user_id = emp.user_id
    if not user_id and emp.email:
        # Fallback: resolve user by email within this tenant
        uid_res = await db.execute(
            select(UserModel.id)
            .join(UserTenantModel, UserModel.id == UserTenantModel.user_id)
            .where(
                UserModel.email.ilike(emp.email),
                UserTenantModel.tenant_id == tenant_id,
            )
        )
        user_id = uid_res.scalar_one_or_none()
        if user_id:
            emp.user_id = user_id  # persist the resolved FK for future cascade calls

    if not user_id:
        return

    # Deactivate the UserTenant membership
    ut_res = await db.execute(
        select(UserTenantModel).where(
            UserTenantModel.user_id == user_id,
            UserTenantModel.tenant_id == tenant_id,
        )
    )
    ut = ut_res.scalar_one_or_none()
    if ut:
        ut.is_active = False
        # Revoke sessions. Session rows FK via user_tenant_id only —
        # there is no Session.user_id or Session.tenant_id column.
        _sqldelete = __import__("sqlalchemy", fromlist=["delete"]).delete
        await db.execute(
            _sqldelete(SessionModel).where(
                SessionModel.user_tenant_id == ut.id,
            )
        )


async def _cascade_employee_reactivate(
    emp: Employee, tenant_id: uuid.UUID, db: AsyncSession
) -> None:
    """
    Reactivate the portal user account linked to an employee.
    Safe to call even if emp.user_id is None.
    """
    if not emp.user_id:
        return
    from app.models.auth import UserTenant as UserTenantModel

    ut_res = await db.execute(
        select(UserTenantModel).where(
            UserTenantModel.user_id == emp.user_id,
            UserTenantModel.tenant_id == tenant_id,
        )
    )
    ut = ut_res.scalar_one_or_none()
    if ut and not ut.is_active:
        ut.is_active = True


async def _cascade_employee_hard_delete(
    emp: Employee, tenant_id: uuid.UUID, db: AsyncSession
) -> None:
    """
    On pre-go-live hard-delete: clean up the portal User and all their UserTenant
    rows IF the user has no activity in any LIVE tenant (test activity doesn't count).

    'Activity in a live tenant' means any UserTenant row whose tenant has
    environment='live'. If found, we only deactivate the current test-tenant
    UserTenant instead of deleting the User entirely.

    Email fallback: same as _cascade_employee_deactivate — if emp.user_id is not
    set we resolve by email so that employees whose user_id was never backfilled
    are still cleaned up correctly on hard-delete.
    """
    from app.models.auth import UserTenant as UserTenantModel, User as UserModel, Tenant

    user_id = emp.user_id
    if not user_id and emp.email:
        uid_res = await db.execute(
            select(UserModel.id)
            .join(UserTenantModel, UserModel.id == UserTenantModel.user_id)
            .where(
                UserModel.email == emp.email,
                UserTenantModel.tenant_id == tenant_id,
            )
        )
        user_id = uid_res.scalar_one_or_none()
        if user_id:
            emp.user_id = user_id

    if not user_id:
        return

    # Check if this user is a member of any LIVE tenant (other than the current one)
    live_check = await db.execute(
        select(UserTenantModel)
        .join(Tenant, Tenant.id == UserTenantModel.tenant_id)
        .where(
            UserTenantModel.user_id == user_id,
            UserTenantModel.tenant_id != tenant_id,
            Tenant.environment == "live",
        )
    )
    has_live_activity = live_check.scalar_one_or_none() is not None

    if has_live_activity:
        # User has live-tenant activity — only deactivate this test tenant's membership
        await _cascade_employee_deactivate(emp, tenant_id, db)
    else:
        # No live-tenant activity — hard-delete the User (cascades to UserTenant, sessions, etc.)
        user_res = await db.execute(select(UserModel).where(UserModel.id == user_id))
        user = user_res.scalar_one_or_none()
        if user:
            await db.delete(user)


@router.delete("/employees/{employee_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_employee(
    employee_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """
    Delete or deactivate an employee. Admin only.

    Go-live gate (BRIEF_coa_remap_golive_gate.md Part C):
    - Pre-go-live (lifecycle_status != 'live'): hard-delete the row. Safe because
      no operational postings exist yet; allows removing test/erroneous records.
    - Post-go-live (lifecycle_status == 'live'): deactivate only (is_active=False),
      regardless of whether the employee has any references anywhere. Never hard-delete
      a live-tenant employee, even if they have zero references.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    emp = await _get_employee_or_404(employee_id, tenant_id, db)

    # Check tenant lifecycle
    from sqlalchemy import select as sa_select
    from app.models.auth import Tenant
    tenant_res = await db.execute(sa_select(Tenant).where(Tenant.id == tenant_id))
    tenant_row = tenant_res.scalar_one_or_none()
    is_live = tenant_row is not None and tenant_row.lifecycle_status == "live"

    if is_live:
        # Post-go-live: soft-delete only — deactivate employee AND their portal account
        emp.is_active = False
        await _cascade_employee_deactivate(emp, tenant_id, db)
    else:
        # Pre-go-live: hard-delete (remove the row entirely)
        # Clean up portal user if they have no activity in any live tenant.
        await _cascade_employee_hard_delete(emp, tenant_id, db)
        # Also cascade-clean dependent rows that have no operational meaning yet
        from app.models.master_data import EmployeeCodeHistory, EmployeeTransfer
        _sqldelete = __import__("sqlalchemy", fromlist=["delete"]).delete
        await db.execute(_sqldelete(EmployeeCodeHistory).where(EmployeeCodeHistory.employee_id == emp.id))
        await db.execute(_sqldelete(EmployeeTransfer).where(EmployeeTransfer.employee_id == emp.id))
        await db.delete(emp)

    await db.flush()


@router.post("/employees/{employee_id}/transfer", response_model=TransferResponse, status_code=status.HTTP_201_CREATED)
async def transfer_employee(
    employee_id: uuid.UUID,
    data: TransferCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> TransferResponse:
    """Transfer an employee to a new cost center. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    emp = await _get_employee_or_404(employee_id, tenant_id, db)

    # Verify target cost center is an active org_structure node for this tenant
    cc_result = await db.execute(
        _cc_nodes_query(tenant_id).where(OrgStructureNode.id == data.to_cost_center_id)
    )
    if not cc_result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cost center not found.")

    transfer = EmployeeTransfer(
        tenant_id=tenant_id,
        employee_id=employee_id,
        from_cost_center_id=emp.cost_center_id,
        to_cost_center_id=data.to_cost_center_id,
        effective_date=data.effective_date,
        notes=data.notes,
        transferred_by=current_user.id,
    )
    db.add(transfer)

    emp.cost_center_id = data.to_cost_center_id
    await db.flush()
    await db.refresh(transfer)

    result = await db.execute(
        select(EmployeeTransfer)
        .where(EmployeeTransfer.id == transfer.id)
        .options(
            selectinload(EmployeeTransfer.from_cost_center),
            selectinload(EmployeeTransfer.to_cost_center),
        )
    )
    return TransferResponse.from_orm(result.scalar_one())


@router.post("/employees/{employee_id}/update-code", response_model=EmployeeResponse)
async def update_employee_code(
    employee_id: uuid.UUID,
    data: CodeUpdateRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeResponse:
    """Update employee code (retrospective or progressive). Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)
    emp = await _get_employee_or_404(employee_id, tenant_id, db)

    # Check uniqueness of new code
    existing = await db.execute(
        select(Employee).where(
            Employee.tenant_id == tenant_id,
            Employee.employee_code == data.new_code,
            Employee.id != employee_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Employee code '{data.new_code}' is already in use.",
        )

    db.add(EmployeeCodeHistory(
        tenant_id=tenant_id,
        employee_id=employee_id,
        old_code=emp.employee_code,
        new_code=data.new_code,
        change_type=data.change_type,
        effective_date=data.effective_date,
        changed_by=current_user.user_id,
        notes=data.notes,
    ))
    emp.employee_code = data.new_code
    emp.employee_code_auto_generated = False
    await db.flush()

    return await _get_employee_or_404(employee_id, tenant_id, db)  # type: ignore[return-value]


@router.get("/employees/{employee_id}/history", response_model=EmployeeHistoryResponse)
async def get_employee_history(
    employee_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeHistoryResponse:
    """Get the code change and transfer history for an employee."""
    tenant_id = _require_tenant(current_user)
    await _get_employee_or_404(employee_id, tenant_id, db)

    code_result = await db.execute(
        select(EmployeeCodeHistory)
        .where(EmployeeCodeHistory.employee_id == employee_id)
        .order_by(EmployeeCodeHistory.effective_date.desc(), EmployeeCodeHistory.changed_at.desc())
    )
    transfers_result = await db.execute(
        select(EmployeeTransfer)
        .where(EmployeeTransfer.employee_id == employee_id)
        .options(
            selectinload(EmployeeTransfer.from_cost_center),
            selectinload(EmployeeTransfer.to_cost_center),
        )
        .order_by(EmployeeTransfer.effective_date.desc())
    )

    return EmployeeHistoryResponse(
        code_history=[CodeHistoryResponse.from_orm(r) for r in code_result.scalars().all()],
        transfers=[TransferResponse.from_orm(t) for t in transfers_result.scalars().all()],
    )


@router.post("/employees/upload", response_model=EmployeeUploadResult)
async def upload_employees(
    file: UploadFile = File(...),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeUploadResult:
    """
    Bulk upload employees from .xlsx or .csv.

    Required columns: First Name, Last Name, Email, Org Role, Resumption Date.
    Optional: Other Name, Preferred Name, Employee Code, Phone, Cost Center Code.
    Duplicate email = update existing record.
    Capacity and cost-centre/role alignment are enforced on upload.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    content = await file.read()
    fname = (file.filename or "").lower()

    headers: list[str] = []
    rows: list[list[str]] = []

    if fname.endswith(".xlsx"):
        try:
            import openpyxl
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="openpyxl not installed.") from exc
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c).strip() if c is not None else "" for c in row])
        if all_rows:
            headers = all_rows[0]
            # Row 1 = header; data starts at row 2 (format standard — no inline instruction rows)
            rows = all_rows[1:]
    elif fname.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows_csv = list(reader)
        if all_rows_csv:
            headers = [h.strip() for h in all_rows_csv[0]]
            rows = [[c.strip() for c in r] for r in all_rows_csv[1:]]
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are supported.")

    h = [hdr.lower().strip("*").strip() for hdr in headers]

    def col(name: str) -> Optional[int]:
        try:
            return h.index(name)
        except ValueError:
            return None

    fn_col = col("first name")
    ln_col = col("last name")
    em_col = col("email")
    on_col = col("other name")
    pn_col = col("preferred name")
    code_col = col("employee code")
    ph_col = col("phone")
    cc_col = col("cost center code")
    res_col = col("resumption date (dd/mm/yyyy)") or col("resumption date")
    org_role_col = col("org role")

    if fn_col is None or ln_col is None or em_col is None:
        raise HTTPException(
            status_code=400,
            detail="File must have 'First Name', 'Last Name', and 'Email' columns.",
        )
    if org_role_col is None:
        raise HTTPException(
            status_code=400,
            detail="File must have an 'Org Role' column. Download the latest template.",
        )

    # Load cost center nodes from org_structure (single source of truth).
    cc_result = await db.execute(_cc_nodes_query(tenant_id))
    cc_by_code = {_cc_node_code(n).lower(): n for n in cc_result.scalars().all() if _cc_node_code(n)}

    # Load org roles for this tenant (name lookup, case-insensitive)
    from app.models.approvals import ApprovalRole as AR
    ar_result = await db.execute(
        select(AR).where(AR.tenant_id == tenant_id, AR.is_active.is_(True))
    )
    all_roles = ar_result.scalars().all()
    # Build display-name → role map (same disambiguation logic as the template)
    # Build a CC name map for the upload parser (id → name)
    cc_name_map_upload: dict[str, str] = {str(n.id): n.name for n in cc_by_code.values()}
    roles_by_name: dict[str, AR] = {}
    for r in all_roles:
        # 1. Plain name (backward compat — old templates still work)
        roles_by_name[r.name.lower().strip()] = r
        # 2. "Name — CC Name" (new format)
        _cc_id_str = str(r.cost_center_id) if r.cost_center_id else None
        _cc_nm = cc_name_map_upload.get(_cc_id_str, "") if _cc_id_str else ""
        if _cc_nm:
            roles_by_name[f"{r.name.lower().strip()} — {_cc_nm.lower()}"] = r
            # 3. "Name — CC Name [Area]" and "Name — CC Name [Area > Sub-area]"
            _parts: list[str] = []
            if r.area:
                _parts.append(r.area)
            if r.sub_area:
                _parts.append(r.sub_area)
            if _parts:
                _suffix = " > ".join(_parts)
                roles_by_name[f"{r.name.lower().strip()} — {_cc_nm.lower()} [{_suffix.lower()}]"] = r
        elif r.area or r.sub_area:
            # No CC but has area/sub-area — "Name [Area > Sub-area]"
            _parts2: list[str] = [x for x in [r.area, r.sub_area] if x]
            roles_by_name[f"{r.name.lower().strip()} [{(' > '.join(_parts2)).lower()}]"] = r

    # Pre-load current occupant counts per role (for capacity enforcement)
    from sqlalchemy import func as sqlfunc
    occ_res = await db.execute(
        select(Employee.approval_role_id, sqlfunc.count(Employee.id))
        .where(Employee.tenant_id == tenant_id, Employee.approval_role_id.isnot(None))
        .group_by(Employee.approval_role_id)
    )
    role_occupant_counts: dict[uuid.UUID, int] = {r: c for r, c in occ_res.all()}
    batch_role_additions: dict[uuid.UUID, int] = {}  # tracks new occupants added in this batch

    # Fetch registration date floor for resumption date validation.
    from app.models.setup import TenantOrgConfig
    org_res = await db.execute(
        select(TenantOrgConfig).where(TenantOrgConfig.tenant_id == tenant_id)
    )
    org_cfg = org_res.scalar_one_or_none()
    _reg_date = org_cfg.date_of_registration if org_cfg else None

    imported = 0
    updated = 0
    errors: list[dict] = []
    email_to_emp: dict[str, Employee] = {}

    from datetime import datetime as _dt3

    for i, row in enumerate(rows, start=4):
        if not any((c or "").strip() for c in row):
            continue

        def get(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        first_name = get(fn_col)
        last_name = get(ln_col)
        email = get(em_col)

        if not first_name:
            errors.append({"row": i, "reason": "Missing First Name."})
            continue
        if not last_name:
            errors.append({"row": i, "reason": "Missing Last Name."})
            continue
        if not email:
            errors.append({"row": i, "reason": "Missing Email."})
            continue

        other_name = get(on_col) or None
        preferred_name = get(pn_col) or None
        employee_code = get(code_col) or None
        phone = get(ph_col) or None
        cc_code = get(cc_col)
        res_str = get(res_col)
        org_role_name = get(org_role_col)

        # Org Role is mandatory
        if not org_role_name:
            errors.append({"row": i, "reason": "Org Role is required. Select a role from the dropdown."})
            continue

        # Resumption Date is mandatory
        if not res_str:
            errors.append({"row": i, "reason": "Resumption Date is required. Enter the date in dd/mm/yyyy format."})
            continue

        cost_center_id = None
        if cc_code:
            cc_node = cc_by_code.get(cc_code.lower())
            if not cc_node:
                errors.append({"row": i, "reason": f"Cost center code '{cc_code}' not found in organisation structure."})
            else:
                cost_center_id = cc_node.id

        resumption_date = None
        if res_str:
            # Normalise: Excel datetime cells come through as "2024-01-04 00:00:00"
            # (openpyxl -> str() adds the time component). Strip the time part so
            # "%Y-%m-%d" can match.  Also handle ISO "T" separator just in case.
            res_date_part = res_str.split(" ")[0].split("T")[0].strip()

            parsed = False
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                try:
                    resumption_date = _dt3.strptime(res_date_part, fmt).date()
                    parsed = True
                    break
                except ValueError:
                    continue

            if not parsed:
                errors.append({"row": i, "reason": f"Invalid Resumption Date: '{res_str}'. Expected dd/mm/yyyy or a real Excel date cell."})
            elif _reg_date and resumption_date < _reg_date:
                errors.append({"row": i, "reason": f"Resumption Date {resumption_date} is before the organisation's registration date {_reg_date}."})
                resumption_date = None  # don't save the floor-failing date

        # Upsert by email
        existing_result = await db.execute(
            select(Employee).where(
                Employee.tenant_id == tenant_id,
                Employee.email == email,
            )
        )
        emp_obj = existing_result.scalar_one_or_none()
        # Resolve org role (mandatory — already checked above)
        matched_role = roles_by_name.get(org_role_name.lower().strip())
        if not matched_role:
            errors.append({"row": i, "reason": f"Org Role '{org_role_name}' not found. Check spelling or create the role first."})
            continue
        org_role_id = matched_role.id

        # Capacity enforcement
        if matched_role.max_occupants is not None:
            current_count = role_occupant_counts.get(matched_role.id, 0)
            batch_count = batch_role_additions.get(matched_role.id, 0)
            is_already_in_role = emp_obj is not None and getattr(emp_obj, "approval_role_id", None) == matched_role.id
            if not is_already_in_role and (current_count + batch_count) >= matched_role.max_occupants:
                cap_label = "occupant" if matched_role.max_occupants == 1 else "occupants"
                errors.append({"row": i, "reason": (
                    f"Role '{matched_role.name}' is at capacity "
                    f"({current_count + batch_count}/{matched_role.max_occupants} {cap_label}). "
                    f"Increase the role's headcount limit or choose a different role."
                )})
                continue

        # Role vs Cost Center validation
        if matched_role.cost_center_id:
            if cost_center_id and cost_center_id != matched_role.cost_center_id:
                expected_cc_code = next(
                    (code.upper() for code, node in cc_by_code.items() if node.id == matched_role.cost_center_id),
                    "the role's assigned cost centre"
                )
                errors.append({"row": i, "reason": (
                    f"Cost Centre mismatch: role '{matched_role.name}' belongs to '{expected_cc_code}'. "
                    f"Remove the Cost Centre Code or enter '{expected_cc_code}'."
                )})
                continue
            if not cost_center_id:
                # Auto-assign cost centre from role
                cost_center_id = matched_role.cost_center_id

        # Track this batch addition for capacity (only if new to this role)
        is_already_in_role = emp_obj is not None and getattr(emp_obj, "approval_role_id", None) == matched_role.id
        if not is_already_in_role:
            batch_role_additions[matched_role.id] = batch_role_additions.get(matched_role.id, 0) + 1

        if emp_obj:
            emp_obj.first_name = first_name
            emp_obj.last_name = last_name
            emp_obj.other_name = other_name
            emp_obj.preferred_name = preferred_name
            emp_obj.phone = phone
            if cost_center_id:
                emp_obj.cost_center_id = cost_center_id
            if resumption_date:
                emp_obj.resumption_date = resumption_date
            if org_role_id is not None:
                emp_obj.approval_role_id = org_role_id
            elif org_role_name == "":
                pass  # blank = leave existing role unchanged
            updated += 1
        else:
            auto_generated = False
            if not employee_code:
                employee_code = await _next_employee_code(tenant_id, db)
                auto_generated = True
            emp_obj = Employee(
                tenant_id=tenant_id,
                first_name=first_name,
                last_name=last_name,
                other_name=other_name,
                preferred_name=preferred_name,
                email=email,
                phone=phone,
                employee_code=employee_code,
                employee_code_auto_generated=auto_generated,
                cost_center_id=cost_center_id,
                resumption_date=resumption_date,
                approval_role_id=org_role_id,
            )
            db.add(emp_obj)
            imported += 1

        await db.flush()
        await db.refresh(emp_obj)
        email_to_emp[email.lower()] = emp_obj

        # Auto-provision portal account for both new and updated employees.
        full_name_emp = f"{emp_obj.first_name} {emp_obj.last_name}".strip()
        await _ensure_portal_account(email, full_name_emp, emp_obj.first_name, tenant_id, db, employee=emp_obj)

    await db.flush()
    skipped = max(0, len([r for r in rows if any((c or "").strip() for c in r)]) - imported - updated)
    return EmployeeUploadResult(
        imported=imported, updated=updated, skipped=skipped,
        errors=errors, head_assignments=0,
    )


@router.post("/employees/sync-portal-accounts")
async def sync_employee_portal_accounts(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Retroactively ensure every active employee in this tenant has a Ziva portal
    account (User + UserTenant). Idempotent — safe to call multiple times.

    Used to backfill accounts for employees that were imported before the
    auto-provision logic was added. After running this, the Impersonate button
    on the Employees page will be enabled for all employees (greyed out only for
    employees who still have no email address on record).

    Admin only.
    """
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(Employee).where(
            Employee.tenant_id == tenant_id,
            Employee.is_active.is_(True),
            Employee.email.isnot(None),
        )
    )
    employees = result.scalars().all()

    for emp in employees:
        full_name = f"{emp.first_name} {emp.last_name}".strip()
        await _ensure_portal_account(emp.email, full_name, emp.first_name, tenant_id, db, employee=emp)

    await db.flush()

    # Backfill head_user_id on cost_center_config rows where it is still null.
    # These are assignments made before the auto-resolve logic was added.
    # We resolve by matching head_employee.email → users.id.
    from app.models.auth import User as _UserModel
    cc_result = await db.execute(
        select(CostCenterConfig)
        .options(selectinload(CostCenterConfig.head_employee))
        .where(
            CostCenterConfig.tenant_id == tenant_id,
            CostCenterConfig.head_employee_id.isnot(None),
            CostCenterConfig.head_user_id.is_(None),
        )
    )
    cc_rows = cc_result.scalars().all()
    head_emails = [cc.head_employee.email for cc in cc_rows if cc.head_employee and cc.head_employee.email]
    if head_emails:
        uid_result = await db.execute(
            select(_UserModel.id, _UserModel.email).where(_UserModel.email.in_(head_emails))
        )
        email_to_uid = {row.email.lower(): row.id for row in uid_result}
        for cc in cc_rows:
            if cc.head_employee and cc.head_employee.email:
                uid = email_to_uid.get(cc.head_employee.email.lower())
                if uid:
                    cc.head_user_id = uid

    await db.flush()
    return {"synced": len(employees), "cc_heads_resolved": len(cc_rows)}


# ── Cost-center dimension helper ──────────────────────────────────────────────

def _cc_nodes_query(tenant_id: uuid.UUID):
    """
    Return a SELECT that yields OrgStructureNode rows that are cost centers for the tenant.

    org_structure (node_type='Cost center', is_active=True) is the single source
    of truth for cost centers — not dimension_values. This query is used by the
    options endpoint, template CC dropdown, upload validator, and transfer/head guards.

    The 'code' field is the business key used for matching in bulk uploads and dropdowns.
    'cost_center_code' is preferred when non-null (same value in practice); falls back to 'code'.
    """
    return (
        select(OrgStructureNode)
        .where(
            OrgStructureNode.tenant_id == tenant_id,
            OrgStructureNode.node_type == "Cost center",
            OrgStructureNode.is_active.is_(True),
        )
        .order_by(OrgStructureNode.sort_order)
    )


def _cc_node_code(node: OrgStructureNode) -> str:
    """Return the canonical code for a cost center node (cost_center_code preferred)."""
    return (node.cost_center_code or node.code or "").strip()


# ═══════════════════════════════════════════════════════════════════════════════
# COST CENTER CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/cost-centers/options", response_model=list[CostCenterOption])
async def list_cost_center_options(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[CostCenterOption]:
    """
    Return the tenant's cost centers as lightweight dropdown options.

    Source of truth: org_structure nodes with node_type='Cost center' and is_active=True.
    Same URL and response shape ({id, code, name}) as before — frontend consumers unchanged.
    'id' is now an org_structure.id UUID, not a dimension_values UUID.
    """
    tenant_id = _require_tenant(current_user)
    result = await db.execute(_cc_nodes_query(tenant_id))
    return [
        CostCenterOption(id=str(n.id), code=_cc_node_code(n), name=n.name)
        for n in result.scalars().all()
        if _cc_node_code(n)
    ]


@router.get("/cost-centers", response_model=list[CostCenterConfigResponse])
async def list_cost_center_configs(
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[CostCenterConfigResponse]:
    """List all cost center head assignments for the tenant.

    head_user_id is resolved dynamically from the head employee's email → users.id
    so the Impersonate button works even for assignments made before the auto-resolve
    logic was added (i.e. where head_user_id is still null in the DB).
    """
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(CostCenterConfig)
        .where(CostCenterConfig.tenant_id == tenant_id)
        .options(
            selectinload(CostCenterConfig.cost_center),
            selectinload(CostCenterConfig.head_employee),
            selectinload(CostCenterConfig.head_user),
        )
        .order_by(CostCenterConfig.created_at)
    )
    configs = result.scalars().all()

    # Batch-resolve head_user_id from employee email → users.id for rows where
    # the stored column is still null (set before auto-resolve was introduced).
    # Wrapped in try/except so any error here degrades gracefully (impersonate
    # buttons just won't activate) rather than crashing the entire endpoint.
    try:
        from app.models.auth import User as _UserModel
        missing = [c for c in configs if c.head_employee_id and not c.head_user_id]
        if missing:
            emails = [c.head_employee.email for c in missing if c.head_employee and c.head_employee.email]
            if emails:
                uid_rows = await db.execute(
                    select(_UserModel.id, _UserModel.email).where(_UserModel.email.in_(emails))
                )
                email_to_uid = {str(row[1]).lower(): str(row[0]) for row in uid_rows.all()}
                for c in missing:
                    if c.head_employee and c.head_employee.email:
                        resolved = email_to_uid.get(c.head_employee.email.lower())
                        if resolved:
                            import uuid as _uuid
                            c.head_user_id = _uuid.UUID(resolved)
    except Exception:
        pass  # Degrade gracefully — impersonate buttons just won't be active

    return [CostCenterConfigResponse.from_orm(c) for c in configs]


@router.put("/cost-centers/{cost_center_id}/head", response_model=CostCenterConfigResponse)
async def set_cost_center_head(
    cost_center_id: uuid.UUID,
    data: CostCenterHeadUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> CostCenterConfigResponse:
    """Set or update the head of a cost center. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    # Verify cost center is an active org_structure node for this tenant
    cc_result = await db.execute(
        _cc_nodes_query(tenant_id).where(OrgStructureNode.id == cost_center_id)
    )
    if not cc_result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cost center not found.")

    # Resolve head_user_id from the employee's email so the impersonate button
    # on the frontend can find the portal user for this cost center head.
    # The frontend only sends head_employee_id; we do the email→user_id lookup here.
    resolved_head_user_id: uuid.UUID | None = None
    if data.head_employee_id:
        from app.models.master_data import Employee as EmployeeModel
        from app.models.auth import User as UserModel
        emp_row = await db.execute(
            select(EmployeeModel.email).where(
                EmployeeModel.id == data.head_employee_id,
                EmployeeModel.tenant_id == tenant_id,
            )
        )
        emp_email = emp_row.scalar_one_or_none()
        if emp_email:
            user_row = await db.execute(
                select(UserModel.id).where(UserModel.email == emp_email)
            )
            resolved_head_user_id = user_row.scalar_one_or_none()

    cfg_result = await db.execute(
        select(CostCenterConfig).where(
            CostCenterConfig.cost_center_id == cost_center_id,
            CostCenterConfig.tenant_id == tenant_id,
        )
    )
    cfg = cfg_result.scalar_one_or_none()
    if cfg:
        cfg.head_employee_id = data.head_employee_id
        cfg.head_user_id = resolved_head_user_id
    else:
        cfg = CostCenterConfig(
            tenant_id=tenant_id,
            cost_center_id=cost_center_id,
            head_employee_id=data.head_employee_id,
            head_user_id=resolved_head_user_id,
        )
        db.add(cfg)

    await db.flush()
    await db.refresh(cfg)

    result2 = await db.execute(
        select(CostCenterConfig)
        .where(CostCenterConfig.id == cfg.id)
        .options(
            selectinload(CostCenterConfig.cost_center),
            selectinload(CostCenterConfig.head_employee),
            selectinload(CostCenterConfig.head_user),
        )
    )
    return CostCenterConfigResponse.from_orm(result2.scalar_one())


# ═══════════════════════════════════════════════════════════════════════════════
# FINANCE REVIEW CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/finance-review", response_model=list[FinanceReviewConfigResponse])
async def list_finance_reviewers(
    module: str = Query(default="expense_retirement"),
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[FinanceReviewConfigResponse]:
    """List finance reviewers for a module, ordered by review level."""
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(FinanceReviewConfig)
        .where(
            FinanceReviewConfig.tenant_id == tenant_id,
            FinanceReviewConfig.module == module,
        )
        .options(
            selectinload(FinanceReviewConfig.reviewer),
            selectinload(FinanceReviewConfig.cost_center),
        )
        .order_by(FinanceReviewConfig.review_level, FinanceReviewConfig.created_at)
    )
    return [FinanceReviewConfigResponse.from_orm(r) for r in result.scalars().all()]


@router.post("/finance-review", response_model=FinanceReviewConfigResponse, status_code=status.HTTP_201_CREATED)
async def add_finance_reviewer(
    data: FinanceReviewerCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FinanceReviewConfigResponse:
    """Add a finance reviewer. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    cfg = FinanceReviewConfig(
        tenant_id=tenant_id,
        module=data.module,
        reviewer_user_id=data.reviewer_user_id,
        review_level=data.review_level,
        cost_center_id=data.cost_center_id,
    )
    db.add(cfg)
    await db.flush()
    await db.refresh(cfg)

    result = await db.execute(
        select(FinanceReviewConfig)
        .where(FinanceReviewConfig.id == cfg.id)
        .options(
            selectinload(FinanceReviewConfig.reviewer),
            selectinload(FinanceReviewConfig.cost_center),
        )
    )
    return FinanceReviewConfigResponse.from_orm(result.scalar_one())


@router.patch("/finance-review/{reviewer_id}", response_model=FinanceReviewConfigResponse)
async def update_finance_reviewer(
    reviewer_id: uuid.UUID,
    data: FinanceReviewerUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> FinanceReviewConfigResponse:
    """Update a reviewer's level or scope. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(FinanceReviewConfig).where(
            FinanceReviewConfig.id == reviewer_id,
            FinanceReviewConfig.tenant_id == tenant_id,
        )
    )
    cfg = result.scalar_one_or_none()
    if not cfg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reviewer config not found.")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(cfg, field, value)
    await db.flush()

    result2 = await db.execute(
        select(FinanceReviewConfig)
        .where(FinanceReviewConfig.id == reviewer_id)
        .options(
            selectinload(FinanceReviewConfig.reviewer),
            selectinload(FinanceReviewConfig.cost_center),
        )
    )
    return FinanceReviewConfigResponse.from_orm(result2.scalar_one())


@router.delete("/finance-review/{reviewer_id}", status_code=status.HTTP_204_NO_CONTENT)
async def remove_finance_reviewer(
    reviewer_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> None:
    """Remove a finance reviewer. Admin only."""
    tenant_id = _require_tenant(current_user)
    _require_admin(current_user)

    result = await db.execute(
        select(FinanceReviewConfig).where(
            FinanceReviewConfig.id == reviewer_id,
            FinanceReviewConfig.tenant_id == tenant_id,
        )
    )
    cfg = result.scalar_one_or_none()
    if not cfg:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reviewer config not found.")

    await db.delete(cfg)
    await db.flush()


# ── Employee Self-onboarding (M8.2) ──────────────────────────────────────────

@router.post("/employees/invite", status_code=201)
async def send_employee_invite(
    data: EmployeeInviteCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    HR creates a basic employee record and sends a self-onboarding invite.

    Creates employee with status='pending_self_onboarding', generates a
    30-day secure token, and logs the onboarding link (email TBD).
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    emp = Employee(
        tenant_id=tenant_id,
        first_name=data.first_name,
        last_name=data.last_name,
        email=data.email,
        cost_center_id=data.cost_center_id,
        resumption_date=data.start_date,
        is_active=False,
    )
    db.add(emp)
    await db.flush()

    token_value = secrets.token_urlsafe(48)
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)

    token = EmployeeOnboardingToken(
        tenant_id=tenant_id,
        employee_id=emp.id,
        token=token_value,
        expires_at=expires_at,
    )
    db.add(token)
    await db.commit()

    # Log invite link at debug level — sensitive, should not appear in production logs
    onboarding_link = f"/onboard/{token_value}"
    logger.debug("[ONBOARDING] Invite created for %s (link suppressed at non-debug level)", data.email)

    return {
        "message": "Invite created successfully.",
        "employee_id": str(emp.id),
        "onboarding_link": onboarding_link,
    }


@router.post("/employees/{employee_id}/approve-onboarding", status_code=200)
async def approve_employee_onboarding(
    employee_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """HR approves a self-onboarding submission. Activates the employee from their start date."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(Employee).where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    emp.is_active = True
    await _cascade_employee_reactivate(emp, tenant_id, db)
    await db.commit()
    return {"message": "Employee onboarding approved. Account is now active."}


@router.post("/employees/{employee_id}/reject-onboarding", status_code=200)
async def reject_employee_onboarding(
    employee_id: uuid.UUID,
    comment: str,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """HR rejects self-onboarding with a comment. Employee stays in pending state."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    result = await db.execute(
        select(Employee).where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
    )
    emp = result.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    logger.info("[ONBOARDING] HR rejected onboarding for employee %s", employee_id)
    return {"message": "Onboarding rejected.", "comment": comment}


# ── Positions API (People v1) ─────────────────────────────────────────────────

async def _resolve_position(role: OrgRole, db: AsyncSession) -> PositionResponse:
    """
    Build a PositionResponse from an ApprovalRole ORM object.
    approval_roles is the single source of truth — positions are a view of it.
    """
    cc_name = cc_code = None
    if role.cost_center_id:
        cc_res = await db.execute(
            select(OrgStructureNode).where(OrgStructureNode.id == role.cost_center_id)
        )
        cc = cc_res.scalar_one_or_none()
        if cc:
            cc_name, cc_code = cc.name, getattr(cc, "code", None)

    parent_name = None
    if role.parent_role_id:
        par_res = await db.execute(select(OrgRole).where(OrgRole.id == role.parent_role_id))
        par = par_res.scalar_one_or_none()
        if par:
            parent_name = par.name

    # Current active assignments (effective_to IS NULL)
    occ_res = await db.execute(
        select(EmployeePositionAssignment, Employee)
        .join(Employee, Employee.id == EmployeePositionAssignment.employee_id)
        .where(
            EmployeePositionAssignment.approval_role_id == role.id,
            EmployeePositionAssignment.effective_to.is_(None),
        )
        .order_by(EmployeePositionAssignment.effective_from)
    )
    occupants = [
        PositionOccupant(
            employee_id=str(emp.id),
            employee_code=emp.employee_code,
            full_name=f"{emp.first_name} {emp.last_name}",
            email=emp.email,
            assignment_type=asgn.assignment_type,
            effective_from=asgn.effective_from,
        )
        for asgn, emp in occ_res.all()
    ]

    return PositionResponse(
        id=str(role.id),
        name=role.name,
        code=role.code,
        grade=role.grade,
        description=role.description,
        display_order=role.display_order,
        is_active=role.is_active,
        parent_role_id=str(role.parent_role_id) if role.parent_role_id else None,
        parent_role_name=parent_name,
        cost_center_id=str(role.cost_center_id) if role.cost_center_id else None,
        cost_center_name=cc_name,
        cost_center_code=cc_code,
        entity_node_id=str(role.entity_node_id) if role.entity_node_id else None,
        max_occupants=role.max_occupants,
        designation=role.designation,
        area=role.area,
        sub_area=role.sub_area,
        employment_type=role.employment_type,
        occupant_count=len(occupants),
        occupants=occupants,
        created_at=role.created_at,
    )


@router.get("/positions", response_model=list[PositionResponse])
async def list_positions(
    cost_center_id: Optional[uuid.UUID] = None,
    is_active: Optional[bool] = None,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[PositionResponse]:
    """
    List all positions (approval_roles) for the tenant.
    Positions and the role hierarchy are the same data — single source of truth.
    """
    tenant_id = _require_tenant(current_user)

    q = select(OrgRole).where(OrgRole.tenant_id == tenant_id)
    if cost_center_id:
        q = q.where(OrgRole.cost_center_id == cost_center_id)
    if is_active is not None:
        q = q.where(OrgRole.is_active.is_(is_active))
    q = q.order_by(OrgRole.display_order, OrgRole.name)

    res = await db.execute(q)
    roles = res.scalars().all()
    return [await _resolve_position(r, db) for r in roles]


@router.post("/positions", response_model=PositionResponse, status_code=201)
async def create_position(
    payload: PositionCreate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PositionResponse:
    """
    Create a new position slot (writes to approval_roles).
    Immediately visible in both the Positions page and the Role Hierarchy page.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    role = OrgRole(
        tenant_id=tenant_id,
        name=payload.name,
        code=payload.code,
        grade=payload.grade,
        description=payload.description,
        display_order=payload.display_order,
        parent_role_id=payload.parent_role_id,
        cost_center_id=payload.cost_center_id,
        entity_node_id=payload.entity_node_id,
        max_occupants=payload.max_occupants,
        designation=payload.designation,
        area=payload.area,
        sub_area=payload.sub_area,
        employment_type=payload.employment_type or "permanent",
    )
    db.add(role)
    await db.commit()
    await db.refresh(role)
    return await _resolve_position(role, db)


@router.get("/positions/{position_id}", response_model=PositionResponse)
async def get_position(
    position_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PositionResponse:
    """Get a single position by ID, including current occupants."""
    tenant_id = _require_tenant(current_user)
    res = await db.execute(
        select(OrgRole).where(OrgRole.id == position_id, OrgRole.tenant_id == tenant_id)
    )
    role = res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Position not found.")
    return await _resolve_position(role, db)


@router.patch("/positions/{position_id}", response_model=PositionResponse)
async def update_position(
    position_id: uuid.UUID,
    payload: PositionUpdate,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PositionResponse:
    """Update position / role metadata. Reflected immediately on both pages."""
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    res = await db.execute(
        select(OrgRole).where(OrgRole.id == position_id, OrgRole.tenant_id == tenant_id)
    )
    role = res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Position not found.")

    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(role, field, value)

    await db.commit()
    await db.refresh(role)
    return await _resolve_position(role, db)


@router.delete("/positions/{position_id}", status_code=200)
async def archive_position(
    position_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Deactivate a position (sets is_active=False on the approval_role).
    Blocked if there are active occupants.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    res = await db.execute(
        select(OrgRole).where(OrgRole.id == position_id, OrgRole.tenant_id == tenant_id)
    )
    role = res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Position not found.")

    # Block if active occupants
    occ_res = await db.execute(
        select(EmployeePositionAssignment).where(
            EmployeePositionAssignment.approval_role_id == position_id,
            EmployeePositionAssignment.effective_to.is_(None),
        ).limit(1)
    )
    if occ_res.scalar_one_or_none():
        raise HTTPException(
            status_code=409,
            detail="Cannot deactivate a position with active occupants. Transfer or end their assignments first.",
        )

    role.is_active = False
    await db.commit()
    return {"message": f"Position '{role.name}' deactivated."}


@router.post("/positions/{position_id}/move", response_model=PositionResponse)
async def move_position(
    position_id: uuid.UUID,
    payload: PositionMoveRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> PositionResponse:
    """
    Move a position to a new parent or cost centre.
    Updates approval_role.parent_role_id and/or cost_center_id.
    Syncs cost_center_id on current substantive occupant's Employee row.
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    res = await db.execute(
        select(OrgRole).where(OrgRole.id == position_id, OrgRole.tenant_id == tenant_id)
    )
    role = res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Position not found.")

    if payload.new_cost_center_id is not None:
        role.cost_center_id = payload.new_cost_center_id
    if payload.new_parent_role_id is not None:
        role.parent_role_id = payload.new_parent_role_id
    if payload.new_name is not None:
        role.name = payload.new_name

    # Sync cost_center_id on current occupant's Employee row (denorm)
    if payload.new_cost_center_id:
        occ_res = await db.execute(
            select(EmployeePositionAssignment).where(
                EmployeePositionAssignment.approval_role_id == position_id,
                EmployeePositionAssignment.effective_to.is_(None),
                EmployeePositionAssignment.assignment_type == "substantive",
            )
        )
        for asgn in occ_res.scalars().all():
            emp_res = await db.execute(select(Employee).where(Employee.id == asgn.employee_id))
            emp = emp_res.scalar_one_or_none()
            if emp:
                emp.cost_center_id = payload.new_cost_center_id

    await db.commit()
    await db.refresh(role)
    return await _resolve_position(role, db)


@router.get("/positions/{position_id}/history")
async def get_position_history(
    position_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list:
    """
    Position history endpoint — returns empty list.
    History tracking for approval_roles will be added via the audit log in a future milestone.
    """
    tenant_id = _require_tenant(current_user)
    res = await db.execute(
        select(OrgRole.id).where(OrgRole.id == position_id, OrgRole.tenant_id == tenant_id)
    )
    if not res.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Position not found.")
    return []


# ── Employee Position Assignment Endpoints ────────────────────────────────────

@router.post("/employees/{employee_id}/assign", response_model=EmployeeAssignmentResponse, status_code=201)
async def assign_employee_to_position(
    employee_id: uuid.UUID,
    payload: EmployeeAssignRequest,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeAssignmentResponse:
    """
    Assign an employee to a position (hire, transfer, promotion, acting, secondment).

    For substantive assignments:
      - Closes any existing active substantive assignment
      - Updates employee.cost_center_id and employee.approval_role_id (denorm sync)

    For acting / secondment:
      - Does NOT close the primary substantive assignment
    """
    _require_admin(current_user)
    tenant_id = _require_tenant(current_user)

    from datetime import timedelta

    # Validate employee
    emp_res = await db.execute(
        select(Employee).where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
    )
    emp = emp_res.scalar_one_or_none()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    # Validate role / position
    role_res = await db.execute(
        select(OrgRole).where(OrgRole.id == payload.approval_role_id, OrgRole.tenant_id == tenant_id)
    )
    role = role_res.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Position not found.")
    if not role.is_active:
        raise HTTPException(status_code=409, detail="Cannot assign to an inactive position.")

    # Check capacity
    if role.max_occupants is not None:
        occ_count_res = await db.execute(
            select(EmployeePositionAssignment).where(
                EmployeePositionAssignment.approval_role_id == payload.approval_role_id,
                EmployeePositionAssignment.effective_to.is_(None),
            )
        )
        active_count = len(occ_count_res.scalars().all())
        if active_count >= role.max_occupants:
            raise HTTPException(
                status_code=409,
                detail=f"Position is at capacity ({role.max_occupants} occupant(s)). "
                       "Increase max_occupants or end an existing assignment first.",
            )

    # Close existing substantive assignment
    if payload.assignment_type == "substantive":
        existing_res = await db.execute(
            select(EmployeePositionAssignment).where(
                EmployeePositionAssignment.employee_id == employee_id,
                EmployeePositionAssignment.assignment_type == "substantive",
                EmployeePositionAssignment.effective_to.is_(None),
            )
        )
        existing = existing_res.scalar_one_or_none()
        if existing:
            existing.effective_to = payload.effective_from - timedelta(days=1)

    # Create new assignment
    asgn = EmployeePositionAssignment(
        tenant_id=tenant_id,
        employee_id=employee_id,
        approval_role_id=payload.approval_role_id,
        effective_from=payload.effective_from,
        effective_to=None,
        assignment_type=payload.assignment_type,
        transfer_reason=payload.transfer_reason,
        is_retrospective=payload.is_retrospective,
        notes=payload.notes,
        created_by=current_user.user_id,
    )
    db.add(asgn)

    # Sync denorm fields on Employee
    if payload.assignment_type == "substantive":
        emp.cost_center_id = role.cost_center_id
        emp.approval_role_id = role.id

    await db.commit()
    await db.refresh(asgn)

    cc_name = None
    if role.cost_center_id:
        cc_res = await db.execute(
            select(OrgStructureNode).where(OrgStructureNode.id == role.cost_center_id)
        )
        cc = cc_res.scalar_one_or_none()
        cc_name = cc.name if cc else None

    return EmployeeAssignmentResponse(
        id=str(asgn.id),
        employee_id=str(asgn.employee_id),
        approval_role_id=str(asgn.approval_role_id) if asgn.approval_role_id else None,
        role_name=role.name,
        cost_center_id=str(role.cost_center_id) if role.cost_center_id else None,
        cost_center_name=cc_name,
        effective_from=asgn.effective_from,
        effective_to=asgn.effective_to,
        assignment_type=asgn.assignment_type,
        transfer_reason=asgn.transfer_reason,
        is_retrospective=asgn.is_retrospective,
        notes=asgn.notes,
        created_at=asgn.created_at,
    )


@router.get("/employees/{employee_id}/assignments", response_model=list[EmployeeAssignmentResponse])
async def get_employee_assignments(
    employee_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> list[EmployeeAssignmentResponse]:
    """Return the full position assignment history for an employee (newest first)."""
    tenant_id = _require_tenant(current_user)

    emp_res = await db.execute(
        select(Employee.id).where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
    )
    if not emp_res.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Employee not found.")

    asgn_res = await db.execute(
        select(EmployeePositionAssignment, OrgRole)
        .join(OrgRole, OrgRole.id == EmployeePositionAssignment.approval_role_id, isouter=True)
        .where(
            EmployeePositionAssignment.employee_id == employee_id,
            EmployeePositionAssignment.tenant_id == tenant_id,
        )
        .order_by(EmployeePositionAssignment.effective_from.desc())
    )

    result = []
    for asgn, role in asgn_res.all():
        cc_name = None
        if role and role.cost_center_id:
            cc_res = await db.execute(
                select(OrgStructureNode).where(OrgStructureNode.id == role.cost_center_id)
            )
            cc = cc_res.scalar_one_or_none()
            cc_name = cc.name if cc else None
        result.append(EmployeeAssignmentResponse(
            id=str(asgn.id),
            employee_id=str(asgn.employee_id),
            approval_role_id=str(asgn.approval_role_id) if asgn.approval_role_id else None,
            role_name=role.name if role else None,
            cost_center_id=str(role.cost_center_id) if (role and role.cost_center_id) else None,
            cost_center_name=cc_name,
            effective_from=asgn.effective_from,
            effective_to=asgn.effective_to,
            assignment_type=asgn.assignment_type,
            transfer_reason=asgn.transfer_reason,
            is_retrospective=asgn.is_retrospective,
            notes=asgn.notes,
            created_at=asgn.created_at,
        ))
    return result


@router.get("/employees/{employee_id}/position", response_model=EmployeeAssignmentResponse | None)
async def get_employee_current_position(
    employee_id: uuid.UUID,
    current_user: CurrentUser = Depends(require_auth),
    db: AsyncSession = Depends(get_db),
) -> EmployeeAssignmentResponse | None:
    """Return the current substantive position assignment for an employee, or null."""
    tenant_id = _require_tenant(current_user)

    asgn_res = await db.execute(
        select(EmployeePositionAssignment, OrgRole)
        .join(OrgRole, OrgRole.id == EmployeePositionAssignment.approval_role_id, isouter=True)
        .where(
            EmployeePositionAssignment.employee_id == employee_id,
            EmployeePositionAssignment.tenant_id == tenant_id,
            EmployeePositionAssignment.assignment_type == "substantive",
            EmployeePositionAssignment.effective_to.is_(None),
        )
        .limit(1)
    )
    row = asgn_res.first()
    if not row:
        return None

    asgn, role = row
    cc_name = None
    if role and role.cost_center_id:
        cc_res = await db.execute(
            select(OrgStructureNode).where(OrgStructureNode.id == role.cost_center_id)
        )
        cc = cc_res.scalar_one_or_none()
        cc_name = cc.name if cc else None

    return EmployeeAssignmentResponse(
        id=str(asgn.id),
        employee_id=str(asgn.employee_id),
        approval_role_id=str(asgn.approval_role_id) if asgn.approval_role_id else None,
        role_name=role.name if role else None,
        cost_center_id=str(role.cost_center_id) if (role and role.cost_center_id) else None,
        cost_center_name=cc_name,
        effective_from=asgn.effective_from,
        effective_to=asgn.effective_to,
        assignment_type=asgn.assignment_type,
        transfer_reason=asgn.transfer_reason,
        is_retrospective=asgn.is_retrospective,
        notes=asgn.notes,
        created_at=asgn.created_at,
    )
