"""
ZivaBI — Posting Batches router.

Available when tenant_org_config.posting_mode = 'connected'.
Provides list, detail, export (CSV/Excel), and mark-exported endpoints.

In Connected Mode, approved transactions (expenses, AP invoices etc.) are
serialised to posting_batches instead of journal_entries. The finance team
downloads a batch file and imports it into their external ERP.

Routes:
    GET  /api/posting-batches               — list batches (paginated, filterable)
    GET  /api/posting-batches/{id}          — single batch detail
    GET  /api/posting-batches/{id}/export   — download as CSV or Excel
    POST /api/posting-batches/{id}/mark-exported — mark as exported

Auth: require_auth. Posting-mode guard returns 400 for non-connected tenants.
"""

import csv
import io
import logging
from datetime import datetime
from typing import Annotated, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.middleware.auth import require_auth
from app.models.auth import UserTenant
from app.models.gl import PostingBatch
from app.models.setup import TenantOrgConfig
from app.schemas.posting import PostingBatchDetail, PostingBatchSummary, PostingTransaction

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/posting-batches", tags=["Posting Batches"])


# ── Guard helper ──────────────────────────────────────────────────────────────

async def _require_connected_mode(db: AsyncSession, tenant_id: UUID) -> None:
    """
    Raise 400 if the tenant is not in Connected posting mode.

    This guard prevents Full ERP / Lite tenants from accidentally using batch
    export endpoints that have no meaning in their context.
    """
    result = await db.execute(
        select(TenantOrgConfig.posting_mode).where(
            TenantOrgConfig.tenant_id == tenant_id
        )
    )
    mode = result.scalar_one_or_none() or "full_erp"
    if mode != "connected":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Posting batches are only available in Connected mode. "
                f"This tenant is configured for '{mode}' mode."
            ),
        )


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("", response_model=list[PostingBatchSummary])
async def list_posting_batches(
    current_user: Annotated[UserTenant, Depends(require_auth)],
    db: AsyncSession = Depends(get_db),
    batch_status: Optional[str] = Query(None, alias="status"),
    module: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
) -> list[PostingBatchSummary]:
    """
    List posting batches for the current tenant.

    Filterable by status (pending/exported/synced) and module.
    Ordered newest first.

    Parameters:
        status    — filter by batch status
        module    — filter by source module (expense, ap, ar …)
        page      — 1-based page number
        page_size — rows per page (max 200)
    """
    tenant_id = current_user.tenant_id
    await _require_connected_mode(db, tenant_id)

    q = (
        select(PostingBatch)
        .where(PostingBatch.tenant_id == tenant_id)
        .order_by(PostingBatch.created_at.desc())
    )
    if batch_status:
        q = q.where(PostingBatch.status == batch_status)
    if module:
        q = q.where(PostingBatch.module == module)

    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    batches = result.scalars().all()

    return [
        PostingBatchSummary(
            id=b.id,
            tenant_id=b.tenant_id,
            batch_ref=b.batch_ref,
            module=b.module,
            status=b.status,
            transaction_count=len(b.transactions) if b.transactions else 0,
            created_at=b.created_at,
            exported_at=b.exported_at,
            synced_at=b.synced_at,
        )
        for b in batches
    ]


@router.get("/{batch_id}", response_model=PostingBatchDetail)
async def get_posting_batch(
    batch_id: UUID,
    current_user: Annotated[UserTenant, Depends(require_auth)],
    db: AsyncSession = Depends(get_db),
) -> PostingBatchDetail:
    """
    Get full detail for a single posting batch, including all transactions JSONB.

    Parameters:
        batch_id — UUID of the posting batch
    """
    tenant_id = current_user.tenant_id
    await _require_connected_mode(db, tenant_id)

    result = await db.execute(
        select(PostingBatch).where(
            PostingBatch.id == batch_id,
            PostingBatch.tenant_id == tenant_id,
        )
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Posting batch not found.")

    return PostingBatchDetail(
        id=batch.id,
        tenant_id=batch.tenant_id,
        batch_ref=batch.batch_ref,
        module=batch.module,
        status=batch.status,
        transactions=[PostingTransaction(**t) for t in (batch.transactions or [])],
        created_at=batch.created_at,
        exported_at=batch.exported_at,
        synced_at=batch.synced_at,
    )


@router.get("/{batch_id}/export")
async def export_posting_batch(
    batch_id: UUID,
    current_user: Annotated[UserTenant, Depends(require_auth)],
    db: AsyncSession = Depends(get_db),
    file_format: str = Query("excel", alias="format", pattern="^(csv|excel)$"),
) -> Response:
    """
    Download a posting batch as a formatted CSV or Excel file.

    The file contains one row per journal line across all transactions in the batch.
    Columns: Batch Ref, Entry Date, Source Module, Source Reference, Description,
             GL Code, GL Name, Debit, Credit, Dimensions.

    This format is designed for direct import into most external ERP systems.
    Auto-marks batch as exported (status='exported') on download.

    Parameters:
        batch_id — UUID of the batch to export
        format   — 'csv' or 'excel' (default: 'excel')
    """
    tenant_id = current_user.tenant_id
    await _require_connected_mode(db, tenant_id)

    result = await db.execute(
        select(PostingBatch).where(
            PostingBatch.id == batch_id,
            PostingBatch.tenant_id == tenant_id,
        )
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Posting batch not found.")

    # Flatten all transaction lines into rows.
    rows: list[dict] = []
    for txn in (batch.transactions or []):
        for line in txn.get("lines", []):
            dimensions_str = "; ".join(
                f"{k}={v}" for k, v in (line.get("dimensions") or {}).items()
            )
            rows.append({
                "Batch Ref": batch.batch_ref,
                "Entry Date": txn.get("entry_date", ""),
                "Source Module": txn.get("source_module", ""),
                "Source Reference": txn.get("source_id", ""),
                "Description": line.get("description") or txn.get("description", ""),
                "GL Code": line.get("gl_code", ""),
                "GL Name": line.get("gl_name", ""),
                "Debit": line.get("debit", 0),
                "Credit": line.get("credit", 0),
                "Dimensions": dimensions_str,
            })

    # Mark as exported if still pending.
    if batch.status == "pending":
        batch.status = "exported"
        batch.exported_at = datetime.utcnow()
        await db.flush()

    if file_format == "csv":
        return _build_csv_response(rows, batch.batch_ref)
    return _build_excel_response(rows, batch.batch_ref)


@router.post("/{batch_id}/mark-exported", status_code=status.HTTP_200_OK)
async def mark_batch_exported(
    batch_id: UUID,
    current_user: Annotated[UserTenant, Depends(require_auth)],
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    Manually mark a posting batch as exported.

    Useful when the batch was transferred outside of the download endpoint
    (e.g. via API sync to an external ERP). Idempotent — calling on an
    already-exported batch updates exported_at to now().

    Parameters:
        batch_id — UUID of the batch
    """
    tenant_id = current_user.tenant_id
    await _require_connected_mode(db, tenant_id)

    result = await db.execute(
        select(PostingBatch).where(
            PostingBatch.id == batch_id,
            PostingBatch.tenant_id == tenant_id,
        )
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Posting batch not found.")

    if batch.status == "synced":
        raise HTTPException(
            status_code=400,
            detail="Batch is already marked as synced. No further status changes allowed.",
        )

    batch.status = "exported"
    batch.exported_at = datetime.utcnow()
    await db.flush()

    return {"id": str(batch_id), "status": "exported", "batch_ref": batch.batch_ref}


# ── Private: file builders ────────────────────────────────────────────────────

def _build_csv_response(rows: list[dict], batch_ref: str) -> Response:
    """Build a CSV file response from the flattened row list."""
    if not rows:
        output = io.StringIO()
        output.write("No data\n")
        content = output.getvalue()
    else:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
        content = output.getvalue()

    filename = f"{batch_ref}.csv"
    return Response(
        content=content.encode("utf-8-sig"),  # BOM for Excel compatibility
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_excel_response(rows: list[dict], batch_ref: str) -> Response:
    """Build an Excel (.xlsx) file response from the flattened row list."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Posting Batch"

    if not rows:
        ws.append(["No data"])
    else:
        headers = list(rows[0].keys())

        # Header row styling
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows — alternate shading
        light_fill = PatternFill(start_color="EFF3FF", end_color="EFF3FF", fill_type="solid")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
                if row_idx % 2 == 0:
                    cell.fill = light_fill
                # Right-align numeric columns
                if key in ("Debit", "Credit"):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(header)),
                max((len(str(row.get(header, ""))) for row in rows), default=0),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{batch_ref}.xlsx"
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
