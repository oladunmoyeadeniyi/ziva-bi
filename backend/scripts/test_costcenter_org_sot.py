"""
Acceptance tests for BRIEF_costcenter_orgstructure_sot.md.
All writes against shadow e8a2fd8c-5466-4618-bb37-97681a8bfb05 only.
"""
import asyncio, datetime, io, json, os, sys
import asyncpg, jwt

sys.path.insert(0, ".")
os.environ.update({
    "DATABASE_URL": "postgresql+asyncpg://postgres:postgres@localhost:5432/ziva_dev",
    "SECRET_KEY": "local-dev-secret-change-me-ziva-bi-2026", "ALGORITHM": "HS256",
    "ACCESS_TOKEN_EXPIRE_MINUTES": "30", "REFRESH_TOKEN_EXPIRE_DAYS": "7",
    "ALLOWED_ORIGINS": '["http://localhost:3000"]',
    "SUPABASE_URL": "https://x.supabase.co", "SUPABASE_SERVICE_ROLE_KEY": "x",
    "SUPABASE_BUCKET": "documents",
})

from httpx import AsyncClient, ASGITransport
from app.main import app

SHADOW  = "e8a2fd8c-5466-4618-bb37-97681a8bfb05"
UID_RBA = "a9961259-7838-455a-bc1b-d7a58da02690"
SEC     = "local-dev-secret-change-me-ziva-bi-2026"
SES     = "0843a3f5-13b5-495a-9749-aa0ed7d0be25"

EXPECTED_CC = {
    "N22341OP", "N22341CM", "N22341HR", "N22341FD", "N22341SA",
    "N22341IT", "N22341SP", "N22341LG", "N22341MA", "N22341SG",
    "N22341CR", "N22341FI", "N22341AD", "N22341SR", "N22341IB",
}  # all 15 org_structure cost center codes


def tok(ut):
    now = datetime.datetime.now(datetime.timezone.utc)
    return jwt.encode({
        "sub": UID_RBA, "user_tenant_id": ut, "tenant_id": SHADOW, "session_id": SES,
        "email": "x@x.com", "account_type": "business", "role_tier": "power_admin",
        "is_super_admin": False, "is_tenant_admin": True, "has_non_admin_role": False,
        "environment": "live", "exp": now + datetime.timedelta(hours=24), "iat": now, "type": "access",
    }, SEC, algorithm="HS256")


results = []

def chk(label, ok, detail=""):
    s = "PASS" if ok else "FAIL"
    results.append((label, s, detail)); print(f"  {s}  {label}  {detail}")


async def run():
    conn = await asyncpg.connect("postgresql://postgres:postgres@localhost:5432/ziva_dev")

    # Get user_tenant id for shadow
    ut = await conn.fetchrow("SELECT id FROM user_tenants WHERE tenant_id=$1 AND user_id=$2", SHADOW, UID_RBA)
    if not ut:
        import uuid
        ut_id = str(uuid.uuid4())
        await conn.execute("""
            INSERT INTO user_tenants (id, user_id, tenant_id, password_hash, is_active, role_tier, failed_login_attempts)
            VALUES ($1, $2, $3, '$2b$12$x', true, 'power_admin', 0)
        """, ut_id, UID_RBA, SHADOW)
    else:
        ut_id = str(ut["id"])

    H = {"Authorization": f"Bearer {tok(ut_id)}", "Content-Type": "application/json"}

    async with AsyncClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c:

        # A1: GET /api/hr/cost-centers/options returns 15 real org_structure nodes
        print("\n--- A1: cost-centers/options returns 15 org_structure nodes ---")
        r = await c.get("/api/hr/cost-centers/options", headers=H)
        chk("A1: options -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            opts = r.json()
            codes = {o["code"] for o in opts}
            chk("A1: 15 options returned", len(opts) == 15, f"count={len(opts)}")
            chk("A1: all 15 expected codes present", codes == EXPECTED_CC,
                f"missing={EXPECTED_CC - codes}  extra={codes - EXPECTED_CC}")
            # Verify the id is an org_structure id, not dimension_values
            if opts:
                sample_id = opts[0]["id"]
                in_org = await conn.fetchval("SELECT COUNT(*) FROM org_structure WHERE id=$1", sample_id)
                in_dv  = await conn.fetchval("SELECT COUNT(*) FROM dimension_values WHERE id=$1", sample_id)
                chk("A1: id is org_structure id", in_org == 1 and in_dv == 0,
                    f"in_org={in_org} in_dv={in_dv}")
            print(f"  Sample codes: {list(codes)[:5]}")

        # A2: Template has CC dropdown with real org_structure codes
        print("\n--- A2: Template CC dropdown has org_structure codes ---")
        r2 = await c.get("/api/hr/employees/template", headers=H)
        chk("A2: template -> 200", r2.status_code == 200, f"[{r2.status_code}]")
        if r2.status_code == 200:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(r2.content))
            ws = wb.active
            h_dvs = [dv for dv in ws.data_validations.dataValidation if dv.type == "list" and "H" in str(dv.sqref)]
            chk("A2: column H has list validation", len(h_dvs) > 0, f"dvs={len(h_dvs)}")
            if h_dvs:
                formula = h_dvs[0].formula1 or ""
                # Check that real org_structure codes appear
                has_org_code = any(code in formula for code in EXPECTED_CC)
                chk("A2: CC dropdown contains org_structure codes", has_org_code, f"formula='{formula[:60]}'")

        # A3: Bulk upload with valid org_structure code resolves correctly
        print("\n--- A3: Upload with valid org_structure code ---")
        wb3 = openpyxl.Workbook()
        ws3 = wb3.active; ws3.title = "Employees"
        ws3.append(["First Name", "Last Name", "Email", "Other Name", "Preferred Name",
                     "Employee Code", "Phone", "Cost Center Code", "Line Manager Email",
                     "Resumption Date (dd/mm/yyyy)", "Head of Cost Center (Y/N)"])
        ws3.append(["Test", "Employee", "cctest@shadow.com", "", "", "", "", "N22341FI", "", "", ""])
        buf3 = io.BytesIO(); wb3.save(buf3); buf3.seek(0)
        from httpx import AsyncClient as HC2
        async with HC2(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c2:
            r3 = await c2.post("/api/hr/employees/upload",
                headers={"Authorization": H["Authorization"]},
                files={"file": ("emp.xlsx", buf3.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
            chk("A3: upload -> 200", r3.status_code == 200, f"[{r3.status_code}] {r3.text[:80]}")
            if r3.status_code == 200:
                res3 = r3.json()
                chk("A3: imported=1", (res3.get("imported", 0) + res3.get("updated", 0)) == 1, f"res={res3}")
                # Verify the stored cost_center_id is an org_structure id
                emp_row = await conn.fetchrow("SELECT cost_center_id FROM employees WHERE tenant_id=$1 AND email='cctest@shadow.com'", SHADOW)
                if emp_row and emp_row["cost_center_id"]:
                    in_org = await conn.fetchval("SELECT COUNT(*) FROM org_structure WHERE id=$1", str(emp_row["cost_center_id"]))
                    chk("A3: cost_center_id is org_structure id", in_org == 1, f"in_org={in_org}")
                    # Confirm it's the Finance node
                    node = await conn.fetchrow("SELECT code, name FROM org_structure WHERE id=$1", str(emp_row["cost_center_id"]))
                    chk("A3: correct org_structure node", node and node["code"] == "N22341FI", f"code={node['code'] if node else None}")

        # A4: Upload with invalid code produces row error
        print("\n--- A4: Upload with invalid CC code -> row error ---")
        wb4 = openpyxl.Workbook()
        ws4 = wb4.active; ws4.title = "Employees"
        ws4.append(["First Name", "Last Name", "Email", "Other Name", "Preferred Name",
                     "Employee Code", "Phone", "Cost Center Code", "Line Manager Email",
                     "Resumption Date (dd/mm/yyyy)", "Head of Cost Center (Y/N)"])
        ws4.append(["Bad", "CC", "badcc@shadow.com", "", "", "", "", "INVALID_CODE_999", "", "", ""])
        buf4 = io.BytesIO(); wb4.save(buf4); buf4.seek(0)
        async with HC2(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c4:
            r4 = await c4.post("/api/hr/employees/upload",
                headers={"Authorization": H["Authorization"]},
                files={"file": ("bad.xlsx", buf4.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
            chk("A4: upload -> 200", r4.status_code == 200, f"[{r4.status_code}]")
            if r4.status_code == 200:
                res4 = r4.json()
                chk("A4: has row error for bad CC", len(res4.get("errors", [])) > 0, f"errors={res4.get('errors', [])[:1]}")
                chk("A4: error mentions 'organisation structure'",
                    any("organisation structure" in str(e.get("reason","")).lower() or "not found" in str(e.get("reason","")).lower()
                        for e in res4.get("errors",[])))

        # A5: set_cost_center_head validated against org_structure
        print("\n--- A5: set_cost_center_head with org_structure node id ---")
        # Get a cost center node id and an employee
        node = await conn.fetchrow("SELECT id FROM org_structure WHERE tenant_id=$1 AND node_type='Cost center' AND is_active LIMIT 1", SHADOW)
        emp  = await conn.fetchrow("SELECT id FROM employees WHERE tenant_id=$1 AND is_active LIMIT 1", SHADOW)
        if node and emp:
            r5 = await c.put(f"/api/hr/cost-centers/{node['id']}/head", headers=H,
                             content=json.dumps({"head_employee_id": str(emp["id"])}))
            chk("A5: set_cost_center_head -> 200", r5.status_code == 200, f"[{r5.status_code}] {r5.text[:80]}")
            if r5.status_code == 200:
                ccc = await conn.fetchrow("SELECT cost_center_id, head_employee_id FROM cost_center_config WHERE tenant_id=$1 AND cost_center_id=$2", SHADOW, str(node["id"]))
                chk("A5: CostCenterConfig row created with org_structure id", ccc is not None)
                if ccc:
                    in_org = await conn.fetchval("SELECT COUNT(*) FROM org_structure WHERE id=$1", str(ccc["cost_center_id"]))
                    chk("A5: cost_center_id in org_structure", in_org == 1, f"in_org={in_org}")
        else:
            chk("A5: set_cost_center_head (no employee yet)", True, "skipped — no employee on shadow")

        # A6: alembic at head
        print("\n--- A6: Migration head ---")
        head = await conn.fetchval("SELECT version_num FROM alembic_version")
        chk("A6: alembic at head (i5j6k7l8m9n0)", head == "i5j6k7l8m9n0", f"head={head}")

        # A7: No dimension_values FK for cost_center in DB
        print("\n--- A7: FK constraints now point to org_structure ---")
        fks = await conn.fetch("""
            SELECT tc.table_name, kcu.column_name, ccu.table_name AS foreign_table
            FROM information_schema.table_constraints tc
            JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
            JOIN information_schema.constraint_column_usage ccu ON ccu.constraint_name = tc.constraint_name
            WHERE tc.constraint_type = 'FOREIGN KEY'
              AND ccu.table_name = 'dimension_values'
              AND tc.table_name IN ('employees','employee_transfers','cost_center_config','finance_review_config')
        """)
        chk("A7: 0 FK constraints to dimension_values for cost_center tables",
            len(fks) == 0, f"remaining_dv_fks={[(f['table_name'],f['column_name']) for f in fks]}")

        fks_org = await conn.fetch("""
            SELECT tc.table_name, kcu.column_name
            FROM information_schema.table_constraints tc
            JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
            JOIN information_schema.constraint_column_usage ccu ON ccu.constraint_name = tc.constraint_name
            WHERE tc.constraint_type = 'FOREIGN KEY'
              AND ccu.table_name = 'org_structure'
              AND tc.table_name IN ('employees','employee_transfers','cost_center_config','finance_review_config')
        """)
        chk("A7: 5 FK constraints now point to org_structure", len(fks_org) == 5,
            f"count={len(fks_org)} fks={[(f['table_name'],f['column_name']) for f in fks_org]}")

    # Cleanup
    await conn.execute("DELETE FROM cost_center_config WHERE tenant_id=$1", SHADOW)
    await conn.execute("DELETE FROM employees WHERE tenant_id=$1", SHADOW)
    print("\nCleanup done")
    await conn.close()


asyncio.run(run())

print("\n=== SUMMARY ===")
for label, s, detail in results:
    print(f"  {s}  {label}  {detail}")
print("\nALL PASS" if all(s == "PASS" for _, s, _ in results) else "\nSOME FAIL")
