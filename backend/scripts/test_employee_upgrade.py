"""
Acceptance tests for BRIEF_employee_costcenter_listupgrade.
Uses Ziva BI Test Tenant (f2aecfab-025f-410f-a7f6-df923172c8a1) — NOT live Red Bull.
Run: python scripts/test_employee_upgrade.py
"""
import asyncio, datetime, io, json, os, sys
import asyncpg, jwt

sys.path.insert(0, ".")
os.environ.update({
    "DATABASE_URL": "postgresql+asyncpg://postgres:postgres@localhost:5432/ziva_dev",
    "SECRET_KEY": "local-dev-secret-change-me-ziva-bi-2026", "ALGORITHM": "HS256",
    "ACCESS_TOKEN_EXPIRE_MINUTES": "30", "REFRESH_TOKEN_EXPIRE_DAYS": "7",
    "ALLOWED_ORIGINS": '["http://localhost:3000"]',
    "SUPABASE_URL": "https://x.supabase.co", "SUPABASE_SERVICE_ROLE_KEY": "x",
    "SUPABASE_BUCKET": "documents",
})

from httpx import AsyncClient, ASGITransport
from app.main import app

# Use Red Bull tenant (has real data) for READ tests; test tenant for WRITE tests
RB_TID   = "bd2c8a25-7467-494a-96fa-30f40b5b5d19"
TEST_TID = "f2aecfab-025f-410f-a7f6-df923172c8a1"
SEC = "local-dev-secret-change-me-ziva-bi-2026"
SES = "0843a3f5-13b5-495a-9749-aa0ed7d0be25"
SA_UID = "7d9a7dab-5b8f-43ac-8d16-18e6bbe9feeb"


def tok(uid, ut, tenant_id, admin=True):
    now = datetime.datetime.now(datetime.timezone.utc)
    return jwt.encode({
        "sub": uid, "user_tenant_id": ut, "tenant_id": tenant_id, "session_id": SES,
        "email": "x@x.com", "account_type": "business", "role_tier": "power_admin" if admin else None,
        "is_super_admin": False, "is_tenant_admin": admin, "has_non_admin_role": not admin,
        "environment": "live", "exp": now + datetime.timedelta(hours=24), "iat": now, "type": "access",
    }, SEC, algorithm="HS256")


# Red Bull tokens (for read-only tests — has real CC data)
UID_RB_A = "a9961259-7838-455a-bc1b-d7a58da02690"
UT_RB_A  = "22da50be-66d7-4068-8010-678386404c4c"
H_RB = {"Authorization": f"Bearer {tok(UID_RB_A, UT_RB_A, RB_TID, True)}", "Content-Type": "application/json"}

results = []


def chk(label, ok, detail=""):
    s = "PASS" if ok else "FAIL"
    results.append((label, s, detail)); print(f"  {s}  {label}  {detail}")


async def setup_test_tenant(conn):
    """Seed the test tenant with a cost_center dimension + values + a user + employee."""
    # Add a cost_center dimension
    await conn.execute("""
        INSERT INTO tenant_dimensions (id, tenant_id, name, code, is_required, is_active, sort_order, value_source)
        VALUES ('d0d0d0d0-1111-1111-1111-000000000001', $1, 'Cost Center', 'cost_center', false, true, 1, 'org_structure')
        ON CONFLICT DO NOTHING
    """, TEST_TID)

    # Add 2 cost center values
    for uid, code, name in [
        ("e1e1e1e1-1111-1111-1111-000000000001", "TC_FIN", "Test Finance"),
        ("e1e1e1e1-1111-1111-1111-000000000002", "TC_HR",  "Test HR"),
    ]:
        await conn.execute("""
            INSERT INTO dimension_values (id, tenant_id, dimension_id, code, name, is_active, sort_order)
            VALUES ($1, $2, 'd0d0d0d0-1111-1111-1111-000000000001', $3, $4, true, 1)
            ON CONFLICT DO NOTHING
        """, uid, TEST_TID, code, name)

    # Add a user_tenant for UID_RB_A on test tenant (to allow writes)
    await conn.execute("""
        INSERT INTO user_tenants (id, user_id, tenant_id, password_hash, is_active, role_tier, failed_login_attempts)
        VALUES ('f0f0f0f0-1111-1111-1111-000000000001', $1, $2, '$2b$12$x', true, 'power_admin', 0)
        ON CONFLICT DO NOTHING
    """, UID_RB_A, TEST_TID)

    # Add accounting period for write tests
    await conn.execute("""
        INSERT INTO accounting_periods (id, tenant_id, fiscal_year, period_no, period_name, start_date, end_date, status)
        VALUES ('a0a0a0a0-1111-1111-1111-000000000001', $1, 'FY2027', 1, 'January 2027', '2027-01-01', '2027-01-31', 'OPEN')
        ON CONFLICT DO NOTHING
    """, TEST_TID)


async def teardown_test_tenant(conn):
    """Remove test data from test tenant."""
    await conn.execute("DELETE FROM cost_center_config WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM employees WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM dimension_values WHERE tenant_id=$1 AND code LIKE 'TC_%'", TEST_TID)
    await conn.execute("DELETE FROM tenant_dimensions WHERE tenant_id=$1 AND code='cost_center'", TEST_TID)
    await conn.execute("DELETE FROM user_tenants WHERE tenant_id=$1 AND user_id=$2", TEST_TID, UID_RB_A)
    await conn.execute("DELETE FROM accounting_periods WHERE tenant_id=$1 AND fiscal_year='FY2027'", TEST_TID)


async def run():
    conn = await asyncpg.connect("postgresql://postgres:postgres@localhost:5432/ziva_dev")
    await setup_test_tenant(conn)

    H_TEST = {"Authorization": f"Bearer {tok(UID_RB_A, UT_RB_A, TEST_TID, True)}", "Content-Type": "application/json"}

    async with AsyncClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c:

        # A1: GET /api/hr/cost-centers/options returns real CC dimension values
        print("\n--- A1: cost-centers/options endpoint ---")
        r = await c.get(f"/api/hr/cost-centers/options", headers=H_RB)
        chk("A1: options -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            opts = r.json()
            chk("A1: returns list", isinstance(opts, list))
            chk("A1: has id/code/name fields", all("id" in o and "code" in o and "name" in o for o in opts[:3]))
            print(f"  Red Bull cost center options: {len(opts)} items")
            if opts: print(f"  Sample: {opts[0]}")

        # A1b: Test tenant cost-centers/options — only test data
        r = await c.get(f"/api/hr/cost-centers/options", headers=H_TEST)
        chk("A1b: test tenant options -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            test_opts = r.json()
            chk("A1b: returns TC_ codes only", all(o["code"].startswith("TC_") for o in test_opts),
                f"opts={[o['code'] for o in test_opts]}")
            cc_id = next((o["id"] for o in test_opts if o["code"] == "TC_FIN"), None)
            chk("A1b: TC_FIN present", cc_id is not None, f"found={cc_id}")

        # A2: Add employee with cost_center_id via POST /api/hr/employees
        print("\n--- A2: Add employee with cost_center_id ---")
        if r.status_code == 200 and test_opts:
            cc_id = test_opts[0]["id"]
            r2 = await c.post("/api/hr/employees", headers=H_TEST, content=json.dumps({
                "first_name": "Test", "last_name": "Employee", "email": "test.employee@testtenant.com",
                "cost_center_id": cc_id,
            }))
            chk("A2: create employee with CC -> 201", r2.status_code == 201, f"[{r2.status_code}] {r2.text[:80]}")
            if r2.status_code == 201:
                emp = r2.json()
                chk("A2: cost_center_id set", emp.get("cost_center_id") == cc_id, f'cc={emp.get("cost_center_id")}')

        # A3: Template download — check it's xlsx and has new column
        print("\n--- A3: Template download ---")
        BASE = os.environ.get("NEXT_PUBLIC_API_URL", "http://localhost:8000")
        import asyncio
        from httpx import AsyncClient as HClient
        async with HClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as hc:
            tr = await hc.get("/api/hr/employees/template", headers=H_TEST)
            chk("A3: template download -> 200", tr.status_code == 200, f"[{tr.status_code}]")
            if tr.status_code == 200:
                content_type = tr.headers.get("content-type", "")
                chk("A3: is xlsx", "spreadsheetml" in content_type, f"ct={content_type}")
                # Check the xlsx contains our new column
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(tr.content))
                ws = wb.active
                headers = [str(ws.cell(1, c).value or "").lower() for c in range(1, ws.max_column + 1)]
                chk("A3: has head of cost center column", any("head" in h for h in headers), f"headers={headers}")
                chk("A3: has cost center code column", any("cost center" in h for h in headers), f"headers={headers}")
                # Check data validations
                dvs = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
                chk("A3: has dropdown validations", len(dvs) > 0, f"dv_count={len(dvs)}")

        # A4: Upload with head column — test two-pass resolution
        print("\n--- A4: Upload with head-of-cost-center column ---")
        if test_opts:
            cc_code = test_opts[0]["code"]
            cc_id_str = test_opts[0]["id"]
            # Build a minimal xlsx upload
            import openpyxl
            wb2 = openpyxl.Workbook()
            ws2 = wb2.active
            ws2.title = "Employees"
            ws2.append(["First Name", "Last Name", "Email", "Other Name", "Preferred Name",
                         "Employee Code", "Phone", "Cost Center Code", "Line Manager Email",
                         "Resumption Date (dd/mm/yyyy)", "Head of Cost Center (Y/N)"])
            ws2.append(["Instr", "Row", "skip", "", "", "", "", "", "", "", ""])
            ws2.append(["Example", "Row", "skip@ex.com", "", "", "", "", "", "", "", ""])
            # Data row: head=Y
            ws2.append(["Head", "Person", "head@testtenant.com", "", "", "", "", cc_code, "", "", "Y"])
            # Data row: not head
            ws2.append(["Normal", "Person", "normal@testtenant.com", "", "", "", "", cc_code, "", "", ""])

            buf = io.BytesIO()
            wb2.save(buf)
            buf.seek(0)

            from httpx import AsyncClient as HClient2
            async with HClient2(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as hc2:
                r_up = await hc2.post("/api/hr/employees/upload",
                    headers={"Authorization": H_TEST["Authorization"]},
                    files={"file": ("test.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
                chk("A4: upload -> 200", r_up.status_code == 200, f"[{r_up.status_code}] {r_up.text[:80]}")
                if r_up.status_code == 200:
                    res = r_up.json()
                    chk("A4: imported > 0", res.get("imported", 0) > 0, f"imported={res.get('imported')}")
                    chk("A4: head_assignments = 1", res.get("head_assignments", 0) == 1, f"head_assignments={res.get('head_assignments')}")

                    # Verify CostCenterConfig was upserted
                    cc_cfg = await conn.fetchrow(
                        "SELECT head_employee_id FROM cost_center_config WHERE tenant_id=$1 AND cost_center_id=$2",
                        TEST_TID, cc_id_str)
                    chk("A4: CostCenterConfig upserted", cc_cfg is not None)
                    if cc_cfg:
                        head_emp = await conn.fetchrow("SELECT email FROM employees WHERE id=$1", cc_cfg["head_employee_id"])
                        chk("A4: head employee email correct", head_emp and head_emp["email"] == "head@testtenant.com",
                            f'email={head_emp["email"] if head_emp else None}')

        # A5: Upload with invalid CC code → row error, but employee still created
        print("\n--- A5: Upload unrecognized CC code → row error ---")
        if test_opts:
            wb3 = openpyxl.Workbook()
            ws3 = wb3.active
            ws3.title = "Employees"
            ws3.append(["First Name", "Last Name", "Email", "Other Name", "Preferred Name",
                         "Employee Code", "Phone", "Cost Center Code", "Line Manager Email",
                         "Resumption Date (dd/mm/yyyy)", "Head of Cost Center (Y/N)"])
            ws3.append(["i", "i", "i", "", "", "", "", "", "", "", ""])
            ws3.append(["e", "e", "e@e.com", "", "", "", "", "", "", "", ""])
            ws3.append(["Bad", "CC", "badcc@testtenant.com", "", "", "", "", "INVALID_CC_CODE", "", "", ""])

            buf3 = io.BytesIO(); wb3.save(buf3); buf3.seek(0)
            async with HClient2(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as hc3:
                r5 = await hc3.post("/api/hr/employees/upload",
                    headers={"Authorization": H_TEST["Authorization"]},
                    files={"file": ("t.xlsx", buf3.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
                chk("A5: upload -> 200", r5.status_code == 200, f"[{r5.status_code}]")
                if r5.status_code == 200:
                    res5 = r5.json()
                    chk("A5: errors > 0 (CC rejection)", len(res5.get("errors", [])) > 0, f"errors={res5.get('errors', [])[:1]}")

        # A6: Transfer uses CC options endpoint (test the transfer endpoint with a valid CC)
        print("\n--- A6: Transfer validates CC dimension type ---")
        if test_opts:
            emp_res = await conn.fetchrow("SELECT id FROM employees WHERE tenant_id=$1 AND email='normal@testtenant.com'", TEST_TID)
            if emp_res:
                r6 = await c.post(f"/api/hr/employees/{emp_res['id']}/transfer", headers=H_TEST,
                                  content=json.dumps({
                                      "to_cost_center_id": test_opts[1]["id"] if len(test_opts) > 1 else test_opts[0]["id"],
                                      "effective_date": "2027-01-01",
                                  }))
                chk("A6: transfer with valid CC -> 201", r6.status_code == 201, f"[{r6.status_code}] {r6.text[:80]}")

        # A7: /api/hr/cost-centers/options only returns CC dimension values (not all dims)
        print("\n--- A7: options only returns cost_center dimension values ---")
        # For Red Bull, check real IO/other dims are not in options
        r7 = await c.get("/api/hr/cost-centers/options", headers=H_RB)
        if r7.status_code == 200:
            opts7 = r7.json()
            # All should be from cost_center dimension. Verify by checking DB.
            if opts7:
                opt_ids = [o["id"] for o in opts7]
                db_check = await conn.fetch("""
                    SELECT d.code FROM dimension_values dv
                    JOIN tenant_dimensions d ON d.id = dv.dimension_id
                    WHERE dv.id = ANY($1::uuid[])
                """, opt_ids[:5])
                chk("A7: all options from cost_center dimension",
                    all(r["code"] == "cost_center" for r in db_check),
                    f"dim_codes={[r['code'] for r in db_check]}")

    await teardown_test_tenant(conn)
    await conn.close()
    print("\nTeardown complete")


asyncio.run(run())

print("\n=== SUMMARY ===")
for label, s, detail in results:
    print(f"  {s}  {label}  {detail}")
print("\nALL PASS" if all(s == "PASS" for _, s, _ in results) else "\nSOME FAIL")
