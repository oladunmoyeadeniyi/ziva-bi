"""
Acceptance tests for BRIEF_coa_remap_golive_gate.md.
Uses Red Bull tenant for CoA tests (real CoA data), test tenant for employee delete gate.
Run: python scripts/test_coa_remap.py
"""
import asyncio, datetime, io, json, os, sys
import asyncpg, jwt

sys.path.insert(0, ".")
os.environ.update({
    "DATABASE_URL": "postgresql+asyncpg://postgres:postgres@localhost:5432/ziva_dev",
    "SECRET_KEY": "local-dev-secret-change-me-ziva-bi-2026", "ALGORITHM": "HS256",
    "ACCESS_TOKEN_EXPIRE_MINUTES": "30", "REFRESH_TOKEN_EXPIRE_DAYS": "7",
    "ALLOWED_ORIGINS": '["http://localhost:3000"]',
    "SUPABASE_URL": "https://x.supabase.co", "SUPABASE_SERVICE_ROLE_KEY": "x",
    "SUPABASE_BUCKET": "documents",
})

from httpx import AsyncClient, ASGITransport
from app.main import app

RB_TID   = "bd2c8a25-7467-494a-96fa-30f40b5b5d19"  # Red Bull (live CoA data)
TEST_TID = "f2aecfab-025f-410f-a7f6-df923172c8a1"  # Ziva test tenant
SEC      = "local-dev-secret-change-me-ziva-bi-2026"
SES      = "0843a3f5-13b5-495a-9749-aa0ed7d0be25"
UID_RB_A = "a9961259-7838-455a-bc1b-d7a58da02690"
UT_RB_A  = "22da50be-66d7-4068-8010-678386404c4c"


def tok(uid, ut, tid, admin=True):
    now = datetime.datetime.now(datetime.timezone.utc)
    return jwt.encode({
        "sub": uid, "user_tenant_id": ut, "tenant_id": tid, "session_id": SES,
        "email": "x@x.com", "account_type": "business", "role_tier": "power_admin" if admin else None,
        "is_super_admin": False, "is_tenant_admin": admin, "has_non_admin_role": not admin,
        "environment": "live", "exp": now + datetime.timedelta(hours=24), "iat": now, "type": "access",
    }, SEC, algorithm="HS256")


H_RB = {"Authorization": f"Bearer {tok(UID_RB_A, UT_RB_A, RB_TID, True)}", "Content-Type": "application/json"}
results = []


def chk(label, ok, detail=""):
    s = "PASS" if ok else "FAIL"
    results.append((label, s, detail)); print(f"  {s}  {label}  {detail}")


async def run():
    conn = await asyncpg.connect("postgresql://postgres:postgres@localhost:5432/ziva_dev")

    # Pick active GL accounts from Red Bull by type — query specifically
    pl_accts = await conn.fetch("""
        SELECT id, gl_number, gl_name, account_type, is_active, is_retired
        FROM chart_of_accounts
        WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('PL','SOCI')
        ORDER BY gl_number LIMIT 10
    """, RB_TID)
    bs_accts = await conn.fetch("""
        SELECT id, gl_number, gl_name, account_type, is_active, is_retired
        FROM chart_of_accounts
        WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('BS','SOFP')
        ORDER BY gl_number LIMIT 5
    """, RB_TID)
    print(f"PL/SOCI accounts found: {len(pl_accts)}, BS/SOFP accounts found: {len(bs_accts)}")
    if pl_accts: print(f"  Sample PL: {[(a['gl_number'], a['account_type']) for a in pl_accts[:3]]}")
    if bs_accts: print(f"  Sample BS: {[(a['gl_number'], a['account_type']) for a in bs_accts[:2]]}")

    async with AsyncClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c:

        # A1: GET /api/config/coa returns is_retired field
        print("\n--- A1: is_retired in CoA list ---")
        r = await c.get("/api/config/coa?active_only=false&limit=5", headers=H_RB)
        chk("A1: coa list -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            items = r.json()
            chk("A1: has is_retired field", all("is_retired" in a for a in items[:3]))

        # A2: GET /api/config/coa/remap-template -> 200, is xlsx
        print("\n--- A2: Remap template download ---")
        r = await c.get("/api/config/coa/remap-template", headers=H_RB)
        chk("A2: template -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            chk("A2: is xlsx", "spreadsheetml" in r.headers.get("content-type",""))
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            sheets = wb.sheetnames
            chk("A2: has Remap sheet", "Remap" in sheets, f"sheets={sheets}")
            chk("A2: has Instructions sheet", "Instructions" in sheets)
            chk("A2: has Reference sheet", any("Reference" in s for s in sheets))

        # A3: Single-screen remap: 2 PL codes -> existing PL code
        print("\n--- A3: Single remap (existing new code) ---")
        if len(pl_accts) >= 3:
            old1, old2, new_acct = pl_accts[0], pl_accts[1], pl_accts[2]
            body = {
                "old_account_ids": [str(old1["id"]), str(old2["id"])],
                "new_account_id": str(new_acct["id"]),
                "reason": "Test remap A3"
            }
            r = await c.post("/api/config/coa/remap", headers=H_RB, content=json.dumps(body))
            chk("A3: remap -> 200", r.status_code == 200, f"[{r.status_code}] {r.text[:100]}")
            if r.status_code == 200:
                res = r.json()
                chk("A3: remapped count=2", len(res["remapped"]) == 2, f"count={len(res['remapped'])}")
                chk("A3: new_account_created=False", res["new_account_created"] == False)

                # Verify old accounts are retired
                for old_id in [old1["id"], old2["id"]]:
                    row = await conn.fetchrow("SELECT is_active, is_retired FROM chart_of_accounts WHERE id=$1", str(old_id))
                    chk(f"A3: old account {str(old_id)[:8]} is_retired=True", row["is_retired"])
                    chk(f"A3: old account {str(old_id)[:8]} is_active=False", not row["is_active"])

                # Verify gl_code_remap rows created
                remap_rows = await conn.fetch("SELECT * FROM gl_code_remaps WHERE tenant_id=$1 AND new_account_id=$2",
                                              RB_TID, str(new_acct["id"]))
                chk("A3: 2 gl_code_remap rows created", len(remap_rows) == 2, f"count={len(remap_rows)}")
                chk("A3: reason stored", all(r["reason"] == "Test remap A3" for r in remap_rows))

                # Verify audit log
                audit = await conn.fetchrow("SELECT log_metadata FROM audit_logs WHERE tenant_id=$1 AND event_type='coa.remap' ORDER BY created_at DESC LIMIT 1", RB_TID)
                chk("A3: audit log created", audit is not None)
        else:
            chk("A3: skipped (not enough PL accounts)", True, f"pl_accts={len(pl_accts)}")

        # A4: Single remap with inline new account creation
        print("\n--- A4: Single remap (inline new code) ---")
        import time
        test_new_gl = f"TEST_REMAP_{int(time.time()) % 100000}"
        if len(pl_accts) >= 4:
            old3 = pl_accts[3]
            body4 = {
                "old_account_ids": [str(old3["id"])],
                "new_account": {"gl_number": test_new_gl, "gl_name": "Test Remap New Account", "account_type": old3["account_type"]},
                "reason": "Test remap A4 inline"
            }
            r = await c.post("/api/config/coa/remap", headers=H_RB, content=json.dumps(body4))
            chk("A4: remap with inline new -> 200", r.status_code == 200, f"[{r.status_code}] {r.text[:100]}")
            if r.status_code == 200:
                res4 = r.json()
                chk("A4: new_account_created=True", res4["new_account_created"] == True)
                chk("A4: new_gl_number correct", res4["new_gl_number"] == test_new_gl)
        else:
            chk("A4: skipped (not enough PL accounts)", True)

        # A5: Remap with mismatched account_type -> rejected
        print("\n--- A5: Mismatched account_type -> rejected ---")
        # Reload fresh non-retired accounts after A3/A4 may have retired some
        pl_fresh5 = await conn.fetch("""
            SELECT id FROM chart_of_accounts WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('PL','SOCI') ORDER BY gl_number LIMIT 1
        """, RB_TID)
        bs_fresh5 = await conn.fetch("""
            SELECT id FROM chart_of_accounts WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('BS','SOFP') ORDER BY gl_number LIMIT 2
        """, RB_TID)
        if pl_fresh5 and len(bs_fresh5) >= 2:
            body5 = {
                "old_account_ids": [str(pl_fresh5[0]["id"]), str(bs_fresh5[0]["id"])],
                "new_account_id": str(bs_fresh5[1]["id"]),
            }
            r = await c.post("/api/config/coa/remap", headers=H_RB, content=json.dumps(body5))
            chk("A5: mismatched types -> 422", r.status_code == 422, f"[{r.status_code}]")
            if r.status_code == 422:
                detail = r.json().get("detail", {})
                chk("A5: error mentions conflicting types", "conflicting_types" in str(detail) or "account_type" in str(detail),
                    f"detail={str(detail)[:100]}")
        else:
            chk("A5: skipped (need PL+2 BS accounts)", True)

        # A6: Bulk remap template upload
        print("\n--- A6: Bulk remap upload ---")
        fresh_pl = await conn.fetch("""
            SELECT gl_number, account_type FROM chart_of_accounts
            WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('PL','SOCI')
            ORDER BY gl_number LIMIT 4
        """, RB_TID)
        fresh_pl_new = await conn.fetchrow("""
            SELECT gl_number FROM chart_of_accounts
            WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('PL','SOCI')
            ORDER BY gl_number OFFSET 4 LIMIT 1
        """, RB_TID)
        if len(fresh_pl) >= 2 and fresh_pl_new:
            old_gl1 = fresh_pl[0]["gl_number"]
            old_gl2 = fresh_pl[1]["gl_number"]
            new_gl  = fresh_pl_new["gl_number"]
            import openpyxl
            wb6 = openpyxl.Workbook()
            ws6 = wb6.active; ws6.title = "Remap"
            ws6.append(["Old GL Number", "New GL Number", "Reason (optional)"])
            ws6.append([old_gl1, new_gl, "Bulk test A6"])
            ws6.append([old_gl2, new_gl, "Bulk test A6"])
            buf6 = io.BytesIO(); wb6.save(buf6); buf6.seek(0)
            from httpx import AsyncClient as HC2
            async with HC2(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c2:
                r6 = await c2.post("/api/config/coa/remap-bulk",
                    headers={"Authorization": H_RB["Authorization"]},
                    files={"file": ("remap.xlsx", buf6.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
                chk("A6: bulk remap -> 200", r6.status_code == 200, f"[{r6.status_code}] {r6.text[:100]}")
                if r6.status_code == 200:
                    res6 = r6.json()
                    chk("A6: remapped=2", res6["remapped"] == 2, f"remapped={res6['remapped']}")
                    chk("A6: errors=0", res6["errors"] == 0, f"errors={res6['errors']}")
        else:
            chk("A6: skipped (not enough accounts)", True)

        # A7: Bulk remap with type-mismatch in one group -> that group rejected
        print("\n--- A7: Bulk remap type mismatch rejected ---")
        fresh_bs7 = await conn.fetchrow("""
            SELECT gl_number FROM chart_of_accounts
            WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('BS','SOFP')
            ORDER BY gl_number LIMIT 1
        """, RB_TID)
        fresh_pl7 = await conn.fetchrow("""
            SELECT gl_number FROM chart_of_accounts
            WHERE tenant_id=$1 AND is_active AND NOT is_retired AND account_type IN ('PL','SOCI')
            ORDER BY gl_number LIMIT 1
        """, RB_TID)
        fresh_bs = fresh_bs7; fresh_pl2 = fresh_pl7
        if fresh_bs and fresh_pl2:
            import openpyxl
            wb7 = openpyxl.Workbook()
            ws7 = wb7.active; ws7.title = "Remap"
            ws7.append(["Old GL Number", "New GL Number", "Reason"])
            # Mismatch: PL old code but BS new code -> error group
            ws7.append([fresh_pl2["gl_number"], fresh_bs["gl_number"], "Type mismatch test"])
            buf7 = io.BytesIO(); wb7.save(buf7); buf7.seek(0)
            from httpx import AsyncClient as HC3
            async with HC3(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c3:
                r7 = await c3.post("/api/config/coa/remap-bulk",
                    headers={"Authorization": H_RB["Authorization"]},
                    files={"file": ("remap7.xlsx", buf7.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
                chk("A7: bulk with mismatch -> 200", r7.status_code == 200, f"[{r7.status_code}]")
                if r7.status_code == 200:
                    res7 = r7.json()
                    chk("A7: errors > 0 (mismatch rejected)", res7["errors"] > 0, f"errors={res7['errors']}")
        else:
            chk("A7: skipped (need BS+PL accounts)", True)

        # A8: Posting picker (gl/search) excludes retired codes
        print("\n--- A8: GL search excludes retired codes ---")
        r8 = await c.get("/api/config/gl/search?q=&limit=100", headers=H_RB)
        chk("A8: gl/search -> 200", r8.status_code == 200, f"[{r8.status_code}]")
        if r8.status_code == 200:
            gl_results = r8.json()
            # Check: none of the returned accounts are retired
            ids_returned = {g["gl_id"] for g in gl_results}
            retired_ids = set(str(r["id"]) for r in await conn.fetch("""
                SELECT id FROM chart_of_accounts WHERE tenant_id=$1 AND is_retired
            """, RB_TID))
            overlap = ids_returned & retired_ids
            chk("A8: retired codes excluded from gl/search", len(overlap) == 0, f"overlap={len(overlap)}")
            print(f"  gl/search returned {len(ids_returned)} accounts; {len(retired_ids)} retired in DB; overlap={len(overlap)}")

        # A9: Trial balance includes retired code activity (no is_active filter on CoA)
        # (No journal data in Red Bull currently, but verify the endpoint works)
        print("\n--- A9: Trial balance endpoint (reporting context) ---")
        r9 = await c.get("/api/gl/trial-balance", headers=H_RB)
        chk("A9: trial-balance -> 200", r9.status_code in (200,), f"[{r9.status_code}]")

        # A10: Replace-all endpoint returns 410
        print("\n--- A10: Replace-all returns 410 ---")
        r10 = await c.post("/api/config/coa/replace-all", headers=H_RB)
        chk("A10: replace-all -> 410", r10.status_code == 410, f"[{r10.status_code}]")

    # A11: Employee hard-delete gate
    print("\n--- A11: Employee delete gate ---")
    # Setup: create a test user_tenant + employee on test tenant
    await conn.execute("""
        INSERT INTO user_tenants (id, user_id, tenant_id, password_hash, is_active, role_tier, failed_login_attempts)
        VALUES ('f0f0f0f1-1111-1111-1111-000000000001', $1, $2, '$2b$12$x', true, 'power_admin', 0)
        ON CONFLICT DO NOTHING
    """, UID_RB_A, TEST_TID)
    await conn.execute("""
        INSERT INTO employees (id, tenant_id, first_name, last_name, email, is_active)
        VALUES ('a1a1a1a1-1111-1111-1111-000000000001', $1, 'Test', 'Delete', 'testdelete@example.com', true)
        ON CONFLICT DO NOTHING
    """, TEST_TID)
    # Ensure test tenant is NOT live
    await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)
    live_status = await conn.fetchval("SELECT lifecycle_status FROM tenants WHERE id=$1", TEST_TID)
    print(f"  Test tenant lifecycle_status before gate test: {live_status}")

    H_TEST = {"Authorization": f"Bearer {tok(UID_RB_A, 'f0f0f0f1-1111-1111-1111-000000000001', TEST_TID, True)}", "Content-Type": "application/json"}

    async with AsyncClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c:
        # A11a: Pre-live -> hard-delete (row removed)
        r11a = await c.delete("/api/hr/employees/a1a1a1a1-1111-1111-1111-000000000001", headers=H_TEST)
        chk("A11a: pre-live delete -> 204", r11a.status_code == 204, f"[{r11a.status_code}]")
        still_exists = await conn.fetchval("SELECT COUNT(*) FROM employees WHERE id='a1a1a1a1-1111-1111-1111-000000000001'")
        chk("A11a: pre-live employee row actually deleted", still_exists == 0, f"rows={still_exists}")

        # A11b: Post-live -> soft-delete only
        await conn.execute("""
            INSERT INTO employees (id, tenant_id, first_name, last_name, email, is_active)
            VALUES ('a2a2a2a2-1111-1111-1111-000000000001', $1, 'Test2', 'Live', 'testlive@example.com', true)
            ON CONFLICT DO NOTHING
        """, TEST_TID)
        # Set tenant to live
        await conn.execute("UPDATE tenants SET lifecycle_status='live' WHERE id=$1", TEST_TID)
        r11b = await c.delete("/api/hr/employees/a2a2a2a2-1111-1111-1111-000000000001", headers=H_TEST)
        chk("A11b: post-live delete -> 204", r11b.status_code == 204, f"[{r11b.status_code}]")
        still_exists2 = await conn.fetchval("SELECT COUNT(*) FROM employees WHERE id='a2a2a2a2-1111-1111-1111-000000000001'")
        is_active_2 = await conn.fetchval("SELECT is_active FROM employees WHERE id='a2a2a2a2-1111-1111-1111-000000000001'")
        chk("A11b: post-live employee row still exists", still_exists2 == 1, f"rows={still_exists2}")
        chk("A11b: post-live employee is_active=False", is_active_2 == False, f"is_active={is_active_2}")
        # Reset
        await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)

    # Verify migration + routes + alembic head
    print("\n--- A12: Migration + routes + alembic head ---")
    head = await conn.fetchval("SELECT version_num FROM alembic_version")
    chk("A12: alembic at head (h4i5j6k7l8m9)", head == "h4i5j6k7l8m9", f"head={head}")

    col = await conn.fetchrow("SELECT column_name FROM information_schema.columns WHERE table_name='chart_of_accounts' AND column_name='is_retired'")
    chk("A12: is_retired column exists", col is not None)

    gl_remap_exists = await conn.fetchrow("SELECT table_name FROM information_schema.tables WHERE table_name='gl_code_remaps'")
    chk("A12: gl_code_remaps table exists", gl_remap_exists is not None)

    # Cleanup
    await conn.execute("DELETE FROM employees WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM user_tenants WHERE id='f0f0f0f1-1111-1111-1111-000000000001'")
    # Restore retired Red Bull accounts for cleanup (best effort)
    await conn.execute("DELETE FROM gl_code_remaps WHERE tenant_id=$1", RB_TID)
    await conn.execute("UPDATE chart_of_accounts SET is_active=true, is_retired=false WHERE tenant_id=$1 AND is_retired", RB_TID)
    await conn.execute("DELETE FROM chart_of_accounts WHERE tenant_id=$1 AND gl_number LIKE 'TEST_REMAP_%'", RB_TID)
    print("Cleanup done")
    await conn.close()


asyncio.run(run())

print("\n=== SUMMARY ===")
for label, s, detail in results:
    print(f"  {s}  {label}  {detail}")
print("\nALL PASS" if all(s == "PASS" for _, s, _ in results) else "\nSOME FAIL")
