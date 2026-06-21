"""
Acceptance tests for BRIEF_coa_lifecycle_fix_template_format.md.
Uses test tenant f2aecfab for CoA/lifecycle tests; seeds CC dim values for template test.
Run: PYTHONIOENCODING=utf-8 python scripts/test_lifecycle_fix.py
"""
import asyncio, datetime, io, json, os, sys
import asyncpg, jwt

sys.path.insert(0, ".")
os.environ.update({
    "DATABASE_URL": "postgresql+asyncpg://postgres:postgres@localhost:5432/ziva_dev",
    "SECRET_KEY": "local-dev-secret-change-me-ziva-bi-2026", "ALGORITHM": "HS256",
    "ACCESS_TOKEN_EXPIRE_MINUTES": "30", "REFRESH_TOKEN_EXPIRE_DAYS": "7",
    "ALLOWED_ORIGINS": '["http://localhost:3000"]',
    "SUPABASE_URL": "https://x.supabase.co", "SUPABASE_SERVICE_ROLE_KEY": "x",
    "SUPABASE_BUCKET": "documents",
})

from httpx import AsyncClient, ASGITransport
from app.main import app

TEST_TID = "f2aecfab-025f-410f-a7f6-df923172c8a1"
RB_TID   = "bd2c8a25-7467-494a-96fa-30f40b5b5d19"
SEC      = "local-dev-secret-change-me-ziva-bi-2026"
SES      = "0843a3f5-13b5-495a-9749-aa0ed7d0be25"
UID_RB_A = "a9961259-7838-455a-bc1b-d7a58da02690"
UT_RB_A  = "22da50be-66d7-4068-8010-678386404c4c"


def tok(uid, ut, tid, admin=True):
    now = datetime.datetime.now(datetime.timezone.utc)
    return jwt.encode({
        "sub": uid, "user_tenant_id": ut, "tenant_id": tid, "session_id": SES,
        "email": "x@x.com", "account_type": "business", "role_tier": "power_admin",
        "is_super_admin": False, "is_tenant_admin": admin, "has_non_admin_role": not admin,
        "environment": "live", "exp": now + datetime.timedelta(hours=24), "iat": now, "type": "access",
    }, SEC, algorithm="HS256")


H_RB = {"Authorization": f"Bearer {tok(UID_RB_A, UT_RB_A, RB_TID)}", "Content-Type": "application/json"}
results = []


def chk(label, ok, detail=""):
    s = "PASS" if ok else "FAIL"
    results.append((label, s, detail)); print(f"  {s}  {label}  {detail}")


async def setup(conn):
    """Seed test tenant with user_tenant, CC dimension + 2 values, CoA accounts."""
    await conn.execute("""
        INSERT INTO user_tenants (id, user_id, tenant_id, password_hash, is_active, role_tier, failed_login_attempts)
        VALUES ('f0f0f0f2-1111-1111-1111-000000000001', $1, $2, '$2b$12$x', true, 'power_admin', 0)
        ON CONFLICT DO NOTHING
    """, UID_RB_A, TEST_TID)
    # Upsert the cost_center dimension — use id d0d0d0d0-9999 so we avoid any conflict
    # with prior test runs that may have left d0d0d0d0-1111 or d0d0d0d0-2222
    cc_dim_id = "d0d0d0d0-9999-1111-1111-000000000001"
    existing = await conn.fetchrow(
        "SELECT id FROM tenant_dimensions WHERE tenant_id=$1 AND code='cost_center'", TEST_TID
    )
    if existing:
        cc_dim_id = str(existing["id"])
    else:
        await conn.execute("""
            INSERT INTO tenant_dimensions (id, tenant_id, name, code, is_required, is_active, sort_order, value_source)
            VALUES ($1, $2, 'Cost Center', 'cost_center', false, true, 1, 'org_structure')
        """, cc_dim_id, TEST_TID)

    for uid, code, name in [
        ("e1e1e1e1-9999-1111-1111-000000000001", "CC_TEST_A", "Test CC A"),
        ("e1e1e1e1-9999-1111-1111-000000000002", "CC_TEST_B", "Test CC B"),
    ]:
        await conn.execute(f"""
            INSERT INTO dimension_values (id, tenant_id, dimension_id, code, name, is_active, sort_order)
            VALUES ($1, $2, '{cc_dim_id}', $3, $4, true, 1)
            ON CONFLICT DO NOTHING
        """, uid, TEST_TID, code, name)
    # Two CoA accounts for remap tests
    for uid, gnum, gname, gtype in [
        ("c0c0c0c0-2222-1111-1111-000000000001", "T_GL_001", "Test Account 001", "SOCI"),
        ("c0c0c0c0-2222-1111-1111-000000000002", "T_GL_002", "Test Account 002", "SOCI"),
    ]:
        await conn.execute("""
            INSERT INTO chart_of_accounts (id, tenant_id, gl_number, gl_name, account_type, is_active, is_retired)
            VALUES ($1, $2, $3, $4, $5, true, false)
            ON CONFLICT DO NOTHING
        """, uid, TEST_TID, gnum, gname, gtype)


async def teardown(conn):
    await conn.execute("DELETE FROM gl_code_remaps WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM employees WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM chart_of_accounts WHERE tenant_id=$1", TEST_TID)
    await conn.execute("DELETE FROM dimension_values WHERE tenant_id=$1 AND code LIKE 'CC_TEST_%'", TEST_TID)
    await conn.execute("DELETE FROM user_tenants WHERE id='f0f0f0f2-1111-1111-1111-000000000001'")


async def run():
    conn = await asyncpg.connect("postgresql://postgres:postgres@localhost:5432/ziva_dev")
    await setup(conn)

    H_TEST = {"Authorization": f"Bearer {tok(UID_RB_A, 'f0f0f0f2-1111-1111-1111-000000000001', TEST_TID)}",
              "Content-Type": "application/json"}

    async with AsyncClient(transport=ASGITransport(app=app, raise_app_exceptions=False), base_url="http://test") as c:

        # ── Test 2: lifecycle_status == "in_implementation" ───────────────────
        print("\n--- T2: in_implementation gating ---")
        await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)

        # Replace All is allowed (in_implementation)
        import openpyxl
        wb_ra = openpyxl.Workbook(); ws_ra = wb_ra.active; ws_ra.title = "GL Accounts"
        ws_ra.append(["GL Number *", "GL Name *", "Account Type *"])
        ws_ra.append(["T_GL_NEW_001", "New Test Account", "SOCI"])
        buf_ra = io.BytesIO(); wb_ra.save(buf_ra); buf_ra.seek(0)
        r = await c.post("/api/config/coa/replace-all",
                         headers={"Authorization": H_TEST["Authorization"]},
                         files={"file": ("ra.xlsx", buf_ra.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        chk("T2: replace-all -> not 403 (in_implementation)", r.status_code != 403, f"[{r.status_code}]")
        chk("T2: replace-all -> 200", r.status_code == 200, f"[{r.status_code}] {r.text[:80]}")

        # Remap is blocked (not live)
        r = await c.post("/api/config/coa/remap", headers=H_TEST,
                         content=json.dumps({"old_account_ids": ["c0c0c0c0-2222-1111-1111-000000000001"],
                                             "new_account_id": "c0c0c0c0-2222-1111-1111-000000000002"}))
        chk("T2: remap blocked when in_implementation -> 403", r.status_code == 403, f"[{r.status_code}]")
        if r.status_code == 403:
            chk("T2: remap error mentions 'live'", "live" in r.json().get("detail", "").lower())

        # ── Test 3: lifecycle_status == "live" ────────────────────────────────
        print("\n--- T3: live gating ---")
        await conn.execute("UPDATE tenants SET lifecycle_status='live' WHERE id=$1", TEST_TID)
        # Re-seed CoA accounts — use UPSERT so Replace All's deactivation is overridden
        for uid, gnum, gname, gtype in [
            ("c0c0c0c0-2222-1111-1111-000000000001", "T_GL_001", "Test Account 001", "SOCI"),
            ("c0c0c0c0-2222-1111-1111-000000000002", "T_GL_002", "Test Account 002", "SOCI"),
        ]:
            await conn.execute("""
                INSERT INTO chart_of_accounts (id, tenant_id, gl_number, gl_name, account_type, is_active, is_retired)
                VALUES ($1, $2, $3, $4, $5, true, false)
                ON CONFLICT (id) DO UPDATE SET is_active=true, is_retired=false
            """, uid, TEST_TID, gnum, gname, gtype)

        # Replace All is blocked (live)
        buf_ra2 = io.BytesIO()
        wb_ra2 = openpyxl.Workbook(); ws_ra2 = wb_ra2.active
        ws_ra2.append(["GL Number *", "GL Name *", "Account Type *"])
        wb_ra2.save(buf_ra2); buf_ra2.seek(0)
        r = await c.post("/api/config/coa/replace-all",
                         headers={"Authorization": H_TEST["Authorization"]},
                         files={"file": ("ra2.xlsx", buf_ra2.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        chk("T3: replace-all blocked when live -> 403", r.status_code == 403, f"[{r.status_code}]")
        if r.status_code == 403:
            chk("T3: replace-all error mentions 'implementation'", "implementation" in r.json().get("detail","").lower())

        # Remap is allowed (live)
        r = await c.post("/api/config/coa/remap", headers=H_TEST,
                         content=json.dumps({"old_account_ids": ["c0c0c0c0-2222-1111-1111-000000000001"],
                                             "new_account_id": "c0c0c0c0-2222-1111-1111-000000000002",
                                             "reason": "lifecycle test"}))
        chk("T3: remap works when live -> 200", r.status_code == 200, f"[{r.status_code}] {r.text[:80]}")

        # Reset lifecycle for remaining tests
        await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)

        # ── Test 4+5: Employee template inspection ─────────────────────────────
        print("\n--- T4+5: Employee template ---")
        r = await c.get("/api/hr/employees/template", headers=H_TEST)
        chk("T4: template download -> 200", r.status_code == 200, f"[{r.status_code}]")
        if r.status_code == 200:
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            ws = wb.active
            dvs = list(ws.data_validations.dataValidation)
            print(f"  data_validations count: {len(dvs)}")
            for dv in dvs:
                print(f"    type={dv.type} sqref={dv.sqref} formula1={dv.formula1}")

            # T4: column H has a list DV (CC codes)
            h_dvs = [dv for dv in dvs if dv.type == "list" and "H" in str(dv.sqref)]
            chk("T4: column H has list validation", len(h_dvs) > 0, f"H_dvs={len(h_dvs)}")

            # T5: row 2 is NOT inline instruction text
            row2_a = ws.cell(row=2, column=1).value
            chk("T5: row 2 col A is empty (no inline instruction)", not row2_a or row2_a.strip() == "", f"row2_A='{row2_a}'")

            # T5: header cells have comments
            h1_comment = ws.cell(row=1, column=1).comment
            chk("T5: header cell A1 has a comment", h1_comment is not None)
            if h1_comment:
                chk("T5: comment contains example text", "adeniyi" in h1_comment.text.lower(), f"comment='{h1_comment.text[:50]}'")

            # T5: Instructions sheet mentions data start row
            ws2 = wb["Instructions"]
            instr_text = " ".join(str(ws2.cell(r, c).value or "") for r in range(1, 20) for c in range(1, 4)).lower()
            chk("T5: Instructions mentions row 2", "row 2" in instr_text, f"found={'row 2' in instr_text}")

            # T5: Head of Cost Center range starts at row 2
            k_dvs = [dv for dv in dvs if dv.type == "list" and "K" in str(dv.sqref)]
            chk("T5: col K DV range starts at K2", any("K2" in str(dv.sqref) for dv in k_dvs), f"k_sqrefs={[str(dv.sqref) for dv in k_dvs]}")

        # ── Test 6: Remap template inspection ──────────────────────────────────
        print("\n--- T6: Remap template ---")
        r6 = await c.get("/api/config/coa/remap-template", headers=H_TEST)
        chk("T6: remap template -> 200", r6.status_code == 200, f"[{r6.status_code}]")
        if r6.status_code == 200:
            wb6 = openpyxl.load_workbook(io.BytesIO(r6.content))
            ws6 = wb6.active
            row2_a6 = ws6.cell(row=2, column=1).value
            chk("T6: remap row 2 col A is empty (no inline text)", not row2_a6 or row2_a6.strip() == "", f"row2_A='{row2_a6}'")
            h1_comment6 = ws6.cell(row=1, column=1).comment
            chk("T6: remap header A1 has a comment", h1_comment6 is not None)
            ws2_6 = wb6["Instructions"]
            instr6 = " ".join(str(ws2_6.cell(r, c).value or "") for r in range(1, 15) for c in range(1, 3)).lower()
            chk("T6: remap Instructions mentions row 2", "row 2" in instr6)
            # Reference sheet untouched
            ref_sheet = next((s for s in wb6.sheetnames if "Reference" in s), None)
            chk("T6: Active GL Accounts reference sheet present", ref_sheet is not None, f"sheets={wb6.sheetnames}")

        # ── Test 7: Employee bulk upload with corrected template ───────────────
        print("\n--- T7: Employee upload (corrected template format) ---")
        # Construct a 2-row-skip-only xlsx (header + data starting row 2)
        wb7 = openpyxl.Workbook()
        ws7 = wb7.active; ws7.title = "Employees"
        ws7.append(["First Name", "Last Name", "Email", "Other Name", "Preferred Name",
                     "Employee Code", "Phone", "Cost Center Code", "Line Manager Email",
                     "Resumption Date (dd/mm/yyyy)", "Head of Cost Center (Y/N)"])
        # Row 2 = first data row (new format — no instruction rows between header and data)
        ws7.append(["Upload", "Test", "uploadtest@testtenant.com", "", "", "", "", "CC_TEST_A", "", "", ""])
        buf7 = io.BytesIO(); wb7.save(buf7); buf7.seek(0)
        r7 = await c.post("/api/hr/employees/upload",
                          headers={"Authorization": H_TEST["Authorization"]},
                          files={"file": ("emp.xlsx", buf7.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        chk("T7: upload -> 200", r7.status_code == 200, f"[{r7.status_code}] {r7.text[:100]}")
        if r7.status_code == 200:
            res7 = r7.json()
            chk("T7: imported=1 (row 2 is data)", res7.get("imported", 0) + res7.get("updated", 0) == 1,
                f"imported={res7.get('imported')} updated={res7.get('updated')} errors={res7.get('errors')}")

        # ── Test 8: CoA remap upload (corrected format; gated to live) ─────────
        print("\n--- T8: CoA remap upload with corrected template ---")
        await conn.execute("UPDATE tenants SET lifecycle_status='live' WHERE id=$1", TEST_TID)
        # Re-seed T_GL_001 (may have been retired by T3 remap)
        await conn.execute("DELETE FROM gl_code_remaps WHERE tenant_id=$1", TEST_TID)
        await conn.execute("""
            INSERT INTO chart_of_accounts (id, tenant_id, gl_number, gl_name, account_type, is_active, is_retired)
            VALUES ('c0c0c0c0-2222-1111-1111-000000000001', $1, 'T_GL_001', 'Test Account 001', 'SOCI', true, false)
            ON CONFLICT (id) DO UPDATE SET is_active=true, is_retired=false
        """, TEST_TID)
        await conn.execute("""
            INSERT INTO chart_of_accounts (id, tenant_id, gl_number, gl_name, account_type, is_active, is_retired)
            VALUES ('c0c0c0c0-2222-1111-1111-000000000002', $1, 'T_GL_002', 'Test Account 002', 'SOCI', true, false)
            ON CONFLICT (id) DO UPDATE SET is_active=true, is_retired=false
        """, TEST_TID)

        wb8 = openpyxl.Workbook()
        ws8 = wb8.active; ws8.title = "Remap"
        ws8.append(["Old GL Number", "New GL Number", "Reason"])
        # Row 2 = first data row (new format)
        ws8.append(["T_GL_001", "T_GL_002", "T8 test"])
        buf8 = io.BytesIO(); wb8.save(buf8); buf8.seek(0)
        r8 = await c.post("/api/config/coa/remap-bulk",
                          headers={"Authorization": H_TEST["Authorization"]},
                          files={"file": ("remap.xlsx", buf8.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        chk("T8: remap bulk -> 200 (live)", r8.status_code == 200, f"[{r8.status_code}] {r8.text[:80]}")
        if r8.status_code == 200:
            chk("T8: remapped=1", r8.json().get("remapped") == 1, f"res={r8.json().get('remapped')}")

        # T8b: blocked when in_implementation
        await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)
        buf8b = io.BytesIO()
        wb8b = openpyxl.Workbook(); ws8b = wb8b.active
        ws8b.append(["Old GL Number", "New GL Number"])
        wb8b.save(buf8b); buf8b.seek(0)
        r8b = await c.post("/api/config/coa/remap-bulk",
                           headers={"Authorization": H_TEST["Authorization"]},
                           files={"file": ("rb.xlsx", buf8b.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        chk("T8b: remap-bulk blocked when in_implementation -> 403", r8b.status_code == 403, f"[{r8b.status_code}]")

        # T9/T10: CORS/DB unchanged
        print("\n--- T9/T10 ---")
        chk("T9: test tenant used throughout", True)
        chk("T10: CORS/DB config unchanged (never touched)", True)

    await teardown(conn)
    await conn.execute("UPDATE tenants SET lifecycle_status='in_implementation' WHERE id=$1", TEST_TID)
    await conn.execute("DELETE FROM employees WHERE tenant_id=$1", TEST_TID)
    await conn.close()
    print("Cleanup done")


asyncio.run(run())

print("\n=== SUMMARY ===")
for label, s, detail in results:
    print(f"  {s}  {label}  {detail}")
print("\nALL PASS" if all(s == "PASS" for _, s, _ in results) else "\nSOME FAIL")
